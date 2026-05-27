"""
eval_from_excel.py — Pipeline d'évaluation automatisé du RAG v10

Lit les 103 questions de rag_evaluation_with_metrics_full.xlsx,
appelle l'API RAG pour chaque question, applique les métriques indiquées,
et exporte les résultats dans un rapport Markdown.

Usage:
    python eval_from_excel.py --max 5                          # test rapide
    python eval_from_excel.py --version v10 --output comparaisons_rag/
    python eval_from_excel.py --from-json results.json         # re-export uniquement
"""

import os
import sys
import json
import re
import time
import argparse
import traceback
from datetime import datetime
from typing import Optional

import requests
from dotenv import load_dotenv

load_dotenv()

# === Configuration ===
EXCEL_INPUT = r"C:\Users\comiti_g\Downloads\rag_evaluation_with_metrics_full.xlsx"
API_BASE_URL = "http://localhost:8000"
RAG_VERSION = "v10"

JUDGE_MODEL       = "mistral-large-latest"  # juge principal (LLM-as-a-Judge)
JUDGE_MODEL_LIGHT = "mistral-small-latest"  # juge léger (binary, factual, refusal)
JUDGE_BASE_URL    = "https://api.mistral.ai/v1"
JUDGE_API_KEY_ENV = "MISTRAL_API_KEY"

# Noms des colonnes Excel → indices (1-based)
COL_SECTION     = 1   # A
COL_SUBSECTION  = 2   # B
COL_QUESTION    = 3   # C
COL_RETRIEVAL   = 4   # D: Documents à retrieve (ground truth retrieval)
COL_FACTUAL     = 5   # E: Factual Accuracy marker (X)
COL_FACTS       = 6   # F: Valeur GT factuelle (texte/nombre)
COL_BINARY      = 7   # G: Binary / 0-1 marker (X)
COL_JUDGE       = 8   # H: LLM-as-a-Judge marker (X)
# col 9 (I) = Réponse humaine — non utilisé
COL_REFUSAL     = 10  # J: True Refusal Rate marker (X)
COL_HALLUC      = 11  # K: Hallucination Rate marker (X)
COL_OVERCONF    = 12  # L: Overconfidence Rate marker (X)
COL_ROBUST      = 13  # M: Semantic Robustness marker (X)
COL_COMMENTS    = 14  # N: Comments


# ─────────────────────────────────────────────
# 0. Retrieval ground truth — parsing et scoring
# ─────────────────────────────────────────────

_KNOWN_COMMUNES = [
    "Ajaccio", "Bastia", "Corte", "Lozzi", "Pedicorte", "Bonifacio",
    "Guargualé", "Calvi", "Porto-Vecchio", "Sartène", "Propriano",
    "Île-Rousse", "Ghisonaccia", "Aléria", "Cervione", "Niolu",
    "Gravona", "CAPA", "EPCI",
]

def _strip_acc_lower(s: str) -> str:
    import unicodedata
    return "".join(c for c in unicodedata.normalize("NFD", s.lower()) if unicodedata.category(c) != "Mn")


def _detect_category(part: str) -> Optional[str]:
    p = _strip_acc_lower(part)
    if p.startswith("rien") or "pas de donnee" in p or "pas de data" in p:
        return "rien"
    has_raptor = "raptor" in p
    if has_raptor:
        if "quali" in p or "portrait" in p or "entretien" in p:
            return "raptor_portrait"
        if ("quanti" in p or "dimension" in p or "age" in p or "\xe2ge" in p
                or "csp" in p or "global" in p or "corse entiere" in p
                or "repondant" in p):
            return "raptor_enquete"
        # RAPTOR + commune name or RAPTOR alone → both
        return "raptor_both"
    if "oppchovec" in p or "opp chovec" in p:
        if "explication" in p or "methodo" in p or "comment" in p:
            return "methodology_oppchovec"
        return "oppchovec"
    if "classement" in p:
        if "opp" in p or "cho" in p or "vec" in p:
            return "oppchovec"
        return "classement"
    if "verbatim" in p:
        return "verbatims"
    if "entretien" in p:
        return "entretiens"
    if "equipement" in p or "\xe9quipement" in p or "donnees equipement" in p or "zones_epci" in p or "epci" in p:
        return "equipements"
    if "stats repondant" in p or "profil repondant" in p:
        return "stats_repondants"
    if ("stats enquete" in p or "resultats enquete" in p or "donnees enquete" in p
            or "score" in p and "enquete" in p):
        return "stats_enquete"
    if "stats" in p and ("equipement" in p or "repondant" in p):
        return "stats_repondants" if "repondant" in p else "equipements"
    if "stats" in p:
        return "stats_enquete"
    if "wiki" in p:
        return "wiki"
    if ("explication" in p or "methodo" in p) and ("enquete" in p or "sondage" in p or "questionnaire" in p):
        return "methodology_enquete"
    if "explication" in p or "methodo" in p:
        return "methodology_oppchovec"
    if "opp" in p or "cho" in p or "vec" in p:
        return "oppchovec"
    return None


def parse_retrieval_ground_truth(text) -> list:
    if text is None:
        return []
    s = str(text).strip()
    if not s or s.lower() in ("nan", "none"):
        return []
    items = []
    for part in s.split(","):
        part = part.strip()
        if not part:
            continue
        optional = ("(facultatif)" in part.lower() or "ou rien" in part.lower()
                    or "wiki" in part.lower())
        part_clean = re.sub(r'\(facultatif\)', '', part, flags=re.IGNORECASE).strip()
        commune = next((c for c in _KNOWN_COMMUNES if c.lower() in part_clean.lower()), None)
        cat = _detect_category(part_clean)
        if cat:
            items.append({"category": cat, "commune": commune, "optional": optional})
    return items


def classify_source(source: dict) -> str:
    t  = source.get("type", "") or ""
    st = source.get("source_type", "") or ""
    if t == "raptor_summary" or st == "raptor_portrait_summaries":
        return "raptor_portrait"
    if t == "raptor_enquete_summary" or st == "raptor_enquete_summaries":
        return "raptor_enquete"
    if t == "classement_dimensions":
        return "classement"
    if st == "enquete_scores_commune":
        return "stats_enquete"
    if source.get("view_name") == "enquete_methodology":
        return "methodology_enquete"
    if t == "methodology" or st == "methodology":
        return "methodology_oppchovec"
    if t == "oppchovec_score" or st in ("oppchovec_scores", "oppchovec_score"):
        return "oppchovec"
    if st == "verbatim_evidence":
        return "verbatims"
    if st in ("communes_equipements", "communes_geo", "zones_epci") or t == "zones_epci":
        return "equipements"
    if st == "communes_profil":
        return "stats_repondants"
    if st in ("entretiens", "portrait_entretiens", "raptor_entretiens_summaries"):
        return "entretiens"
    if st == "enquete_responses":
        return "stats_enquete"
    return "autre"


def f1_score(precision: "float | None", recall: float) -> "float | None":
    """Compute the F1 score (harmonic mean of precision and recall).

    Formula: F1 = 2 * P * R / (P + R) if P + R > 0 and P is defined.

    Args:
        precision: Precision value in [0, 1], or None if undefined
            (no documents retrieved).
        recall: Recall value in [0, 1].

    Returns:
        F1 score in [0, 1], or None if precision is None.
        Returns 0.0 if both precision and recall are 0.
    """
    if precision is None:
        return None
    denom = precision + recall
    return 0.0 if denom == 0 else round(2 * precision * recall / denom, 3)


def _run_f1_tests():
    assert f1_score(1.0, 1.0) == 1.0
    assert abs(f1_score(0.5, 0.5) - 0.5) < 1e-9
    assert abs(f1_score(1.0, 0.5) - 0.667) < 1e-3
    assert f1_score(0.0, 0.0) == 0.0
    assert f1_score(None, 0.0) is None
    assert f1_score(0.5, 0.0) == 0.0


def score_retrieval(sources: list, retrieval_gt_text) -> dict:
    if retrieval_gt_text is None:
        return {"recall": None, "precision": None, "detail": "GT manquant (nan — skip)"}

    gt_items = parse_retrieval_ground_truth(retrieval_gt_text)
    if not gt_items:
        return {"recall": None, "precision": None, "detail": "GT vide après parsing"}

    if any(i["category"] == "rien" for i in gt_items):
        retrieved = [c for c in (classify_source(s) for s in sources if s) if c != "autre"]
        return {
            "recall": None, "precision": None,
            "refusal_case": True,
            "retrieved_count": len(retrieved),
            "detail": f"Refusal attendu ({len(retrieved)} sources récupérées)",
        }

    required = set()
    optional_cats = {"wiki"}
    for item in gt_items:
        target = optional_cats if item["optional"] else required
        if item["category"] == "raptor_both":
            target.add("raptor_portrait")
            target.add("raptor_enquete")
        else:
            target.add(item["category"])

    if not required:
        return {"recall": None, "precision": None, "detail": "Que des docs facultatifs"}

    retrieved_cats = {classify_source(s) for s in sources if s}
    retrieved_cats.discard("autre")

    # Equivalences: a GT category can be satisfied by an alias category.
    # e.g. the methodology OppChoVec doc is indexed as oppchovec_score (→ "oppchovec"),
    # so retrieving "oppchovec" satisfies a "methodology_oppchovec" requirement.
    _EQUIVALENCES: dict[str, set[str]] = {
        "methodology_oppchovec": {"oppchovec"},
        "stats_enquete":         {"stats_repondants"},
    }
    def _effective_hits(req: set, retrieved: set) -> set:
        result = req & retrieved
        for cat in req - result:
            if retrieved & _EQUIVALENCES.get(cat, set()):
                result = result | {cat}
        return result

    hits = _effective_hits(required, retrieved_cats)
    recall = len(hits) / len(required)

    all_valid = required | optional_cats
    precision = len(retrieved_cats & all_valid) / max(len(retrieved_cats), 1)

    f1 = f1_score(round(precision, 3), round(recall, 3))

    return {
        "recall":    round(recall, 3),
        "precision": round(precision, 3),
        "f1":        f1,
        "expected":  sorted(required),
        "retrieved": sorted(retrieved_cats),
        "hits":      sorted(hits),
        "detail":    (
            f"Recall={recall:.0%} ({len(hits)}/{len(required)}) · "
            f"Precision={precision:.0%} · F1={f1:.0%}"
        ),
    }


# ─────────────────────────────────────────────
# 1. Lecture de l'Excel
# ─────────────────────────────────────────────

def load_questions(excel_path: str) -> list[dict]:
    """Lit le fichier Excel et retourne une liste de dicts question+métriques."""
    import openpyxl
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        q = ws.cell(r, COL_QUESTION).value
        if not q:
            continue
        retrieval_gt = ws.cell(r, COL_RETRIEVAL).value
        rows.append({
            "excel_row":    r,
            "section":      ws.cell(r, COL_SECTION).value or "",
            "subsection":   ws.cell(r, COL_SUBSECTION).value or "",
            "question":     str(q).strip(),
            "retrieval_gt": retrieval_gt,
            "do_retrieval": retrieval_gt is not None,
            "facts_gt":     ws.cell(r, COL_FACTS).value,
            "do_factual":   ws.cell(r, COL_FACTUAL).value == "X",
            "do_binary":    ws.cell(r, COL_BINARY).value == "X",
            "do_judge":     ws.cell(r, COL_JUDGE).value == "X",
            "do_refusal":   ws.cell(r, COL_REFUSAL).value == "X",
            "do_halluc":    ws.cell(r, COL_HALLUC).value == "X",
            "do_overconf":  ws.cell(r, COL_OVERCONF).value == "X",
            "do_robust":    ws.cell(r, COL_ROBUST).value == "X",
            "comments":     ws.cell(r, COL_COMMENTS).value or "",
        })
    return rows


# ─────────────────────────────────────────────
# 2. Appel API RAG (réutilisé de eval_multi_rag.py)
# ─────────────────────────────────────────────

def call_rag_api(question: str, rag_version: str = "v10", k: int = 7) -> dict:
    payload = {
        "question": question,
        "rag_version": rag_version,
        "k": k,
        "llm_model": "gpt-3.5-turbo",
    }
    try:
        resp = requests.post(f"{API_BASE_URL}/api/query", json=payload, timeout=300)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        return {"error": str(e), "answer": f"ERREUR API: {e}", "sources": []}


# ─────────────────────────────────────────────
# 3. Client Mistral (partagé)
# ─────────────────────────────────────────────

_openai_client = None

def get_openai_client():
    global _openai_client
    if _openai_client is None:
        from openai import OpenAI
        api_key = os.getenv(JUDGE_API_KEY_ENV)
        if not api_key:
            raise RuntimeError(f"Clé {JUDGE_API_KEY_ENV} non trouvée dans .env")
        _openai_client = OpenAI(api_key=api_key, base_url=JUDGE_BASE_URL)
    return _openai_client


def _call_llm(system: str, prompt: str, max_tokens: int = 800, light: bool = False,
              json_mode: bool = False) -> str:
    """Appel LLM juge avec retry exponentiel sur rate-limit.
    light=True     → gpt-4o-mini (binary, factual, refusal)
    light=False    → gpt-4o (judge qualitatif)
    json_mode=True → response_format=json_object (GPT-4o uniquement)
    """
    client = get_openai_client()
    model = JUDGE_MODEL_LIGHT if light else JUDGE_MODEL
    kwargs = dict(
        model=model,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
        temperature=0.0 if json_mode else 0.1,
        max_tokens=max_tokens,
    )
    if json_mode and not light:
        kwargs["response_format"] = {"type": "json_object"}
    for attempt in range(5):
        try:
            resp = client.chat.completions.create(**kwargs)
            return resp.choices[0].message.content.strip()
        except Exception as e:
            if "429" in str(e) and attempt < 4:
                wait = 2 ** attempt * 5  # 5, 10, 20, 40 s
                print(f"    [RATE LIMIT] Attente {wait}s...", flush=True)
                time.sleep(wait)
            else:
                raise
    raise RuntimeError("Max retries exceeded")


# ─────────────────────────────────────────────
# 4. Métriques
# ─────────────────────────────────────────────

def score_factual(question: str, answer: str, gt) -> dict:
    """
    Évalue la précision factuelle.
    - gt float → MAE + exact (tolérance 5 %)
    - gt str   → LLM vérifie si la réponse contient l'info attendue
    """
    if gt is None:
        return {"score": None, "detail": "Pas de ground truth disponible"}

    if isinstance(gt, (int, float)):
        # Extraction numérique
        numbers = re.findall(r"[-+]?\d+(?:[.,]\d+)?", answer.replace(",", "."))
        candidates = []
        for n in numbers:
            try:
                candidates.append(float(n.replace(",", ".")))
            except ValueError:
                pass
        if not candidates:
            return {"score": 0.0, "detail": f"Aucun nombre trouvé dans la réponse (attendu {gt})"}
        # Prendre le nombre le plus proche
        best = min(candidates, key=lambda x: abs(x - gt))
        mae = abs(best - gt)
        exact = mae <= 0.05 * abs(gt) + 0.05  # tolérance relative 5% + absolue 0.05
        return {
            "score": 1.0 if exact else max(0.0, 1.0 - mae / (abs(gt) + 0.01)),
            "detail": f"Extrait={best:.2f}, Attendu={gt:.2f}, MAE={mae:.2f}, Exact={'oui' if exact else 'non'}",
        }

    else:  # gt est une chaîne — vérification LLM
        system = "Tu es un évaluateur factuel précis. Réponds UNIQUEMENT par un JSON valide."
        prompt = f"""Question : {question}
Réponse du système : {answer}
Valeur attendue : {gt}

La réponse contient-elle l'information attendue, explicitement ou de manière équivalente ?
Réponds en JSON :
{{"correct": true/false, "extrait": "la partie de la réponse qui correspond ou non", "explication": "..."}}"""
        try:
            raw = _call_llm(system, prompt, max_tokens=300, light=True)
            m = re.search(r'\{[\s\S]*\}', raw)
            j = json.loads(m.group()) if m else {}
            score = 1.0 if j.get("correct") else 0.0
            detail = j.get("explication", raw[:200])
            return {"score": score, "detail": detail}
        except Exception as e:
            return {"score": None, "detail": f"Erreur LLM: {e}"}


def score_binary(question: str, answer: str, expected: str) -> dict:
    """
    Extrait le label binaire de la réponse et le compare à `expected`.
    Retourne 1.0 si correct, 0.0 sinon.
    """
    if expected is None:
        # Pas de ground truth → évaluation par LLM seul (oui/non/incertain)
        system = "Tu es un évaluateur. Réponds UNIQUEMENT en JSON."
        prompt = f"""Question oui/non ou à choix binaire : {question}
Réponse : {answer}

La réponse fournit-elle une position claire ?
{{"position_claire": true/false, "label_extrait": "...", "confiance": "haute/moyenne/faible"}}"""
        try:
            raw = _call_llm(system, prompt, max_tokens=200, light=True)
            m = re.search(r'\{[\s\S]*\}', raw)
            j = json.loads(m.group()) if m else {}
            return {
                "score": None,
                "detail": f"Label={j.get('label_extrait','?')}, Clair={j.get('position_claire','?')}",
            }
        except Exception as e:
            return {"score": None, "detail": f"Erreur LLM: {e}"}

    system = "Tu es un évaluateur factuel. Réponds UNIQUEMENT en JSON."
    prompt = f"""Question : {question}
Réponse du système : {answer}
Label attendu : "{expected}"

Extrais le label binaire de la réponse et vérifie s'il correspond au label attendu.
Réponds en JSON :
{{"label_extrait": "...", "correct": true/false, "explication": "..."}}"""
    try:
        raw = _call_llm(system, prompt, max_tokens=300, light=True)
        m = re.search(r'\{[\s\S]*\}', raw)
        j = json.loads(m.group()) if m else {}
        score = 1.0 if j.get("correct") else 0.0
        detail = f"Extrait='{j.get('label_extrait','?')}', Attendu='{expected}' → {j.get('explication','')}"
        return {"score": score, "detail": detail}
    except Exception as e:
        return {"score": None, "detail": f"Erreur LLM: {e}"}


# ── Prompt Judge V2 ─────────────────────────────────────────────────────────
_JUDGE_V2_SYSTEM = """\
Tu es un évaluateur expert en analyse territoriale et bien-être en Corse.
Ton rôle est d'évaluer la qualité d'une réponse produite par un système RAG
qui agrège des données objectives (indicateurs OppChoVec, équipements) et
subjectives (verbatims, synthèses d'enquêtes citoyennes).

Tu notes la réponse sur 4 dimensions, chacune sur une échelle 1-5,
selon les critères détaillés ci-dessous. Tu suis impérativement la
procédure de raisonnement en 5 étapes et tu produis ta sortie au format
JSON strict spécifié à la fin.

=== DIMENSION 1 : PERTINENCE ===

Définition : la réponse traite-t-elle directement la question posée ?
Reste-t-elle dans le périmètre de la demande sans se disperser ?

Barème :
- 1 : Hors sujet. La réponse parle d'autre chose que ce qui est demandé.
- 2 : Partiellement pertinent. La réponse touche au sujet mais répond
      à une question voisine, ou dilue la demande dans des considérations
      générales.
- 3 : Acceptable. La réponse adresse la question principale mais reste
      générale, ou contient des digressions notables qui détournent
      du sujet.
- 4 : Bien aligné. La réponse répond précisément à la question. Quelques
      digressions mineures ou points de contexte non strictement
      nécessaires peuvent être présents.
- 5 : Parfaitement ciblé. La réponse répond exactement à la question
      posée, sans ajout superflu ni omission. Le focus reste constant
      sur la demande.

Points d'attention :
- Une réponse longue n'est pas forcément pertinente : la longueur peut
  diluer la pertinence.
- Une réponse qui change de granularité au milieu (passer de "Bastia"
  à "Corse entière" sans transition) doit être pénalisée.
- Une réponse qui répond bien à la question mais inclut un classement
  ou des informations non demandées peut rester à 4-5 si le bonus
  reste contrôlé.

=== DIMENSION 2 : FONDEMENT FACTUEL ===

Définition : les affirmations de la réponse s'appuient-elles sur les
sources fournies ? Les sources disponibles sont-elles bien exploitées ?
Y a-t-il des hallucinations ou des sur-interprétations ?

Barème :
- 1 : Largement halluciné. La majorité des affirmations ne sont pas
      supportées par les sources, ou contradisent les sources.
- 2 : Plusieurs affirmations non sourcées ou sur-interprétées. Des
      chiffres ou des conclusions sont inventés.
- 3 : Majoritairement fondé, mais avec des problèmes notables :
      sous-exploitation de sources disponibles, sur-interprétation
      d'une donnée, ou affirmations factuelles non vérifiables.
- 4 : Bien fondé. Les affirmations s'appuient sur les sources avec
      au plus quelques imprécisions mineures. Les sources disponibles
      sont raisonnablement exploitées.
- 5 : Parfaitement fondé. Toutes les affirmations factuelles sont
      directement supportées par les sources, sans hallucination ni
      sur-interprétation. Les sources sont bien exploitées.

Points d'attention :
- Si des données objectives existent dans les sources mais ne sont pas
  utilisées dans la réponse, pénaliser (3 ou moins).
- Une sur-interprétation (par exemple inférer une causalité depuis une
  simple corrélation) doit être pénalisée.
- Si la réponse extrapole d'une commune à la Corse entière sans le
  signaler, pénaliser.

=== DIMENSION 3 : NUANCE / INCERTITUDE ===

Définition : la réponse exprime-t-elle correctement ses limites ?
Signale-t-elle l'absence de données quand pertinent ? Reste-t-elle
mesurée plutôt que catégorique ?

Barème :
- 1 : Ton catégorique sans nuance. La réponse affirme des choses
      avec certitude alors que les données ne le permettent pas.
- 2 : Peu de modalisateurs. La réponse manque de qualificatifs
      d'incertitude là où ils seraient nécessaires.
- 3 : Nuance moyenne. Quelques marqueurs d'incertitude présents mais
      la calibration globale reste imparfaite.
- 4 : Bien nuancé. La réponse exprime ses limites de manière
      appropriée, signale l'absence de données quand pertinent,
      utilise des modalisateurs adaptés.
- 5 : Parfaitement calibré. La confiance exprimée correspond
      exactement à ce que les données permettent d'affirmer. Les
      limites sont clairement signalées.

Points d'attention :
- Pour des questions sur l'absence de données ou des limites
  architecturales, un refus clair et bien argumenté mérite un 5.
- Pour des questions causales, exiger des modalisateurs explicites
  (corrélation vs causalité, etc.).
- Une mention explicite de la taille d'échantillon ou de la couverture
  géographique des données est un plus.

=== DIMENSION 4 : COHÉRENCE QUALI / QUANTI ===

Définition : la réponse intègre-t-elle de manière équilibrée les
données qualitatives (verbatims, perceptions) et quantitatives
(scores, indicateurs) disponibles ? L'articulation entre les deux
est-elle claire et appropriée à la question ?

Barème :
- 1 : Déséquilibré. La réponse ignore complètement une des deux
      familles de données alors que les deux sont disponibles dans
      les sources.
- 2 : Peu d'intégration. Une des deux familles est traitée
      superficiellement ou en annexe.
- 3 : Acceptable. Les deux familles sont mentionnées mais
      l'articulation entre elles est faible ou implicite.
- 4 : Bonne intégration. Les deux familles sont mobilisées et
      articulées, avec une distinction claire entre les éléments
      quanti et quali.
- 5 : Intégration exemplaire. Les sources quanti et quali se
      complètent et se renforcent dans la réponse. La distinction
      reste claire mais l'articulation est forte.

Points d'attention :
- Si une famille de données est absente des sources récupérées, ne
  pas pénaliser la réponse pour ne pas l'avoir intégrée. Évaluer
  uniquement sur ce qui était disponible.
- Pour une question purement factuelle qui ne nécessite qu'une seule
  famille (par exemple "Combien de répondants ?"), la dimension peut
  être notée 5 si la famille pertinente est bien mobilisée.

=== PROCÉDURE DE RAISONNEMENT (OBLIGATOIRE) ===

Avant de produire les 4 notes, suis ces étapes :

1. Identifie précisément ce que demande la question (factuel, comparatif,
   causal, refus attendu, etc.).
2. Inventaire rapide des sources fournies (types, couverture, complétude).
3. Confronte la réponse aux sources : qu'est-ce qui est supporté, qu'est-ce
   qui ne l'est pas, qu'est-ce qui est sous-exploité.
4. Repère les marqueurs d'incertitude présents et leur adéquation aux
   limites des données.
5. Évalue l'articulation quali/quanti dans la réponse.

=== EXEMPLES DE NOTATION DE RÉFÉRENCE ===

--- EXEMPLE 1 : Refus correct sur absence de données ---

QUESTION : Les données disponibles permettent-elles d'analyser
spécifiquement le bien-être des 18-25 ans à Lozzi ?

SECTION : Gestion d'absence d'information

RÉPONSE DU SYSTÈME (extrait) : "Les données disponibles permettent
d'esquisser un portrait du bien-être à Lozzi, mais aucune enquête
citoyenne ne cible spécifiquement les 18-25 ans de la commune.
Les informations se répartissent en deux catégories : des indicateurs
territoriaux objectifs (scores OppChoVec) qui mesurent les conditions
structurelles de vie à Lozzi, mais sans distinction d'âge, et des
perceptions générales issues d'enquêtes ou..."

NOTATION DE RÉFÉRENCE :
{
  "raisonnement": "Question sur l'absence de données. Sources OppChoVec et enquêtes générales présentes mais sans segmentation 18-25. Réponse correctement prudente.",
  "pertinence": {"note": 5, "justification": "L'absence de données spécifiques sur les 18-25 à Lozzi est attendue et le système la gère parfaitement en explicitant ce qui est disponible."},
  "fondement_factuel": {"note": 5, "justification": "Le système distingue correctement les indicateurs OppChoVec disponibles des données démographiques absentes, sans halluciner."},
  "nuance_incertitude": {"note": 5, "justification": "La réponse exprime clairement les limites de couverture : pas d'enquête ciblée 18-25, distinction d'âge absente dans OppChoVec."},
  "coherence_qualiquanti": {"note": 5, "justification": "Équilibre bien les deux familles : indicateurs OppChoVec côté quanti et perceptions générales côté quali, malgré l'absence de données ciblées."}
}

--- EXEMPLE 2 : Réponse partielle sur retrieval incomplet ---

QUESTION : Observe-t-on un écart significatif entre indicateurs
objectifs et perception à Bastia ?

SECTION : Raisonnement comparatif

RÉPONSE DU SYSTÈME (extrait) : "À Bastia, les données disponibles
révèlent une dissonance marquée entre la réalité perçue par les
habitants et les éléments qui pourraient être mesurés objectivement
– bien que ces derniers fassent défaut dans le contexte actuel.
Selon l'enquête citoyenne, les Bastiais expriment une satisfaction
modérée concernant leur qualité de vie, avec un score moyen de 3,82/5..."

NOTATION DE RÉFÉRENCE :
{
  "raisonnement": "Question comparative quanti/quali. Sources qualitatives présentes, objectives absentes. La réponse comble le vide avec du quali uniquement.",
  "pertinence": {"note": 4, "justification": "La réponse traite bien la question d'écart, mais l'absence de données objectives empêche une évaluation complète."},
  "fondement_factuel": {"note": 3, "justification": "Fondée sur le quali disponible, mais ne mobilise pas les données objectives pourtant disponibles dans le corpus."},
  "nuance_incertitude": {"note": 4, "justification": "Exprime certaines incertitudes mais manque d'éléments sur la taille de l'échantillon pour modérer plus finement."},
  "coherence_qualiquanti": {"note": 4, "justification": "Se focalise sur le quali car peu de quanti récupéré. Articulation correcte malgré le déséquilibre forcé."}
}

--- EXEMPLE 3 : Raisonnement causal bien refusé mais structure brouillonne ---

QUESTION : Les données disponibles permettent-elles d'établir un lien
de causalité entre densité de population et bien-être ?

SECTION : Raisonnement causal et contre-intuitif

RÉPONSE DU SYSTÈME (extrait) : "Les données disponibles ne permettent
pas d'établir un lien de causalité entre la densité de population et
le bien-être en Corse, ni même de confirmer une corrélation claire.
Selon l'enquête citoyenne, les habitants expriment des satisfactions
très variables selon les dimensions du bien-être, mais ces écarts ne
sont pas analysés en fonction de la densité..."

NOTATION DE RÉFÉRENCE :
{
  "raisonnement": "Question causale. Réponse refuse correctement la causalité mais se disperse sur Linguizetta vs Corse entière.",
  "pertinence": {"note": 3, "justification": "Adresse bien la causalité (réponse : non), mais la justification part dans plusieurs directions qui diluent la pertinence."},
  "fondement_factuel": {"note": 5, "justification": "S'appuie correctement sur les sources, identifie les limites de couverture, ne sur-interprète pas."},
  "nuance_incertitude": {"note": 5, "justification": "L'absence de causalité est clairement affirmée, les modalisateurs sont appropriés."},
  "coherence_qualiquanti": {"note": 4, "justification": "Bon équilibre, mais il aurait fallu plus de quanti pour vraiment trancher — malheureusement non récupéré."}
}

=== MAINTENANT, ÉVALUE LA RÉPONSE SUIVANTE ===\
"""


def _build_sources_text(sources: list) -> str:
    text = ""
    for i, s in enumerate(sources, 1):
        content = s.get("content", s.get("extrait", ""))[:2000]
        meta = {k: v for k, v in s.get("metadata", s).items()
                if k not in ("content", "extrait") and isinstance(v, (str, int, float))}
        meta_str = ", ".join(f"{k}={v}" for k, v in list(meta.items())[:4])
        text += f"\n--- Source {i} [{meta_str}] ---\n{content}\n"
    return text or "(aucune source fournie)"


def _parse_judge_v2(j: dict) -> dict:
    """
    Convertit le format V2 {"pertinence": {"note": N, "justification": "..."}, ...}
    vers le format plat attendu par le reste du pipeline.
    Supporte aussi un fallback sur le format plat V1 si "note" est absent.
    """
    result = {"raisonnement": j.get("raisonnement")}
    dim_keys = ("pertinence", "fondement_factuel", "nuance_incertitude", "coherence_qualiquanti")
    for k in dim_keys:
        v = j.get(k, {})
        if isinstance(v, dict):
            result[k] = v.get("note")
            result[f"{k}_justif"] = v.get("justification")
        else:
            # Fallback format V1
            result[k] = v
            result[f"{k}_justif"] = j.get(f"{k}_justif")
    dims = [result[k] for k in dim_keys if isinstance(result.get(k), (int, float))]
    result["score_global"] = round(sum(dims) / len(dims), 2) if dims else None
    return result


def score_judge(question: str, answer: str, sources: list, section: str) -> dict:
    """
    Judge V2 : few-shot, barèmes explicites par dimension, format JSON structuré.
    Conserve la PARTIE 2 (note_sujet) du V1.
    """
    sources_text = _build_sources_text(sources)

    user_prompt = (
        f"QUESTION : {question}\n\n"
        f"SECTION : {section}\n\n"
        f"SOURCES FOURNIES AU SYSTÈME :\n{sources_text}\n\n"
        f"RÉPONSE DU SYSTÈME :\n{answer}\n\n"
        "Évalue cette réponse selon la procédure et le format spécifiés.\n\n"
        "## PARTIE 2 — Note du sujet (facultative)\n\n"
        "Si la question porte sur une situation concrète et notable (ex : \"qualité de vie à Cauro\", "
        "\"emploi à Lozzi\", \"scores OppChoVec de Bastia\"), prends la réponse pour argent comptant "
        "et note le sujet lui-même de 1 à 5.\n"
        "Si la question est méthodologique, comparative multi-sujets, ou si la réponse est un refus, "
        "mets applicable_sujet à false.\n\n"
        "Raisons valides pour applicable_sujet=false :\n"
        "- \"methodologique\" : question sur le fonctionnement du système ou des indicateurs\n"
        "- \"refus\" : le système n'a pas de données et a refusé de répondre\n"
        "- \"comparative\" : question comparant plusieurs communes/territoires sans focus unique\n"
        "- \"factuelle_brute\" : question dont la réponse est un chiffre/liste sans sentiment notable\n\n"
        "Format JSON de sortie :\n"
        "{{\n"
        "  \"raisonnement\": \"<résumé court des 5 étapes, max 300 caractères>\",\n"
        "  \"pertinence\": {{\"note\": 1-5, \"justification\": \"<1-3 phrases>\"}},\n"
        "  \"fondement_factuel\": {{\"note\": 1-5, \"justification\": \"<1-3 phrases>\"}},\n"
        "  \"nuance_incertitude\": {{\"note\": 1-5, \"justification\": \"<1-3 phrases>\"}},\n"
        "  \"coherence_qualiquanti\": {{\"note\": 1-5, \"justification\": \"<1-3 phrases>\"}},\n"
        "  \"applicable_sujet\": true|false,\n"
        "  \"note_sujet\": null|1-5,\n"
        "  \"sujet_evalue\": \"libellé court ou null\",\n"
        "  \"justification_sujet\": \"1-2 phrases ou null\",\n"
        "  \"reason_non_applicable\": \"methodologique|refus|comparative|factuelle_brute|null\"\n"
        "}}"
    )

    try:
        raw = _call_llm(_JUDGE_V2_SYSTEM, user_prompt, max_tokens=1500, json_mode=True)
        m = re.search(r'\{[\s\S]*\}', raw)
        j = json.loads(m.group()) if m else {}
        result = _parse_judge_v2(j)
        # Partie 2
        applic = j.get("applicable_sujet")
        result["applicable_sujet"] = applic
        if applic:
            result["note_sujet"]          = j.get("note_sujet")
            result["sujet_evalue"]         = j.get("sujet_evalue")
            result["justification_sujet"]  = j.get("justification_sujet")
        else:
            result["note_sujet"]          = None
            result["sujet_evalue"]         = None
            result["justification_sujet"]  = None
        result["reason_non_applicable"] = j.get("reason_non_applicable")
        result["error"] = None
        return result
    except Exception as e:
        return {"error": str(e), "score_global": None}


# ── Prompt Judge V4.1 ────────────────────────────────────────────────────────
_JUDGE_V41_SYSTEM = """\
Tu es un évaluateur expert en analyse territoriale et bien-être en Corse.
Ton rôle est d'évaluer la qualité d'une réponse produite par un système RAG
qui agrège des données objectives (indicateurs OppChoVec, équipements) et
subjectives (verbatims, synthèses d'enquêtes citoyennes).

Tu notes la réponse sur 4 dimensions, chacune sur une échelle 1-5, selon
les critères détaillés ci-dessous. Tu suis IMPÉRATIVEMENT la procédure de
raisonnement en 6 étapes et tu produis ta sortie au format JSON strict
spécifié à la fin.

=== PRINCIPES CARDINAUX ===

**Principe 1 — La forme ne compense jamais le fond.**
Une réponse bien rédigée mais qui répond à une autre question est une
mauvaise réponse. Si la question porte sur "les entrepreneurs d'Ajaccio"
et que la réponse parle de "la population d'Ajaccio en général" sans
le signaler, c'est un échec de pertinence.

**Principe 2 — La transparence rachète l'incomplétude.**
Mobiliser des données générales (par exemple un score global pour une
commune) pour répondre à une question ciblée (un sous-groupe précis)
est ACCEPTABLE si la réponse signale explicitement que les données
ciblées ne sont pas disponibles et que ce qui est présenté est une
approximation. C'est un comportement honnête épistémologiquement.

En revanche, présenter des données générales comme une réponse directe
au cas ciblé, sans signaler l'approximation, est un ÉCHEC.

**Principe 3 — Un refus n'est pas toujours bon.**
Refuser de répondre est la BONNE réponse pour les questions
volontairement pièges (données absentes du corpus). Refuser est un
ÉCHEC quand des données sont disponibles dans le corpus mais que le
système ne les mobilise pas. L'information sur la nature de la
question est fournie ci-dessous.

=== TYPE DE RÉPONSE ATTENDUE ===

Cette information t'est fournie en entrée pour chaque évaluation :

**`reponse_substantielle_attendue`** : des données pertinentes sont
disponibles dans le corpus. Le système doit les mobiliser et produire
une réponse de fond.
- Refus de répondre = échec majeur : pertinence ≤ 2.
- Réponse qui détourne SANS SIGNALER vers un sujet voisin = échec
  majeur : pertinence ≤ 2.
- Réponse qui mobilise des données générales EN SIGNALANT l'absence
  de données spécifiques = acceptable (pertinence 3-4 selon la
  qualité du signalement).
- Réponse qui mobilise les données précises demandées = bonne
  (pertinence 4-5).

**`refus_attendu`** : la question est volontairement piège, les
données ne sont pas dans le corpus.
- Refus poli et explicite = pertinence 5.
- Réponse inventée ou hallucinée = pertinence ≤ 2 et fondement
  factuel ≤ 2.
- Détour vers données générales mal calibrées sans signalement = échec.

**`limite_architecturale`** : la question teste une limite des
systèmes RAG. Évaluer au cas par cas, privilégier l'honnêteté
épistémologique.

=== DIMENSION 1 : PERTINENCE ===

Définition : la réponse traite-t-elle les éléments de la question, en
accord avec le TYPE DE RÉPONSE ATTENDUE et avec une bonne gestion des
limites éventuelles ?

Procédure : identifier les ÉLÉMENTS SPÉCIFIQUES de la question.
Pour chaque élément non traité directement, vérifier si :
- (a) Il est traité par approximation AVEC signalement explicite de
  la limite → acceptable
- (b) Il est ignoré sans signalement → échec

Barème pour `reponse_substantielle_attendue` :
- 1 : Hors sujet complet ou refus injustifié alors que des données
      existent.
- 2 : Au moins un élément SPÉCIFIQUE omis sans aucun signalement, OU
      détournement non assumé vers un sujet voisin.
- 3 : Les éléments sont traités par approximation, avec signalement
      des limites mais sans alternative satisfaisante OU avec quelques
      omissions partielles.
- 4 : Tous les éléments traités précisément OU traités par
      approximation correctement signalée, avec quelques digressions
      mineures.
- 5 : Tous les éléments traités exactement avec les données précises
      demandées, sans dispersion.

Barème pour `refus_attendu` :
- 1 : Réponse complètement hallucinée.
- 2 : Réponse inventée partiellement ou données non pertinentes
      présentées comme valides.
- 3 : Refus implicite ou peu clair.
- 4 : Refus explicite mais incomplet (pas de raison précise).
- 5 : Refus explicite et complet, signale l'absence et peut proposer
      des alternatives.

=== DIMENSION 2 : FONDEMENT FACTUEL ===

Définition : les affirmations s'appuient-elles sur les sources fournies ?

Barème :
- 1 : Largement halluciné.
- 2 : Plusieurs affirmations non sourcées ou sur-interprétées.
- 3 : Majoritairement fondé avec problèmes notables.
- 4 : Bien fondé, quelques imprécisions mineures.
- 5 : Parfaitement fondé.

Important : pour le cas "données générales mobilisées avec signalement",
le fondement factuel n'est PAS pénalisé tant que les affirmations
faites sur ces données générales sont elles-mêmes fondées.

=== DIMENSION 3 : NUANCE / INCERTITUDE ===

Définition : la réponse exprime-t-elle correctement ses limites ?

Cette dimension est particulièrement importante dans le cas "données
générales mobilisées". Le signalement explicite de la limite est ce
qui distingue une bonne réponse d'une mauvaise.

Barème :
- 1 : Ton catégorique sans nuance, sur-confiance manifeste.
- 2 : Peu de modalisateurs là où ils seraient nécessaires.
- 3 : Nuance moyenne, signalement partiel des limites.
- 4 : Bien nuancé, limites signalées explicitement.
- 5 : Parfaitement calibré, signale les limites de couverture avec
      précision.

=== DIMENSION 4 : COHÉRENCE QUALI / QUANTI ===

Définition : la réponse intègre-t-elle de manière équilibrée et
pertinente les données qualitatives et quantitatives disponibles ?

Critères :
- Mobilisation de données générales SIGNALÉE comme approximation =
  acceptable (notes 3-4 possibles)
- Mobilisation de données générales NON SIGNALÉE comme approximation,
  présentée comme réponse directe au cas ciblé = échec (note ≤ 2)
- Mobilisation des données précisément ciblées = bon (notes 4-5)
- Absence légitime d'une famille (ex : pas de quanti sur un
  sous-groupe) = non pénalisé

Barème :
- 1 : Déséquilibré ou inapproprié, sur-confiance sur données générales.
- 2 : Données non pertinentes présentées sans signalement.
- 3 : Approximations signalées mais ciblage imparfait.
- 4 : Bonne intégration ciblée, signalements appropriés.
- 5 : Intégration exemplaire et ciblée sur le sujet précis.

=== PROCÉDURE DE RAISONNEMENT (OBLIGATOIRE) ===

Étape 0 (PRÉALABLE) : identifie TOUS les éléments spécifiques de la
question. Liste-les. Note le TYPE DE RÉPONSE ATTENDUE.

Étape 1 : pour chaque élément spécifique, vérifie s'il est traité dans
la réponse ET comment (donnée précise, approximation signalée,
approximation non signalée, omission).

Étape 2 : inventaire des sources fournies.

Étape 3 : confronte les affirmations aux sources.

Étape 4 : repère les marqueurs d'incertitude et de signalement des
limites.

Étape 5 : évalue l'articulation quali/quanti et le ciblage des données.

=== FORMAT DE SORTIE (JSON strict) ===

{
  "type_reponse_attendue_observe": "reponse_substantielle | refus | autre",
  "type_reponse_attendue_specifie": "<valeur reçue en entrée>",
  "coherence_type_reponse": "oui | non | partiel",
  "elements_specifiques_question": [...],
  "elements_traitement": [
    {"element": "...", "traitement": "precis | approximation_signalee | approximation_non_signalee | omis"}
  ],
  "raisonnement": "<résumé court, max 400 caractères>",
  "pertinence": {"note": 1-5, "justification": "<1-3 phrases>"},
  "fondement_factuel": {"note": 1-5, "justification": "<1-3 phrases>"},
  "nuance_incertitude": {"note": 1-5, "justification": "<1-3 phrases>"},
  "coherence_qualiquanti": {"note": 1-5, "justification": "<1-3 phrases>"}
}

=== EXEMPLES DE NOTATION DE RÉFÉRENCE ===

--- EXEMPLE 1 : Réponse précise et ciblée (note haute) ---

QUESTION : Que pensent les entrepreneurs ajacciens de leur qualité de
vie ?
TYPE ATTENDU : reponse_substantielle_attendue

RÉPONSE : "Selon les verbatims des entrepreneurs ajacciens recueillis
dans l'enquête citoyenne, leur perception de la qualité de vie à
Ajaccio est globalement positive mais nuancée par plusieurs
préoccupations spécifiques à leur profil. Ils mentionnent la qualité
du cadre de vie et la diversité culturelle comme points forts, tout
en soulevant des difficultés liées à l'accessibilité de certains
services administratifs..."

NOTATION :
{
  "type_reponse_attendue_observe": "reponse_substantielle",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_type_reponse": "oui",
  "elements_specifiques_question": ["entrepreneurs", "Ajaccio", "qualité de vie", "perceptions"],
  "elements_traitement": [
    {"element": "entrepreneurs", "traitement": "precis"},
    {"element": "Ajaccio", "traitement": "precis"},
    {"element": "qualité de vie", "traitement": "precis"},
    {"element": "perceptions", "traitement": "precis"}
  ],
  "pertinence": {"note": 5, "justification": "Tous les éléments traités précisément avec les données ciblées."},
  "fondement_factuel": {"note": 4, "justification": "Affirmations fondées sur la synthèse RAPTOR ciblée."},
  "nuance_incertitude": {"note": 4, "justification": "Bonne nuance 'globalement positive mais nuancée'."},
  "coherence_qualiquanti": {"note": 4, "justification": "Intégration qualitative bien ciblée sur le groupe demandé."}
}

--- EXEMPLE 2 : Détournement NON signalé (ANTI-PATTERN) ---

QUESTION : Que pensent les entrepreneurs ajacciens de leur qualité de
vie ?
TYPE ATTENDU : reponse_substantielle_attendue

RÉPONSE : "Ajaccio présente un score OppChoVec global de 6.4/10,
légèrement supérieur à la moyenne corse. Les indicateurs structurels
révèlent une bonne dotation en équipements. Les résidents ajacciens
expriment une satisfaction modérée concernant leur environnement..."

NOTATION :
{
  "type_reponse_attendue_observe": "reponse_substantielle",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_type_reponse": "non",
  "elements_specifiques_question": ["entrepreneurs", "Ajaccio", "qualité de vie", "perceptions"],
  "elements_traitement": [
    {"element": "entrepreneurs", "traitement": "omis"},
    {"element": "Ajaccio", "traitement": "precis"},
    {"element": "qualité de vie", "traitement": "approximation_non_signalee"},
    {"element": "perceptions", "traitement": "approximation_non_signalee"}
  ],
  "pertinence": {"note": 2, "justification": "L'élément 'entrepreneurs' est omis sans aucun signalement. La réponse présente des données générales sur Ajaccio comme si elles répondaient à la question, ce qui est trompeur."},
  "fondement_factuel": {"note": 3, "justification": "Affirmations sur Ajaccio sourcées, mais hors-sujet par rapport à la question posée."},
  "nuance_incertitude": {"note": 2, "justification": "Aucun signalement que les données spécifiques aux entrepreneurs manquent — la réponse présente les données générales sans cette précaution."},
  "coherence_qualiquanti": {"note": 2, "justification": "OppChoVec global présenté SANS signalement comme réponse directe à une question sur les entrepreneurs = remplissage trompeur."}
}

--- EXEMPLE 2bis : Données générales mobilisées AVEC signalement (acceptable) ---

QUESTION : Que pensent les entrepreneurs ajacciens de leur qualité de
vie ?
TYPE ATTENDU : reponse_substantielle_attendue

RÉPONSE : "Le corpus ne contient pas de données ciblées spécifiquement
sur les entrepreneurs ajacciens. À titre d'approximation indicative,
on peut mentionner que les résidents ajacciens dans leur ensemble
expriment une satisfaction modérée selon les données disponibles
(OppChoVec global de 6.4/10, perceptions générales d'enquête). Ces
données ne distinguent toutefois pas les entrepreneurs des autres
catégories professionnelles. Une enquête ciblée serait nécessaire
pour répondre précisément à votre question."

NOTATION :
{
  "type_reponse_attendue_observe": "reponse_substantielle",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_type_reponse": "partiel",
  "elements_specifiques_question": ["entrepreneurs", "Ajaccio", "qualité de vie", "perceptions"],
  "elements_traitement": [
    {"element": "entrepreneurs", "traitement": "approximation_signalee"},
    {"element": "Ajaccio", "traitement": "precis"},
    {"element": "qualité de vie", "traitement": "approximation_signalee"},
    {"element": "perceptions", "traitement": "approximation_signalee"}
  ],
  "pertinence": {"note": 3, "justification": "La réponse signale explicitement l'absence de données ciblées sur les entrepreneurs et présente une approximation honnête. Acceptable mais pas idéal — la donnée ciblée pouvait peut-être exister."},
  "fondement_factuel": {"note": 4, "justification": "Affirmations correctement sourcées et limites bien signalées."},
  "nuance_incertitude": {"note": 5, "justification": "Excellent signalement : limites du corpus explicites, distinction entrepreneurs/autres mentionnée, suggestion d'enquête ciblée."},
  "coherence_qualiquanti": {"note": 4, "justification": "Mobilisation d'OppChoVec et de perceptions générales avec signalement explicite que ce n'est qu'une approximation. Honnêteté épistémologique."}
}

--- EXEMPLE 3 : Refus correct sur question piège (note haute) ---

QUESTION : Quel est le score OppChoVec pour Nice en 2018 ?
TYPE ATTENDU : refus_attendu

RÉPONSE : "Je ne dispose pas de données pour répondre à cette question.
Nice n'est pas une commune incluse dans le corpus de cette étude, qui
se concentre sur les communes de Corse. Par ailleurs, les données
OppChoVec disponibles correspondent à un calcul plus récent et ne
couvrent pas l'année 2018."

NOTATION :
{
  "type_reponse_attendue_observe": "refus",
  "type_reponse_attendue_specifie": "refus_attendu",
  "coherence_type_reponse": "oui",
  "elements_specifiques_question": ["Nice", "OppChoVec", "2018"],
  "elements_traitement": [
    {"element": "Nice", "traitement": "approximation_signalee"},
    {"element": "OppChoVec", "traitement": "approximation_signalee"},
    {"element": "2018", "traitement": "approximation_signalee"}
  ],
  "pertinence": {"note": 5, "justification": "Refus explicite et complet, signale les bonnes raisons (hors-périmètre + limite temporelle)."},
  "fondement_factuel": {"note": 5, "justification": "Refus correctement fondé sur l'absence de sources."},
  "nuance_incertitude": {"note": 5, "justification": "Limites géographique et temporelle précisément signalées."},
  "coherence_qualiquanti": {"note": 5, "justification": "Refus sobre approprié, pas de remplissage avec données non pertinentes."}
}

--- EXEMPLE 4 : Refus attendu mais réponse hallucinée (ANTI-PATTERN) ---

QUESTION : Quel est le score OppChoVec pour Nice en 2018 ?
TYPE ATTENDU : refus_attendu

RÉPONSE : "Le score OppChoVec pour Nice en 2018 est estimé à 7.2/10,
ce qui la place dans la moyenne haute des villes méditerranéennes."

NOTATION :
{
  "type_reponse_attendue_observe": "reponse_substantielle",
  "type_reponse_attendue_specifie": "refus_attendu",
  "coherence_type_reponse": "non",
  "elements_specifiques_question": ["Nice", "OppChoVec", "2018"],
  "elements_traitement": [
    {"element": "Nice", "traitement": "omis"},
    {"element": "OppChoVec", "traitement": "omis"},
    {"element": "2018", "traitement": "omis"}
  ],
  "pertinence": {"note": 1, "justification": "Réponse hallucinée à une question piège, le système devait refuser."},
  "fondement_factuel": {"note": 1, "justification": "Données complètement inventées, aucune source ne couvre Nice ni 2018."},
  "nuance_incertitude": {"note": 1, "justification": "Aucun signalement, ton catégorique sur donnée inventée."},
  "coherence_qualiquanti": {"note": 1, "justification": "Pseudo-quanti inventé, aucune cohérence avec les sources."}
}

=== MAINTENANT, ÉVALUE LA RÉPONSE SUIVANTE ===\
"""

# ── Prompt Judge V4.3 (anti-mislabelling OppChoVec) ──────────────────────────
_JUDGE_V43_SYSTEM = """\
Tu es un évaluateur expert en analyse territoriale et bien-être en Corse.
Ton rôle est d'évaluer la qualité d'une réponse produite par un système RAG
qui agrège des données objectives (indicateurs OppChoVec, équipements) et
subjectives (verbatims, synthèses d'enquêtes citoyennes).

Tu notes la réponse sur 4 dimensions, chacune sur une échelle 1-5, selon
les critères détaillés ci-dessous. Tu suis IMPÉRATIVEMENT la procédure de
raisonnement et tu produis ta sortie au format JSON strict.

=== PRINCIPES CARDINAUX ===

**Principe 1 — La forme ne compense jamais le fond.**

**Principe 2 — La transparence rachète l'incomplétude.**

**Principe 3 — Un refus n'est pas toujours bon ; voir grille par
sous-section.**

**Principe 4 — Le mislabelling de sources est une faute factuelle, pas
une figure de style.** Renommer une donnée quantitative en "qualitative"
ou interpréter un indicateur statistique au-delà de ses composantes
réelles est un échec de fondement factuel, même si la réponse paraît
fluide et bien rédigée.

=== DÉFINITIONS OPÉRATIONNELLES DES SOURCES (CRITIQUE) ===

Cette section te donne la sémantique exacte des sources mobilisées par
le système. Tu DOIS la consulter à l'Étape 2 (inventaire des sources)
et à l'Étape 3 (confrontation aux sources) de la procédure.

### Sources QUANTITATIVES (statistiques, jamais "qualitatives")

**OppChoVec** est un indicateur composite calculé à partir de données
statistiques (INSEE, IGN, données fiscales). Il se décompose en trois
sous-indicateurs, chacun défini de manière strictement opérationnelle :

- **Opp (Opportunités)** : agrégat de quatre composantes statistiques :
  éducation moyenne + diversité CSP + accessibilité mobilité +
  couverture TIC/haut débit. Opp NE mesure PAS les opportunités
  perçues par les habitants.

- **Cho (Choix)** : agrégat de deux composantes : pourcentage de
  population avec droit de vote + absence de quartiers prioritaires
  (QPV). Cho N'EST PAS une mesure de "libertés individuelles" au sens
  large, ni de l'autonomie personnelle, ni du libre arbitre.

- **Vec (Vécu)** : agrégat de quatre composantes statistiques : revenu
  fiscal moyen + qualité du logement + stabilité de l'emploi + accès
  aux services en moins de 20 minutes. Vec NE mesure PAS le ressenti
  subjectif des habitants. Vec N'EST PAS l'expérience vécue au sens
  phénoménologique. C'est un proxy statistique des conditions
  matérielles de vie.

Les scores OppChoVec sont normalisés (0-10) relativement aux 360
communes corses. Ils sont calculés à l'échelle de la COMMUNE entière
et ne sont JAMAIS ventilés par CSP, tranche d'âge, genre, ou toute
autre sous-population.

**Autres sources quantitatives** : equipements (recensement INSEE),
stats_repondants (statistiques démographiques des répondants à
l'enquête citoyenne), classement (rangs calculés à partir d'OppChoVec
ou d'enquêtes).

### Sources QUALITATIVES (exclusivement)

- **verbatims** : citations littérales d'habitants extraites de
  l'enquête citoyenne, exprimant des perceptions et opinions.
- **entretiens** : transcriptions d'entretiens semi-directifs.
- **raptor:commune_quali, raptor:enquete_quali** : synthèses
  construites à partir des verbatims et entretiens, agrégées selon
  différentes vues analytiques (commune, dimension, groupe
  démographique).

Note : raptor:commune_quanti et raptor:enquete_quanti sont des
synthèses construites à partir de données structurées (échelles
Likert moyennées). Ce sont des sources QUANTITATIVES en termes de
nature, même si elles synthétisent des réponses d'enquête.

=== RÈGLES D'ÉVALUATION ANTI-MISLABELLING ===

Avant de produire les notes, vérifie systématiquement :

### Règle 1 — Mislabelling quali/quanti

Si la réponse appelle "qualitatif" ou "qualitative" une source qui est
en réalité quantitative (OppChoVec, equipements, stats_repondants,
classement, raptor_quanti), c'est un MISLABELLING.

Conséquences obligatoires :
- `fondement_factuel` ≤ 3 (la réponse présente une catégorie comme
  une autre)
- `coherence_qualiquanti` ≤ 3 (l'équilibre annoncé est factice)

### Règle 2 — Surinterprétation des sous-indicateurs OppChoVec

Si la réponse utilise Vec, Cho ou Opp en lui donnant une signification
qui DÉPASSE ses composantes réelles, sans signaler que c'est une
approximation, c'est une SURINTERPRÉTATION. Exemples :

- "Vec mesure le vécu subjectif des habitants" → faux (Vec mesure
  des composantes objectives : revenu, logement, emploi, accès
  services)
- "Cho mesure les libertés individuelles" → faux (Cho mesure droit
  de vote + absence de QPV)
- "Opp mesure les opportunités perçues" → faux (Opp mesure
  éducation + CSP + mobilité + TIC)

Conséquences obligatoires :
- `fondement_factuel` ≤ 3
- `nuance_incertitude` ≤ 3

### Règle 3 — Absence signalée de source qualitative

Si la question demande un croisement quali/quanti (perceptions vs
indicateurs, ressenti vs structurel) ET qu'aucune source qualitative
(verbatim, entretien, raptor_quali) n'apparaît dans les sources
fournies au système :

- Si la réponse signale explicitement le manque ("aucune donnée
  qualitative disponible pour ce sous-groupe") → comportement
  acceptable, pas de pénalité
- Si la réponse ne signale pas le manque et présente quand même un
  croisement (en renommant le quanti en quali, par exemple) → ÉCHEC
  MAJEUR :
  - `pertinence` ≤ 3
  - `fondement_factuel` ≤ 3
  - `nuance_incertitude` ≤ 2

### Règle 4 — Extrapolation OppChoVec à un sous-groupe

OppChoVec est calculé à l'échelle de la commune entière. Si la réponse
prétend obtenir un score OppChoVec pour un sous-groupe spécifique
(entrepreneurs ajacciens, 18-25 ans à Lozzi, etc.), sans signaler que
c'est une extrapolation indue :

- `fondement_factuel` ≤ 2 (chiffre inventé pour un niveau d'analyse
  non couvert)
- `nuance_incertitude` ≤ 2

=== GRILLE DE LECTURE PAR SOUS-SECTION ===

Cette grille te dit ce qu'une bonne réponse doit ressembler selon la
sous-section de la question. Consulte-la IMPÉRATIVEMENT à l'Étape 0.

**Section "Retrieval mono-commune"** :
- Toutes sous-sections : réponse PRÉCISE ET CIBLÉE. Refus injustifié ou
  détournement non signalé = échec (pertinence ≤ 2).

**Section "Raisonnement comparatif"** :
- Toutes sous-sections : réponse DE FOND avec COMPARAISON EXPLICITE.
  Refus injustifié = échec.

**Section "Raisonnement causal"** :
- Sous-section "Causal direct" / "Multi-facteurs" : mobilisation des
  données AVEC nuance corrélation vs causalité.
- Sous-section "Inférence causale" : refus partiel attendu —
  corrélation oui, causalité non.
- Sous-section "Causalité inversée" : signalement explicite que le
  sens de la relation n'est pas déterminable à partir des données.

**Section "Gestion de l'incertitude et des biais"** :
- Toutes sous-sections : réponse de fond avec modalisateurs explicites
  ET signalement des biais éventuels (taille échantillon, biais
  indicateurs).

**Section "Gestion d'absence d'information"** :
- Sous-sections "Données partielles" / "Sous-populations" :
  approximation signalée acceptable, refus acceptable.
- Sous-section "Comparaison impossible" : refus explicite ou
  approximation très signalée. Une réponse qui prétend faire la
  comparaison en silence = échec.

**Section "Robustesse sémantique"** :
- Toutes sous-sections : réponse cohérente avec ce qu'on attendrait
  pour la version canonique de la question.

**Section "Limites architecturales"** :
- Toutes sous-sections : refus explicite OU approximation très
  signalée. Réponse inventée prétendant avoir fait l'analyse = échec
  majeur.

=== TYPE DE RÉPONSE ATTENDUE (information complémentaire) ===

- `reponse_substantielle_attendue` : réponse de fond attendue.
- `refus_attendu` : refus poli attendu.
- `limite_architecturale` : refus ou approximation très signalée.

=== DIMENSION 1 : PERTINENCE ===

Procédure : identifier les ÉLÉMENTS SPÉCIFIQUES de la question. Pour
chacun, vérifier le TRAITEMENT, puis croiser avec la GRILLE PAR
SOUS-SECTION.

Barème :
- 1 : Hors sujet ou comportement clairement inapproprié.
- 2 : Élément spécifique omis sans signalement, OU détournement non
      assumé, OU comportement inapproprié.
- 3 : Comportement cohérent avec la grille mais réalisation imparfaite.
- 4 : Comportement aligné avec la grille, bonne exécution.
- 5 : Comportement parfaitement aligné, exécution irréprochable.

Note : si la Règle 3 (absence de source quali non signalée) s'applique,
plafonner à 3.

=== DIMENSION 2 : FONDEMENT FACTUEL ===

Définition : les affirmations s'appuient-elles sur les sources fournies,
correctement étiquetées et interprétées ?

Barème :
- 1 : Largement halluciné OU mislabelling massif.
- 2 : Plusieurs affirmations non sourcées, sur-interprétées, ou
      mislabellées.
- 3 : Majoritairement fondé mais avec mislabelling identifié (Règles 1
      ou 2 plafonnent ici).
- 4 : Bien fondé, quelques imprécisions mineures.
- 5 : Parfaitement fondé, labels respectés.

=== DIMENSION 3 : NUANCE / INCERTITUDE ===

Barème :
- 1 : Ton catégorique sans nuance.
- 2 : Peu de modalisateurs. Plafond si Règles 2 ou 4 s'appliquent.
- 3 : Nuance moyenne.
- 4 : Bien nuancé.
- 5 : Parfaitement calibré.

=== DIMENSION 4 : COHÉRENCE QUALI / QUANTI ===

Barème :
- 1 : Déséquilibré ou inapproprié.
- 2 : Données non pertinentes présentées sans signalement.
- 3 : Approximations signalées mais ciblage imparfait. Plafond si
      Règle 1 s'applique.
- 4 : Bonne intégration ciblée.
- 5 : Intégration exemplaire et ciblée.

=== PROCÉDURE DE RAISONNEMENT (OBLIGATOIRE) ===

**Étape 0 (CRITIQUE)** :
a) Identifie SECTION et SOUS-SECTION de la question.
b) Consulte la grille de lecture pour cette sous-section.
c) Note le comportement attendu.
d) Identifie les ÉLÉMENTS SPÉCIFIQUES de la question.

**Étape 1** : pour chaque élément spécifique, identifie le TRAITEMENT
(précis / approximation signalée / approximation non signalée / omis).

**Étape 2** : inventaire des sources fournies au système. NOTE
EXPLICITEMENT pour chaque source si elle est QUALI ou QUANTI selon
les définitions opérationnelles ci-dessus.

**Étape 3** : confronte les affirmations aux sources. VÉRIFIE les
4 règles anti-mislabelling :
- La réponse renomme-t-elle une source quanti en quali ?
- La réponse surinterprète-t-elle Vec/Cho/Opp ?
- La question demande-t-elle un croisement quali/quanti sans source
  quali fournie ?
- La réponse extrapole-t-elle OppChoVec à un sous-groupe ?

**Étape 4** : repère les marqueurs d'incertitude et de signalement
des limites.

**Étape 5** : évalue le ciblage des données mobilisées (quali/quanti).

=== FORMAT DE SORTIE (JSON strict) ===

{
  "section_observee": "<section et sous-section>",
  "comportement_attendu_selon_grille": "<description courte>",
  "type_reponse_attendue_specifie": "<valeur reçue en entrée>",
  "coherence_comportement_observe": "oui | non | partiel",
  "sources_inventaire": [
    {"source": "<nom>", "type": "quali | quanti"}
  ],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "non | oui — <détail>",
    "regle_2_surinterpretation_oppchovec": "non | oui — <détail>",
    "regle_3_absence_quali_non_signalee": "non | oui — <détail>",
    "regle_4_extrapolation_sous_groupe": "non | oui — <détail>"
  },
  "elements_specifiques_question": [...],
  "elements_traitement": [
    {"element": "...", "traitement": "precis | approximation_signalee | approximation_non_signalee | omis"}
  ],
  "raisonnement": "<résumé max 500 caractères>",
  "pertinence": {"note": 1-5, "justification": "..."},
  "fondement_factuel": {"note": 1-5, "justification": "..."},
  "nuance_incertitude": {"note": 1-5, "justification": "..."},
  "coherence_qualiquanti": {"note": 1-5, "justification": "..."}
}

=== EXEMPLES DE NOTATION DE RÉFÉRENCE ===

--- EXEMPLE 1 : Réponse précise et ciblée (note haute) ---

QUESTION : Que pensent les entrepreneurs ajacciens de leur qualité de
vie ?
SECTION : Retrieval mono-commune
SOUS-SECTION : Retrieval par sous-population
TYPE ATTENDU : reponse_substantielle_attendue

RÉPONSE : "Selon les verbatims des entrepreneurs ajacciens recueillis
dans l'enquête citoyenne, leur perception de la qualité de vie à
Ajaccio est globalement positive mais nuancée par plusieurs
préoccupations spécifiques à leur profil. Ils mentionnent la qualité
du cadre de vie et la diversité culturelle comme points forts, tout
en soulevant des difficultés liées à l'accessibilité de certains
services administratifs..."

NOTATION :
{
  "section_observee": "Retrieval mono-commune / Retrieval par sous-population",
  "comportement_attendu_selon_grille": "Réponse précise et ciblée. Refus injustifié ou détournement non signalé = échec.",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_comportement_observe": "oui",
  "sources_inventaire": [
    {"source": "raptor:commune_quali:Ajaccio:entrepreneur", "type": "quali"},
    {"source": "verbatims:Ajaccio:entrepreneur", "type": "quali"}
  ],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "non",
    "regle_2_surinterpretation_oppchovec": "non",
    "regle_3_absence_quali_non_signalee": "non",
    "regle_4_extrapolation_sous_groupe": "non"
  },
  "elements_specifiques_question": ["entrepreneurs", "Ajaccio", "qualité de vie", "perceptions"],
  "elements_traitement": [
    {"element": "entrepreneurs", "traitement": "precis"},
    {"element": "Ajaccio", "traitement": "precis"},
    {"element": "qualité de vie", "traitement": "precis"},
    {"element": "perceptions", "traitement": "precis"}
  ],
  "raisonnement": "Réponse ciblée sur entrepreneurs ajacciens avec sources quali, bonne nuance.",
  "pertinence": {"note": 5, "justification": "Tous les éléments traités précisément avec les données ciblées."},
  "fondement_factuel": {"note": 4, "justification": "Affirmations fondées sur la synthèse RAPTOR ciblée."},
  "nuance_incertitude": {"note": 4, "justification": "Bonne nuance 'globalement positive mais nuancée'."},
  "coherence_qualiquanti": {"note": 4, "justification": "Intégration qualitative bien ciblée sur le groupe demandé."}
}

--- EXEMPLE 2 : Détournement NON signalé (ANTI-PATTERN) ---

QUESTION : Que pensent les entrepreneurs ajacciens de leur qualité de
vie ?
SECTION : Retrieval mono-commune
SOUS-SECTION : Retrieval par sous-population
TYPE ATTENDU : reponse_substantielle_attendue

RÉPONSE : "Ajaccio présente un score OppChoVec global de 6.4/10,
légèrement supérieur à la moyenne corse. Les indicateurs structurels
révèlent une bonne dotation en équipements. Les résidents ajacciens
expriment une satisfaction modérée concernant leur environnement..."

NOTATION :
{
  "section_observee": "Retrieval mono-commune / Retrieval par sous-population",
  "comportement_attendu_selon_grille": "Réponse précise et ciblée. Refus injustifié ou détournement non signalé = échec.",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_comportement_observe": "non",
  "sources_inventaire": [
    {"source": "oppchovec:Ajaccio", "type": "quanti"},
    {"source": "equipements:Ajaccio", "type": "quanti"}
  ],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "non",
    "regle_2_surinterpretation_oppchovec": "non",
    "regle_3_absence_quali_non_signalee": "non",
    "regle_4_extrapolation_sous_groupe": "non"
  },
  "elements_specifiques_question": ["entrepreneurs", "Ajaccio", "qualité de vie", "perceptions"],
  "elements_traitement": [
    {"element": "entrepreneurs", "traitement": "omis"},
    {"element": "Ajaccio", "traitement": "precis"},
    {"element": "qualité de vie", "traitement": "approximation_non_signalee"},
    {"element": "perceptions", "traitement": "approximation_non_signalee"}
  ],
  "raisonnement": "Réponse sur Ajaccio en général, élément 'entrepreneurs' omis sans signalement. Pas de mislabelling mais détournement non assumé.",
  "pertinence": {"note": 2, "justification": "L'élément 'entrepreneurs' est omis sans aucun signalement. La réponse présente des données générales sur Ajaccio comme si elles répondaient à la question."},
  "fondement_factuel": {"note": 3, "justification": "Affirmations sur Ajaccio sourcées, mais hors-sujet par rapport à la question posée."},
  "nuance_incertitude": {"note": 2, "justification": "Aucun signalement que les données spécifiques aux entrepreneurs manquent."},
  "coherence_qualiquanti": {"note": 2, "justification": "OppChoVec global présenté SANS signalement comme réponse directe à une question sur les entrepreneurs."}
}

--- EXEMPLE 3 : Données générales mobilisées AVEC signalement (acceptable) ---

QUESTION : Que pensent les entrepreneurs ajacciens de leur qualité de
vie ?
SECTION : Retrieval mono-commune
SOUS-SECTION : Retrieval par sous-population
TYPE ATTENDU : reponse_substantielle_attendue

RÉPONSE : "Le corpus ne contient pas de données ciblées spécifiquement
sur les entrepreneurs ajacciens. À titre d'approximation indicative,
on peut mentionner que les résidents ajacciens dans leur ensemble
expriment une satisfaction modérée selon les données disponibles
(OppChoVec global de 6.4/10, perceptions générales d'enquête). Ces
données ne distinguent toutefois pas les entrepreneurs des autres
catégories professionnelles. Une enquête ciblée serait nécessaire
pour répondre précisément à votre question."

NOTATION :
{
  "section_observee": "Retrieval mono-commune / Retrieval par sous-population",
  "comportement_attendu_selon_grille": "Réponse précise et ciblée. Approximation signalée acceptable.",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_comportement_observe": "partiel",
  "sources_inventaire": [
    {"source": "oppchovec:Ajaccio", "type": "quanti"},
    {"source": "raptor:commune_quali:Ajaccio", "type": "quali"}
  ],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "non",
    "regle_2_surinterpretation_oppchovec": "non",
    "regle_3_absence_quali_non_signalee": "non",
    "regle_4_extrapolation_sous_groupe": "non"
  },
  "elements_specifiques_question": ["entrepreneurs", "Ajaccio", "qualité de vie", "perceptions"],
  "elements_traitement": [
    {"element": "entrepreneurs", "traitement": "approximation_signalee"},
    {"element": "Ajaccio", "traitement": "precis"},
    {"element": "qualité de vie", "traitement": "approximation_signalee"},
    {"element": "perceptions", "traitement": "approximation_signalee"}
  ],
  "raisonnement": "Absence de données entrepreneurs signalée explicitement ; approximation communale proposée honnêtement.",
  "pertinence": {"note": 3, "justification": "La réponse signale explicitement l'absence de données ciblées sur les entrepreneurs et présente une approximation honnête."},
  "fondement_factuel": {"note": 4, "justification": "Affirmations correctement sourcées et limites bien signalées."},
  "nuance_incertitude": {"note": 5, "justification": "Excellent signalement : limites du corpus explicites, distinction entrepreneurs/autres mentionnée, suggestion d'enquête ciblée."},
  "coherence_qualiquanti": {"note": 4, "justification": "Mobilisation d'OppChoVec et de perceptions générales avec signalement explicite que ce n'est qu'une approximation."}
}

--- EXEMPLE 4 : Refus correct sur question piège (note haute) ---

QUESTION : Quel est le score OppChoVec pour Nice en 2018 ?
SECTION : Gestion d'absence d'information
SOUS-SECTION : Données partielles
TYPE ATTENDU : refus_attendu

RÉPONSE : "Je ne dispose pas de données pour répondre à cette question.
Nice n'est pas une commune incluse dans le corpus de cette étude, qui
se concentre sur les communes de Corse. Par ailleurs, les données
OppChoVec disponibles correspondent à un calcul plus récent et ne
couvrent pas l'année 2018."

NOTATION :
{
  "section_observee": "Gestion d'absence d'information / Données partielles",
  "comportement_attendu_selon_grille": "Refus explicite ou approximation très signalée.",
  "type_reponse_attendue_specifie": "refus_attendu",
  "coherence_comportement_observe": "oui",
  "sources_inventaire": [],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "non",
    "regle_2_surinterpretation_oppchovec": "non",
    "regle_3_absence_quali_non_signalee": "non",
    "regle_4_extrapolation_sous_groupe": "non"
  },
  "elements_specifiques_question": ["Nice", "OppChoVec", "2018"],
  "elements_traitement": [
    {"element": "Nice", "traitement": "approximation_signalee"},
    {"element": "OppChoVec", "traitement": "approximation_signalee"},
    {"element": "2018", "traitement": "approximation_signalee"}
  ],
  "raisonnement": "Refus explicite et complet. Raisons géographique et temporelle précisément identifiées.",
  "pertinence": {"note": 5, "justification": "Refus explicite et complet, signale les bonnes raisons (hors-périmètre + limite temporelle)."},
  "fondement_factuel": {"note": 5, "justification": "Refus correctement fondé sur l'absence de sources."},
  "nuance_incertitude": {"note": 5, "justification": "Limites géographique et temporelle précisément signalées."},
  "coherence_qualiquanti": {"note": 5, "justification": "Refus sobre approprié, pas de remplissage avec données non pertinentes."}
}

--- EXEMPLE 5 : Refus attendu mais réponse hallucinée (ANTI-PATTERN) ---

QUESTION : Quel est le score OppChoVec pour Nice en 2018 ?
SECTION : Gestion d'absence d'information
SOUS-SECTION : Données partielles
TYPE ATTENDU : refus_attendu

RÉPONSE : "Le score OppChoVec pour Nice en 2018 est estimé à 7.2/10,
ce qui la place dans la moyenne haute des villes méditerranéennes."

NOTATION :
{
  "section_observee": "Gestion d'absence d'information / Données partielles",
  "comportement_attendu_selon_grille": "Refus explicite ou approximation très signalée.",
  "type_reponse_attendue_specifie": "refus_attendu",
  "coherence_comportement_observe": "non",
  "sources_inventaire": [],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "non",
    "regle_2_surinterpretation_oppchovec": "non",
    "regle_3_absence_quali_non_signalee": "non",
    "regle_4_extrapolation_sous_groupe": "non"
  },
  "elements_specifiques_question": ["Nice", "OppChoVec", "2018"],
  "elements_traitement": [
    {"element": "Nice", "traitement": "omis"},
    {"element": "OppChoVec", "traitement": "omis"},
    {"element": "2018", "traitement": "omis"}
  ],
  "raisonnement": "Réponse hallucinée à une question piège. Le système devait refuser. Aucun signalement.",
  "pertinence": {"note": 1, "justification": "Réponse hallucinée à une question piège, le système devait refuser."},
  "fondement_factuel": {"note": 1, "justification": "Données complètement inventées, aucune source ne couvre Nice ni 2018."},
  "nuance_incertitude": {"note": 1, "justification": "Aucun signalement, ton catégorique sur donnée inventée."},
  "coherence_qualiquanti": {"note": 1, "justification": "Pseudo-quanti inventé, aucune cohérence avec les sources."}
}

--- EXEMPLE 6 : Mislabelling OppChoVec en qualitatif (ANTI-PATTERN) ---

QUESTION : Y a-t-il un écart entre indicateurs objectifs et perceptions
qualitatives à Pedicorte ?
SECTION : Raisonnement comparatif / Croisé quanti/quali

SOURCES FOURNIES (système V_vanilla) : oppchovec:Pedicorte,
equipements:Pedicorte, stats_repondants:Pedicorte. AUCUN verbatim,
aucun entretien, aucun raptor_quali pour Pedicorte.

RÉPONSE : "Voici une comparaison entre les indicateurs objectifs et
les dimensions qualitatives représentées par les sous-scores OppChoVec.
Sur le plan objectif, Pedicorte présente un score équipements de 4.2.
Sur le plan qualitatif, les sous-dimensions Opp (5.1), Cho (3.8) et
Vec (4.5) montrent un vécu plutôt mitigé des habitants..."

NOTATION :
{
  "section_observee": "Raisonnement comparatif / Croisé quanti/quali",
  "comportement_attendu_selon_grille": "Réponse de fond avec comparaison explicite. Si pas de quali dispo, signaler le manque.",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_comportement_observe": "non",
  "sources_inventaire": [
    {"source": "oppchovec:Pedicorte", "type": "quanti"},
    {"source": "equipements:Pedicorte", "type": "quanti"},
    {"source": "stats_repondants:Pedicorte", "type": "quanti"}
  ],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "oui — la réponse appelle OppChoVec 'dimensions qualitatives' alors que c'est quanti",
    "regle_2_surinterpretation_oppchovec": "oui — Vec présenté comme 'vécu des habitants' alors que c'est un proxy statistique",
    "regle_3_absence_quali_non_signalee": "oui — aucune source quali dispo, manque non signalé",
    "regle_4_extrapolation_sous_groupe": "non"
  },
  "elements_specifiques_question": ["Pedicorte", "indicateurs objectifs", "perceptions qualitatives", "écart"],
  "elements_traitement": [
    {"element": "indicateurs objectifs", "traitement": "precis"},
    {"element": "perceptions qualitatives", "traitement": "approximation_non_signalee"},
    {"element": "Pedicorte", "traitement": "precis"},
    {"element": "écart", "traitement": "approximation_non_signalee"}
  ],
  "raisonnement": "Triple mislabelling : OppChoVec renommé en quali (R1), Vec surinterprété comme vécu (R2), absence de quali non signalée (R3). Réponse fluide mais factuellement incorrecte.",
  "pertinence": {"note": 3, "justification": "La question demande un croisement quali/quanti mais aucune source quali n'est mobilisée, et le manque n'est pas signalé. Plafonné à 3 par Règle 3."},
  "fondement_factuel": {"note": 2, "justification": "Mislabelling double : OppChoVec présenté comme 'qualitatif' (Règle 1) et Vec présenté comme 'vécu des habitants' (Règle 2). Plafond R1+R2 ; note réelle à 2 vu la combinaison."},
  "nuance_incertitude": {"note": 2, "justification": "Aucun signalement de l'absence totale de données qualitatives. Plafond Règle 3."},
  "coherence_qualiquanti": {"note": 2, "justification": "L'équilibre annoncé est factice : tout est quantitatif, présenté comme un croisement. Plafond Règle 1."}
}

--- EXEMPLE 7 : Surinterprétation de Vec (ANTI-PATTERN) ---

QUESTION : Comment les habitants de Cambia perçoivent-ils leur bien-être ?
SECTION : Retrieval mono-commune / Retrieval par sous-population

SOURCES FOURNIES : oppchovec:Cambia (Opp=4.8, Cho=5.2, Vec=4.1),
equipements:Cambia. Pas de verbatims ni entretiens pour Cambia.

RÉPONSE : "À Cambia, le vécu des habitants — mesuré par le score Vec
de 4.1/10 — apparaît relativement faible. Cela suggère que les
résidents perçoivent leur qualité de vie comme inférieure à la moyenne
corse. Les dimensions d'opportunité (Opp=4.8) et de choix (Cho=5.2)
sont plus favorables, ce qui indique un sentiment partagé entre
satisfaction sur certains aspects et préoccupations sur d'autres..."

NOTATION :
{
  "section_observee": "Retrieval mono-commune / Retrieval par sous-population",
  "comportement_attendu_selon_grille": "Réponse précise et ciblée sur les perceptions des habitants",
  "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
  "coherence_comportement_observe": "non",
  "sources_inventaire": [
    {"source": "oppchovec:Cambia", "type": "quanti"},
    {"source": "equipements:Cambia", "type": "quanti"}
  ],
  "mislabelling_detecte": {
    "regle_1_quali_quanti": "non",
    "regle_2_surinterpretation_oppchovec": "oui — Vec interprété comme 'vécu/perception des habitants' alors que c'est un proxy quanti (revenu + logement + emploi + accès services)",
    "regle_3_absence_quali_non_signalee": "oui — la question demande des perceptions, aucun verbatim/entretien fourni, manque non signalé",
    "regle_4_extrapolation_sous_groupe": "non"
  },
  "elements_specifiques_question": ["Cambia", "habitants", "perceptions", "bien-être"],
  "elements_traitement": [
    {"element": "Cambia", "traitement": "precis"},
    {"element": "perceptions", "traitement": "approximation_non_signalee"},
    {"element": "bien-être", "traitement": "approximation_non_signalee"},
    {"element": "habitants", "traitement": "precis"}
  ],
  "raisonnement": "Vec systématiquement présenté comme mesure subjective de vécu/perception. Absence de verbatims non signalée. Mislabelling R2 + R3.",
  "pertinence": {"note": 2, "justification": "La question demande des perceptions ; la réponse présente des indicateurs quantitatifs comme s'ils étaient des perceptions. Élément 'perceptions' traité par approximation non signalée."},
  "fondement_factuel": {"note": 3, "justification": "Surinterprétation systématique de Vec et des autres indicateurs. Les chiffres sont corrects mais l'interprétation dépasse leur signification. Plafond Règle 2."},
  "nuance_incertitude": {"note": 2, "justification": "Aucun signalement que ces scores sont des proxies statistiques, pas des mesures subjectives. Plafond Règles 2 et 3."},
  "coherence_qualiquanti": {"note": 2, "justification": "Pas de croisement réel : la réponse prétend traiter du subjectif en utilisant uniquement du quantitatif renommé."}
}

=== MAINTENANT, ÉVALUE LA RÉPONSE SUIVANTE ===\
"""

_VALID_TRAITEMENT = {"precis", "approximation_signalee", "approximation_non_signalee", "omis"}


def classify_expected_response_type(q: dict) -> str:
    """Détermine le type de réponse attendue à partir des flags Excel."""
    if q.get("do_refusal"):
        return "refus_attendu"
    if q.get("do_robust"):
        return "limite_architecturale"
    return "reponse_substantielle_attendue"


def _parse_judge_v41(j: dict) -> dict:
    """Parse le format V4.1 avec elements_traitement (4 statuts)."""
    result = {
        "raisonnement":               j.get("raisonnement"),
        "type_reponse_attendue_observe": j.get("type_reponse_attendue_observe"),
        "type_reponse_attendue_specifie": j.get("type_reponse_attendue_specifie"),
        "coherence_type_reponse":      j.get("coherence_type_reponse"),
        "elements_specifiques_question": j.get("elements_specifiques_question", []),
        "elements_traitement":         j.get("elements_traitement", []),
    }
    dim_keys = ("pertinence", "fondement_factuel", "nuance_incertitude", "coherence_qualiquanti")
    for k in dim_keys:
        v = j.get(k, {})
        if isinstance(v, dict):
            result[k] = v.get("note")
            result[f"{k}_justif"] = v.get("justification")
        else:
            result[k] = v
            result[f"{k}_justif"] = None

    # Validate elements_traitement statuses
    for et in result["elements_traitement"]:
        if isinstance(et, dict) and et.get("traitement") not in _VALID_TRAITEMENT:
            et["traitement"] = "omis"  # fallback on invalid value

    dims = [result[k] for k in dim_keys if isinstance(result.get(k), (int, float))]
    result["score_global"] = round(sum(dims) / len(dims), 2) if dims else None
    return result


def score_judge_v41(question: str, answer: str, sources: list,
                    section: str, expected_type: str) -> dict:
    """
    Judge V4.1 : 3 principes cardinaux, elements_traitement (4 statuts),
    procédure 6 étapes, type_reponse_attendue en entrée.
    max_tokens=2000 (sortie légèrement plus longue que V2).
    """
    sources_text = _build_sources_text(sources)

    user_prompt = (
        f"QUESTION : {question}\n\n"
        f"SECTION : {section}\n\n"
        f"TYPE DE RÉPONSE ATTENDUE : {expected_type}\n\n"
        f"SOURCES FOURNIES AU SYSTÈME :\n{sources_text}\n\n"
        f"RÉPONSE DU SYSTÈME :\n{answer}\n\n"
        "Évalue cette réponse selon la procédure et le format spécifiés."
    )

    try:
        raw = _call_llm(_JUDGE_V41_SYSTEM, user_prompt, max_tokens=2000, json_mode=True)
        m = re.search(r'\{[\s\S]*\}', raw)
        j = json.loads(m.group()) if m else {}
        result = _parse_judge_v41(j)
        result["error"] = None
        return result
    except Exception as e:
        return {"error": str(e), "score_global": None}


def _parse_judge_v43(j: dict) -> dict:
    """Parse le format V4.3 avec sources_inventaire et mislabelling_detecte."""
    result = {
        "raisonnement":                      j.get("raisonnement"),
        "section_observee":                  j.get("section_observee"),
        "comportement_attendu_selon_grille": j.get("comportement_attendu_selon_grille"),
        "type_reponse_attendue_specifie":    j.get("type_reponse_attendue_specifie"),
        "coherence_comportement_observe":    j.get("coherence_comportement_observe"),
        "sources_inventaire":               j.get("sources_inventaire", []),
        "mislabelling_detecte":             j.get("mislabelling_detecte", {}),
        "mislabelling_flag":                any(
            str(v).lower() not in ("non", "false", "", "null", "none")
            for v in j.get("mislabelling_detecte", {}).values()
        ),
        "elements_specifiques_question":    j.get("elements_specifiques_question", []),
        "elements_traitement":              j.get("elements_traitement", []),
    }
    dim_keys = ("pertinence", "fondement_factuel", "nuance_incertitude", "coherence_qualiquanti")
    for k in dim_keys:
        v = j.get(k, {})
        if isinstance(v, dict):
            result[k] = v.get("note")
            result[f"{k}_justif"] = v.get("justification")
        else:
            result[k] = v
            result[f"{k}_justif"] = None

    for et in result["elements_traitement"]:
        if isinstance(et, dict) and et.get("traitement") not in _VALID_TRAITEMENT:
            et["traitement"] = "omis"

    dims = [result[k] for k in dim_keys if isinstance(result.get(k), (int, float))]
    result["score_global"] = round(sum(dims) / len(dims), 2) if dims else None
    return result


def score_judge_v43(question: str, answer: str, sources: list,
                    section: str, subsection: str, expected_type: str) -> dict:
    """
    Judge V4.3 : 4 principes cardinaux, définitions opérationnelles OppChoVec,
    4 règles anti-mislabelling, grille par sous-section.
    max_tokens=3000 (sortie enrichie avec sources_inventaire + mislabelling_detecte).
    """
    sources_text = _build_sources_text(sources)

    user_prompt = (
        f"QUESTION : {question}\n\n"
        f"SECTION : {section}\n\n"
        f"SOUS-SECTION : {subsection}\n\n"
        f"TYPE DE RÉPONSE ATTENDUE : {expected_type}\n\n"
        f"SOURCES FOURNIES AU SYSTÈME :\n{sources_text}\n\n"
        f"RÉPONSE DU SYSTÈME :\n{answer}\n\n"
        "Évalue cette réponse selon la procédure et le format spécifiés.\n"
        "Consulte les définitions opérationnelles et la grille AVANT de noter."
    )

    try:
        raw = _call_llm(_JUDGE_V43_SYSTEM, user_prompt, max_tokens=3000, json_mode=True)
        m = re.search(r'\{[\s\S]*\}', raw)
        j = json.loads(m.group()) if m else {}
        result = _parse_judge_v43(j)
        result["error"] = None
        return result
    except Exception as e:
        return {"error": str(e), "score_global": None}


# ── Tests unitaires parser V4.1 ──────────────────────────────────────────────

def _run_v41_parser_tests():
    sample = {
        "type_reponse_attendue_observe": "reponse_substantielle",
        "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
        "coherence_type_reponse": "oui",
        "elements_specifiques_question": ["entrepreneurs", "Ajaccio"],
        "elements_traitement": [
            {"element": "entrepreneurs", "traitement": "precis"},
            {"element": "Ajaccio", "traitement": "approximation_signalee"},
        ],
        "raisonnement": "Réponse ciblée.",
        "pertinence": {"note": 4, "justification": "OK"},
        "fondement_factuel": {"note": 5, "justification": "OK"},
        "nuance_incertitude": {"note": 4, "justification": "OK"},
        "coherence_qualiquanti": {"note": 3, "justification": "OK"},
    }
    r = _parse_judge_v41(sample)
    assert r["score_global"] == 4.0, f"score_global={r['score_global']}"
    assert r["pertinence"] == 4
    assert r["coherence_type_reponse"] == "oui"
    assert len(r["elements_traitement"]) == 2
    assert r["elements_traitement"][0]["traitement"] == "precis"

    # Invalid status should be coerced to "omis"
    bad = {"elements_traitement": [{"element": "x", "traitement": "invalid"}],
           "pertinence": {"note": 3, "justification": ""},
           "fondement_factuel": {"note": 3, "justification": ""},
           "nuance_incertitude": {"note": 3, "justification": ""},
           "coherence_qualiquanti": {"note": 3, "justification": ""}}
    r2 = _parse_judge_v41(bad)
    assert r2["elements_traitement"][0]["traitement"] == "omis"

    # V1 flat format fallback
    flat = {"pertinence": 3, "fondement_factuel": 4,
            "nuance_incertitude": 2, "coherence_qualiquanti": 5}
    r3 = _parse_judge_v41(flat)
    assert r3["score_global"] == 3.5

    print("  V4.1 parser tests : OK")


def _run_v43_parser_tests():
    sample = {
        "section_observee": "Raisonnement comparatif / Croisé quanti/quali",
        "comportement_attendu_selon_grille": "Comparaison explicite, signaler manque quali",
        "type_reponse_attendue_specifie": "reponse_substantielle_attendue",
        "coherence_comportement_observe": "non",
        "sources_inventaire": [
            {"source": "oppchovec:Pedicorte", "type": "quanti"},
        ],
        "mislabelling_detecte": {
            "regle_1_quali_quanti": "oui — OppChoVec appelé qualitatif",
            "regle_2_surinterpretation_oppchovec": "oui — Vec présenté comme vécu",
            "regle_3_absence_quali_non_signalee": "oui — manque non signalé",
            "regle_4_extrapolation_sous_groupe": "non",
        },
        "elements_specifiques_question": ["Pedicorte", "écart"],
        "elements_traitement": [
            {"element": "Pedicorte", "traitement": "precis"},
            {"element": "écart", "traitement": "approximation_non_signalee"},
        ],
        "raisonnement": "Triple mislabelling.",
        "pertinence": {"note": 3, "justification": "Plafond R3"},
        "fondement_factuel": {"note": 2, "justification": "R1+R2"},
        "nuance_incertitude": {"note": 2, "justification": "R3"},
        "coherence_qualiquanti": {"note": 2, "justification": "R1"},
    }
    r = _parse_judge_v43(sample)
    assert r["score_global"] == round((3 + 2 + 2 + 2) / 4, 2), f"score_global={r['score_global']}"
    assert r["section_observee"] == "Raisonnement comparatif / Croisé quanti/quali"
    assert r["coherence_comportement_observe"] == "non"
    assert len(r["sources_inventaire"]) == 1
    assert r["sources_inventaire"][0]["type"] == "quanti"
    assert "regle_1_quali_quanti" in r["mislabelling_detecte"]
    assert r["fondement_factuel"] == 2

    # Invalid traitement coerced to "omis"
    bad = {
        "elements_traitement": [{"element": "x", "traitement": "invalid"}],
        "pertinence": {"note": 3, "justification": ""},
        "fondement_factuel": {"note": 3, "justification": ""},
        "nuance_incertitude": {"note": 3, "justification": ""},
        "coherence_qualiquanti": {"note": 3, "justification": ""},
    }
    r2 = _parse_judge_v43(bad)
    assert r2["elements_traitement"][0]["traitement"] == "omis"
    assert r2["score_global"] == 3.0

    print("  V4.3 parser tests : OK")


def score_refusal_hallucination(question: str, answer: str) -> dict:
    """
    Pour les questions sur des données absentes (Nice, Haïti, 2018...).
    Détecte :
    - refusal_ok (bool) : le modèle refuse correctement de répondre
    - hallucination (bool) : le modèle invente une réponse factuelle
    - overconfidence (bool) : le modèle répond sans exprimer d'incertitude
    """
    # Heuristiques rapides avant d'appeler le LLM
    refusal_markers = [
        "n'ai pas", "ne dispose pas", "pas de données", "pas d'information",
        "hors de ma", "non disponible", "aucune donnée", "ne couvre pas",
        "je ne suis pas en mesure", "impossible de répondre",
        "données limitées", "pas dans les sources",
    ]
    lower = answer.lower()
    quick_refusal = any(m in lower for m in refusal_markers)

    system = "Tu es un évaluateur de systèmes RAG. Réponds UNIQUEMENT en JSON."
    prompt = f"""Question : {question}
Réponse du système : {answer}

Cette question porte sur des données inexistantes ou hors périmètre (ex : commune étrangère, année non couverte).
Le comportement attendu est de REFUSER de répondre, ou d'exprimer clairement l'absence de données.

Évalue :
{{
  "refusal_ok": true/false,  // le système refuse ou indique l'absence de données
  "hallucination": true/false,  // le système invente une réponse factuelle
  "overconfidence": true/false,  // le système répond sans exprimer d'incertitude
  "explication": "..."
}}"""
    try:
        raw = _call_llm(system, prompt, max_tokens=300, light=True)
        m = re.search(r'\{[\s\S]*\}', raw)
        j = json.loads(m.group()) if m else {}
        # Override si heuristique détecte un refus clair
        if quick_refusal and not j.get("refusal_ok"):
            j["refusal_ok"] = True
        return j
    except Exception as e:
        # Fallback sur heuristique
        return {
            "refusal_ok": quick_refusal,
            "hallucination": not quick_refusal,
            "overconfidence": not quick_refusal,
            "explication": f"Heuristique (LLM échoué: {e})",
        }


# ─────────────────────────────────────────────
# 5. Robustesse sémantique (post-traitement)
# ─────────────────────────────────────────────

def compute_semantic_robustness(groups: dict, row_answers: dict) -> dict:
    """
    Pour chaque groupe, embed les réponses et calcule la similarité cosinus intra-groupe.
    groups: {nom_groupe: [row1, row2, ...]}
    row_answers: {excel_row: answer_str}
    Retourne: {nom_groupe: {"mean_sim": float, "min_sim": float, "answers": [...]}}
    """
    try:
        from sentence_transformers import SentenceTransformer
        import numpy as np

        model_path = "./model_cache/models--BAAI--bge-m3/snapshots/5617a9f61b028005a4858fdac845db406aefb181"
        if not os.path.exists(model_path):
            model_path = "BAAI/bge-m3"

        print("  Chargement BGE-M3 pour robustesse semantique...", flush=True)
        model = SentenceTransformer(model_path)

        results = {}
        for group_name, rows in groups.items():
            answers = [row_answers.get(r, "") for r in rows]
            non_empty = [a for a in answers if a and not a.startswith("ERREUR")]
            if len(non_empty) < 2:
                results[group_name] = {
                    "mean_sim": None, "min_sim": None,
                    "detail": "Moins de 2 réponses disponibles",
                    "answers": answers,
                }
                continue

            embeddings = model.encode(
                [f"passage: {a}" for a in non_empty],
                batch_size=4, show_progress_bar=False
            )
            # Cosinus pairwise
            sims = []
            n = len(embeddings)
            for i in range(n):
                for j in range(i + 1, n):
                    a, b = embeddings[i], embeddings[j]
                    cos = float(np.dot(a, b) / (np.linalg.norm(a) * np.linalg.norm(b) + 1e-9))
                    sims.append(cos)

            results[group_name] = {
                "mean_sim": round(sum(sims) / len(sims), 4) if sims else None,
                "min_sim":  round(min(sims), 4) if sims else None,
                "detail":   f"{len(non_empty)} réponses, {len(sims)} paires",
                "answers":  answers,
            }
        return results

    except Exception as e:
        traceback.print_exc()
        return {g: {"mean_sim": None, "min_sim": None, "detail": str(e), "answers": []}
                for g in groups}


# ─────────────────────────────────────────────
# 6. Export Excel (3 feuilles)
# ─────────────────────────────────────────────

def export_to_markdown(results: list[dict], robustness: dict, output_path: str,
                        metadata: dict = None):
    from collections import defaultdict

    def avg(lst): return round(sum(lst) / len(lst), 3) if lst else None
    def pct(v): return f"{v*100:.1f}%" if v is not None else "—"
    def fmt(v, suffix=""): return f"{v}{suffix}" if v is not None else "—"

    # ── Agrégation par section ──
    import statistics as _stats

    sec_data = defaultdict(lambda: {
        "n": 0, "factual_scores": [], "binary_scores": [],
        "judge_scores": [], "refusals": [], "hallucinations": [], "overconfs": [],
        "f1_scores": [],
    })
    for r in results:
        sec = r.get("section", "Inconnue")
        d = sec_data[sec]
        d["n"] += 1
        f = r.get("scores", {})
        if f.get("factual", {}).get("score") is not None:
            d["factual_scores"].append(f["factual"]["score"])
        if f.get("binary", {}).get("score") is not None:
            d["binary_scores"].append(f["binary"]["score"])
        if f.get("judge", {}).get("score_global") is not None:
            d["judge_scores"].append(f["judge"]["score_global"])
        rs = f.get("refusal", {})
        if rs.get("refusal_ok") is not None:
            d["refusals"].append(1 if rs["refusal_ok"] else 0)
            d["hallucinations"].append(1 if rs.get("hallucination") else 0)
            d["overconfs"].append(1 if rs.get("overconfidence") else 0)
        f1v = f.get("retrieval", {}).get("f1")
        if f1v is not None:
            d["f1_scores"].append(f1v)

    md = []

    # ── En-tête ──
    ts = (metadata or {}).get("timestamp", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    version = (metadata or {}).get("rag_version", "?")
    judge_model = (metadata or {}).get("judge_model", "?")
    total_q = (metadata or {}).get("total_questions", len(results))

    md.append(f"# Rapport d'évaluation RAG — {version.upper()}\n")
    md.append(f"**Date :** {ts}  \n**Modèle juge :** {judge_model}  \n**Questions évaluées :** {total_q}\n")
    md.append("\n---\n")

    # ── Résumé global ──
    factual_ok    = [r for r in results if r.get("scores", {}).get("factual",   {}).get("score") is not None]
    binary_ok     = [r for r in results if r.get("scores", {}).get("binary",    {}).get("score") is not None]
    judge_ok      = [r for r in results if r.get("scores", {}).get("judge",     {}).get("score_global") is not None]
    judge_v41_ok  = [r for r in results if r.get("scores", {}).get("judge_v41", {}).get("score_global") is not None]
    refusal_ok    = [r for r in results if r.get("scores", {}).get("refusal",   {}).get("refusal_ok") is not None]
    retrieval_ok  = [r for r in results if r.get("scores", {}).get("retrieval", {}).get("recall") is not None]
    errors        = [r for r in results if r.get("answer", "").startswith("ERREUR")]

    md.append("## Résumé global\n")
    md.append("| Métrique | Valeur | N |\n|---------|--------|---|\n")
    if retrieval_ok:
        mean_rec  = avg([r["scores"]["retrieval"]["recall"]    for r in retrieval_ok])
        mean_prec = avg([r["scores"]["retrieval"]["precision"] for r in retrieval_ok])
        f1_vals   = [r["scores"]["retrieval"]["f1"] for r in retrieval_ok
                     if r["scores"]["retrieval"].get("f1") is not None]
        mean_f1   = avg(f1_vals)
        med_f1    = round(_stats.median(f1_vals), 3) if f1_vals else None
        std_f1    = round(_stats.stdev(f1_vals), 3) if len(f1_vals) >= 2 else None
        md.append(f"| Recall retrieval | {pct(mean_rec)} | {len(retrieval_ok)} |\n")
        md.append(f"| Precision retrieval | {pct(mean_prec)} | {len(retrieval_ok)} |\n")
        md.append(f"| F1 retrieval (moyen) | {pct(mean_f1)} | {len(f1_vals)} |\n")
        md.append(f"| F1 retrieval (médian) | {pct(med_f1)} | {len(f1_vals)} |\n")
        md.append(f"| F1 retrieval (std) | {fmt(std_f1)} | {len(f1_vals)} |\n")
    if factual_ok:
        md.append(f"| Factual accuracy | {pct(avg([r['scores']['factual']['score'] for r in factual_ok]))} | {len(factual_ok)} |\n")
    if binary_ok:
        md.append(f"| Binary accuracy | {pct(avg([r['scores']['binary']['score'] for r in binary_ok]))} | {len(binary_ok)} |\n")
    if judge_ok:
        score = avg([r['scores']['judge']['score_global'] for r in judge_ok])
        md.append(f"| Juge V2 (moy.) | {fmt(score, '/5')} | {len(judge_ok)} |\n")
        sujet_ok = [r for r in judge_ok if r['scores']['judge'].get('note_sujet') is not None]
        if sujet_ok:
            score_sujet = avg([r['scores']['judge']['note_sujet'] for r in sujet_ok])
            md.append(f"| Note sujet (moy.) | {fmt(score_sujet, '/5')} | {len(sujet_ok)} |\n")
    if judge_v41_ok:
        score_v41 = avg([r['scores']['judge_v41']['score_global'] for r in judge_v41_ok])
        md.append(f"| Juge V4.1 (moy.) | {fmt(score_v41, '/5')} | {len(judge_v41_ok)} |\n")
    if refusal_ok:
        ref_rate = avg([1 if r['scores']['refusal']['refusal_ok'] else 0 for r in refusal_ok])
        hal_rate = avg([1 if r['scores']['refusal'].get('hallucination') else 0 for r in refusal_ok])
        md.append(f"| Refus correct | {pct(ref_rate)} | {len(refusal_ok)} |\n")
        md.append(f"| Hallucination | {pct(hal_rate)} | {len(refusal_ok)} |\n")
    if errors:
        md.append(f"| Erreurs API | {len(errors)} | {len(results)} |\n")
    md.append("\n")

    # ── Résumé par section ──
    md.append("## Résumé par section\n")
    md.append("| Section | N | Juge moy. | Binary acc. | Factual acc. | Refusal |\n")
    md.append("|---------|---|-----------|------------|--------------|--------|\n")
    for sec, d in sorted(sec_data.items()):
        md.append(
            f"| {sec} | {d['n']} "
            f"| {fmt(avg(d['judge_scores']), '/5')} "
            f"| {pct(avg(d['binary_scores']))} "
            f"| {pct(avg(d['factual_scores']))} "
            f"| {pct(avg(d['refusals']))} |\n"
        )
    md.append("\n")

    # ── F1 retrieval par section ──
    f1_sections = {sec: d["f1_scores"] for sec, d in sec_data.items() if d["f1_scores"]}
    if f1_sections:
        md.append("## F1 retrieval par section\n")
        md.append("| Section | N | F1 moyen | F1 médian | Std |\n")
        md.append("|---------|---|----------|-----------|-----|\n")
        for sec, vals in sorted(f1_sections.items()):
            m   = avg(vals)
            med = round(_stats.median(vals), 3)
            std = round(_stats.stdev(vals), 3) if len(vals) >= 2 else None
            md.append(f"| {sec} | {len(vals)} | {pct(m)} | {pct(med)} | {fmt(std)} |\n")
        md.append("\n")

    # ── Robustesse sémantique ──
    if robustness:
        md.append("## Robustesse sémantique\n")
        md.append("| Groupe | N | Sim. moy. | Sim. min. |\n|--------|---|-----------|----------|\n")
        for group, data in robustness.items():
            answers = data.get("answers", [])
            n_ok = len([a for a in answers if a and not a.startswith("ERREUR")])
            md.append(
                f"| {group} | {n_ok} "
                f"| {fmt(data.get('mean_sim'))} "
                f"| {fmt(data.get('min_sim'))} |\n"
            )
        md.append("\n> Interprétation : sim > 0.85 = très cohérent · 0.70–0.85 = acceptable · < 0.70 = fragile\n\n")

    # ── Résultats détaillés ──
    md.append("---\n\n## Résultats détaillés\n")

    current_section = None
    for r in results:
        sec = r.get("section", "")
        if sec != current_section:
            current_section = sec
            md.append(f"\n### {sec}\n")

        row = r.get("excel_row", "?")
        subsec = r.get("subsection", "")
        question = r.get("question", "")
        answer = r.get("answer", "")
        f = r.get("scores", {})
        metrics = r.get("metrics", [])

        md.append(f"\n#### R{row} — {subsec}\n\n")
        md.append(f"> {question}\n\n")

        # Badges métriques
        badge_map = {"factual": "📐 Factual", "binary": "🔢 Binary", "judge": "🧑‍⚖️ Juge",
                     "refusal": "🚫 Refusal", "halluc": "👻 Halluc.", "overconf": "⚠️ Overconf.", "robust": "🔄 Robustesse"}
        badges = " · ".join(badge_map.get(m, m) for m in metrics if m in badge_map)
        if badges:
            md.append(f"**Métriques :** {badges}\n\n")

        if answer.startswith("ERREUR"):
            md.append(f"❌ **{answer}**\n\n")
        else:
            md.append("**Réponse RAG :**\n\n")
            md.append(answer.strip() + "\n\n")

            # Scores
            scores_lines = []
            ret = f.get("retrieval", {})
            if ret.get("recall") is not None:
                exp = ", ".join(ret.get("expected", []))
                got = ", ".join(ret.get("retrieved", []))
                f1v = ret.get("f1")
                f1_str = f" · F1={f1v:.0%}" if f1v is not None else ""
                scores_lines.append(
                    f"**📦 Retrieval :** {ret['detail']}{f1_str}  \n"
                    f"  expected=[{exp}] · retrieved=[{got}]"
                )
            elif ret.get("refusal_case"):
                scores_lines.append(
                    f"**📦 Retrieval :** {ret.get('detail', '')} *(refusal attendu)*"
                )
            fs = f.get("factual", {})
            if fs.get("score") is not None:
                scores_lines.append(f"**📐 Factual :** {pct(fs['score'])} — {fs.get('detail', '')}")
            bs = f.get("binary", {})
            if bs.get("score") is not None:
                label = "✅" if bs["score"] == 1 else "❌"
                scores_lines.append(f"**🔢 Binary :** {label} {bs.get('detail', '')}")
            js = f.get("judge", {})
            if js.get("score_global") is not None:
                scores_lines.append(
                    f"**🧑‍⚖️ Juge V2 :** {js['score_global']:.1f}/5 "
                    f"— pertinence: {fmt(js.get('pertinence'))} · "
                    f"fondement: {fmt(js.get('fondement_factuel'))} · "
                    f"nuance: {fmt(js.get('nuance_incertitude'))} · "
                    f"cohérence: {fmt(js.get('coherence_qualiquanti'))}"
                )
                dim_keys = [
                    ("pertinence",          "Pertinence"),
                    ("fondement_factuel",    "Fondement factuel"),
                    ("nuance_incertitude",   "Nuance / incertitude"),
                    ("coherence_qualiquanti","Cohérence quali-quanti"),
                ]
                for key, label in dim_keys:
                    justif = js.get(f"{key}_justif", "")
                    note   = js.get(key)
                    if justif:
                        scores_lines.append(f"  - *{label} ({fmt(note)}) :* {justif}")
                if js.get("applicable_sujet") and js.get("note_sujet") is not None:
                    sujet = js.get("sujet_evalue", "sujet")
                    scores_lines.append(
                        f"**🌍 Note sujet** ({sujet}) : **{js['note_sujet']:.1f}/5**"
                    )
                    if js.get("justification_sujet"):
                        scores_lines.append(f"  - *{js['justification_sujet']}*")
                elif js.get("reason_non_applicable"):
                    scores_lines.append(f"*Note sujet : non applicable ({js['reason_non_applicable']})*")
            js41 = f.get("judge_v41", {})
            if js41.get("score_global") is not None:
                scores_lines.append(
                    f"**🧑‍⚖️ Juge V4.1 :** {js41['score_global']:.1f}/5 "
                    f"[{js41.get('coherence_type_reponse','?')}] "
                    f"— pertinence: {fmt(js41.get('pertinence'))} · "
                    f"fondement: {fmt(js41.get('fondement_factuel'))} · "
                    f"nuance: {fmt(js41.get('nuance_incertitude'))} · "
                    f"cohérence: {fmt(js41.get('coherence_qualiquanti'))}"
                )
                for key, label in [
                    ("pertinence", "Pertinence"), ("fondement_factuel", "Fondement factuel"),
                    ("nuance_incertitude", "Nuance"), ("coherence_qualiquanti", "Cohérence"),
                ]:
                    justif = js41.get(f"{key}_justif", "")
                    note   = js41.get(key)
                    if justif:
                        scores_lines.append(f"  - *{label} ({fmt(note)}) :* {justif}")
                ets = js41.get("elements_traitement", [])
                if ets:
                    et_str = " · ".join(
                        f"{e['element']}={e['traitement']}" for e in ets if isinstance(e, dict)
                    )
                    scores_lines.append(f"  *Traitement éléments :* {et_str}")
            rs = f.get("refusal", {})
            if rs.get("refusal_ok") is not None:
                ref_icon = "✅" if rs["refusal_ok"] else "❌"
                hal_icon = "⚠️" if rs.get("hallucination") else "✅"
                oc_icon  = "⚠️" if rs.get("overconfidence") else "✅"
                scores_lines.append(
                    f"**🚫 Refusal :** {ref_icon} · "
                    f"Hallucination : {hal_icon} · "
                    f"Overconfidence : {oc_icon}"
                )
                if rs.get("explication"):
                    scores_lines.append(f"*{rs['explication'][:200]}*")
            rb = r.get("robustness_group", "")
            if rb:
                scores_lines.append(f"**🔄 Groupe robustesse :** `{rb}`")
            if r.get("comments"):
                scores_lines.append(f"*{r['comments']}*")

            if scores_lines:
                md.append("\n".join(scores_lines) + "\n\n")

        md.append("---\n")

    with open(output_path, "w", encoding="utf-8") as f:
        f.write("".join(md))
    print(f"Markdown sauvegarde : {output_path}")


# ─────────────────────────────────────────────
# 7. Pipeline principal
# ─────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Evaluation RAG depuis Excel — pipeline complet")
    parser.add_argument("--input",   default=EXCEL_INPUT, help="Fichier Excel d'entrée")
    parser.add_argument("--version", default=RAG_VERSION, help="Version RAG à tester (default: v10)")
    parser.add_argument("--k",       type=int, default=7, help="Nombre de chunks à récupérer")
    parser.add_argument("--max",     type=int, default=0, help="Limiter à N questions (0=toutes)")
    parser.add_argument("--rows",    default="", help="Lignes Excel à évaluer, ex: '2,5,7-12' (vide=toutes)")
    parser.add_argument("--output",  default="comparaisons_rag", help="Dossier de sortie")
    parser.add_argument("--no-judge",   action="store_true", help="Désactiver le juge LLM")
    parser.add_argument("--no-robust",  action="store_true", help="Désactiver la robustesse sémantique")
    parser.add_argument("--from-json",  default=None, help="Re-exporter depuis un JSON existant")
    args = parser.parse_args()

    os.makedirs(args.output, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # ── Mode re-export ──
    if args.from_json:
        with open(args.from_json, "r", encoding="utf-8") as f:
            saved = json.load(f)
        results = saved["results"]
        robustness = saved.get("robustness", {})
        md_path = args.from_json.replace(".json", "_reexport.md")
        export_to_markdown(results, robustness, md_path, metadata=saved.get("metadata", {}))
        return

    # ── Chargement ground truth ──
    from eval_ground_truth import get_full_ground_truth, BINARY_EXPECTED, ROBUSTNESS_GROUPS, find_ground_truth
    gt = get_full_ground_truth()
    print(f"Ground truth charge : {len(gt)} entrees")

    # ── Lecture Excel ──
    questions = load_questions(args.input)

    # Filtrage par numéros de lignes (--rows "2,5,7-12")
    if args.rows.strip():
        selected_rows = set()
        for part in args.rows.split(","):
            part = part.strip()
            if "-" in part:
                a, b = part.split("-", 1)
                selected_rows.update(range(int(a), int(b) + 1))
            elif part.isdigit():
                selected_rows.add(int(part))
        questions = [q for q in questions if q["excel_row"] in selected_rows]

    if args.max > 0:
        questions = questions[:args.max]
    total = len(questions)

    print("=" * 70)
    print(f"EVALUATION RAG — {total} questions — version {args.version}")
    print(f"Juge LLM : {JUDGE_MODEL}")
    print("=" * 70)

    results = []
    row_answers = {}  # excel_row → answer (pour robustesse)

    for i, q in enumerate(questions, 1):
        question   = q["question"]
        excel_row  = q["excel_row"]
        section    = q["section"]

        print(f"\n[{i}/{total}] R{excel_row} {question[:65]}...", flush=True)

        # ── 1. Appel RAG ──
        print("  RAG...", end=" ", flush=True)
        api_resp = call_rag_api(question, args.version, k=args.k)
        if "error" in api_resp and api_resp.get("answer", "").startswith("ERREUR"):
            print(f"ERREUR: {api_resp['error']}")
            results.append({**q, "answer": api_resp["answer"], "sources": [], "scores": {}})
            row_answers[excel_row] = api_resp["answer"]
            continue

        answer  = api_resp.get("answer", "")
        sources = api_resp.get("sources", [])
        print(f"OK ({len(sources)} sources)", flush=True)
        row_answers[excel_row] = answer

        scores = {}

        # ── 2. Retrieval Recall/Precision ──
        if q.get("do_retrieval"):
            print("  Retrieval...", end=" ", flush=True)
            scores["retrieval"] = score_retrieval(sources, q["retrieval_gt"])
            print(scores["retrieval"].get("detail", ""), flush=True)

        # ── 4. Factual Accuracy ──
        if q["do_factual"]:
            gt_val = q.get("facts_gt")
            found = gt_val is not None
            if not found:
                gt_val, found = find_ground_truth(question, gt)
            comment = "" if found else "(pas de ground truth — skippé)"
            print(f"  Factual {'[GT='+str(gt_val)[:30]+']' if found else '[GT manquant]'}...",
                  end=" ", flush=True)
            if found:
                scores["factual"] = score_factual(question, answer, gt_val)
                print(f"score={scores['factual'].get('score')}", flush=True)
            else:
                scores["factual"] = {"score": None, "detail": "Ground truth non disponible"}
                q["comments"] = (q.get("comments") or "") + " " + comment
                print("skip", flush=True)

        # ── 5. Binary ──
        if q["do_binary"]:
            expected = BINARY_EXPECTED.get(question)
            print(f"  Binary [attendu={expected}]...", end=" ", flush=True)
            scores["binary"] = score_binary(question, answer, expected)
            print(f"score={scores['binary'].get('score')}", flush=True)

        # ── 6. LLM-as-a-Judge (V2 + V4.1) ──
        if q["do_judge"] and not args.no_judge:
            print("  Judge V2...", end=" ", flush=True)
            scores["judge"] = score_judge(question, answer, sources, section)
            sg = scores["judge"].get("score_global")
            print(f"global={sg}/5" if sg else f"erreur={scores['judge'].get('error')}", flush=True)

            print("  Judge V4.1...", end=" ", flush=True)
            expected_type = classify_expected_response_type(q)
            scores["judge_v41"] = score_judge_v41(question, answer, sources, section, expected_type)
            sg41 = scores["judge_v41"].get("score_global")
            print(f"global={sg41}/5" if sg41 else f"erreur={scores['judge_v41'].get('error')}", flush=True)
        elif q["do_judge"]:
            scores["judge"]     = {"score_global": None, "detail": "Judge desactive"}
            scores["judge_v41"] = {"score_global": None, "detail": "Judge desactive"}

        # ── 7. Refusal / Hallucination / Overconfidence ──
        if q["do_refusal"] or q["do_halluc"] or q["do_overconf"]:
            print("  Refusal/Halluc/Overconf...", end=" ", flush=True)
            scores["refusal"] = score_refusal_hallucination(question, answer)
            r_ok = scores["refusal"].get("refusal_ok")
            hall = scores["refusal"].get("hallucination")
            print(f"refusal={r_ok}, halluc={hall}", flush=True)

        # ── Robustesse : pas encore calculée (post-traitement) ──
        rob_group = ""
        for gname, rows in ROBUSTNESS_GROUPS.items():
            if excel_row in rows:
                rob_group = gname
                break

        results.append({
            **q,
            "answer":          answer,
            "sources":         sources,
            "scores":          scores,
            "robustness_group": rob_group,
        })

        # Pause anti-rate-limit
        time.sleep(3)

    # ── 6. Robustesse sémantique ──
    robustness = {}
    if not args.no_robust:
        print("\n[Robustesse semantique]")
        robustness = compute_semantic_robustness(ROBUSTNESS_GROUPS, row_answers)
        for g, d in robustness.items():
            sim = d.get("mean_sim")
            print(f"  {g}: sim_moy={sim}")

    # ── 7. Export ──
    base = f"eval_from_excel_{args.version}_{timestamp}"

    json_path = os.path.join(args.output, f"{base}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        # Rendre les sources sérialisables
        for r in results:
            r["sources"] = [
                {k: v for k, v in s.items() if isinstance(v, (str, int, float, bool, type(None)))}
                if isinstance(s, dict) else str(s)
                for s in r.get("sources", [])
            ]
        json.dump({
            "metadata": {
                "timestamp": timestamp,
                "rag_version": args.version,
                "judge_model": JUDGE_MODEL,
                "total_questions": total,
            },
            "results": results,
            "robustness": robustness,
        }, f, ensure_ascii=False, indent=2)
    print(f"\nJSON sauvegarde : {json_path}")

    md_path = os.path.join(args.output, f"{base}.md")
    export_to_markdown(results, robustness, md_path, metadata={
        "timestamp": timestamp,
        "rag_version": args.version,
        "judge_model": JUDGE_MODEL,
        "total_questions": total,
    })

    # ── 8. Résumé terminal ──
    print("\n" + "=" * 70)
    print("RESUME")
    print("=" * 70)
    factual_ok    = [r for r in results if r.get("scores", {}).get("factual",   {}).get("score") is not None]
    binary_ok     = [r for r in results if r.get("scores", {}).get("binary",    {}).get("score") is not None]
    judge_ok      = [r for r in results if r.get("scores", {}).get("judge",     {}).get("score_global") is not None]
    refusal_ok    = [r for r in results if r.get("scores", {}).get("refusal",   {}).get("refusal_ok") is not None]
    retrieval_ok  = [r for r in results if r.get("scores", {}).get("retrieval", {}).get("recall") is not None]

    def safe_avg(lst, key_fn): return round(sum(key_fn(r) for r in lst) / len(lst), 3) if lst else "N/A"

    if retrieval_ok:
        rec  = safe_avg(retrieval_ok, lambda r: r["scores"]["retrieval"]["recall"])
        prec = safe_avg(retrieval_ok, lambda r: r["scores"]["retrieval"]["precision"])
        f1_vals = [r["scores"]["retrieval"]["f1"] for r in retrieval_ok
                   if r["scores"]["retrieval"].get("f1") is not None]
        f1_m = safe_avg(f1_vals, lambda v: v) if f1_vals else "N/A"
        print(f"  Retrieval : {len(retrieval_ok)} questions, recall={rec}, precision={prec}, f1={f1_m}")
    print(f"  Factual   : {len(factual_ok)} questions, score moyen = {safe_avg(factual_ok, lambda r: r['scores']['factual']['score'])}")
    print(f"  Binary    : {len(binary_ok)} questions, accuracy = {safe_avg(binary_ok, lambda r: r['scores']['binary']['score'])}")
    print(f"  Judge V2  : {len(judge_ok)} questions, score moyen = {safe_avg(judge_ok, lambda r: r['scores']['judge']['score_global'])}/5")
    judge_v41_ok  = [r for r in results if r.get("scores", {}).get("judge_v41", {}).get("score_global") is not None]
    print(f"  Judge V4.1: {len(judge_v41_ok)} questions, score moyen = {safe_avg(judge_v41_ok, lambda r: r['scores']['judge_v41']['score_global'])}/5")
    refusal_rate = safe_avg(refusal_ok, lambda r: 1 if r['scores']['refusal']['refusal_ok'] else 0)
    halluc_rate  = safe_avg(refusal_ok, lambda r: 1 if r['scores']['refusal'].get('hallucination') else 0)
    print(f"  Refusal   : {len(refusal_ok)} questions, taux refus correct = {refusal_rate}, hallucination = {halluc_rate}")
    for g, d in robustness.items():
        print(f"  Robust [{g[:25]}] : sim_moy={d.get('mean_sim')}, sim_min={d.get('min_sim')}")
    print("=" * 70)


if __name__ == "__main__":
    main()
