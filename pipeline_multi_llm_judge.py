"""
Pipeline Multi-LLM avec Agent Juge (Mistral Large)
- Interroge 5 LLM (modèles économiques) sur les questions de questions_communes.json
- Utilise Mistral Large comme juge pour noter /10, classer et synthétiser les réponses
- Exporte les résultats dans un Excel structuré
"""

import os
import sys
import json
import time
import argparse
import re
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional
from dotenv import load_dotenv

load_dotenv()

# --- Configuration des modèles (1 petit modèle par provider, versions les plus récentes) ---
MODELS_CONFIG = {
    "OpenAI": {
        "api_key_env": "OPENAI_API_KEY",
        "model": "gpt-4.1-mini",          # Avril 2025, remplace gpt-4o-mini
        "base_url": None
    },
    "Grok": {
        "api_key_env": "GROK_API_KEY",
        "model": "grok-3-mini-beta",       # ID officiel xAI API
        "base_url": "https://api.x.ai/v1"
    },
    "Kimi": {
        "api_key_env": "KIMIK2_API_KEY",
        "model": "kimi-k2-0905-preview",   # K2 (juillet 2025), remplace moonshot-v1
        "base_url": "https://api.moonshot.ai/v1"
    },
    "Claude": {
        "api_key_env": "CLAUDE_API_KEY",
        "model": "claude-haiku-4-5-20251001",  # Haiku 4.5 (octobre 2025), remplace 3-haiku
        "base_url": None,
        "use_anthropic": True
    },
    "Mistral": {
        "api_key_env": "MISTRAL_API_KEY",
        "model": "mistral-small-latest",   # Alias auto-mis a jour par Mistral
        "base_url": "https://api.mistral.ai/v1"
    }
}

# Modèle juge
JUDGE_MODEL = "mistral-large-latest"
JUDGE_API_KEY_ENV = "MISTRAL_API_KEY"
JUDGE_BASE_URL = "https://api.mistral.ai/v1"

# Lettres pour anonymiser les réponses
LETTERS = ["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"]


# ============================================================
# Fonctions d'appel LLM (reprises de compare_multi_llm.py)
# ============================================================

def call_openai_compatible(api_key: str, model: str, prompt: str, system_prompt: str,
                           base_url: Optional[str] = None) -> str:
    from openai import OpenAI
    client_kwargs = {"api_key": api_key}
    if base_url:
        client_kwargs["base_url"] = base_url
    client = OpenAI(**client_kwargs)
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": prompt}
        ],
        temperature=0.7,
        max_tokens=1000
    )
    return response.choices[0].message.content


def call_anthropic(api_key: str, model: str, prompt: str, system_prompt: str) -> str:
    import anthropic
    client = anthropic.Anthropic(api_key=api_key)
    response = client.messages.create(
        model=model,
        max_tokens=1000,
        system=system_prompt,
        messages=[{"role": "user", "content": prompt}]
    )
    return response.content[0].text


# ============================================================
# Adaptateurs RAG (retrieval-only, pas de generation LLM)
# ============================================================

AVAILABLE_RAG_VERSIONS = ["baseline", "v2", "v6", "v7", "raptor"]


def _format_sources(documents: list, metadatas: list) -> List[Dict]:
    """Formate les sources depuis des listes documents/metadatas brutes."""
    sources = []
    for i, (doc, meta) in enumerate(zip(documents, metadatas), 1):
        sources.append({
            "rank": i,
            "commune": meta.get("nom", meta.get("commune", "N/A")),
            "genre": meta.get("genre", "N/A"),
            "age": meta.get("age_exact", "N/A"),
            "profession": meta.get("profession", "N/A"),
            "dimension": meta.get("dimension", "N/A"),
            "source_type": meta.get("source_type", "dense"),
            "extrait": doc[:200] + "..."
        })
    return sources


class RAGAdapter:
    """Interface commune pour tous les backends RAG."""

    def __init__(self, version: str):
        self.version = version
        self._pipeline = None
        self._embed_model = None
        self._collection = None

    def init(self):
        """Initialise le backend RAG (chargement modeles, connexions)."""
        if self.version == "baseline":
            self._init_baseline()
        elif self.version == "v2":
            self._init_v2()
        elif self.version == "v6":
            self._init_v6()
        elif self.version == "v7":
            self._init_v7()
        elif self.version == "raptor":
            self._init_raptor()
        else:
            raise ValueError(f"Version RAG inconnue: {self.version}. "
                             f"Disponibles: {AVAILABLE_RAG_VERSIONS}")

    def get_context(self, question: str, k: int = 5) -> tuple:
        """Retourne (context_str, sources_list) pour une question."""
        if self.version == "baseline":
            return self._retrieve_baseline(question, k)
        elif self.version == "v2":
            return self._retrieve_v2(question, k)
        elif self.version == "v6":
            return self._retrieve_v6(question, k)
        elif self.version == "v7":
            return self._retrieve_v7(question, k)
        elif self.version == "raptor":
            return self._retrieve_raptor(question, k)

    def close(self):
        """Libere les ressources."""
        if hasattr(self._pipeline, 'close'):
            self._pipeline.close()

    # --- baseline : ChromaDB direct (ancien get_rag_context) ---

    def _init_baseline(self):
        from sentence_transformers import SentenceTransformer
        import chromadb
        self._embed_model = SentenceTransformer("BAAI/bge-m3")
        client = chromadb.PersistentClient(path="./chroma_portrait")
        self._collection = client.get_collection("portrait_verbatims")

    def _retrieve_baseline(self, question: str, k: int) -> tuple:
        query_embedding = self._embed_model.encode(f"query: {question}").tolist()
        results = self._collection.query(
            query_embeddings=[query_embedding],
            n_results=k,
            include=["documents", "metadatas"]
        )
        docs = results['documents'][0]
        metas = results['metadatas'][0]
        context_parts = [f"[Source {i}] {doc[:300]}..." for i, doc in enumerate(docs, 1)]
        sources = _format_sources(docs, metas)
        return "\n\n".join(context_parts), sources

    # --- v2 : PortraitRAGPipeline (hybrid + reranking) ---

    def _init_v2(self):
        from rag_v2_2_portrait import PortraitRAGPipeline
        openai_key = os.getenv("OPENAI_API_KEY")
        self._pipeline = PortraitRAGPipeline(
            chroma_path="./chroma_portrait",
            collection_name="portrait_verbatims",
            embedding_model="BAAI/bge-m3",
            reranker_model="BAAI/bge-reranker-v2-m3",
            llm_model="gpt-4o-mini",
            openai_api_key=openai_key
        )

    def _retrieve_v2(self, question: str, k: int) -> tuple:
        # query() retourne (response_llm, retrieval_results, detected_filters)
        # On ignore response_llm car chaque LLM du pipeline generera sa propre reponse
        _, retrieval_results, _ = self._pipeline.query(
            question, k=k, use_reranking=True, include_quantitative=True
        )
        context_parts = []
        sources = []
        for i, rr in enumerate(retrieval_results, 1):
            context_parts.append(f"[Source {i} ({rr.source_type})] {rr.text[:300]}...")
            meta = rr.metadata
            sources.append({
                "rank": i,
                "commune": meta.get("nom", "N/A"),
                "genre": meta.get("genre", "N/A"),
                "age": meta.get("age_exact", "N/A"),
                "profession": meta.get("profession", "N/A"),
                "dimension": meta.get("dimension", "N/A"),
                "source_type": rr.source_type,
                "score": round(rr.score, 4),
                "extrait": rr.text[:200] + "..."
            })
        return "\n\n".join(context_parts), sources

    # --- v6 : G-Retriever (GNN heterogene + Neo4j + ChromaDB) ---

    def _init_v6(self):
        from rag_v6_gretriever import GRetrieverRAGPipeline
        openai_key = os.getenv("OPENAI_API_KEY")
        neo4j_password = os.getenv("NEO4J_PASSWORD", "password")
        self._pipeline = GRetrieverRAGPipeline(
            neo4j_uri="bolt://localhost:7687",
            neo4j_user="neo4j",
            neo4j_password=neo4j_password,
            chroma_path="./chroma_portrait",
            collection_name="portrait_verbatims",
            openai_api_key=openai_key
        )

    def _retrieve_v6(self, question: str, k: int) -> tuple:
        # query() retourne (response_str, List[RetrievalResult])
        # On ignore response_str, on extrait le contexte des RetrievalResult
        _, retrieval_results = self._pipeline.query(question, k=k, use_gnn=True)
        context_parts = []
        sources = []
        for i, rr in enumerate(retrieval_results, 1):
            context_parts.append(f"[Source {i} ({rr.source_type})] {rr.text[:300]}...")
            meta = rr.metadata
            sources.append({
                "rank": i,
                "commune": meta.get("nom", meta.get("commune", "N/A")),
                "genre": meta.get("genre", "N/A"),
                "age": meta.get("age_exact", "N/A"),
                "profession": meta.get("profession", "N/A"),
                "dimension": meta.get("dimension", "N/A"),
                "source_type": rr.source_type,
                "score": round(rr.score, 4),
                "extrait": rr.text[:200] + "..."
            })
        return "\n\n".join(context_parts), sources

    # --- v7 : LlamaIndex (vector + graph via Neo4j) ---

    def _init_v7(self):
        from rag_v7_llamaindex import LlamaIndexRAGPipeline
        openai_key = os.getenv("OPENAI_API_KEY")
        neo4j_password = os.getenv("NEO4J_PASSWORD", "password")
        self._pipeline = LlamaIndexRAGPipeline(
            openai_api_key=openai_key,
            chroma_path="./chroma_v2",
            neo4j_uri="bolt://localhost:7687",
            neo4j_user="neo4j",
            neo4j_password=neo4j_password
        )

    def _retrieve_v7(self, question: str, k: int) -> tuple:
        # query() retourne (response_str, sources_list)
        # response_str est genere par le LLM interne de LlamaIndex, on l'ignore
        # mais on recupere les sources pour construire le contexte
        _, sources_raw = self._pipeline.query(
            question, mode="hybrid", k=k
        )
        context_parts = []
        sources = []
        for i, src in enumerate(sources_raw, 1):
            text = src.get("content", "")
            meta = src.get("metadata", {})
            src_type = src.get("source_type", "unknown")
            context_parts.append(f"[Source {i} ({src_type})] {text[:300]}...")
            sources.append({
                "rank": i,
                "commune": meta.get("nom", meta.get("commune", "N/A")),
                "genre": meta.get("genre", "N/A"),
                "age": meta.get("age_exact", "N/A"),
                "profession": meta.get("profession", "N/A"),
                "dimension": meta.get("dimension", "N/A"),
                "source_type": src_type,
                "score": round(src.get("score", 0), 4),
                "extrait": text[:200] + "..."
            })
        return "\n\n".join(context_parts), sources

    # --- raptor : RAPTOR-lite (syntheses analytiques hierarchiques) ---

    def _init_raptor(self):
        from rag_v9_raptor import RaptorRetriever
        self._pipeline = RaptorRetriever(
            chroma_path="./chroma_portrait",
            source_collection="portrait_verbatims",
            summary_collection="raptor_summaries",
        )
        self._pipeline.init()

    def _retrieve_raptor(self, question: str, k: int) -> tuple:
        return self._pipeline.query(question, k=k)


def build_prompt(question: str, context: str, rag_version: str = "baseline") -> tuple:
    system_prompt = """Tu es un assistant specialise dans l'analyse de donnees sur la qualite de vie en Corse.
Tu as acces a des donnees issues d'enquetes aupres d'habitants (verbatims, donnees quantitatives)
et potentiellement a une ontologie du bien-etre (concepts, dimensions, indicateurs).
Base tes reponses UNIQUEMENT sur les informations fournies dans le contexte.
Si les donnees sont insuffisantes, indique-le clairement.
Reponds de maniere concise et factuelle."""

    user_prompt = f"""=== CONTEXTE (sources RAG {rag_version}) ===
{context}

=== QUESTION ===
{question}

Reponds a la question en te basant sur le contexte ci-dessus."""

    return system_prompt, user_prompt


# ============================================================
# Test d'un modèle
# ============================================================

def test_model(provider: str, question: str, context: str, sources: List[Dict],
               rag_version: str = "baseline") -> Dict[str, Any]:
    config = MODELS_CONFIG[provider]
    model = config["model"]
    api_key = os.getenv(config["api_key_env"])

    if not api_key:
        return {
            "provider": provider,
            "model": model,
            "question": question,
            "answer": f"ERREUR: Cle API {config['api_key_env']} non trouvee",
            "time_seconds": 0,
            "sources": sources,
            "error": True
        }

    system_prompt, user_prompt = build_prompt(question, context, rag_version)
    start_time = time.time()

    try:
        if config.get("use_anthropic"):
            answer = call_anthropic(api_key, model, user_prompt, system_prompt)
        else:
            answer = call_openai_compatible(api_key, model, user_prompt, system_prompt, config.get("base_url"))

        elapsed = time.time() - start_time
        return {
            "provider": provider,
            "model": model,
            "question": question,
            "answer": answer,
            "time_seconds": round(elapsed, 2),
            "sources": sources,
            "error": False
        }
    except Exception as e:
        elapsed = time.time() - start_time
        return {
            "provider": provider,
            "model": model,
            "question": question,
            "answer": f"ERREUR: {str(e)}",
            "time_seconds": round(elapsed, 2),
            "sources": sources,
            "error": True
        }


# ============================================================
# Agent Juge (Mistral Large)
# ============================================================

def call_judge(question: str, sources: List[Dict], responses: List[Dict]) -> Dict[str, Any]:
    """
    Appelle Mistral Large pour évaluer les réponses.
    Les réponses sont anonymisées (A, B, C...) pour éviter le biais.
    """
    api_key = os.getenv(JUDGE_API_KEY_ENV)
    if not api_key:
        return {"error": f"Clé API {JUDGE_API_KEY_ENV} non trouvée pour le juge"}

    # Construire le mapping anonymisé
    provider_to_letter = {}
    letter_to_provider = {}
    valid_responses = [r for r in responses if not r["error"]]
    error_responses = [r for r in responses if r["error"]]

    for i, r in enumerate(valid_responses):
        letter = LETTERS[i]
        provider_to_letter[r["provider"]] = letter
        letter_to_provider[letter] = r["provider"]

    if not valid_responses:
        return {
            "error": "Aucune réponse valide à évaluer",
            "evaluations": [],
            "classement": [],
            "meilleure_reponse": None,
            "synthese": "Aucune réponse valide.",
            "provider_mapping": {}
        }

    # Formater les sources pour le juge
    def _format_source_line(s):
        rank = s.get('rank', '?')
        label = s.get('commune', s.get('view', s.get('type', 'source')))
        extrait = s.get('extrait', '')
        return f"- Source {rank}: [{label}] {extrait}"

    sources_text = "\n".join([_format_source_line(s) for s in sources])

    # Formater les réponses anonymisées
    responses_text = "\n\n".join([
        f"=== RÉPONSE {provider_to_letter[r['provider']]} ===\n{r['answer']}"
        for r in valid_responses
    ])

    judge_system = """Tu es un évaluateur expert en analyse de données qualitatives sur le bien-être territorial.
Tu dois évaluer des réponses générées par différents modèles de langage à partir des mêmes sources.

Critères d'évaluation :
1. FIDÉLITÉ aux sources : la réponse s'appuie-t-elle sur les données fournies sans inventer ?
2. PERTINENCE : la réponse répond-elle bien à la question posée ?
3. COMPLÉTUDE : les informations clés des sources sont-elles exploitées ?
4. CLARTÉ : la réponse est-elle bien structurée et compréhensible ?
5. NUANCE : la réponse évite-t-elle les généralisations abusives ?

Tu DOIS répondre UNIQUEMENT en JSON valide, sans texte avant ni après."""

    judge_prompt = f"""=== QUESTION POSÉE ===
{question}

=== SOURCES RAG FOURNIES ===
{sources_text}

=== RÉPONSES À ÉVALUER ===
{responses_text}

Évalue chaque réponse selon les 5 critères et produis le JSON suivant :
{{
  "evaluations": [
    {{"id": "LETTRE", "note": NOTE_SUR_10, "justification": "Explication courte (2-3 phrases)"}}
  ],
  "classement": ["LETTRE_1er", "LETTRE_2e", ...],
  "meilleure_reponse": "LETTRE",
  "synthese": "Réponse idéale combinant les meilleurs éléments de chaque réponse (3-5 phrases)"
}}"""

    try:
        raw_response = call_openai_compatible(
            api_key, JUDGE_MODEL, judge_prompt, judge_system, JUDGE_BASE_URL
        )

        # Parser le JSON (gérer le cas où le juge met du texte autour)
        json_match = re.search(r'\{[\s\S]*\}', raw_response)
        if json_match:
            judgment = json.loads(json_match.group())
        else:
            judgment = json.loads(raw_response)

        # Rétablir le mapping provider
        judgment["provider_mapping"] = letter_to_provider
        judgment["error"] = None

        # Ajouter les providers en erreur
        judgment["providers_en_erreur"] = [
            {"provider": r["provider"], "erreur": r["answer"]}
            for r in error_responses
        ]

        return judgment

    except Exception as e:
        return {
            "error": f"Erreur juge: {str(e)}",
            "evaluations": [],
            "classement": [],
            "meilleure_reponse": None,
            "synthese": "",
            "provider_mapping": letter_to_provider,
            "providers_en_erreur": [
                {"provider": r["provider"], "erreur": r["answer"]}
                for r in error_responses
            ]
        }


# ============================================================
# Pipeline principale
# ============================================================

def run_pipeline(questions: List[str], rag_adapter: 'RAGAdapter',
                 max_questions: Optional[int] = None) -> tuple:
    """Execute la pipeline complete : RAG -> LLMs -> Juge."""
    if max_questions:
        questions = questions[:max_questions]

    print("=" * 70)
    print("PIPELINE MULTI-LLM AVEC AGENT JUGE")
    print(f"  {len(questions)} questions x {len(MODELS_CONFIG)} providers")
    print(f"  RAG : {rag_adapter.version}")
    print(f"  Juge : {JUDGE_MODEL}")
    print("=" * 70)

    # Initialisation du backend RAG
    print(f"\nInitialisation RAG ({rag_adapter.version})...")
    rag_adapter.init()
    print("OK")

    all_results = []       # Toutes les reponses LLM
    all_judgments = []      # Tous les jugements
    all_sources = {}        # Sources par question

    total_steps = len(questions) * (len(MODELS_CONFIG) + 1)  # +1 pour le juge
    current_step = 0

    for q_idx, question in enumerate(questions, 1):
        print(f"\n{'=' * 70}")
        print(f"QUESTION {q_idx}/{len(questions)}: {question[:70]}...")
        print("=" * 70)

        # Etape 1 : Contexte RAG
        print(f"  [RAG {rag_adapter.version}] Recuperation du contexte...", end=" ", flush=True)
        context, sources = rag_adapter.get_context(question)
        all_sources[question] = sources
        print(f"OK ({len(sources)} sources)")

        # Etape 2 : Interrogation des LLMs
        question_results = []
        for provider in MODELS_CONFIG:
            current_step += 1
            model_name = MODELS_CONFIG[provider]["model"]
            print(f"  [{current_step}/{total_steps}] {provider}/{model_name}...", end=" ", flush=True)

            result = test_model(provider, question, context, sources, rag_adapter.version)
            all_results.append(result)
            question_results.append(result)

            if result["error"]:
                print(f"ERREUR ({result['time_seconds']}s)")
            else:
                print(f"OK ({result['time_seconds']}s)")

        # Étape 3 : Agent Juge
        current_step += 1
        print(f"  [{current_step}/{total_steps}] JUGE ({JUDGE_MODEL})...", end=" ", flush=True)

        judgment = call_judge(question, sources, question_results)
        judgment["question"] = question

        if judgment.get("error"):
            print(f"ERREUR: {judgment['error']}")
        else:
            best_letter = judgment.get("meilleure_reponse", "?")
            best_provider = judgment.get("provider_mapping", {}).get(best_letter, "?")
            print(f"OK -> Meilleur: {best_provider} ({best_letter})")

        all_judgments.append(judgment)

    return all_results, all_judgments, all_sources


# ============================================================
# Export Excel
# ============================================================

def export_to_excel(results: List[Dict], judgments: List[Dict],
                    sources_by_question: Dict, output_path: str):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    print(f"\nExport vers {output_path}...")

    providers = list(MODELS_CONFIG.keys())

    # === Couleurs par provider ===
    PROVIDER_FILLS = {
        "OpenAI": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "Grok": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "Kimi": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "Claude": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "Mistral": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
    }
    SYNTHESE_FILL = PatternFill(start_color="D5A6E6", end_color="D5A6E6", fill_type="solid")
    QUESTION_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    QUESTION_FONT = Font(bold=True, color="FFFFFF", size=12)
    HEADER_FONT = Font(bold=True, size=10)
    WRAP = Alignment(wrap_text=True, vertical="top")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="999999"),
        right=Side(style="thin", color="CCCCCC"),
    )
    THICK_BORDER = Border(bottom=Side(style="medium", color="333333"))

    # ============================================================
    # Feuille principale : RECAP (une ligne par LLM par question)
    # ============================================================
    import openpyxl
    wb = openpyxl.Workbook()
    ws_recap = wb.active
    ws_recap.title = "Recap"

    # Colonnes : Provider | Modele | Note /10 | Rang | Reponse | Commentaire juge
    col_headers = ["Provider", "Modele", "Note /10", "Rang", "Reponse", "Commentaire du juge"]
    col_widths = [14, 28, 10, 8, 80, 50]

    for i, w in enumerate(col_widths, 1):
        ws_recap.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    row = 1
    for j in judgments:
        question = j["question"]
        mapping = j.get("provider_mapping", {})
        reverse_mapping = {v: k for k, v in mapping.items()}

        # --- Ligne question (merged) ---
        ws_recap.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(col_headers))
        cell = ws_recap.cell(row=row, column=1, value=question)
        cell.font = QUESTION_FONT
        cell.fill = QUESTION_FILL
        cell.alignment = Alignment(vertical="center")
        ws_recap.row_dimensions[row].height = 30
        row += 1

        # --- En-tetes colonnes ---
        for ci, h in enumerate(col_headers, 1):
            cell = ws_recap.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.border = THICK_BORDER
        ws_recap.row_dimensions[row].height = 20
        row += 1

        # --- Lignes par provider (triees par rang) ---
        q_results = [r for r in results if r["question"] == question]
        entries = []
        for r in q_results:
            provider = r["provider"]
            letter = reverse_mapping.get(provider)
            note = ""
            rang = ""
            justification = ""
            if letter:
                for ev in j.get("evaluations", []):
                    if ev["id"] == letter:
                        note = ev["note"]
                        justification = ev.get("justification", "")
                classement = j.get("classement", [])
                if letter in classement:
                    rang = classement.index(letter) + 1
            entries.append({
                "provider": provider,
                "model": r["model"],
                "note": note,
                "rang": rang if rang else 99,
                "answer": r["answer"] if not r["error"] else f"ERREUR: {r['answer']}",
                "justification": justification,
                "error": r["error"],
                "rang_display": rang if rang else "ERR"
            })
        entries.sort(key=lambda x: x["rang"] if isinstance(x["rang"], int) else 99)

        for entry in entries:
            values = [
                entry["provider"],
                entry["model"],
                entry["note"],
                entry["rang_display"],
                entry["answer"],
                entry["justification"]
            ]
            fill = PROVIDER_FILLS.get(entry["provider"])
            for ci, val in enumerate(values, 1):
                cell = ws_recap.cell(row=row, column=ci, value=val)
                cell.alignment = WRAP
                cell.border = THIN_BORDER
                if fill:
                    cell.fill = fill
            ws_recap.row_dimensions[row].height = 100
            row += 1

        # --- Ligne synthese du juge ---
        best_letter = j.get("meilleure_reponse")
        best_provider = mapping.get(best_letter, "?") if best_letter else "?"
        synthese = j.get("synthese", "")

        ws_recap.merge_cells(start_row=row, start_column=5, end_row=row, end_column=6)
        ws_recap.cell(row=row, column=1, value="JUGE")
        ws_recap.cell(row=row, column=2, value="mistral-large-latest")
        ws_recap.cell(row=row, column=3, value="SYNTHESE")
        ws_recap.cell(row=row, column=4, value=f"Best: {best_provider}")
        ws_recap.cell(row=row, column=5, value=synthese)
        for ci in range(1, len(col_headers) + 1):
            cell = ws_recap.cell(row=row, column=ci)
            cell.fill = SYNTHESE_FILL
            cell.alignment = WRAP
            cell.font = Font(italic=True)
            cell.border = THICK_BORDER
        ws_recap.row_dimensions[row].height = 80
        row += 1

        # Ligne vide de separation
        row += 1

    # ============================================================
    # Feuille 2 : Classement global
    # ============================================================
    ws_classement = wb.create_sheet("Classement_global")
    global_scores = {p: {"notes": [], "first_places": 0, "errors": 0} for p in providers}

    for j in judgments:
        mapping = j.get("provider_mapping", {})
        for ev in j.get("evaluations", []):
            provider = mapping.get(ev["id"])
            if provider:
                global_scores[provider]["notes"].append(ev["note"])
        best_letter = j.get("meilleure_reponse")
        if best_letter and best_letter in mapping:
            global_scores[mapping[best_letter]]["first_places"] += 1
        for err in j.get("providers_en_erreur", []):
            if err["provider"] in global_scores:
                global_scores[err["provider"]]["errors"] += 1

    classement_headers = ["Provider", "Modele", "Note moyenne", "Note min", "Note max",
                          "Nb 1eres places", "Nb erreurs", "Nb reponses"]
    for ci, h in enumerate(classement_headers, 1):
        cell = ws_classement.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT

    classement_rows = []
    for provider in providers:
        sc = global_scores[provider]
        notes = sc["notes"]
        classement_rows.append({
            "provider": provider,
            "model": MODELS_CONFIG[provider]["model"],
            "avg": round(sum(notes) / len(notes), 2) if notes else 0,
            "min": min(notes) if notes else "N/A",
            "max": max(notes) if notes else "N/A",
            "wins": sc["first_places"],
            "errors": sc["errors"],
            "count": len(notes)
        })
    classement_rows.sort(key=lambda x: x["avg"], reverse=True)

    for ri, cr in enumerate(classement_rows, 2):
        vals = [cr["provider"], cr["model"], cr["avg"], cr["min"], cr["max"],
                cr["wins"], cr["errors"], cr["count"]]
        fill = PROVIDER_FILLS.get(cr["provider"])
        for ci, val in enumerate(vals, 1):
            cell = ws_classement.cell(row=ri, column=ci, value=val)
            if fill:
                cell.fill = fill

    for col in ws_classement.columns:
        max_len = max(len(str(c.value or "")) for c in col)
        ws_classement.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

    # ============================================================
    # Sauvegarde
    # ============================================================
    wb.save(output_path)
    print(f"OK - Fichier exporte: {output_path}")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description="Pipeline Multi-LLM avec Agent Juge")
    parser.add_argument("--questions-file", default="questions_communes.json",
                        help="Fichier JSON contenant les questions (defaut: questions_communes.json)")
    parser.add_argument("--max-questions", type=int, default=None,
                        help="Nombre max de questions a traiter (pour tests)")
    parser.add_argument("--output-dir", default="comparaisons_rag",
                        help="Dossier de sortie (defaut: comparaisons_rag)")
    parser.add_argument("--rag-version", default="baseline",
                        choices=AVAILABLE_RAG_VERSIONS,
                        help=f"Version RAG a utiliser (defaut: baseline). "
                             f"Disponibles: {', '.join(AVAILABLE_RAG_VERSIONS)}")
    args = parser.parse_args()

    # Charger les questions
    with open(args.questions_file, "r", encoding="utf-8") as f:
        data = json.load(f)
    questions = data.get("questions", data) if isinstance(data, dict) else data
    print(f"Charge {len(questions)} questions depuis {args.questions_file}")

    # Creer l'adaptateur RAG
    rag = RAGAdapter(args.rag_version)

    # Executer la pipeline
    results, judgments, sources = run_pipeline(questions, rag, args.max_questions)

    # Export Excel
    # Liberer les ressources RAG
    rag.close()

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    os.makedirs(args.output_dir, exist_ok=True)
    output_path = os.path.join(args.output_dir,
                               f"pipeline_judge_{args.rag_version}_{timestamp}.xlsx")

    export_to_excel(results, judgments, sources, output_path)

    # Export JSON complet
    json_path = output_path.replace(".xlsx", ".json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "timestamp": timestamp,
            "questions_file": args.questions_file,
            "rag_version": args.rag_version,
            "nb_questions": len(questions),
            "models": {p: c["model"] for p, c in MODELS_CONFIG.items()},
            "judge_model": JUDGE_MODEL,
            "judgments": judgments,
            "results_summary": [
                {k: v for k, v in r.items() if k != "sources"}
                for r in results
            ]
        }, f, ensure_ascii=False, indent=2)
    print(f"JSON exporté: {json_path}")

    # Résumé final
    print("\n" + "=" * 70)
    print("RÉSUMÉ FINAL")
    print("=" * 70)

    # Classement
    provider_notes = {}
    provider_wins = {}
    for j in judgments:
        mapping = j.get("provider_mapping", {})
        for ev in j.get("evaluations", []):
            prov = mapping.get(ev["id"])
            if prov:
                provider_notes.setdefault(prov, []).append(ev["note"])
        best = j.get("meilleure_reponse")
        if best and best in mapping:
            prov = mapping[best]
            provider_wins[prov] = provider_wins.get(prov, 0) + 1

    print("\nClassement par note moyenne:")
    sorted_providers = sorted(
        provider_notes.items(),
        key=lambda x: sum(x[1]) / len(x[1]),
        reverse=True
    )
    for rank, (prov, notes) in enumerate(sorted_providers, 1):
        avg = sum(notes) / len(notes)
        wins = provider_wins.get(prov, 0)
        print(f"  {rank}. {prov}/{MODELS_CONFIG[prov]['model']}: "
              f"{avg:.2f}/10 (1ère place: {wins}x)")

    errors = [r for r in results if r["error"]]
    if errors:
        print(f"\nErreurs ({len(errors)}):")
        for e in errors:
            print(f"  - {e['provider']}/{e['model']}: {e['answer'][:60]}...")

    print(f"\nRésultats: {output_path}")
    print(f"JSON:      {json_path}")


if __name__ == "__main__":
    main()
