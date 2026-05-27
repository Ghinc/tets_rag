"""
run_agreement_sonnet.py — Accord humain-juge : Claude Sonnet + V4.3 sur 20 questions.

Lit annotation_editee_2026-05-12.xlsx (réponses v10, sources texte brut, scores humains),
applique le judge V4.3 via claude-sonnet-4-5-20250929, calcule Pearson / Spearman / MAE /
biais / exact% / ±1% par dimension + global, et sauvegarde les résultats.

Usage:
    python run_agreement_sonnet.py
    python run_agreement_sonnet.py --dry-run   # 3 premières questions seulement
"""
import argparse, json, re, time, os, sys
from pathlib import Path
from datetime import datetime

sys.stdout.reconfigure(encoding="utf-8")

import openpyxl, anthropic, dotenv
from scipy.stats import pearsonr, spearmanr
import numpy as np

dotenv.load_dotenv()

sys.path.insert(0, str(Path(__file__).parent))
from eval_from_excel import _JUDGE_V43_SYSTEM, _parse_judge_v43

ANNO_FILE    = r"C:\Users\comiti_g\Downloads\annotation_editee_2026-05-12.xlsx"
OUT_DIR      = Path("comparaisons_rag")
CLAUDE_MODEL = "claude-sonnet-4-5-20250929"

# Colonnes XLSX (1-based)
C_LIGNE        = 1
C_SECTION      = 2
C_SUBSECTION   = 3
C_QUESTION     = 4
C_ANSWER       = 5
C_SOURCES      = 6
C_H_PERT       = 7
C_H_FOND       = 11
C_H_NUAN       = 15
C_H_COHE       = 19
C_H_GLOBAL     = 23

DIMS = ["pertinence", "fondement_factuel", "nuance_incertitude", "coherence_qualiquanti", "score_global"]

api_key = os.getenv("ANTHROPIC_API_KEY") or os.getenv("CLAUDE_API_KEY")
_client = anthropic.Anthropic(api_key=api_key)


def _expected_type(section: str) -> str:
    s = (section or "").lower()
    if "limite" in s and "architect" in s:
        return "limite_architecturale"
    return "reponse_substantielle_attendue"


def call_claude(system: str, prompt: str, max_tokens: int = 3000) -> str:
    for attempt in range(5):
        try:
            msg = _client.messages.create(
                model=CLAUDE_MODEL,
                max_tokens=max_tokens,
                system=system,
                messages=[{"role": "user", "content": prompt}],
            )
            return msg.content[0].text.strip()
        except Exception as e:
            if attempt < 4 and ("529" in str(e) or "overloaded" in str(e).lower()
                                 or "rate" in str(e).lower()):
                wait = 2 ** attempt * 5
                print(f"    [RATE LIMIT] Attente {wait}s...", flush=True)
                time.sleep(wait)
            else:
                raise


def judge_v43(question: str, answer: str, sources_text: str,
              section: str, subsection: str, expected_type: str) -> dict:
    user_prompt = (
        f"QUESTION : {question}\n\n"
        f"SECTION : {section}\n\n"
        f"SOUS-SECTION : {subsection}\n\n"
        f"TYPE DE RÉPONSE ATTENDUE : {expected_type}\n\n"
        f"SOURCES FOURNIES AU SYSTÈME :\n{sources_text}\n\n"
        f"RÉPONSE DU SYSTÈME :\n{answer}\n\n"
        "Évalue cette réponse selon la procédure et le format spécifiés.\n"
        "Consulte les définitions opérationnelles et la grille AVANT de noter.\n"
        "Réponds UNIQUEMENT avec le JSON demandé, sans texte avant ni après."
    )
    try:
        raw = call_claude(_JUDGE_V43_SYSTEM, user_prompt)
        m = re.search(r'\{[\s\S]*\}', raw)
        j = json.loads(m.group()) if m else {}
        result = _parse_judge_v43(j)
        result["error"] = None
        return result
    except Exception as e:
        return {"error": str(e), "score_global": None}


def compute_metrics(human: list, judge: list) -> dict:
    """Pearson, Spearman, MAE, biais, exact%, ±1% sur paires valides."""
    pairs = [(h, j) for h, j in zip(human, judge) if h is not None and j is not None]
    if len(pairs) < 2:
        return {"n": len(pairs), "error": "insufficient data"}
    H = [p[0] for p in pairs]
    J = [p[1] for p in pairs]
    pr, pp = pearsonr(H, J)
    sr, sp = spearmanr(H, J)
    mae  = float(np.mean(np.abs(np.array(H) - np.array(J))))
    bias = float(np.mean(np.array(H) - np.array(J)))
    exact = sum(1 for h, j in pairs if abs(h - j) < 0.01) / len(pairs) * 100
    pm1   = sum(1 for h, j in pairs if abs(h - j) <= 1.0) / len(pairs) * 100
    return {
        "n": len(pairs),
        "pearson_r": round(float(pr), 3),
        "pearson_p": round(float(pp), 3),
        "spearman_r": round(float(sr), 3),
        "spearman_p": round(float(sp), 3),
        "mae": round(mae, 3),
        "bias_h_minus_j": round(bias, 3),
        "exact_pct": round(exact, 1),
        "within1_pct": round(pm1, 1),
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true", help="Test sur 3 premières questions")
    args = parser.parse_args()

    # ── Chargement annotation ────────────────────────────────────────────────
    wb = openpyxl.load_workbook(ANNO_FILE)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        rows.append({
            "excel_row":   ws.cell(r, C_LIGNE).value,
            "section":     ws.cell(r, C_SECTION).value or "",
            "subsection":  ws.cell(r, C_SUBSECTION).value or "",
            "question":    ws.cell(r, C_QUESTION).value or "",
            "answer":      ws.cell(r, C_ANSWER).value or "",
            "sources_text": ws.cell(r, C_SOURCES).value or "",
            "h_pertinence":           ws.cell(r, C_H_PERT).value,
            "h_fondement_factuel":    ws.cell(r, C_H_FOND).value,
            "h_nuance_incertitude":   ws.cell(r, C_H_NUAN).value,
            "h_coherence_qualiquanti": ws.cell(r, C_H_COHE).value,
            "h_score_global":         ws.cell(r, C_H_GLOBAL).value,
        })
    if args.dry_run:
        rows = rows[:3]
        print(f"[DRY-RUN] 3 questions seulement\n")

    print(f"Model : {CLAUDE_MODEL}")
    print(f"Judge : V4.3")
    print(f"N     : {len(rows)} questions\n")

    # ── Jugement ─────────────────────────────────────────────────────────────
    results = []
    for i, row in enumerate(rows, 1):
        etype = _expected_type(row["section"])
        print(f"  [{i:2}/{len(rows)}] R{row['excel_row']:3}  {row['section'][:35]:<35}", end="  ", flush=True)
        t0 = time.time()
        scores = judge_v43(
            question     = row["question"],
            answer       = row["answer"][:4000],
            sources_text = row["sources_text"][:8000],
            section      = row["section"],
            subsection   = row["subsection"],
            expected_type= etype,
        )
        elapsed = round(time.time() - t0, 1)

        sg    = scores.get("score_global")
        sg_h  = row["h_score_global"]
        delta = round(sg - sg_h, 2) if sg is not None and sg_h is not None else "N/A"
        print(f"J={sg}  H={sg_h}  Δ={delta}  ({elapsed}s)", flush=True)

        entry = {**row, "judge_model": CLAUDE_MODEL, "judge_version": "V4.3",
                 "elapsed_s": elapsed, **{f"j_{k}": v for k, v in scores.items()}}
        results.append(entry)
        time.sleep(0.5)

    # ── Sauvegarde JSON ──────────────────────────────────────────────────────
    OUT_DIR.mkdir(exist_ok=True)
    ts       = datetime.now().strftime("%Y%m%d_%H%M%S")
    json_out = OUT_DIR / f"judge_scores_sonnet_v4_3_{len(rows)}q_{ts}.json"
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    print(f"\nJSON → {json_out}")

    # ── Métriques accord humain-juge ─────────────────────────────────────────
    print(f"\n{'─'*70}")
    print(f"ACCORD HUMAIN ↔ JUGE V4.3 ({CLAUDE_MODEL})   N={len(rows)}")
    print(f"{'─'*70}")
    print(f"{'Dimension':<28}  {'Pearson':>8}  {'Spearman':>9}  {'MAE':>6}  {'Biais':>7}  {'Exact':>6}  {'±1':>5}")
    print(f"{'─'*70}")

    metrics_all = {}
    dim_map = {
        "pertinence":             ("h_pertinence",            "j_pertinence"),
        "fondement_factuel":      ("h_fondement_factuel",     "j_fondement_factuel"),
        "nuance_incertitude":     ("h_nuance_incertitude",    "j_nuance_incertitude"),
        "coherence_qualiquanti":  ("h_coherence_qualiquanti", "j_coherence_qualiquanti"),
        "score_global":           ("h_score_global",          "j_score_global"),
    }
    for dim, (hk, jk) in dim_map.items():
        human  = [r[hk] for r in results]
        judge_ = [r[jk] for r in results]
        m = compute_metrics(human, judge_)
        metrics_all[dim] = m
        pr  = f"{m['pearson_r']:+.3f}" if "pearson_r" in m else "-"
        sr  = f"{m['spearman_r']:+.3f}" if "spearman_r" in m else "-"
        mae = f"{m['mae']:.3f}" if "mae" in m else "-"
        bi  = f"{m['bias_h_minus_j']:+.3f}" if "bias_h_minus_j" in m else "-"
        ex  = f"{m['exact_pct']:.0f}%" if "exact_pct" in m else "-"
        p1  = f"{m['within1_pct']:.0f}%" if "within1_pct" in m else "-"
        print(f"  {dim:<26}  {pr:>8}  {sr:>9}  {mae:>6}  {bi:>7}  {ex:>6}  {p1:>5}")

    print(f"{'─'*70}")

    # ── Comparaison vs V4.1 ──────────────────────────────────────────────────
    print(f"\n--- COMPARAISON V4.1 (GPT-4o, N=20) vs V4.3 (Sonnet, N={len(rows)}) ---")
    v41_ref = {
        "pertinence":            {"pearson_r": 0.189, "mae": 0.550, "bias_h_minus_j": -0.150, "within1_pct": 90.0},
        "fondement_factuel":     {"pearson_r": 0.491, "mae": 0.800, "bias_h_minus_j": +0.800, "within1_pct": 85.0},
        "nuance_incertitude":    {"pearson_r": 0.131, "mae": 0.700, "bias_h_minus_j": +0.300, "within1_pct": 75.0},
        "coherence_qualiquanti": {"pearson_r": 0.024, "mae": 0.800, "bias_h_minus_j": +0.300, "within1_pct": 90.0},
        "score_global":          {"pearson_r": 0.404, "mae": 0.613, "bias_h_minus_j": +0.312, "within1_pct": 90.0},
    }
    print(f"{'Dimension':<28}  {'V4.1 r':>8}  {'V4.3 r':>8}  {'V4.1 MAE':>9}  {'V4.3 MAE':>9}  {'V4.1 Bias':>10}  {'V4.3 Bias':>10}")
    print(f"{'─'*100}")
    for dim in DIMS:
        m43 = metrics_all.get(dim, {})
        m41 = v41_ref.get(dim, {})
        r41 = f"{m41.get('pearson_r', float('nan')):+.3f}"
        r43 = f"{m43.get('pearson_r', float('nan')):+.3f}"
        mae41 = f"{m41.get('mae', float('nan')):.3f}"
        mae43 = f"{m43.get('mae', float('nan')):.3f}"
        b41 = f"{m41.get('bias_h_minus_j', float('nan')):+.3f}"
        b43 = f"{m43.get('bias_h_minus_j', float('nan')):+.3f}"
        print(f"  {dim:<26}  {r41:>8}  {r43:>8}  {mae41:>9}  {mae43:>9}  {b41:>10}  {b43:>10}")

    # ── Export métriques JSON ────────────────────────────────────────────────
    metrics_out = OUT_DIR / f"judge_agreement_sonnet_v4_3_{len(rows)}q_{ts}.json"
    with open(metrics_out, "w", encoding="utf-8") as f:
        json.dump({
            "model": CLAUDE_MODEL, "judge_version": "V4.3",
            "n_questions": len(rows), "timestamp": ts,
            "metrics": metrics_all,
            "v41_reference": v41_ref,
        }, f, ensure_ascii=False, indent=2)
    print(f"\nMétriques → {metrics_out}")


if __name__ == "__main__":
    main()
