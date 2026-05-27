"""
Crée annotation_humaine_20q_final_v3_avec_juge.xlsx :
- Même contenu que v2
- Colonnes interleaved : pour chaque dimension, note_humain | note_juge | justif_humain | justif_juge
- Le juge GPT-4o évalue chaque réponse RAG
"""

import openpyxl, os, sys, json, re, time, io
from openai import OpenAI
from dotenv import load_dotenv

load_dotenv()
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

SRC_PATH = r'C:\Users\comiti_g\Downloads\annotation_humaine_20q_final_v2.xlsx'
DST_PATH = r'C:\Users\comiti_g\Downloads\annotation_humaine_20q_final_v3_avec_juge.xlsx'

JUDGE_MODEL = "gpt-4o"
client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))

# Nouvelle structure colonnes :
# C1-C6  : inchangé (Ligne Excel, Section, Subsection, Question, Réponse RAG, Sources)
# C7     : pertinence humain       C8  : pertinence juge
# C9     : justif_pertinence h     C10 : justif_pertinence juge
# C11    : fondement_factuel h     C12 : fondement_factuel juge
# C13    : justif_fondement h      C14 : justif_fondement juge
# C15    : nuance_incertitude h    C16 : nuance_incertitude juge
# C17    : justif_nuance h         C18 : justif_nuance juge
# C19    : coherence_qualiquanti h C20 : coherence_qualiquanti juge
# C21    : justif_coherence h      C22 : justif_coherence juge
# C23    : score_global h (auto)   C24 : score_global juge (auto)
# C25    : applicable_sujet h      C26 : applicable_sujet juge
# C27    : note_sujet h            C28 : note_sujet juge
# C29    : sujet_evalue h          C30 : sujet_evalue juge
# C31    : justif_sujet h          C32 : justif_sujet juge

NEW_HEADERS = [
    'Ligne Excel', 'Section', 'Subsection', 'Question',
    'Réponse RAG (v10)', 'Sources RAG (texte brut)',
    'pertinence\n(humain 1-5)',        'pertinence\n(juge 1-5)',
    'justif_pertinence\n(humain)',      'justif_pertinence\n(juge)',
    'fondement_factuel\n(humain 1-5)', 'fondement_factuel\n(juge 1-5)',
    'justif_fondement\n(humain)',       'justif_fondement\n(juge)',
    'nuance_incertitude\n(humain 1-5)','nuance_incertitude\n(juge 1-5)',
    'justif_nuance\n(humain)',          'justif_nuance\n(juge)',
    'coherence_qualiquanti\n(humain 1-5)', 'coherence_qualiquanti\n(juge 1-5)',
    'justif_coherence\n(humain)',       'justif_coherence\n(juge)',
    'score_global\n(humain auto)',      'score_global\n(juge auto)',
    'applicable_sujet\n(humain)',       'applicable_sujet\n(juge)',
    'note_sujet\n(humain 1-5)',         'note_sujet\n(juge 1-5)',
    'sujet_evalue\n(humain)',           'sujet_evalue\n(juge)',
    'justif_sujet\n(humain)',           'justif_sujet\n(juge)',
]

# Mapping : ancien col dans v2 → nouveau col dans v3
# old C7(pertinence)→new C7, old C8(justif_pert)→new C9,
# old C9(fondement)→new C11, old C10(justif_fond)→new C13,
# old C11(nuance)→new C15,   old C12(justif_nuan)→new C17,
# old C13(coherence)→new C19, old C14(justif_coh)→new C21,
# old C16(applicable)→new C25, old C17(note_sujet)→new C27,
# old C18(sujet_eval)→new C29, old C19(justif_suj)→new C31
OLD_TO_NEW = {
    7: 7, 8: 9, 9: 11, 10: 13,
    11: 15, 12: 17, 13: 19, 14: 21,
    16: 25, 17: 27, 18: 29, 19: 31,
}


def parse_sources_for_judge(src_text):
    if not src_text:
        return []
    parts = re.split(r'--- Source \d+ \[.*?\] ---\n?', str(src_text))
    return [p.strip() for p in parts if p.strip()]


def call_judge(question, answer, sources_text, section):
    src_parts = parse_sources_for_judge(sources_text)
    sources_str = ""
    for i, p in enumerate(src_parts[:10], 1):
        sources_str += f"\n--- Source {i} ---\n{p[:600]}\n"

    system = (
        "Tu es un évaluateur expert en analyse territoriale et bien-être en Corse. "
        "Réponds UNIQUEMENT en JSON valide, sans texte avant ni après."
    )
    prompt = f"""=== QUESTION ===
{question}

=== SECTION THÉMATIQUE ===
{section}

=== SOURCES RAG (extraits) ===
{sources_str or "(aucune source fournie)"}

=== RÉPONSE DU SYSTÈME ===
{answer[:3000]}

---
Note chaque dimension de 1 à 5 ET justifie en 1 phrase.

- pertinence : la réponse traite bien la question posée
- fondement_factuel : les affirmations s'appuient sur les sources fournies
- nuance_incertitude : le modèle exprime ses limites de façon appropriée
- coherence_qualiquanti : équilibre quali/quanti adapté au type de question

Si la question porte sur une situation concrète et notable (commune, score, groupe),
note le sujet lui-même (applicable_sujet=true, note_sujet 1-5).
Si méthodologique, comparative multi-sujets, factuelle brute → applicable_sujet=false.

JSON :
{{
  "pertinence": N, "pertinence_justif": "1 phrase",
  "fondement_factuel": N, "fondement_factuel_justif": "1 phrase",
  "nuance_incertitude": N, "nuance_incertitude_justif": "1 phrase",
  "coherence_qualiquanti": N, "coherence_qualiquanti_justif": "1 phrase",
  "applicable_sujet": true/false,
  "note_sujet": N_ou_null,
  "sujet_evalue": "libellé court ou null",
  "justification_sujet": "1-2 phrases ou null"
}}"""

    for attempt in range(3):
        try:
            resp = client.chat.completions.create(
                model=JUDGE_MODEL,
                messages=[
                    {"role": "system", "content": system},
                    {"role": "user", "content": prompt},
                ],
                temperature=0.1,
                max_tokens=700,
            )
            raw = resp.choices[0].message.content.strip()
            m = re.search(r'\{[\s\S]*\}', raw)
            if not m:
                raise ValueError(f"Pas de JSON trouvé dans: {raw[:200]}")
            j = json.loads(m.group())
            dims = [j.get(k) for k in
                    ("pertinence", "fondement_factuel", "nuance_incertitude", "coherence_qualiquanti")]
            dims_ok = [d for d in dims if isinstance(d, (int, float))]
            j["score_global"] = round(sum(dims_ok) / len(dims_ok), 2) if dims_ok else None
            return j
        except Exception as e:
            if "429" in str(e) and attempt < 2:
                print(f"    [RATE LIMIT] attente 15s...", flush=True)
                time.sleep(15)
            else:
                print(f"    Erreur juge (tentative {attempt+1}): {e}", flush=True)
                if attempt == 2:
                    return {}
    return {}


def main():
    wb_src = openpyxl.load_workbook(SRC_PATH)
    ws_src = wb_src.active

    wb_dst = openpyxl.Workbook()
    ws_dst = wb_dst.active
    ws_dst.title = "Annotations"

    # En-têtes
    for c, h in enumerate(NEW_HEADERS, 1):
        cell = ws_dst.cell(1, c)
        cell.value = h
        cell.font = openpyxl.styles.Font(bold=True)
        cell.alignment = openpyxl.styles.Alignment(wrap_text=True)

    # Colorier colonnes juge (colonnes paires 8,10,12,...32) en bleu clair
    from openpyxl.styles import PatternFill
    judge_fill = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
    human_fill = PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid")
    judge_cols = {8, 10, 12, 14, 16, 18, 20, 22, 24, 26, 28, 30, 32}
    human_cols = {7, 9, 11, 13, 15, 17, 19, 21, 23, 25, 27, 29, 31}
    for c in judge_cols:
        ws_dst.cell(1, c).fill = judge_fill
    for c in human_cols:
        ws_dst.cell(1, c).fill = human_fill

    total_rows = sum(1 for r in range(2, ws_src.max_row + 1) if ws_src.cell(r, 4).value)
    done = 0

    for r_src in range(2, ws_src.max_row + 1):
        q = ws_src.cell(r_src, 4).value
        if not q:
            continue
        r_dst = r_src

        # C1-C6 copie directe
        for c in range(1, 7):
            ws_dst.cell(r_dst, c).value = ws_src.cell(r_src, c).value

        # Données humaines
        for old_c, new_c in OLD_TO_NEW.items():
            ws_dst.cell(r_dst, new_c).value = ws_src.cell(r_src, old_c).value

        # Formules score_global (G=7→C7 humain, H=8→C8 juge, K=11, L=12, O=15, P=16, S=19, T=20)
        ws_dst.cell(r_dst, 23).value = f'=IFERROR(AVERAGE(G{r_dst},K{r_dst},O{r_dst},S{r_dst}),"")'
        ws_dst.cell(r_dst, 24).value = f'=IFERROR(AVERAGE(H{r_dst},L{r_dst},P{r_dst},T{r_dst}),"")'

        # Appel juge
        section = str(ws_src.cell(r_src, 2).value or "")
        answer = str(ws_src.cell(r_src, 5).value or "")
        sources_text = str(ws_src.cell(r_src, 6).value or "")
        question_str = str(q).strip()

        done += 1
        print(f"[{done}/{total_rows}] R{r_dst}: {question_str[:60]}...", flush=True)
        j = call_judge(question_str, answer, sources_text, section)

        if j:
            ws_dst.cell(r_dst,  8).value = j.get("pertinence")
            ws_dst.cell(r_dst, 10).value = j.get("pertinence_justif")
            ws_dst.cell(r_dst, 12).value = j.get("fondement_factuel")
            ws_dst.cell(r_dst, 14).value = j.get("fondement_factuel_justif")
            ws_dst.cell(r_dst, 16).value = j.get("nuance_incertitude")
            ws_dst.cell(r_dst, 18).value = j.get("nuance_incertitude_justif")
            ws_dst.cell(r_dst, 20).value = j.get("coherence_qualiquanti")
            ws_dst.cell(r_dst, 22).value = j.get("coherence_qualiquanti_justif")
            applic = j.get("applicable_sujet")
            ws_dst.cell(r_dst, 26).value = "oui" if applic else "non"
            if applic:
                ws_dst.cell(r_dst, 28).value = j.get("note_sujet")
                ws_dst.cell(r_dst, 30).value = j.get("sujet_evalue")
                ws_dst.cell(r_dst, 32).value = j.get("justification_sujet")
            print(f"   -> score_global juge={j.get('score_global')}", flush=True)
        else:
            print("   -> juge ECHEC", flush=True)

        # Colorer les cellules juge de la ligne
        for c in judge_cols:
            ws_dst.cell(r_dst, c).fill = judge_fill

        time.sleep(2)

    # Largeurs colonnes
    ws_dst.column_dimensions['A'].width = 8
    ws_dst.column_dimensions['B'].width = 14
    ws_dst.column_dimensions['C'].width = 18
    ws_dst.column_dimensions['D'].width = 45
    ws_dst.column_dimensions['E'].width = 60
    ws_dst.column_dimensions['F'].width = 40
    for col_letter in ['G','H','K','L','O','P','S','T']:
        ws_dst.column_dimensions[col_letter].width = 12
    for col_letter in ['I','J','M','N','Q','R','U','V']:
        ws_dst.column_dimensions[col_letter].width = 35

    wb_dst.save(DST_PATH)
    print(f"\nFichier sauvegardé : {DST_PATH}")


if __name__ == "__main__":
    main()
