"""
compute_judge_agreement.py
Calcule l'accord humain / LLM-judge sur les annotations RAG (20 questions).
"""

import json
import logging
import sys
import io
import warnings
from pathlib import Path

import numpy as np
import pandas as pd
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from scipy.stats import pearsonr, spearmanr
import openpyxl

warnings.filterwarnings("ignore", category=RuntimeWarning)
logging.basicConfig(level=logging.INFO, format="%(levelname)s  %(message)s")
log = logging.getLogger(__name__)

# ── Chemins ─────────────────────────────────────────────────────────────────
EXCEL_PATH  = Path(r"C:\Users\comiti_g\Downloads\annotation_humaine_20q_final_v3_avec_juge.xlsx")
REPORT_PATH = Path("judge_agreement_report.md")
SCATTER_PATH = Path("judge_agreement_scatter.png")
JSON_PATH   = Path("judge_agreement_results.json")

# ── Mapping colonnes ─────────────────────────────────────────────────────────
# On lit par numéro de colonne (0-based après read_excel) pour éviter les \n
# C1=0, C4=3(Question), C7=6(pert_h), C8=7(pert_j), C11=10, C12=11,
# C15=14, C16=15, C19=18, C20=19
COL_IDX = {
    "question":            3,   # C4
    "pertinence_h":        6,   # C7
    "pertinence_j":        7,   # C8
    "fondement_factuel_h": 10,  # C11
    "fondement_factuel_j": 11,  # C12
    "nuance_incertitude_h":14,  # C15
    "nuance_incertitude_j":15,  # C16
    "coherence_qualiquanti_h":18, # C19
    "coherence_qualiquanti_j":19, # C20
    "ligne_excel":         0,   # C1
}

DIMS = ["pertinence", "fondement_factuel", "nuance_incertitude", "coherence_qualiquanti"]


# ════════════════════════════════════════════════════════════════════════════
# Tests unitaires basiques
# ════════════════════════════════════════════════════════════════════════════

def _run_unit_tests():
    a = np.array([1, 2, 3, 4, 5], dtype=float)

    # Accord parfait
    r, _ = pearsonr(a, a)
    assert abs(r - 1.0) < 1e-9, "Pearson parfait doit être 1"
    mae = float(np.abs(a - a).mean())
    assert mae == 0.0, "MAE parfaite doit être 0"
    exact = float((a == a).mean() * 100)
    assert exact == 100.0, "Accord exact parfait = 100%"

    # Désaccord total (décalage de 3)
    b = np.array([4, 5, 1, 2, 3], dtype=float)
    mae_bad = float(np.abs(a - b).mean())
    assert mae_bad > 1.5, "MAE désaccord total doit être > 1.5"

    log.info("Tests unitaires : OK")


# ════════════════════════════════════════════════════════════════════════════
# Étape 1 — Chargement
# ════════════════════════════════════════════════════════════════════════════

def load_data() -> pd.DataFrame:
    log.info(f"Lecture : {EXCEL_PATH}")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    rows = []
    for r in range(2, ws.max_row + 1):
        q = ws.cell(r, 4).value           # C4 Question
        if not q:
            continue
        row = {
            "ligne_excel": ws.cell(r, 1).value,
            "question":    str(q).strip(),
        }
        for dim in DIMS:
            row[f"{dim}_h"] = ws.cell(r, COL_IDX[f"{dim}_h"] + 1).value
            row[f"{dim}_j"] = ws.cell(r, COL_IDX[f"{dim}_j"] + 1).value
        rows.append(row)

    df = pd.DataFrame(rows)

    # Convertir en float
    score_cols = [f"{d}_{s}" for d in DIMS for s in ("h", "j")]
    df[score_cols] = df[score_cols].apply(pd.to_numeric, errors="coerce")

    # Score global (moyenne 4 dims, recalculé en Python)
    df["score_global_h"] = df[[f"{d}_h" for d in DIMS]].mean(axis=1)
    df["score_global_j"] = df[[f"{d}_j" for d in DIMS]].mean(axis=1)

    # Vérifier les NaN
    missing_mask = df[score_cols].isna().any(axis=1)
    if missing_mask.any():
        log.warning(f"{missing_mask.sum()} ligne(s) avec valeurs manquantes — exclues : "
                    f"{df.loc[missing_mask, 'ligne_excel'].tolist()}")
        df = df[~missing_mask].reset_index(drop=True)

    log.info(f"{len(df)} lignes valides chargées.")
    return df


# ════════════════════════════════════════════════════════════════════════════
# Étape 2+3 — Métriques par dimension
# ════════════════════════════════════════════════════════════════════════════

def compute_metrics(h: pd.Series, j: pd.Series) -> dict:
    h, j = h.astype(float), j.astype(float)
    r, p_r  = pearsonr(h, j)
    rho, p_s = spearmanr(h, j)
    mae     = float((h - j).abs().mean())
    bias    = float((h - j).mean())
    exact   = float((h == j).mean() * 100)
    w1      = float(((h - j).abs() <= 1).mean() * 100)
    return dict(
        pearson=round(r, 4), p_pearson=round(p_r, 4),
        spearman=round(rho, 4), p_spearman=round(p_s, 4),
        mae=round(mae, 3),
        bias=round(bias, 3),
        exact_pct=round(exact, 1),
        within_1_pct=round(w1, 1),
    )


def compute_all_metrics(df: pd.DataFrame) -> dict:
    results = {}
    all_dims = DIMS + ["score_global"]
    for dim in all_dims:
        results[dim] = compute_metrics(df[f"{dim}_h"], df[f"{dim}_j"])
    return results


# ════════════════════════════════════════════════════════════════════════════
# Étape 4 — Rapport Markdown
# ════════════════════════════════════════════════════════════════════════════

def _p_fmt(p: float) -> str:
    if p < 0.001: return "p<0.001"
    if p < 0.01:  return "p<0.01"
    if p < 0.05:  return "p<0.05"
    return f"p={p:.3f}"

def _bias_phrase(b: float) -> str:
    if abs(b) < 0.05: return "pas de biais notable"
    direction = "l'humain note plus haut" if b > 0 else "le juge note plus haut"
    return f"biais de {b:+.2f} ({direction})"

def build_report(df: pd.DataFrame, metrics: dict, outliers: list) -> str:
    n = len(df)
    lines = [
        "# Rapport d'accord humain / LLM-judge\n",
        f"**N = {n} questions** | Modèle juge : GPT-4o | Échelle : 1–5\n",
        "\n---\n",
        "## Table principale\n",
        "| Dimension | Pearson | p | Spearman | MAE | Biais | Exact % | ±1 % |",
        "|-----------|---------|---|----------|-----|-------|---------|-------|",
    ]
    all_dims = DIMS + ["score_global"]
    dim_labels = {
        "pertinence":            "Pertinence",
        "fondement_factuel":     "Fondement factuel",
        "nuance_incertitude":    "Nuance / incertitude",
        "coherence_qualiquanti": "Cohérence quali-quanti",
        "score_global":          "**Score global**",
    }
    for dim in all_dims:
        m = metrics[dim]
        lines.append(
            f"| {dim_labels[dim]} "
            f"| {m['pearson']:.3f} | {_p_fmt(m['p_pearson'])} "
            f"| {m['spearman']:.3f} "
            f"| {m['mae']:.3f} "
            f"| {m['bias']:+.3f} "
            f"| {m['exact_pct']:.0f}% "
            f"| {m['within_1_pct']:.0f}% |"
        )

    lines += ["\n\n## Interprétation par dimension\n"]
    for dim in all_dims:
        m = metrics[dim]
        label = dim_labels[dim].replace("**","")
        r_interp = (
            "accord fort" if m["pearson"] >= 0.7
            else "accord modéré" if m["pearson"] >= 0.4
            else "accord faible"
        )
        lines.append(
            f"- **{label}** : Pearson={m['pearson']:.3f} ({_p_fmt(m['p_pearson'])}) — {r_interp}. "
            f"MAE={m['mae']:.3f} sur échelle 1–5. "
            f"{_bias_phrase(m['bias'])}. "
            f"Accord à ±1 : {m['within_1_pct']:.0f}%."
        )

    # Conclusion globale
    sg = metrics["score_global"]
    lines += [
        "\n\n## Conclusion globale\n",
        "> L'accord entre annotations humaines (N={n}) et notes du LLM-judge "
        "sur le **score global** est de Pearson={r:.3f} ({p}), "
        "MAE={mae:.3f} sur une échelle 1–5, "
        "avec {w1:.0f}% d'accord à ±1 point près. "
        "Ces résultats {valid} l'usage du LLM-as-judge pour l'évaluation à grande échelle.".format(
            n=n,
            r=sg["pearson"], p=_p_fmt(sg["p_pearson"]),
            mae=sg["mae"], w1=sg["within_1_pct"],
            valid="valident" if sg["within_1_pct"] >= 80 else "nuancent",
        ),
        "\n",
    ]

    # Outliers
    lines += ["\n\n## Diagnostic des outliers (|écart score global| > 1)\n"]
    if outliers:
        lines.append("| # | Question (100 cars) | Score humain | Score juge | Pire dimension | Écart |")
        lines.append("|---|---------------------|-------------|-----------|---------------|-------|")
        for o in outliers:
            lines.append(
                f"| L{o['ligne']} "
                f"| {o['question'][:100]} "
                f"| {o['score_h']:.2f} "
                f"| {o['score_j']:.2f} "
                f"| {o['worst_dim']} "
                f"| {o['abs_diff']:.2f} |"
            )
    else:
        lines.append("*Aucun outlier (|écart| ≤ 1 sur tous les exemples).*")

    # Moyennes par dimension
    lines += ["\n\n## Moyennes par dimension\n",
              "| Dimension | Moy. humain | Moy. juge | Δ |",
              "|-----------|-------------|-----------|---|"]
    for dim in all_dims:
        mh = df[f"{dim}_h"].mean()
        mj = df[f"{dim}_j"].mean()
        lines.append(f"| {dim_labels[dim].replace('**','')} | {mh:.2f} | {mj:.2f} | {mh-mj:+.2f} |")

    return "\n".join(lines) + "\n"


# ════════════════════════════════════════════════════════════════════════════
# Étape 5 — Scatter plot
# ════════════════════════════════════════════════════════════════════════════

def plot_scatter(df: pd.DataFrame, metrics: dict):
    all_dims = DIMS + ["score_global"]
    dim_labels = {
        "pertinence":            "Pertinence",
        "fondement_factuel":     "Fondement factuel",
        "nuance_incertitude":    "Nuance",
        "coherence_qualiquanti": "Cohérence quali-quanti",
        "score_global":          "Score global",
    }
    fig, axes = plt.subplots(2, 3, figsize=(14, 9))
    axes_flat = axes.flatten()

    for i, dim in enumerate(all_dims):
        ax = axes_flat[i]
        h = df[f"{dim}_h"].values
        j = df[f"{dim}_j"].values
        m = metrics[dim]

        ax.scatter(h, j, alpha=0.7, s=60, color="#2563eb", zorder=3)

        # Ligne y = x
        lo = min(h.min(), j.min()) - 0.2
        hi = max(h.max(), j.max()) + 0.2
        ax.plot([lo, hi], [lo, hi], "k--", lw=1, alpha=0.5, label="y = x")

        # Numéroter les points
        for idx, (xi, yi) in enumerate(zip(h, j)):
            ax.annotate(str(int(df.iloc[idx]["ligne_excel"])),
                        (xi, yi), fontsize=6, alpha=0.6,
                        xytext=(3, 3), textcoords="offset points")

        ax.set_xlabel("Note humaine", fontsize=9)
        ax.set_ylabel("Note juge", fontsize=9)
        ax.set_title(
            f"{dim_labels[dim]}\nPearson={m['pearson']:.3f}  MAE={m['mae']:.3f}",
            fontsize=10
        )
        ax.set_xlim(lo, hi)
        ax.set_ylim(lo, hi)
        ax.grid(True, alpha=0.3)

    # Cacher la 6e case (2x3 = 6, on a 5 dimensions)
    axes_flat[5].set_visible(False)

    fig.suptitle("Accord humain / LLM-judge — 20 questions RAG", fontsize=13, fontweight="bold")
    plt.tight_layout()
    plt.savefig(SCATTER_PATH, dpi=150)
    plt.close()
    log.info(f"Scatter sauvegardé : {SCATTER_PATH}")


# ════════════════════════════════════════════════════════════════════════════
# Main
# ════════════════════════════════════════════════════════════════════════════

def main():
    _run_unit_tests()

    # Étape 1
    df = load_data()
    print("\n─── Étape 1 : DataFrame préparé ───")
    display_cols = ["ligne_excel", "question"] + \
                   [f"{d}_{s}" for d in DIMS for s in ("h","j")] + \
                   ["score_global_h", "score_global_j"]
    with pd.option_context("display.max_colwidth", 50, "display.width", 160):
        print(df[display_cols].to_string(index=False))

    print("\nMoyennes :")
    for dim in DIMS + ["score_global"]:
        mh = df[f"{dim}_h"].mean()
        mj = df[f"{dim}_j"].mean()
        print(f"  {dim:30s}  humain={mh:.2f}  juge={mj:.2f}  Δ={mh-mj:+.2f}")

    # Étapes 2+3
    print("\n─── Étapes 2+3 : Métriques d'accord ───")
    metrics = compute_all_metrics(df)
    for dim, m in metrics.items():
        print(f"  {dim:30s}  "
              f"Pearson={m['pearson']:.3f}({_p_fmt(m['p_pearson'])})  "
              f"Spearman={m['spearman']:.3f}  "
              f"MAE={m['mae']:.3f}  bias={m['bias']:+.3f}  "
              f"exact={m['exact_pct']:.0f}%  ±1={m['within_1_pct']:.0f}%")

    # Outliers
    df["abs_diff_global"] = (df["score_global_h"] - df["score_global_j"]).abs()
    threshold = 1.0
    outlier_rows = df[df["abs_diff_global"] > threshold].copy()
    outliers = []
    for _, row in outlier_rows.iterrows():
        dim_diffs = {d: abs(row[f"{d}_h"] - row[f"{d}_j"]) for d in DIMS}
        worst = max(dim_diffs, key=dim_diffs.get)
        outliers.append({
            "ligne":     int(row["ligne_excel"]),
            "question":  row["question"][:100],
            "score_h":   round(float(row["score_global_h"]), 2),
            "score_j":   round(float(row["score_global_j"]), 2),
            "abs_diff":  round(float(row["abs_diff_global"]), 2),
            "worst_dim": worst,
        })
    log.info(f"{len(outliers)} outlier(s) (|écart global| > {threshold})")

    # Étape 4 — Rapport
    report = build_report(df, metrics, outliers)
    REPORT_PATH.write_text(report, encoding="utf-8")
    log.info(f"Rapport sauvegardé : {REPORT_PATH}")
    print("\n─── Rapport ───")
    print(report)

    # Étape 5 — Scatter
    plot_scatter(df, metrics)

    # Étape 6 — JSON
    result_json = {
        "n_samples": len(df),
        "dimensions": {
            dim: {**metrics[dim],
                  "mean_h": round(float(df[f"{dim}_h"].mean()), 3),
                  "mean_j": round(float(df[f"{dim}_j"].mean()), 3)}
            for dim in DIMS + ["score_global"]
        },
        "outliers": outliers,
    }
    JSON_PATH.write_text(json.dumps(result_json, ensure_ascii=False, indent=2), encoding="utf-8")
    log.info(f"JSON sauvegardé : {JSON_PATH}")


if __name__ == "__main__":
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")
    main()
