"""
Script de comparaison des configurations RAG avec différentes ontologies

Compare RAG v1 et RAG v2 avec 3 configurations d'ontologie :
1. Données seules (sans ontologie)
2. Données + ontology_be_2010.ttl
3. Données + ontology_be_2010.ttl + onto_be_instances.ttl

Total: 6 systèmes à comparer (2 RAG x 3 configs)

Questions de test (CQ - Competency Questions):
- "quelles sont les dimensions du bien-être subjectif mentionnés par les habitants de la commune de Grossetto Prugna"
- "quels indicateurs subjectifs sont disponibles pour grossetto prugna et que mesurent-ils"
- "au sein de la commune de Luri, les sources qualitatives et quantitatives s'accordent-elles sur les différents sujets abordés ?"

Sortie: Fichier Excel comparable à compare_rag_systems.py
"""

import os
import sys
import json
import time
from datetime import datetime
from typing import List, Dict, Tuple, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import io
import contextlib

# Imports des systèmes RAG
import chromadb
from sentence_transformers import SentenceTransformer
import openai
from openai import OpenAI

# Configuration de la clé API OpenAI depuis les variables d'environnement
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("❌ Clé API OpenAI non trouvée. Définissez la variable d'environnement OPENAI_API_KEY")
openai.api_key = OPENAI_API_KEY


# ============================================================================
# CONFIGURATION DES ONTOLOGIES
# ============================================================================

class OntologyConfig:
    """Configuration d'ontologie"""
    def __init__(self, name: str, base_ontology: Optional[str] = None,
                 instances_ontology: Optional[str] = None):
        self.name = name
        self.base_ontology = base_ontology
        self.instances_ontology = instances_ontology

    def has_ontology(self) -> bool:
        """Vérifie si cette config utilise une ontologie"""
        return self.base_ontology is not None

    def __str__(self):
        return self.name


# Configurations à tester
ONTOLOGY_CONFIGS = [
    OntologyConfig("No Ontology", None, None),
    OntologyConfig("Base Ontology", "ontology_be_2010_bilingue_fr_en.ttl", None),
    OntologyConfig("Base + Instances", "ontology_be_2010_bilingue_fr_en.ttl", "onto_be_instances.ttl")
]


# ============================================================================
# WRAPPER POUR RAG V1 AVEC ONTOLOGIE
# ============================================================================

class RAGv1WithOntology:
    """
    Wrapper pour RAG v1 avec support optionnel de l'ontologie
    """

    def __init__(self, ontology_config: OntologyConfig,
                 collection_name: str = "communes_corses_txt"):
        self.config = ontology_config
        self.collection_name = collection_name

        # Charger l'ontologie si 
        self.query_enricher = None
        if self.config.has_ontology():
            from ontology_parser import OntologyParser
            from query_enricher import QueryEnricher

            print(f"  Chargement ontologie: {self.config.base_ontology}")
            parser = OntologyParser(self.config.base_ontology)
            parser.extract_all()

            # Si on a les instances, on pourrait les charger ici
            # TODO: Implémenter le chargement des instances si nécessaire

            self.query_enricher = QueryEnricher(parser)

    def query(self, question: str, n_chunks: int = 5) -> Dict:
        """
        Interroge RAG v1 avec enrichissement optionnel par ontologie
        """
        start_time = time.time()

        try:
            # Enrichir la question si ontologie disponible
            query_to_use = question
            ontology_metadata = None

            if self.query_enricher:
                enrichment = self.query_enricher.enrich_query(question)
                query_to_use = enrichment["enriched_query"]
                ontology_metadata = enrichment["metadata"]

            # Charger la collection ChromaDB
            chroma_client = chromadb.PersistentClient(path="./chroma_txt")
            collection = chroma_client.get_collection(name=self.collection_name)

            # Encoder la question (enrichie)
            embed_model = SentenceTransformer("intfloat/e5-base-v2")
            query_embedding = embed_model.encode([f"query: {query_to_use}"]).tolist()

            # Recherche
            results = collection.query(
                query_embeddings=query_embedding,
                n_results=n_chunks
            )

            # Récupérer les documents
            retrieved_docs = results["documents"][0]
            retrieved_metadata = results["metadatas"][0]
            retrieved_context = "\n\n".join(retrieved_docs)

            # Construire le prompt
            prompt = f"""Tu es un conseiller municipal. Ton but est de donner des informations sur la qualité de vie dans les communes Corses, pour guider les politiques publiques, en te basant uniquement sur les informations suivantes :

{retrieved_context}

Question : {question}
Réponse :
"""

            # Appeler OpenAI
            client = OpenAI(api_key=OPENAI_API_KEY)

            response = client.chat.completions.create(
                model="gpt-3.5-turbo",
                messages=[
                    {"role": "system", "content": "Tu es un assistant utile et factuel. Ne réponds qu'avec les informations données."},
                    {"role": "user", "content": prompt}
                ]
            )

            answer = response.choices[0].message.content

            execution_time = time.time() - start_time

            return {
                "version": f"RAG v1 - {self.config.name}",
                "answer": answer,
                "retrieved_chunks": retrieved_docs,
                "metadata": retrieved_metadata,
                "execution_time": execution_time,
                "num_chunks": len(retrieved_docs),
                "ontology_used": self.config.has_ontology(),
                "ontology_metadata": ontology_metadata
            }

        except Exception as e:
            return {
                "version": f"RAG v1 - {self.config.name}",
                "answer": f"Erreur: {str(e)}",
                "error": str(e),
                "execution_time": time.time() - start_time,
                "ontology_used": self.config.has_ontology()
            }


# ============================================================================
# WRAPPER POUR RAG V2 AVEC ONTOLOGIE
# ============================================================================

class RAGv2WithOntology:
    """
    Wrapper pour RAG v2 avec support optionnel de l'ontologie
    """

    def __init__(self, ontology_config: OntologyConfig,
                 collection_name: str = "communes_corses_v2"):
        self.config = ontology_config
        self.collection_name = collection_name

        # Charger l'ontologie si nécessaire
        self.query_enricher = None
        if self.config.has_ontology():
            from ontology_parser import OntologyParser
            from query_enricher import QueryEnricher

            print(f"  Chargement ontologie: {self.config.base_ontology}")
            parser = OntologyParser(self.config.base_ontology)
            parser.extract_all()

            self.query_enricher = QueryEnricher(parser)

        # Initialiser le pipeline RAG v2
        from rag_v2_improved import ImprovedRAGPipeline

        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            self.rag = ImprovedRAGPipeline(
                chroma_path="./chroma_v2",
                collection_name=self.collection_name,
                embedding_model="dangvantuan/sentence-camembert-large",
                reranker_model="antoinelouis/crossencoder-camembert-base-mmarcoFR",
                llm_model="gpt-3.5-turbo",
                openai_api_key=OPENAI_API_KEY
            )

    def query(self, question: str, k: int = 5) -> Dict:
        """
        Interroge RAG v2 avec enrichissement optionnel par ontologie
        """
        start_time = time.time()

        try:
            # Enrichir la question si ontologie disponible
            query_to_use = question
            ontology_metadata = None

            if self.query_enricher:
                enrichment = self.query_enricher.enrich_query(question)
                query_to_use = enrichment["enriched_query"]
                ontology_metadata = enrichment["metadata"]

            # Vérifier que le pipeline a des documents
            if not self.rag.documents or self.rag.hybrid_retriever is None:
                return {
                    "version": f"RAG v2 - {self.config.name}",
                    "answer": "Le pipeline RAG v2 n'a pas de documents indexés.",
                    "error": "No documents indexed",
                    "retrieved_chunks": [],
                    "execution_time": time.time() - start_time,
                    "ontology_used": self.config.has_ontology()
                }

            # Encoder la requête (enrichie)
            query_embedding = self.rag.embed_model.encode_query(query_to_use)

            # Retrieval hybride
            results = self.rag.hybrid_retriever.retrieve(
                query_to_use,
                query_embedding,
                k=k*2
            )

            # Reranking (avec la question originale)
            results = self.rag.reranker.rerank(question, results, top_k=k)

            # Récupérer données quantitatives si disponibles
            commune = None
            if results:
                commune = results[0].metadata.get('nom') or results[0].metadata.get('commune')

            quant_data = None
            stats = None
            if commune:
                quant_data = self.rag.quant_handler.query_structured_data(commune=commune)
                if not quant_data.empty:
                    stats = self.rag.quant_handler.extract_statistics(quant_data)

            # Construire le prompt
            from rag_v2_improved import ImprovedPromptBuilder
            prompt = ImprovedPromptBuilder.build_rag_prompt(
                question, results, quant_data, stats
            )

            # Générer la réponse
            response = self.rag._generate_response(prompt)

            execution_time = time.time() - start_time

            return {
                "version": f"RAG v2 - {self.config.name}",
                "answer": response,
                "retrieved_chunks": [r.text for r in results],
                "metadata": [r.metadata for r in results],
                "scores": [r.score for r in results],
                "execution_time": execution_time,
                "num_chunks": len(results),
                "ontology_used": self.config.has_ontology(),
                "ontology_metadata": ontology_metadata
            }

        except Exception as e:
            return {
                "version": f"RAG v2 - {self.config.name}",
                "answer": f"Erreur: {str(e)}",
                "error": str(e),
                "execution_time": time.time() - start_time,
                "ontology_used": self.config.has_ontology()
            }


# ============================================================================
# FONCTION DE COMPARAISON
# ============================================================================

def compare_all_configurations(question: str) -> List[Dict]:
    """
    Compare toutes les configurations pour une question donnée

    Args:
        question: Question à tester

    Returns:
        Liste de résultats pour chaque configuration
    """
    results = []

    print(f"\n{'='*80}")
    print(f"Question: {question}")
    print(f"{'='*80}")

    # Tester chaque configuration
    for config in ONTOLOGY_CONFIGS:
        print(f"\n--- Testing: RAG v1 - {config.name} ---")
        try:
            rag_v1 = RAGv1WithOntology(config)
            result_v1 = rag_v1.query(question)
            results.append(result_v1)
            print(f"  OK Termine en {result_v1['execution_time']:.2f}s")
        except Exception as e:
            print(f"  ERREUR: {e}")
            results.append({
                "version": f"RAG v1 - {config.name}",
                "answer": f"Erreur: {str(e)}",
                "error": str(e),
                "execution_time": 0
            })

        print(f"\n--- Testing: RAG v2 - {config.name} ---")
        try:
            rag_v2 = RAGv2WithOntology(config)
            result_v2 = rag_v2.query(question)
            results.append(result_v2)
            print(f"  OK Termine en {result_v2['execution_time']:.2f}s")
        except Exception as e:
            print(f"  ERREUR: {e}")
            results.append({
                "version": f"RAG v2 - {config.name}",
                "answer": f"Erreur: {str(e)}",
                "error": str(e),
                "execution_time": 0
            })

    return results


# ============================================================================
# EXPORT VERS EXCEL
# ============================================================================

def export_to_excel(all_results: Dict[str, List[Dict]], output_file: str):
    """
    Exporte les résultats vers Excel

    Args:
        all_results: Dict {question: [results]}
        output_file: Chemin du fichier Excel
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparaison RAG Ontology"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    config_colors = {
        "No Ontology": "E2EFDA",
        "Base Ontology": "FCE4D6",
        "Base + Instances": "FFF2CC"
    }

    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    # En-têtes de colonnes
    headers = ["Question"]

    # Pour chaque configuration
    for config in ONTOLOGY_CONFIGS:
        for rag_version in ["v1", "v2"]:
            system_name = f"RAG {rag_version} - {config.name}"
            headers.extend([
                f"{system_name}\nRéponse",
                f"{system_name}\nTemps (s)",
                f"{system_name}\nChunks"
            ])

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    ws.row_dimensions[1].height = 40

    # Remplir les données
    row_idx = 2
    for question, results in all_results.items():
        col_idx = 1

        # Question
        cell = ws.cell(row=row_idx, column=col_idx, value=question)
        cell.alignment = wrap_alignment
        col_idx += 1

        # Résultats pour chaque système
        for result in results:
            config_name = result['version'].split(' - ')[1] if ' - ' in result['version'] else "No Ontology"
            fill_color = config_colors.get(config_name, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")

            # Réponse
            cell = ws.cell(row=row_idx, column=col_idx, value=result.get('answer', ''))
            cell.fill = fill
            cell.alignment = wrap_alignment
            col_idx += 1

            # Temps
            cell = ws.cell(row=row_idx, column=col_idx, value=f"{result.get('execution_time', 0):.2f}")
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

            # Nombre de chunks
            cell = ws.cell(row=row_idx, column=col_idx, value=result.get('num_chunks', 0))
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

        row_idx += 1

    # Ajuster les largeurs de colonnes
    ws.column_dimensions['A'].width = 60  # Question
    for col_idx in range(2, len(headers) + 1):
        col_letter = ws.cell(row=1, column=col_idx).column_letter
        if (col_idx - 2) % 3 == 0:  # Réponses
            ws.column_dimensions[col_letter].width = 50
        else:  # Temps et chunks
            ws.column_dimensions[col_letter].width = 12

    # Sauvegarder
    wb.save(output_file)
    print(f"\nOK Fichier Excel cree: {output_file}")


# ============================================================================
# FONCTION PRINCIPALE
# ============================================================================

def main():
    """
    Fonction principale
    """
    print("="*80)
    print("COMPARAISON DES CONFIGURATIONS RAG AVEC ONTOLOGIE")
    print("="*80)
    print(f"\nConfigurations testées:")
    for i, config in enumerate(ONTOLOGY_CONFIGS, 1):
        print(f"  {i}. {config.name}")
        if config.base_ontology:
            print(f"     - Base: {config.base_ontology}")
        if config.instances_ontology:
            print(f"     - Instances: {config.instances_ontology}")

    print(f"\nVersions RAG: v1, v2")
    print(f"Total de systèmes: {len(ONTOLOGY_CONFIGS) * 2}")

    # Questions CQ spécifiques
    test_questions = [
        "Quelles sont les dimensions du bien-être subjectif mentionnés par les habitants de la commune de Grossetto Prugna ?",
        "quels indicateurs subjectifs sont disponibles pour grossetto prugna et que mesurent-ils ?",
        "au sein de la commune de Luri, les sources qualitatives et quantitatives s'accordent-elles sur les différents sujets abordés ?", 
        "Quelle est la différence entre bien-être subjectif et objectif ?",
        "Quels sont les divergences et les convergences entre les mesures de bine-être subjectives et objectives à Ajaccio ?",
        "De quelles données disposons-nous pour évaluer le bien-être subjectif à Ajaccio ? ",
        "Comment améliorer le bien-être à Ajaccio en relation avec les retours des habitants ?"
    ]

    print(f"\nQuestions de test (CQ):")
    for i, q in enumerate(test_questions, 1):
        print(f"  {i}. {q}")

    # Tester toutes les configurations
    all_results = {}

    for question in test_questions:
        results = compare_all_configurations(question)
        all_results[question] = results

    # Exporter vers Excel
    output_dir = "comparaisons_rag"
    os.makedirs(output_dir, exist_ok=True)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_file = os.path.join(output_dir, f"comparison_ontology_configs_{timestamp}.xlsx")

    export_to_excel(all_results, excel_file)

    # Sauvegarder aussi en JSON
    json_file = os.path.join(output_dir, f"comparison_ontology_configs_{timestamp}.json")
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)
    print(f"OK Fichier JSON cree: {json_file}")

    # Statistiques globales
    print(f"\n{'='*80}")
    print("STATISTIQUES GLOBALES")
    print(f"{'='*80}")

    for config in ONTOLOGY_CONFIGS:
        print(f"\n{config.name}:")

        for rag_version in ["v1", "v2"]:
            version_name = f"RAG {rag_version} - {config.name}"
            times = []

            for question, results in all_results.items():
                for result in results:
                    if result['version'] == version_name:
                        times.append(result.get('execution_time', 0))

            if times:
                avg_time = sum(times) / len(times)
                print(f"  {version_name}: {avg_time:.2f}s (moyenne)")

    print(f"\n{'='*80}")
    print("COMPARAISON TERMINÉE")
    print(f"{'='*80}")
    print(f"Fichiers générés:")
    print(f"  - {excel_file}")
    print(f"  - {json_file}")


if __name__ == "__main__":
    main()
