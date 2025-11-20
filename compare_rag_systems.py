"""
Script de comparaison des 3 systèmes RAG (v1 vs v2 vs v3)

v1: RAG simple avec embeddings e5-base-v2
v2: RAG amélioré avec hybrid retrieval + reranking + données quantitatives
v3: RAG v2 + enrichissement par ontologie du bien-être

Appelle les 3 scripts avec les mêmes questions et compare les résultats
"""

import os
import sys
import json
import time
from datetime import datetime
from typing import List, Dict, Tuple
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# Imports des 2 systèmes RAG
# Note: Les imports nécessitent que les modules soient correctement installés
import chromadb
from sentence_transformers import SentenceTransformer
import openai

# Configuration de la clé API OpenAI depuis les variables d'environnement
OPENAI_API_KEY = os.environ.get("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("❌ Clé API OpenAI non trouvée. Définissez la variable d'environnement OPENAI_API_KEY")
openai.api_key = OPENAI_API_KEY


# ============================================================================
# PROMPTS UTILISÉS PAR LES 2 SYSTÈMES
# ============================================================================

# Prompt RAG v1 (rag_v1_2904.py - lignes 662-668)
PROMPT_V1_TEMPLATE = """Tu es un conseiller municipal. Ton but est de donner des informations sur la qualité de vie dans les communes Corses, pour guider les politiques publiques, en te basant uniquement sur les informations suivantes :

{retrieved_context}

Question : {question}
Réponse :
"""

SYSTEM_PROMPT_V1 = "Tu es un assistant utile et factuel. Ne réponds qu'avec les informations données."


# Prompt RAG v2 (rag_v2_improved.py - lignes 455-469 et 471-540)
SYSTEM_PROMPT_V2 = """Tu es un expert en analyse qualitative d'entretiens semi-directifs sur la qualité de vie en Corse.

Tu as accès à:
1. Des extraits d'entretiens avec des habitants
2. Des données quantitatives sur des indicateurs de qualité de vie
3. Des informations contextuelles (Wikipedia) sur les communes

PRINCIPES D'ANALYSE:
- Base tes réponses UNIQUEMENT sur les informations fournies
- Cite systématiquement tes sources (commune, numéro d'entretien)
- Identifie les thèmes récurrents et les divergences d'opinion
- Distingue clairement données qualitatives et quantitatives
- Si les données sont insuffisantes, indique explicitement les limites
- Nuance tes propos : évite les généralisations hâtives
"""

PROMPT_V2_INSTRUCTIONS = """
=== INSTRUCTIONS ===
Réponds à la question en te basant sur les informations ci-dessus.
Structure ta réponse ainsi:
1. Synthèse des éléments qualitatifs (entretiens)
2. Éclairage quantitatif (si disponible)
3. Analyse croisée et nuances
4. Limites et données manquantes (le cas échéant)

Cite systématiquement tes sources (commune, n° entretien).
"""


# ============================================================================
# FONCTIONS POUR INTERROGER RAG V1
# ============================================================================

def query_rag_v1(question: str, collection_name: str = "communes_corses_txt", n_chunks: int = 5) -> Dict:
    """
    Interroge le système RAG v1 (simple)

    Args:
        question: Question à poser
        collection_name: Nom de la collection ChromaDB
        n_chunks: Nombre de chunks à récupérer

    Returns:
        Dict avec la réponse, les chunks récupérés, le temps d'exécution
    """
    start_time = time.time()

    try:
        # Charger la collection ChromaDB
        chroma_client = chromadb.PersistentClient(path="./chroma_txt")
        collection = chroma_client.get_collection(name=collection_name)

        # Encoder la question
        embed_model = SentenceTransformer("intfloat/e5-base-v2")
        query_embedding = embed_model.encode([f"query: {question}"]).tolist()

        # Recherche
        results = collection.query(
            query_embeddings=query_embedding,
            n_results=n_chunks
        )

        # Récupérer les documents
        retrieved_docs = results["documents"][0]
        retrieved_metadata = results["metadatas"][0]
        retrieved_context = "\n\n".join(retrieved_docs)

        # Construire le prompt v1
        prompt = PROMPT_V1_TEMPLATE.format(
            retrieved_context=retrieved_context,
            question=question
        )

        # Appeler OpenAI
        from openai import OpenAI
        client = OpenAI(api_key=OPENAI_API_KEY)

        response = client.chat.completions.create(
            model="gpt-3.5-turbo",
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT_V1},
                {"role": "user", "content": prompt}
            ]
        )

        answer = response.choices[0].message.content

        execution_time = time.time() - start_time

        return {
            "version": "RAG v1",
            "answer": answer,
            "retrieved_chunks": retrieved_docs,
            "metadata": retrieved_metadata,
            "execution_time": execution_time,
            "num_chunks": len(retrieved_docs)
        }

    except Exception as e:
        return {
            "version": "RAG v1",
            "answer": f"Erreur: {str(e)}",
            "error": str(e),
            "execution_time": time.time() - start_time
        }


# ============================================================================
# FONCTIONS POUR INTERROGER RAG V2
# ============================================================================

def query_rag_v2(question: str, collection_name: str = "communes_corses_v2", k: int = 5) -> Dict:
    """
    Interroge le système RAG v2 (amélioré)

    Args:
        question: Question à poser
        collection_name: Nom de la collection ChromaDB
        k: Nombre de résultats à récupérer

    Returns:
        Dict avec la réponse, les chunks récupérés, le temps d'exécution
    """
    start_time = time.time()

    # Variables pour stocker les résultats partiels
    retrieved_chunks = []
    metadata = []
    scores = []
    response = None
    error_msg = None

    try:
        # Rediriger stdout/stderr pour éviter les erreurs d'encodage Windows
        import io
        import contextlib

        # Importer le pipeline v2
        from rag_v2_improved import ImprovedRAGPipeline

        # Initialiser le pipeline (utilise le cache s'il existe)
        # Rediriger la sortie pour éviter les problèmes d'encodage
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            rag = ImprovedRAGPipeline(
                chroma_path="./chroma_v2",
                collection_name=collection_name,
                embedding_model="dangvantuan/sentence-camembert-large",
                reranker_model="antoinelouis/crossencoder-camembert-base-mmarcoFR",
                llm_model="gpt-3.5-turbo",
                openai_api_key=OPENAI_API_KEY
            )

        # Si le pipeline n'a pas de documents, on ne peut pas faire la requête
        if not rag.documents or rag.hybrid_retriever is None:
            return {
                "version": "RAG v2",
                "answer": "Le pipeline RAG v2 n'a pas de documents indexes. Veuillez executer rag_v2_improved.py en mode principal (if __name__ == '__main__') pour creer le cache embeddings_v2.pkl, ou verifiez que ce fichier existe dans le repertoire courant.",
                "error": "No documents indexed",
                "retrieved_chunks": [],
                "execution_time": time.time() - start_time
            }

        # Effectuer la requête avec redirection de sortie
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            response, results = rag.query(
                question,
                k=k,
                use_reranking=True,
                include_quantitative=True  # Activer toutes les sources de données
            )

        # Extraire les informations des résultats
        retrieved_chunks = [r.text for r in results]
        metadata = [r.metadata for r in results]
        scores = [r.score for r in results]

    except Exception as e:
        error_msg = str(e)
        if not response:
            response = f"Erreur: {error_msg}"

    execution_time = time.time() - start_time

    result = {
        "version": "RAG v2",
        "answer": response if response else f"Erreur: {error_msg}",
        "retrieved_chunks": retrieved_chunks,
        "metadata": metadata,
        "scores": scores,
        "execution_time": execution_time,
        "num_chunks": len(retrieved_chunks)
    }

    if error_msg:
        result["error"] = error_msg

    return result


# ============================================================================
# FONCTIONS POUR INTERROGER RAG V3
# ============================================================================

def query_rag_v3(question: str, collection_name: str = "communes_corses_v2", k: int = 5) -> Dict:
    """
    Interroge le système RAG v3 (avec ontologie)

    Args:
        question: Question à poser
        collection_name: Nom de la collection ChromaDB
        k: Nombre de résultats à récupérer

    Returns:
        Dict avec la réponse, les chunks récupérés, le temps d'exécution
    """
    start_time = time.time()

    # Variables pour stocker les résultats partiels
    retrieved_chunks = []
    metadata = []
    scores = []
    response = None
    error_msg = None

    try:
        # Rediriger stdout/stderr pour éviter les erreurs d'encodage Windows
        import io
        import contextlib

        # Importer le pipeline v3
        from rag_v3_ontology import RAGPipelineWithOntology

        # Initialiser le pipeline (utilise le cache s'il existe)
        # Rediriger la sortie pour éviter les problèmes d'encodage
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            rag = RAGPipelineWithOntology(
                chroma_path="./chroma_v2",
                collection_name=collection_name,
                embedding_model="dangvantuan/sentence-camembert-large",
                reranker_model="antoinelouis/crossencoder-camembert-base-mmarcoFR",
                llm_model="gpt-3.5-turbo",
                openai_api_key=OPENAI_API_KEY
            )

        # Si le pipeline n'a pas de documents, on ne peut pas faire la requête
        if not rag.documents or rag.hybrid_retriever is None:
            return {
                "version": "RAG v3",
                "answer": "Le pipeline RAG v3 n'a pas de documents indexes. Veuillez executer rag_v2_improved.py en mode principal (if __name__ == '__main__') pour creer le cache embeddings_v2.pkl, ou verifiez que ce fichier existe dans le repertoire courant.",
                "error": "No documents indexed",
                "retrieved_chunks": [],
                "execution_time": time.time() - start_time
            }

        # Effectuer la requête avec redirection de sortie
        with contextlib.redirect_stdout(io.StringIO()), contextlib.redirect_stderr(io.StringIO()):
            response, results = rag.query(
                question,
                k=k,
                use_reranking=True,
                include_quantitative=True,
                use_ontology_enrichment=True  # Activer l'enrichissement ontologique
            )

        # Extraire les informations des résultats
        retrieved_chunks = [r.text for r in results]
        metadata = [r.metadata for r in results]
        scores = [r.score for r in results]

    except Exception as e:
        error_msg = str(e)
        if not response:
            response = f"Erreur: {error_msg}"

    execution_time = time.time() - start_time

    result = {
        "version": "RAG v3",
        "answer": response if response else f"Erreur: {error_msg}",
        "retrieved_chunks": retrieved_chunks,
        "metadata": metadata,
        "scores": scores,
        "execution_time": execution_time,
        "num_chunks": len(retrieved_chunks)
    }

    if error_msg:
        result["error"] = error_msg

    return result


# ============================================================================
# FONCTIONS DE COMPARAISON
# ============================================================================

def compare_results(question: str, result_v1: Dict, result_v2: Dict, result_v3: Dict = None) -> Dict:
    """
    Compare les résultats des 3 systèmes RAG

    Args:
        question: Question posée
        result_v1: Résultat du RAG v1
        result_v2: Résultat du RAG v2
        result_v3: Résultat du RAG v3 (optionnel)

    Returns:
        Dict avec l'analyse comparative
    """
    comparison = {
        "question": question,
        "timestamp": datetime.now().isoformat(),
        "v1": {
            "answer": result_v1.get("answer", ""),
            "num_chunks": result_v1.get("num_chunks", 0),
            "execution_time": result_v1.get("execution_time", 0),
            "error": result_v1.get("error", None)
        },
        "v2": {
            "answer": result_v2.get("answer", ""),
            "num_chunks": result_v2.get("num_chunks", 0),
            "execution_time": result_v2.get("execution_time", 0),
            "has_scores": "scores" in result_v2,
            "error": result_v2.get("error", None)
        },
        "comparison": {
            "time_difference_v2_v1": result_v2.get("execution_time", 0) - result_v1.get("execution_time", 0),
            "v2_slower_by_factor": (result_v2.get("execution_time", 1) / result_v1.get("execution_time", 1)) if result_v1.get("execution_time", 0) > 0 else None,
            "answer_length_v1": len(result_v1.get("answer", "")),
            "answer_length_v2": len(result_v2.get("answer", ""))
        },
        # Ajouter les extraits de texte (chunks) pour l'export Excel
        "retrieved_chunks_v1": result_v1.get("retrieved_chunks", []),
        "retrieved_chunks_v2": result_v2.get("retrieved_chunks", [])
    }

    # Ajouter v3 si disponible
    if result_v3:
        comparison["v3"] = {
            "answer": result_v3.get("answer", ""),
            "num_chunks": result_v3.get("num_chunks", 0),
            "execution_time": result_v3.get("execution_time", 0),
            "has_scores": "scores" in result_v3,
            "error": result_v3.get("error", None)
        }
        comparison["retrieved_chunks_v3"] = result_v3.get("retrieved_chunks", [])
        comparison["comparison"]["time_difference_v3_v2"] = result_v3.get("execution_time", 0) - result_v2.get("execution_time", 0)
        comparison["comparison"]["v3_slower_by_factor"] = (result_v3.get("execution_time", 1) / result_v2.get("execution_time", 1)) if result_v2.get("execution_time", 0) > 0 else None
        comparison["comparison"]["answer_length_v3"] = len(result_v3.get("answer", ""))

    return comparison


def format_comparison_text(comparison: Dict) -> str:
    """
    Formate la comparaison en texte lisible

    Args:
        comparison: Dict de comparaison

    Returns:
        Texte formaté
    """
    lines = []
    lines.append("=" * 100)
    lines.append(f"QUESTION: {comparison['question']}")
    lines.append("=" * 100)
    lines.append("")

    # RAG v1
    lines.append("-" * 100)
    lines.append("RAG V1 (Simple)")
    lines.append("-" * 100)
    lines.append(f"Temps d'exécution: {comparison['v1']['execution_time']:.2f}s")
    lines.append(f"Nombre de chunks: {comparison['v1']['num_chunks']}")
    if comparison['v1']['error']:
        lines.append(f"ERREUR: {comparison['v1']['error']}")
    lines.append("")
    lines.append("RÉPONSE:")
    lines.append(comparison['v1']['answer'])
    lines.append("")

    # RAG v2
    lines.append("-" * 100)
    lines.append("RAG V2 (Amélioré)")
    lines.append("-" * 100)
    lines.append(f"Temps d'exécution: {comparison['v2']['execution_time']:.2f}s")
    lines.append(f"Nombre de chunks: {comparison['v2']['num_chunks']}")
    lines.append(f"Reranking: {'Oui' if comparison['v2']['has_scores'] else 'Non'}")
    if comparison['v2']['error']:
        lines.append(f"ERREUR: {comparison['v2']['error']}")
    lines.append("")
    lines.append("RÉPONSE:")
    lines.append(comparison['v2']['answer'])
    lines.append("")

    # RAG v3 (si disponible)
    if 'v3' in comparison:
        lines.append("-" * 100)
        lines.append("RAG V3 (Avec Ontologie)")
        lines.append("-" * 100)
        lines.append(f"Temps d'execution: {comparison['v3']['execution_time']:.2f}s")
        lines.append(f"Nombre de chunks: {comparison['v3']['num_chunks']}")
        lines.append(f"Reranking: {'Oui' if comparison['v3']['has_scores'] else 'Non'}")
        if comparison['v3']['error']:
            lines.append(f"ERREUR: {comparison['v3']['error']}")
        lines.append("")
        lines.append("REPONSE:")
        lines.append(comparison['v3']['answer'])
        lines.append("")

    # Analyse comparative
    lines.append("-" * 100)
    lines.append("ANALYSE COMPARATIVE")
    lines.append("-" * 100)
    lines.append(f"Difference de temps V2-V1: {comparison['comparison']['time_difference_v2_v1']:.2f}s")
    if comparison['comparison']['v2_slower_by_factor']:
        lines.append(f"V2 est {comparison['comparison']['v2_slower_by_factor']:.2f}x plus lent que V1")
    lines.append(f"Longueur reponse V1: {comparison['comparison']['answer_length_v1']} caracteres")
    lines.append(f"Longueur reponse V2: {comparison['comparison']['answer_length_v2']} caracteres")

    # Ajouter v3 si disponible
    if 'v3' in comparison:
        lines.append(f"\nComparaison avec V3 (ontologie):")
        lines.append(f"Difference de temps V3-V2: {comparison['comparison'].get('time_difference_v3_v2', 0):.2f}s")
        if comparison['comparison'].get('v3_slower_by_factor'):
            lines.append(f"V3 est {comparison['comparison']['v3_slower_by_factor']:.2f}x plus lent que V2")
        lines.append(f"Longueur reponse V3: {comparison['comparison'].get('answer_length_v3', 0)} caracteres")

    lines.append("")
    lines.append("=" * 100)
    lines.append("")
    lines.append("")

    return "\n".join(lines)


def display_prompts():
    """
    Affiche les prompts utilisés par les 2 systèmes
    """
    lines = []
    lines.append("=" * 100)
    lines.append("PROMPTS UTILISÉS PAR LES 2 SYSTÈMES RAG")
    lines.append("=" * 100)
    lines.append("")

    # Prompt V1
    lines.append("-" * 100)
    lines.append("RAG V1 - SYSTEM PROMPT")
    lines.append("-" * 100)
    lines.append(SYSTEM_PROMPT_V1)
    lines.append("")
    lines.append("-" * 100)
    lines.append("RAG V1 - USER PROMPT TEMPLATE")
    lines.append("-" * 100)
    lines.append(PROMPT_V1_TEMPLATE)
    lines.append("")

    # Prompt V2
    lines.append("-" * 100)
    lines.append("RAG V2 - SYSTEM PROMPT")
    lines.append("-" * 100)
    lines.append(SYSTEM_PROMPT_V2)
    lines.append("")
    lines.append("-" * 100)
    lines.append("RAG V2 - INSTRUCTIONS ADDITIONNELLES")
    lines.append("-" * 100)
    lines.append(PROMPT_V2_INSTRUCTIONS)
    lines.append("")
    lines.append("Note: Le prompt V2 inclut également un formatage détaillé des extraits d'entretiens")
    lines.append("avec métadonnées (commune, source, pertinence, etc.) - voir build_rag_prompt() dans rag_v2_improved.py")
    lines.append("")
    lines.append("=" * 100)
    lines.append("")

    return "\n".join(lines)


def export_to_excel(all_comparisons: List[Dict], output_file: str):
    """
    Exporte les résultats de comparaison vers un fichier Excel
    avec les réponses et les extraits de texte utilisés par chaque modèle

    Args:
        all_comparisons: Liste des dictionnaires de comparaison
        output_file: Chemin du fichier Excel de sortie
    """
    # Créer un nouveau workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparaison RAG"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    v1_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    v2_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    # Déterminer le nombre maximum d'extraits
    max_chunks_v1 = max(len(comp.get('retrieved_chunks_v1', [])) for comp in all_comparisons) if all_comparisons else 5
    max_chunks_v2 = max(len(comp.get('retrieved_chunks_v2', [])) for comp in all_comparisons) if all_comparisons else 5

    # Créer les en-têtes
    headers = ["Question"]

    # En-têtes RAG v1
    headers.append("Réponse RAG v1")
    for i in range(max_chunks_v1):
        headers.append(f"Extrait {i+1} RAG v1")
    headers.append("Temps exec. v1 (s)")

    # En-têtes RAG v2
    headers.append("Réponse RAG v2")
    for i in range(max_chunks_v2):
        headers.append(f"Extrait {i+1} RAG v2")
    headers.append("Temps exec. v2 (s)")

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    # Ajuster la hauteur de la ligne d'en-tête
    ws.row_dimensions[1].height = 30

    # Remplir les données
    for row_idx, comp in enumerate(all_comparisons, 2):
        col_idx = 1

        # Question
        cell = ws.cell(row=row_idx, column=col_idx, value=comp['question'])
        cell.alignment = wrap_alignment
        col_idx += 1

        # RAG v1 - Réponse
        cell = ws.cell(row=row_idx, column=col_idx, value=comp['v1']['answer'])
        cell.fill = v1_fill
        cell.alignment = wrap_alignment
        col_idx += 1

        # RAG v1 - Extraits
        v1_chunks = comp.get('retrieved_chunks_v1', [])
        for i in range(max_chunks_v1):
            chunk_text = v1_chunks[i] if i < len(v1_chunks) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=chunk_text)
            cell.fill = v1_fill
            cell.alignment = wrap_alignment
            col_idx += 1

        # RAG v1 - Temps
        cell = ws.cell(row=row_idx, column=col_idx, value=f"{comp['v1']['execution_time']:.2f}")
        cell.fill = v1_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1

        # RAG v2 - Réponse
        cell = ws.cell(row=row_idx, column=col_idx, value=comp['v2']['answer'])
        cell.fill = v2_fill
        cell.alignment = wrap_alignment
        col_idx += 1

        # RAG v2 - Extraits
        v2_chunks = comp.get('retrieved_chunks_v2', [])
        for i in range(max_chunks_v2):
            chunk_text = v2_chunks[i] if i < len(v2_chunks) else ""
            cell = ws.cell(row=row_idx, column=col_idx, value=chunk_text)
            cell.fill = v2_fill
            cell.alignment = wrap_alignment
            col_idx += 1

        # RAG v2 - Temps
        cell = ws.cell(row=row_idx, column=col_idx, value=f"{comp['v2']['execution_time']:.2f}")
        cell.fill = v2_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 50  # Question

    col_idx = 2
    # RAG v1
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 60  # Réponse v1
    col_idx += 1
    for i in range(max_chunks_v1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 50  # Extraits v1
        col_idx += 1
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15  # Temps v1
    col_idx += 1

    # RAG v2
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 60  # Réponse v2
    col_idx += 1
    for i in range(max_chunks_v2):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 50  # Extraits v2
        col_idx += 1
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15  # Temps v2

    # Sauvegarder
    wb.save(output_file)
    print(f"  OK Fichier Excel cree : {output_file}")


# ============================================================================
# FONCTION PRINCIPALE
# ============================================================================

def main():
    """
    Fonction principale de comparaison
    """
    print("=" * 100)
    print("COMPARAISON DES SYSTÈMES RAG V1 vs V2")
    print("=" * 100)
    print("")

    # Afficher les prompts
    print(display_prompts())

    # Liste de questions à tester - communes spécifiques
    test_questions = [
        # Questions réelles sur des communes spécifiques
        "Combien y a-t-il de médecins à Afa ?",
        "Comment est la qualité de vie à Ajaccio ?",
        "Quels sont les problèmes de transport à Bastia ?",
        "Quelles sont les infrastructures éducatives à Corte ?",
        "Comment est l'accès aux services publics à Porto-Vecchio ?",
        "Quels sont les enjeux de logement à Ajaccio ?",
        "Comment les habitants d'Afa perçoivent-ils leur cadre de vie ?",

        # Questions pièges pour détecter les hallucinations
        "Combien y a-t-il de stations de ski à Ajaccio ?",
        "Quel est le nombre de lignes de métro à Bastia ?",
        "Combien de tours Eiffel y a-t-il à Porto-Vecchio ?"
    ]

    print("Questions de test:")
    for i, q in enumerate(test_questions, 1):
        print(f"  {i}. {q}")
    print("")

    # Créer le répertoire de sortie
    output_dir = "comparaisons_rag"
    os.makedirs(output_dir, exist_ok=True)

    # Fichier de sortie principal
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"comparison_{timestamp}.txt")

    # Ouvrir le fichier en écriture
    with open(output_file, 'w', encoding='utf-8') as f:
        # Écrire les prompts
        f.write(display_prompts())
        f.write("\n\n")

        # Pour chaque question
        all_comparisons = []

        for i, question in enumerate(test_questions, 1):
            print(f"\n[{i}/{len(test_questions)}] Traitement de la question: {question}")
            print("-" * 100)

            # Interroger RAG v1
            print("  - Interrogation RAG v1...")
            result_v1 = query_rag_v1(question)
            print(f"    OK Termine en {result_v1.get('execution_time', 0):.2f}s")

            # Interroger RAG v2
            print("  - Interrogation RAG v2...")
            result_v2 = query_rag_v2(question)
            print(f"    OK Termine en {result_v2.get('execution_time', 0):.2f}s")

            # Interroger RAG v3 (avec ontologie)
            print("  - Interrogation RAG v3 (avec ontologie)...")
            result_v3 = query_rag_v3(question)
            print(f"    OK Termine en {result_v3.get('execution_time', 0):.2f}s")

            # Comparer les 3 versions
            comparison = compare_results(question, result_v1, result_v2, result_v3)
            all_comparisons.append(comparison)

            # Écrire dans le fichier
            comparison_text = format_comparison_text(comparison)
            f.write(comparison_text)

            # Afficher un résumé
            print(f"  - Difference de temps V2-V1: {comparison['comparison']['time_difference_v2_v1']:.2f}s")
            if 'time_difference_v3_v2' in comparison['comparison']:
                print(f"  - Difference de temps V3-V2: {comparison['comparison']['time_difference_v3_v2']:.2f}s")
            print("")

        # Sauvegarder aussi en JSON
        json_file = os.path.join(output_dir, f"comparison_{timestamp}.json")
        with open(json_file, 'w', encoding='utf-8') as jf:
            json.dump(all_comparisons, jf, indent=2, ensure_ascii=False)

        # Sauvegarder en Excel avec les extraits
        excel_file = os.path.join(output_dir, f"comparison_{timestamp}.xlsx")
        print("\nCréation du fichier Excel avec les extraits...")
        export_to_excel(all_comparisons, excel_file)

    print("\n" + "=" * 100)
    print("COMPARAISON TERMINÉE")
    print("=" * 100)
    print(f"Résultats sauvegardés dans:")
    print(f"  - {output_file}")
    print(f"  - {json_file}")
    print(f"  - {excel_file}")
    print("")

    # Statistiques globales
    total_time_v1 = sum(c['v1']['execution_time'] for c in all_comparisons)
    total_time_v2 = sum(c['v2']['execution_time'] for c in all_comparisons)
    total_time_v3 = sum(c.get('v3', {}).get('execution_time', 0) for c in all_comparisons)

    print("Statistiques globales:")
    print(f"  - Temps total V1: {total_time_v1:.2f}s")
    print(f"  - Temps total V2: {total_time_v2:.2f}s")
    print(f"  - Temps total V3: {total_time_v3:.2f}s")
    print(f"  - V2 vs V1: {total_time_v2/total_time_v1:.2f}x plus lent")
    print(f"  - V3 vs V2: {total_time_v3/total_time_v2:.2f}x plus lent")
    print("")

    return all_comparisons


if __name__ == "__main__":
    main()
