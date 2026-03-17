"""
Évaluation multi-RAG : envoie chaque question à la version RAG appropriée
(v2.2 pour communes, v6 pour ontologie, v9/RAPTOR pour population),
puis fait noter par Mistral Large la réponse et la qualité du contexte.
Exporte en Excel.
"""

import os
import json
import re
import time
import argparse
import requests
import pandas as pd
from datetime import datetime
from typing import List, Dict, Optional
from dotenv import load_dotenv

load_dotenv()

# === Configuration ===
API_BASE_URL = "http://localhost:8000"
LLM_MODEL = "gpt-3.5-turbo"

JUDGE_MODEL = "mistral-large-latest"
JUDGE_BASE_URL = "https://api.mistral.ai/v1"
JUDGE_API_KEY_ENV = "MISTRAL_API_KEY"


def call_rag_api(question: str, rag_version: str, k: int = 5) -> Dict:
    """Appelle l'API serveur pour une question avec une version RAG donnée."""
    payload = {
        "question": question,
        "rag_version": rag_version,
        "k": k,
        "llm_model": LLM_MODEL,
    }
    try:
        resp = requests.post(f"{API_BASE_URL}/api/query", json=payload, timeout=120)
        resp.raise_for_status()
        return resp.json()
    except Exception as e:
        return {"error": str(e), "answer": f"ERREUR API: {e}", "sources": []}


def call_judge(question: str, answer: str, sources: List[Dict],
               rag_version: str, category: str) -> Dict:
    """Appelle Mistral Large pour noter la réponse et la qualité du contexte."""
    api_key = os.getenv(JUDGE_API_KEY_ENV)
    if not api_key:
        return {"error": f"Clé {JUDGE_API_KEY_ENV} non trouvée"}

    # Formater les sources pour le juge
    sources_text = ""
    for i, s in enumerate(sources, 1):
        content = s.get("content", "")[:500]
        meta = s.get("metadata", {})
        meta_str = ", ".join(f"{k}={v}" for k, v in meta.items() if k not in ("content",))
        sources_text += f"\n--- Source {i} [{meta_str}] ---\n{content}\n"

    judge_system = """Tu es un évaluateur expert en analyse de données qualitatives sur le bien-être territorial en Corse.
Tu dois évaluer la qualité d'une réponse générée par un système RAG (Retrieval-Augmented Generation).

Tu DOIS répondre UNIQUEMENT en JSON valide, sans texte avant ni après."""

    judge_prompt = f"""=== QUESTION POSÉE ===
{question}

=== CATÉGORIE ===
{category}

=== VERSION RAG UTILISÉE ===
{rag_version}

=== SOURCES RÉCUPÉRÉES PAR LE RAG ===
{sources_text}

=== RÉPONSE GÉNÉRÉE ===
{answer}

Évalue selon les critères suivants et produis le JSON :
{{
  "note_reponse": NOTE_SUR_10,
  "note_contexte": NOTE_SUR_10,
  "fidelite": NOTE_SUR_10,
  "pertinence": NOTE_SUR_10,
  "completude": NOTE_SUR_10,
  "clarte": NOTE_SUR_10,
  "nuance": NOTE_SUR_10,
  "justification_reponse": "Explication courte de la note de la réponse (2-3 phrases)",
  "justification_contexte": "Évaluation de la pertinence des sources récupérées par rapport à la question (2-3 phrases)",
  "points_forts": "Ce qui est bien dans la réponse",
  "points_faibles": "Ce qui pourrait être amélioré"
}}

Critères détaillés :
- note_reponse : qualité globale de la réponse (fidélité, pertinence, complétude, clarté, nuance)
- note_contexte : pertinence des sources récupérées par le RAG par rapport à la question posée
- fidelite : la réponse s'appuie-t-elle sur les sources sans inventer ?
- pertinence : la réponse répond-elle à la question ?
- completude : les informations clés des sources sont-elles exploitées ?
- clarte : structure et lisibilité
- nuance : évite les généralisations abusives ?"""

    from openai import OpenAI
    client = OpenAI(api_key=api_key, base_url=JUDGE_BASE_URL)

    for attempt in range(3):
        try:
            response = client.chat.completions.create(
                model=JUDGE_MODEL,
                messages=[
                    {"role": "system", "content": judge_system},
                    {"role": "user", "content": judge_prompt}
                ],
                temperature=0.3,
                max_tokens=1000
            )
            raw = response.choices[0].message.content

            # Parser le JSON
            json_match = re.search(r'\{[\s\S]*\}', raw)
            if json_match:
                judgment = json.loads(json_match.group())
            else:
                judgment = json.loads(raw)

            judgment["error"] = None
            return judgment

        except Exception as e:
            if "429" in str(e) and attempt < 2:
                wait = 2 ** attempt * 3
                print(f"    [RATE LIMIT] Attente {wait}s...")
                time.sleep(wait)
            else:
                return {"error": str(e)}

    return {"error": "Max retries exceeded"}


def export_to_excel(results: List[Dict], output_path: str):
    """Exporte les résultats dans un Excel structuré."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    # === Feuille Recap ===
    ws = wb.active
    ws.title = "Résultats"

    headers = [
        "ID", "Catégorie", "Version RAG", "Question", "Réponse",
        "Note Réponse /10", "Note Contexte /10",
        "Fidélité /10", "Pertinence /10", "Complétude /10",
        "Clarté /10", "Nuance /10",
        "Justification Réponse", "Justification Contexte",
        "Points forts", "Points faibles",
        "Nb Sources"
    ]
    col_widths = [5, 25, 10, 60, 80, 12, 12, 10, 10, 10, 10, 10, 50, 50, 40, 40, 10]

    # Styles
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    WRAP = Alignment(wrap_text=True, vertical="top")
    THIN_BORDER = Border(
        bottom=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
    )

    # Couleurs par catégorie
    CAT_FILLS = {
        "2_communes_specifiques": PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid"),
        "3_comparaison_communes": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        "4_communes_adjacentes": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        "5_ontologie": PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid"),
        "6_pieges": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
        "7_population": PatternFill(start_color="D5A6E6", end_color="D5A6E6", fill_type="solid"),
    }

    # En-têtes
    for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = WRAP
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w

    # Données
    for ri, r in enumerate(results, 2):
        j = r.get("judgment", {})
        values = [
            r.get("id", ""),
            r.get("category", ""),
            r.get("rag_version", ""),
            r.get("question", ""),
            r.get("answer", ""),
            j.get("note_reponse", ""),
            j.get("note_contexte", ""),
            j.get("fidelite", ""),
            j.get("pertinence", ""),
            j.get("completude", ""),
            j.get("clarte", ""),
            j.get("nuance", ""),
            j.get("justification_reponse", ""),
            j.get("justification_contexte", ""),
            j.get("points_forts", ""),
            j.get("points_faibles", ""),
            r.get("nb_sources", 0),
        ]
        cat = r.get("category", "")
        cat_fill = CAT_FILLS.get(cat)

        for ci, val in enumerate(values, 1):
            # Convertir les listes en string (le juge peut retourner des listes)
            if isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False) if isinstance(val, dict) else "\n".join(str(v) for v in val)
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.alignment = WRAP
            cell.border = THIN_BORDER
            if cat_fill:
                cell.fill = cat_fill

        ws.row_dimensions[ri].height = 80

    # Fixer la première ligne
    ws.freeze_panes = "A2"

    # === Feuille Statistiques ===
    ws_stats = wb.create_sheet("Statistiques")
    ws_stats.column_dimensions["A"].width = 30
    ws_stats.column_dimensions["B"].width = 15
    ws_stats.column_dimensions["C"].width = 15
    ws_stats.column_dimensions["D"].width = 15
    ws_stats.column_dimensions["E"].width = 10

    # Stats par catégorie
    stats_headers = ["Catégorie", "Moy. Réponse", "Moy. Contexte", "Version RAG", "Nb Q."]
    for ci, h in enumerate(stats_headers, 1):
        cell = ws_stats.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    categories = {}
    for r in results:
        cat = r.get("category", "?")
        j = r.get("judgment", {})
        if cat not in categories:
            categories[cat] = {"notes_rep": [], "notes_ctx": [], "rag": set()}
        nr = j.get("note_reponse")
        nc = j.get("note_contexte")
        if isinstance(nr, (int, float)):
            categories[cat]["notes_rep"].append(nr)
        if isinstance(nc, (int, float)):
            categories[cat]["notes_ctx"].append(nc)
        categories[cat]["rag"].add(r.get("rag_version", "?"))

    row = 2
    for cat, data in sorted(categories.items()):
        avg_rep = sum(data["notes_rep"]) / len(data["notes_rep"]) if data["notes_rep"] else 0
        avg_ctx = sum(data["notes_ctx"]) / len(data["notes_ctx"]) if data["notes_ctx"] else 0
        ws_stats.cell(row=row, column=1, value=cat)
        ws_stats.cell(row=row, column=2, value=round(avg_rep, 2))
        ws_stats.cell(row=row, column=3, value=round(avg_ctx, 2))
        ws_stats.cell(row=row, column=4, value=", ".join(data["rag"]))
        ws_stats.cell(row=row, column=5, value=len(data["notes_rep"]))
        if cat in CAT_FILLS:
            for ci in range(1, 6):
                ws_stats.cell(row=row, column=ci).fill = CAT_FILLS[cat]
        row += 1

    # Ligne totale
    all_rep = [n for d in categories.values() for n in d["notes_rep"]]
    all_ctx = [n for d in categories.values() for n in d["notes_ctx"]]
    ws_stats.cell(row=row + 1, column=1, value="TOTAL").font = Font(bold=True)
    ws_stats.cell(row=row + 1, column=2, value=round(sum(all_rep) / len(all_rep), 2) if all_rep else 0)
    ws_stats.cell(row=row + 1, column=3, value=round(sum(all_ctx) / len(all_ctx), 2) if all_ctx else 0)
    ws_stats.cell(row=row + 1, column=5, value=len(all_rep))

    wb.save(output_path)
    print(f"Excel sauvegardé : {output_path}")


def main():
    parser = argparse.ArgumentParser(description="Évaluation multi-RAG avec juge Mistral Large")
    parser.add_argument("--questions-file", default="questions_evaluation_multi_rag.json")
    parser.add_argument("--max-questions", type=int, default=0, help="Limiter le nombre de questions (0=toutes)")
    parser.add_argument("--output-dir", default="comparaisons_rag")
    parser.add_argument("--llm-model", default="gpt-3.5-turbo")
    parser.add_argument("--k", type=int, default=5)
    parser.add_argument("--from-json", type=str, default=None,
                        help="Recharger les résultats depuis un JSON existant (skip API/juge, re-export Excel)")
    args = parser.parse_args()

    # Mode re-export depuis JSON existant
    if args.from_json:
        with open(args.from_json, "r", encoding="utf-8") as f:
            saved = json.load(f)
        results = saved["results"]
        xlsx_path = args.from_json.replace(".json", ".xlsx")
        export_to_excel(results, xlsx_path)
        return

    global LLM_MODEL
    LLM_MODEL = args.llm_model

    # Charger les questions
    with open(args.questions_file, "r", encoding="utf-8") as f:
        data = json.load(f)

    questions = data["questions"]
    if args.max_questions > 0:
        questions = questions[:args.max_questions]

    total = len(questions)
    print(f"=" * 70)
    print(f"ÉVALUATION MULTI-RAG")
    print(f"=" * 70)
    print(f"Questions : {total}")
    print(f"LLM génération : {LLM_MODEL}")
    print(f"Juge : {JUDGE_MODEL}")
    print(f"Versions RAG : {sorted(set(q['rag_version'] for q in questions))}")
    print(f"=" * 70)

    results = []

    for i, q in enumerate(questions, 1):
        qid = q["id"]
        question = q["question"]
        rag_version = q["rag_version"]
        category = q["category"]

        print(f"\n[{i}/{total}] Q{qid} ({rag_version}) {question[:60]}...")

        # 1. Appel RAG
        print(f"  RAG {rag_version}...", end=" ", flush=True)
        api_response = call_rag_api(question, rag_version, k=args.k)

        if "error" in api_response and api_response.get("answer", "").startswith("ERREUR"):
            print(f"ERREUR: {api_response['error']}")
            results.append({
                "id": qid, "category": category, "rag_version": rag_version,
                "question": question, "answer": api_response["answer"],
                "nb_sources": 0, "judgment": {"error": api_response["error"]}
            })
            continue

        answer = api_response.get("answer", "")
        sources = api_response.get("sources", [])
        print(f"OK ({len(sources)} sources)")

        # 2. Appel juge
        print(f"  Juge ({JUDGE_MODEL})...", end=" ", flush=True)
        judgment = call_judge(question, answer, sources, rag_version, category)

        if judgment.get("error"):
            print(f"ERREUR: {judgment['error']}")
        else:
            nr = judgment.get("note_reponse", "?")
            nc = judgment.get("note_contexte", "?")
            print(f"OK (réponse={nr}/10, contexte={nc}/10)")

        results.append({
            "id": qid,
            "category": category,
            "rag_version": rag_version,
            "question": question,
            "answer": answer,
            "nb_sources": len(sources),
            "sources": sources,
            "judgment": judgment,
        })

        # Pause anti-rate-limit
        time.sleep(1)

    # Export
    os.makedirs(args.output_dir, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    base_name = f"eval_multi_rag_{timestamp}"

    # JSON
    json_path = os.path.join(args.output_dir, f"{base_name}.json")
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump({
            "metadata": {
                "timestamp": timestamp,
                "llm_model": LLM_MODEL,
                "judge_model": JUDGE_MODEL,
                "total_questions": total,
            },
            "results": results,
        }, f, ensure_ascii=False, indent=2)
    print(f"\nJSON sauvegardé : {json_path}")

    # Excel
    xlsx_path = os.path.join(args.output_dir, f"{base_name}.xlsx")
    export_to_excel(results, xlsx_path)

    # Résumé
    print(f"\n{'=' * 70}")
    print("RÉSUMÉ PAR CATÉGORIE")
    print(f"{'=' * 70}")
    categories = {}
    for r in results:
        cat = r["category"]
        j = r.get("judgment", {})
        if cat not in categories:
            categories[cat] = {"rep": [], "ctx": []}
        nr = j.get("note_reponse")
        nc = j.get("note_contexte")
        if isinstance(nr, (int, float)):
            categories[cat]["rep"].append(nr)
        if isinstance(nc, (int, float)):
            categories[cat]["ctx"].append(nc)

    for cat, data in sorted(categories.items()):
        avg_r = sum(data["rep"]) / len(data["rep"]) if data["rep"] else 0
        avg_c = sum(data["ctx"]) / len(data["ctx"]) if data["ctx"] else 0
        print(f"  {cat:30s}  Réponse: {avg_r:.1f}/10  Contexte: {avg_c:.1f}/10  (n={len(data['rep'])})")

    all_r = [n for d in categories.values() for n in d["rep"]]
    all_c = [n for d in categories.values() for n in d["ctx"]]
    if all_r:
        print(f"\n  {'GLOBAL':30s}  Réponse: {sum(all_r)/len(all_r):.1f}/10  Contexte: {sum(all_c)/len(all_c):.1f}/10  (n={len(all_r)})")
    print(f"{'=' * 70}")


if __name__ == "__main__":
    main()
