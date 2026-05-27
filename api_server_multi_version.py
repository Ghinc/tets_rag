"""
API FastAPI multi-versions pour les systèmes RAG

Ce serveur permet de choisir dynamiquement entre les versions v1, v2, v3 et v4 du RAG
via l'interface graphique ou les requêtes API.

Endpoints:
- POST /api/query - Poser une question au chatbot (avec choix de version)
- GET /api/health - Vérifier l'état du serveur
- GET /api/versions - Liste des versions disponibles
- GET /docs - Documentation Swagger automatique

Auteur: Claude Code
Date: 2025-11-17
"""

import os
import sys
import json
import asyncio
import uuid

# Forcer UTF-8 sur Windows pour éviter les UnicodeEncodeError sur les print() du pipeline
if sys.stdout.encoding and sys.stdout.encoding.lower() != "utf-8":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if sys.stderr.encoding and sys.stderr.encoding.lower() != "utf-8":
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import tempfile
import subprocess
import shutil
from typing import Optional, List, Dict, Literal
from datetime import datetime
from contextlib import asynccontextmanager
from fastapi import FastAPI, HTTPException, BackgroundTasks, UploadFile, File, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import HTMLResponse, StreamingResponse, FileResponse
from pydantic import BaseModel, Field, ConfigDict
import uvicorn
from dotenv import load_dotenv

# SentenceTransformer DOIT être importé avant chromadb (conflit libs natives sur Windows).
# rag_v1_class importe chromadb au niveau module → pré-charger ST ici pour éviter le segfault.
from sentence_transformers import SentenceTransformer as _ST_preload  # noqa: F401

# Imports des différentes versions RAG
from rag_v1_class import BasicRAGPipeline, RetrievalResult as RetrievalResult_v1
from rag_v2_boosted import ImprovedRAGPipeline, RetrievalResult as RetrievalResult_v2
try:
    from rag_v2_1_llamaindex import RAGv2_1_LlamaIndex
    V2_1_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: RAG v2.1 non disponible ({e})")
    RAGv2_1_LlamaIndex = None
    V2_1_AVAILABLE = False
from rag_v2_2_portrait import PortraitRAGPipeline
from rag_v3_ontology import RAGPipelineWithOntology, RetrievalResult as RetrievalResult_v3
from rag_v4_cross_analysis import ImprovedRAGPipeline as CrossAnalysisRAGPipeline, RetrievalResult as RetrievalResult_v4
try:
    from rag_v5_graphrag_neo4j import GraphRAGPipeline
    V5_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: RAG v5 non disponible ({e})")
    GraphRAGPipeline = None
    V5_AVAILABLE = False

# Import optionnel de v6 (nécessite torch_geometric)
try:
    from rag_v6_gretriever import GRetrieverRAGPipeline
    V6_AVAILABLE = True
except ImportError:
    print("AVERTISSEMENT: torch_geometric non installé, RAG v6 désactivé")
    GRetrieverRAGPipeline = None
    V6_AVAILABLE = False

# Import optionnel de v7 (nécessite llama-index, graphes optionnels)
try:
    from rag_v7_llamaindex import LlamaIndexRAGPipeline
    V7_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: LlamaIndex non installé, RAG v7 désactivé ({e})")
    LlamaIndexRAGPipeline = None
    V7_AVAILABLE = False

# Import optionnel de v8 (nécessite llama-index + graphes OBLIGATOIRES)
try:
    from rag_v8_llamaindex_full import LlamaIndexRAGPipeline as LlamaIndexRAGPipelineFull
    V8_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: LlamaIndex non installé, RAG v8 désactivé ({e})")
    LlamaIndexRAGPipelineFull = None
    V8_AVAILABLE = False

# Import optionnel de v9 (RAPTOR-lite)
try:
    from rag_v9_raptor import RaptorRetriever
    V9_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: RAPTOR non disponible ({e})")
    RaptorRetriever = None
    V9_AVAILABLE = False

# Import optionnel de v10 (RAPTOR + Sous-questions)
try:
    from rag_v10_raptor_subq import RaptorSubQuestionPipeline
    V10_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: RAG v10 non disponible ({e})")
    RaptorSubQuestionPipeline = None
    V10_AVAILABLE = False

# Import optionnel de v11 (Agentic RAG : ReAct + CRAG gate)
try:
    from rag_v11_agentic import AgenticRAGPipeline
    V11_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: RAG v11 non disponible ({e})")
    AgenticRAGPipeline = None
    V11_AVAILABLE = False

# Import optionnel des ablations (V_vanilla, V_decomp, V_decomp_raptor)
try:
    from rag_ablations import VanillaRAG, DecompOnlyRAG
    ABLATIONS_AVAILABLE = True
except ImportError as e:
    print(f"AVERTISSEMENT: ablations non disponibles ({e})")
    VanillaRAG = DecompOnlyRAG = None
    ABLATIONS_AVAILABLE = False

# Charger les variables d'environnement
load_dotenv()

# === MODELS PYDANTIC ===

class QueryRequest(BaseModel):
    """Modèle de requête pour poser une question"""
    question: str = Field(..., description="Question à poser au chatbot", min_length=1)
    rag_version: Literal["v1", "v2", "v2.1", "v2.2", "v3", "v4", "v5", "v6", "v7", "v8", "v9", "v10", "v11",
                         "v_vanilla_k10", "v_vanilla_k25", "v_decomp", "v_decomp_raptor"] = Field("v2", description="Version du RAG à utiliser")
    k: int = Field(5, description="Nombre de documents à récupérer", ge=1, le=100)
    use_reranking: bool = Field(True, description="Utiliser le reranking (v2/v3/v4 uniquement)")
    include_quantitative: bool = Field(True, description="Inclure les données quantitatives (v2/v3/v4 uniquement)")
    commune_filter: Optional[str] = Field(None, description="Filtrer par commune spécifique")
    use_ontology_enrichment: bool = Field(True, description="Utiliser l'enrichissement ontologique (v3 uniquement)")
    use_cross_analysis: bool = Field(True, description="Activer l'analyse croisée automatique (v4 uniquement)")
    query_mode: Literal["router", "sub_question", "hybrid"] = Field("router", description="Mode de query pour v7/v8 (router/sub_question/hybrid)")
    llm_model: str = Field("gpt-3.5-turbo", description="Modèle LLM à utiliser (gpt-3.5-turbo, gpt-4, gpt-4o, gpt-4o-mini)")
    # Filtres portrait pour v2.2
    auto_detect_portrait: bool = Field(True, description="Détecter automatiquement les filtres portrait (v2.2)")
    portrait_age_min: Optional[int] = Field(None, description="Âge minimum pour filtrer les verbatims (v2.2)", ge=15, le=100)
    portrait_age_max: Optional[int] = Field(None, description="Âge maximum pour filtrer les verbatims (v2.2)", ge=15, le=100)
    portrait_genre: Optional[str] = Field(None, description="Genre pour filtrer les verbatims: Homme/Femme (v2.2)")
    portrait_profession: Optional[str] = Field(None, description="Profession pour filtrer les verbatims (v2.2)")
    portrait_dimension: Optional[str] = Field(None, description="Dimension qualité de vie pour filtrer (v2.2)")
    n_subquestions: int = Field(5, ge=1, le=8, description="Nombre de sous-questions à générer (v10 uniquement)")
    max_iterations: int = Field(5, ge=1, le=8, description="Nombre max d'itérations ReAct (v11 uniquement)")
    use_fast_path: bool = Field(True, description="Fast path pour questions simples (v11 uniquement)")

    model_config = ConfigDict(
        json_schema_extra={
            "example": {
                "question": "Quelles sont les communes avec le meilleur bien-être ?",
                "rag_version": "v2",
                "k": 5,
                "use_reranking": True,
                "include_quantitative": True,
                "use_ontology_enrichment": True
            }
        }
    )


class Source(BaseModel):
    """Modèle pour une source de document"""
    content: str = Field(..., description="Contenu du document")
    score: float = Field(..., description="Score de pertinence")
    metadata: Dict = Field(..., description="Métadonnées du document (commune, source, etc.)")


class QueryResponse(BaseModel):
    """Modèle de réponse à une question"""
    answer: str = Field(..., description="Réponse générée par le chatbot")
    sources: List[Source] = Field(..., description="Sources utilisées pour générer la réponse")
    metadata: Dict = Field(..., description="Métadonnées de la requête")
    rag_version_used: str = Field(..., description="Version du RAG utilisée")
    timestamp: str = Field(..., description="Horodatage de la réponse")
    context: Optional[str] = Field(None, description="Contexte complet passé au LLM (pour debug/export)")
    sub_questions: Optional[List[Dict]] = Field(None, description="Sous-questions et réponses intermédiaires (v10)")


class HealthResponse(BaseModel):
    """Modèle de réponse pour le health check"""
    status: str = Field(..., description="État du serveur")
    rag_v1_initialized: bool = Field(..., description="RAG v1 initialisé")
    rag_v2_initialized: bool = Field(..., description="RAG v2 initialisé")
    rag_v2_1_initialized: bool = Field(..., description="RAG v2.1 initialisé")
    rag_v2_2_initialized: bool = Field(..., description="RAG v2.2 (Portrait) initialisé")
    rag_v3_initialized: bool = Field(..., description="RAG v3 initialisé")
    rag_v4_initialized: bool = Field(..., description="RAG v4 initialisé")
    rag_v5_initialized: bool = Field(..., description="RAG v5 initialisé")
    rag_v6_initialized: bool = Field(..., description="RAG v6 initialisé")
    rag_v7_initialized: bool = Field(..., description="RAG v7 initialisé")
    rag_v8_initialized: bool = Field(..., description="RAG v8 initialisé")
    rag_v9_initialized: bool = Field(..., description="RAG v9 (RAPTOR) initialisé")
    rag_v10_initialized: bool = Field(..., description="RAG v10 (RAPTOR + sous-questions) initialisé")
    rag_v11_initialized: bool = Field(..., description="RAG v11 (Agentic ReAct+CRAG) initialisé")
    timestamp: str = Field(..., description="Horodatage du check")
    version: str = Field(..., description="Version de l'API")


class VersionInfo(BaseModel):
    """Information sur une version du RAG"""
    version: str
    name: str
    description: str
    available: bool
    features: List[str]


# === GESTION DES PIPELINES RAG ===

rag_pipelines = {
    "v1": None,
    "v2": None,
    "v2.1": None,
    "v2.2": None,
    "v3": None,
    "v4": None,
    "v5": None,
    "v6": None,
    "v7": None,
    "v8": None,
    "v9": None,
    "v10": None,
    "v11": None,
    # Ablations
    "v_vanilla_k10":    None,
    "v_vanilla_k25":    None,
    "v_decomp":         None,
    "v_decomp_raptor":  None,
}

# Mots-clés déclenchant la recherche dans oppchovec_scores
_OPPCHOVEC_KEYWORDS = [
    "oppchovec", "score", "scores",
    "opportunités", "opportunite", "opportunites",
    "choix", "vécu", "vecu",
    "indice", "indicateur", "indicateurs",
    "classement", "classé", "classee", "rang",
    "bien placé", "bien place", "mal placé", "mal place",
    "meilleur", "moins bon", "moins bonne",
    "mieux noté", "moins bien noté",
]


def _is_oppchovec_question(question: str) -> bool:
    """Retourne True si la question porte (en partie) sur les scores OppChoVec."""
    q = question.lower()
    return any(kw in q for kw in _OPPCHOVEC_KEYWORDS)


_BIENEETRE_KEYWORDS = [
    "bien-être", "bien être", "bienetre", "bien-etre",
    "qualité de vie", "qualite de vie",
    "comment se porte", "comment vivent", "comment vit",
    "portrait global", "portrait de",
    "satisfaction globale", "conditions de vie",
]

# Si ces mots-clés sont présents, la question est purement qualitative/perception :
# ne pas déclencher force_mixed ni l'injection OppChoVec
_QUALI_OVERRIDE_KEYWORDS = [
    "ressenti", "ressentent", "ressent", "ressens",
    "perçoivent", "perçoit", "perception", "perceptions",
    "que pensent", "que pense", "avis des", "opinion",
    "témoignages", "verbatims", "verbatim",
    "enquête citoyenne", "enquete citoyenne",
    "comment vivent-ils", "comment vivent-elles",
    "comment les habitants", "comment les résidents",
    "comment les gens",
]


def _is_bieneetre_question(question: str) -> bool:
    """Retourne True si la question porte sur le bien-être/QdV global d'une commune
    (mixte quali+quanti attendu). Exclut les questions purement qualitatives/perception
    et les questions purement factuelles/OppChoVec."""
    q = question.lower()
    # Si la question est explicitement sur le ressenti/perception, rester en mode quali pur
    if any(kw in q for kw in _QUALI_OVERRIDE_KEYWORDS):
        return False
    return any(kw in q for kw in _BIENEETRE_KEYWORDS)


def initialize_all_rags():
    """
    Initialise toutes les versions du RAG disponibles

    Tente d'initialiser v1, v2, v3 et v4. Si une version échoue,
    elle reste à None mais les autres sont disponibles.
    """
    global rag_pipelines

    print("\n" + "="*60)
    print("INITIALISATION MULTI-VERSION RAG")
    print("="*60 + "\n")

    openai_api_key = os.getenv("OPENAI_API_KEY")
    if not openai_api_key:
        raise RuntimeError(
            "OPENAI_API_KEY non trouvée dans les variables d'environnement. "
            "Veuillez créer un fichier .env avec votre clé API."
        )

    # === INITIALISER RAG v1 ===
    print("\n[1/10] Initialisation RAG v1 (basique)...")
    try:
        rag_pipelines["v1"] = BasicRAGPipeline(
            openai_api_key=openai_api_key,
            chroma_path="./chroma_portrait",
            collection_name="portrait_verbatims",
            llm_model="gpt-4o-mini",
            embedding_model="BAAI/bge-m3",
        )
        print("OK RAG v1 initialise")
    except Exception as e:
        print(f"AVERTISSEMENT: RAG v1 non disponible : {e}")
        rag_pipelines["v1"] = None

    # === INITIALISER RAG v2 ===
    # Désactivé : chroma_v2 corrompu (Rust panic sur sqlite bindings)
    print("\n[2/10] RAG v2 désactivé (chroma_v2 corrompu)")
    rag_pipelines["v2"] = None

    # === INITIALISER RAG v2.1 (LlamaIndex équivalent v2) ===
    print("\n[2.1/9] Initialisation RAG v2.1 (LlamaIndex)...")
    if not V2_1_AVAILABLE:
        print("AVERTISSEMENT: RAG v2.1 non disponible (llama-index manquant)")
        rag_pipelines["v2.1"] = None
    else:
        try:
            rag_pipelines["v2.1"] = RAGv2_1_LlamaIndex(
                chroma_path="./chroma_v2",
                collection_name="communes_corses_v2",
                embedding_model="BAAI/bge-m3",
                rerank_model="BAAI/bge-reranker-v2-m3",
                llm_model="gpt-3.5-turbo",
                top_k=5,
                use_reranker=True,
                use_hybrid=True
            )
            print("OK RAG v2.1 initialisé")
        except BaseException as e:
            print(f"AVERTISSEMENT: RAG v2.1 non disponible: {e}")
            import traceback
            traceback.print_exc()
            rag_pipelines["v2.1"] = None

    # === RAG v2.2 ===
    # Désactivé : chromadb.PersistentClient provoque un Rust panic fatal (pyo3_runtime.PanicException)
    # qui corrompt l'état interne partagé de ChromaDB pour tout le processus,
    # empêchant v9/v10/v11 de s'initialiser ensuite.
    print("\n[2.2/10] RAG v2.2 désactivé (chroma_portrait — Rust panic ChromaDB)")
    rag_pipelines["v2.2"] = None

    # === RAG v3, v4, v5 ===
    # Désactivés : utilisent chroma_v2 qui provoque un Rust panic fatal
    print("\n[3-5/10] RAG v3, v4, v5 désactivés (chroma_v2 corrompu)")
    rag_pipelines["v3"] = None
    rag_pipelines["v4"] = None
    rag_pipelines["v5"] = None

    # === INITIALISER RAG v6 (G-Retriever GNN) ===
    print("\n[6/9] Initialisation RAG v6 (G-Retriever)...")
    if not V6_AVAILABLE:
        print("AVERTISSEMENT: RAG v6 non disponible (torch_geometric manquant)")
        rag_pipelines["v6"] = None
    else:
        try:
            # Utiliser None si NEO4J_PASSWORD n'est pas défini (connexion sans auth)
            neo4j_password = os.getenv("NEO4J_PASSWORD", None)
            if neo4j_password == "":
                neo4j_password = None

            rag_pipelines["v6"] = GRetrieverRAGPipeline(
                neo4j_uri="bolt://localhost:7687",
                neo4j_user="neo4j",
                neo4j_password=neo4j_password,
                chroma_path="./chroma_portrait",
                collection_name="portrait_verbatims",
                openai_api_key=openai_api_key
            )
            print("OK RAG v6 initialisé")
        except BaseException as e:
            print(f"AVERTISSEMENT: RAG v6 non disponible: {e}")
            rag_pipelines["v6"] = None

    # [7/9] Initialisation RAG v7 (LlamaIndex - graphes optionnels)
    print("\n[7/9] Initialisation RAG v7 (LlamaIndex - graphes optionnels)...")
    if not V7_AVAILABLE:
        print("AVERTISSEMENT: RAG v7 non disponible (llama-index manquant)")
        rag_pipelines["v7"] = None
    else:
        try:
            rag_pipelines["v7"] = LlamaIndexRAGPipeline(
                openai_api_key=openai_api_key,
                chroma_path="./chroma_v2",
                collection_name="communes_corses_v2",
                neo4j_uri="bolt://localhost:7687",
                neo4j_user="neo4j",
                neo4j_password=""
            )
            print("OK RAG v7 initialisé")
        except BaseException as e:
            print(f"AVERTISSEMENT: RAG v7 non disponible: {e}")
            import traceback
            traceback.print_exc()
            rag_pipelines["v7"] = None

    # [8/9] Initialisation RAG v8 (LlamaIndex FULL - graphes OBLIGATOIRES)
    print("\n[8/9] Initialisation RAG v8 (LlamaIndex FULL - graphes OBLIGATOIRES)...")
    if not V8_AVAILABLE:
        print("AVERTISSEMENT: RAG v8 non disponible (llama-index manquant)")
        rag_pipelines["v8"] = None
    else:
        try:
            rag_pipelines["v8"] = LlamaIndexRAGPipelineFull(
                openai_api_key=openai_api_key,
                chroma_path="./chroma_v2",
                collection_name="communes_corses_v2",
                neo4j_uri="bolt://localhost:7687",
                neo4j_user="neo4j",
                neo4j_password=""
            )
            print("OK RAG v8 initialisé")
        except BaseException as e:
            print(f"AVERTISSEMENT: RAG v8 non disponible: {e}")
            print("  Note: v8 nécessite APOC Extended installé dans Neo4j")
            import traceback
            traceback.print_exc()
            rag_pipelines["v8"] = None

    # [9/11] Initialisation RAG v9 (RAPTOR-lite)
    print("\n[9/11] Initialisation RAG v9 (RAPTOR-lite)...")
    if not V9_AVAILABLE:
        print("AVERTISSEMENT: RAG v9 non disponible (import échoué)")
        rag_pipelines["v9"] = None
    else:
        try:
            raptor = RaptorRetriever(
                chroma_path="./chroma_portrait",
                source_collection="portrait_verbatims",
                summary_collection="raptor_summaries",
            )
            raptor.init()
            rag_pipelines["v9"] = raptor
            print(f"OK RAG v9 initialisé ({raptor.summary_count} synthèses RAPTOR)")
        except BaseException as e:
            print(f"AVERTISSEMENT: RAG v9 non disponible: {e}")
            rag_pipelines["v9"] = None

    # [10/11] Initialisation RAG v10 (RAPTOR + Sous-questions)
    print("\n[10/11] Initialisation RAG v10 (RAPTOR + Sous-questions)...")
    if not V10_AVAILABLE:
        print("AVERTISSEMENT: RAG v10 non disponible (import échoué)")
        rag_pipelines["v10"] = None
    else:
        try:
            v10 = RaptorSubQuestionPipeline(
                chroma_path="./chroma_portrait",
                source_collection="portrait_verbatims",
                summary_collection="raptor_summaries",
            )
            v10.init()
            rag_pipelines["v10"] = v10
            print(f"OK RAG v10 initialisé")
        except BaseException as e:
            print(f"AVERTISSEMENT: RAG v10 non disponible: {e}")
            rag_pipelines["v10"] = None

    # [11/11] Initialisation RAG v11 (Agentic ReAct+CRAG)
    print("\n[11/11] Initialisation RAG v11 (Agentic ReAct+CRAG)...")
    if not V11_AVAILABLE:
        print("AVERTISSEMENT: RAG v11 non disponible (import échoué)")
        rag_pipelines["v11"] = None
    else:
        try:
            v11 = AgenticRAGPipeline(chroma_path="./chroma_portrait")
            v11.init()
            rag_pipelines["v11"] = v11
            print("OK RAG v11 initialisé (Agentic ReAct+CRAG)")
        except BaseException as e:
            print(f"AVERTISSEMENT: RAG v11 non disponible: {e}")
            import traceback
            traceback.print_exc()
            rag_pipelines["v11"] = None

    # [Ablations] V_vanilla_k10, V_vanilla_k25, V_decomp, V_decomp_raptor
    print("\n[Ablations] Initialisation des configs d'ablation...")
    if not ABLATIONS_AVAILABLE:
        print("AVERTISSEMENT: ablations non disponibles (import échoué)")
    else:
        try:
            abl_vanilla = VanillaRAG(chroma_path="./chroma_portrait")
            abl_vanilla.init()
            rag_pipelines["v_vanilla_k10"] = abl_vanilla
            rag_pipelines["v_vanilla_k25"] = abl_vanilla  # même instance, k passé à query()
            print("OK V_vanilla initialisé (k10 + k25 partagent la même instance)")
        except BaseException as e:
            print(f"AVERTISSEMENT: V_vanilla non disponible: {e}")

        try:
            abl_decomp = DecompOnlyRAG(chroma_path="./chroma_portrait")
            abl_decomp.init()
            rag_pipelines["v_decomp"] = abl_decomp
            print("OK V_decomp initialisé")
        except BaseException as e:
            print(f"AVERTISSEMENT: V_decomp non disponible: {e}")

        # V_decomp_raptor = v10 avec use_bilan=False — réutilise le pipeline v10 déjà initialisé
        if rag_pipelines.get("v10") is not None:
            rag_pipelines["v_decomp_raptor"] = rag_pipelines["v10"]
            print("OK V_decomp_raptor initialisé (réutilise pipeline v10 avec use_bilan=False)")
        else:
            print("AVERTISSEMENT: V_decomp_raptor non disponible (v10 absent)")

    # Résumé
    print("\n" + "="*60)
    available = [v for v, p in rag_pipelines.items() if p is not None]
    print(f"SYSTEMES RAG DISPONIBLES: {', '.join(available) if available else 'AUCUN'}")
    print("="*60 + "\n")

    if not available:
        raise RuntimeError("Aucune version du RAG n'a pu être initialisée")


# === GESTION DU CYCLE DE VIE ===

@asynccontextmanager
async def lifespan(app: FastAPI):
    """Gestion du cycle de vie de l'application"""
    # Startup
    try:
        initialize_all_rags()
    except BaseException as e:
        print(f"AVERTISSEMENT: Initialisation RAG partielle ou échouée: {e}")
        import traceback
        traceback.print_exc()
    yield
    # Shutdown (si nécessaire)
    pass


# === APPLICATION FASTAPI ===

app = FastAPI(
    title="API Chatbot RAG Multi-Version - Qualité de vie en Corse",
    description="API REST permettant de choisir entre les versions v1, v2, v2.1, v2.2, v3, v4, v5, v6, v7, v8 du système RAG",
    version="2.1.0",
    docs_url="/docs",
    redoc_url="/redoc",
    lifespan=lifespan
)

# === CONFIGURATION CORS ===

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # En production: spécifier les domaines autorisés
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# === FICHIER HTML FRONTEND ===

@app.get("/frontend", response_class=HTMLResponse, tags=["Frontend"])
async def serve_frontend():
    """Sert l'interface graphique HTML"""
    html_path = os.path.join(os.path.dirname(__file__), "example_frontend_multi_version.html")
    with open(html_path, "r", encoding="utf-8") as f:
        return HTMLResponse(content=f.read())


# === ENDPOINTS ===

@app.get("/", tags=["Root"])
async def root():
    """Endpoint racine - Redirige vers la documentation"""
    return {
        "message": "API Chatbot RAG Multi-Version - Qualité de vie en Corse",
        "documentation": "/docs",
        "health_check": "/api/health",
        "versions": "/api/versions",
        "query_endpoint": "/api/query"
    }


@app.get("/api/health", response_model=HealthResponse, tags=["Health"])
async def health_check():
    """
    Vérifie l'état de santé du serveur et des différentes versions RAG

    Returns:
        HealthResponse avec le statut de chaque version
    """
    return HealthResponse(
        status="healthy",
        rag_v1_initialized=rag_pipelines["v1"] is not None,
        rag_v2_initialized=rag_pipelines["v2"] is not None,
        rag_v2_1_initialized=rag_pipelines["v2.1"] is not None,
        rag_v2_2_initialized=rag_pipelines["v2.2"] is not None,
        rag_v3_initialized=rag_pipelines["v3"] is not None,
        rag_v4_initialized=rag_pipelines["v4"] is not None,
        rag_v5_initialized=rag_pipelines["v5"] is not None,
        rag_v6_initialized=rag_pipelines["v6"] is not None,
        rag_v7_initialized=rag_pipelines["v7"] is not None,
        rag_v8_initialized=rag_pipelines["v8"] is not None,
        rag_v9_initialized=rag_pipelines["v9"] is not None,
        rag_v10_initialized=rag_pipelines["v10"] is not None,
        rag_v11_initialized=rag_pipelines["v11"] is not None,
        timestamp=datetime.now().isoformat(),
        version="2.2.0"
    )


@app.get("/api/versions", response_model=List[VersionInfo], tags=["Versions"])
async def get_versions():
    """
    Liste les versions disponibles du RAG avec leurs caractéristiques

    Returns:
        Liste des versions avec leurs fonctionnalités
    """
    versions = [
        VersionInfo(
            version="v1",
            name="RAG Basique",
            description="Retrieval vectoriel simple avec génération LLM",
            available=rag_pipelines["v1"] is not None,
            features=[
                "Retrieval vectoriel (e5-base-v2)",
                "Génération avec GPT-3.5-turbo",
                "Simple et rapide"
            ]
        ),
        VersionInfo(
            version="v2",
            name="RAG Amélioré + Boost",
            description="Hybrid retrieval avec boost intelligent pour questionnaires",
            available=rag_pipelines["v2"] is not None,
            features=[
                "Hybrid retrieval (BM25 + Vector)",
                "Reranking avec cross-encoder",
                "Boost intelligent questionnaires",
                "Données quantitatives",
                "Meilleure précision"
            ]
        ),
        VersionInfo(
            version="v2.1",
            name="RAG LlamaIndex (équivalent v2)",
            description="Version LlamaIndex du RAG v2 pour comparaison",
            available=rag_pipelines["v2.1"] is not None,
            features=[
                "Framework LlamaIndex",
                "Hybrid retrieval (BM25 + Vector)",
                "Reranking avec SentenceTransformer",
                "Boost intelligent questionnaires",
                "Comparaison avec v2 'fait main'"
            ]
        ),
        VersionInfo(
            version="v2.2",
            name="RAG Portrait (filtres démographiques)",
            description="v2 + filtrage par profil démographique (âge, genre, profession)",
            available=rag_pipelines["v2.2"] is not None,
            features=[
                "Toutes les fonctionnalités v2",
                "Filtrage par tranche d'âge (15-24, 25-34, 35-49, 50-64, 65+)",
                "Filtrage par genre (Homme/Femme)",
                "Filtrage par profession (9 catégories)",
                "Filtrage par dimension qualité de vie",
                "Auto-détection des filtres dans la question",
                "Requêtes ciblées: 'Que pensent les jeunes de la santé ?'"
            ]
        ),
        VersionInfo(
            version="v3",
            name="RAG avec Ontologie",
            description="v2 + enrichissement sémantique via ontologie",
            available=rag_pipelines["v3"] is not None,
            features=[
                "Toutes les fonctionnalités v2",
                "Enrichissement de requête via ontologie",
                "Compréhension sémantique avancée",
                "Meilleure couverture thématique"
            ]
        ),
        VersionInfo(
            version="v4",
            name="RAG avec Analyse Croisée",
            description="v2 + décomposition de requêtes et analyse croisée multi-sources",
            available=rag_pipelines["v4"] is not None,
            features=[
                "Toutes les fonctionnalités v2",
                "Décomposition automatique de requêtes complexes",
                "Analyse croisée quantitatif/qualitatif",
                "Diversité de sources garantie",
                "Synthèse multi-perspectives"
            ]
        ),
        VersionInfo(
            version="v5",
            name="Graph-RAG avec Neo4j",
            description="RAG avancé avec graphe de connaissances Neo4j",
            available=rag_pipelines["v5"] is not None,
            features=[
                "Graphe de connaissances Neo4j",
                "Retrieval hybride (Vector + Graph)",
                "Embeddings BGE-M3 état de l'art",
                "Enrichissement ontologique",
                "Raisonnement sur graphe"
            ]
        ),
        VersionInfo(
            version="v6",
            name="G-Retriever (GNN)",
            description="v5 + Graph Neural Networks pour retrieval avancé",
            available=rag_pipelines["v6"] is not None,
            features=[
                "Toutes les fonctionnalités v5",
                "Graph Neural Networks (GraphSAGE)",
                "Embeddings de graphe appris",
                "Retrieval basé sur la structure",
                "Performance maximale"
            ]
        ),
        VersionInfo(
            version="v7",
            name="LlamaIndex Pipeline (graphes optionnels)",
            description="Framework LlamaIndex avec routing intelligent - fonctionne sans APOC Extended",
            available=rag_pipelines["v7"] is not None,
            features=[
                "Framework LlamaIndex complet",
                "3 modes: Router / Sub-question / Hybrid",
                "Graphes Neo4j optionnels (fallback vector-only)",
                "Fonctionne sans APOC Extended",
                "Fusion pondérée vector+graph (0.4/0.6) si graphe dispo",
                "Citations natives avec source_nodes",
                "Code simplifié (~200 lignes vs 500+)"
            ]
        ),
        VersionInfo(
            version="v8",
            name="LlamaIndex Pipeline FULL (graphes OBLIGATOIRES)",
            description="Framework LlamaIndex complet avec PropertyGraphIndex Neo4j - NÉCESSITE APOC Extended",
            available=rag_pipelines["v8"] is not None,
            features=[
                "Toutes les fonctionnalités v7",
                "PropertyGraphIndex Neo4j OBLIGATOIRE",
                "Routing automatique (LLM choisit vector ou graph)",
                "Décomposition automatique questions complexes",
                "Fusion pondérée vector+graph (0.4/0.6)",
                "Exploitation complète du graphe de connaissances",
                "NÉCESSITE: APOC Extended installé dans Neo4j"
            ]
        ),
        VersionInfo(
            version="v9",
            name="RAPTOR-lite (synthèses analytiques)",
            description="Synthèses pré-calculées par groupes démographiques avec fallback hiérarchique",
            available=rag_pipelines["v9"] is not None,
            features=[
                "349 synthèses pré-calculées (6 vues analytiques)",
                "Vues 1D: âge, profession, commune",
                "Vues 2D: âge×profession, âge×commune, profession×commune",
                "Détection automatique des dimensions dans la question",
                "Fallback hiérarchique (2D → 1D → sémantique)",
                "Synthèse + verbatims evidence",
                "Idéal pour questions analytiques (ex: 'Que pensent les jeunes ?')"
            ]
        ),
        VersionInfo(
            version="v10",
            name="RAPTOR + Sous-questions (pipeline 3 étapes)",
            description="Décomposition Mistral Large → réponses Claude Haiku par sous-question → synthèse Mistral Large",
            available=rag_pipelines["v10"] is not None,
            features=[
                "Décomposition automatique en N sous-questions (Mistral Large)",
                "Retrieval RAPTOR-lite par sous-question",
                "Réponse ciblée par sous-question (Claude Haiku)",
                "Synthèse finale avec filtrage du bruit (Mistral Large)",
                "Configurable : n_subquestions (1-8, défaut 5)",
            ]
        ),
        VersionInfo(
            version="v11",
            name="Agentic RAG (ReAct + CRAG gate)",
            description="Boucle ReAct itérative (Claude Sonnet) avec 5 outils, CRAG gate cosine BGE-M3, fast path v9 pour questions simples",
            available=rag_pipelines["v11"] is not None,
            features=[
                "Boucle ReAct dynamique (max 5 itérations)",
                "5 outils : summary_search, verbatim_search, score_lookup, geo_neighbors, decompose",
                "CRAG gate : évaluation cosine BGE-M3 après chaque retrieval",
                "Fast path v9 direct pour questions simples (classifier regex)",
                "Sélection dynamique d'outils par Claude Sonnet 4.6",
                "Reformulation automatique si contexte non pertinent (CRAG score < 0.2)",
                "Aucune dépendance supplémentaire (Anthropic Tool Use natif)",
            ]
        ),
    ]
    return versions


@app.post("/api/query", response_model=QueryResponse, tags=["Query"])
def query_rag(request: QueryRequest):
    """
    Pose une question au système RAG avec choix de version

    Args:
        request: QueryRequest contenant la question et la version souhaitée

    Returns:
        QueryResponse avec la réponse et les sources

    Raises:
        HTTPException 400: Version non disponible
        HTTPException 500: Erreur lors du traitement
    """
    # Vérifier que la version demandée est disponible
    rag = rag_pipelines.get(request.rag_version)

    if rag is None:
        available_versions = [v for v, p in rag_pipelines.items() if p is not None]
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Version '{request.rag_version}' non disponible. Versions disponibles: {', '.join(available_versions)}"
        )

    try:
        print(f"\n[{datetime.now().isoformat()}] Requête {request.rag_version} (LLM: {request.llm_model}): {request.question}")
        v10_scoring = {"applicable": False}  # valeur par defaut pour les versions != v10
        v10_sub_qa = None

        # Changer le modèle LLM si différent du défaut
        if hasattr(rag, 'llm_model'):
            rag.llm_model = request.llm_model
        if hasattr(rag, 'llm') and hasattr(rag.llm, 'model'):
            rag.llm.model = request.llm_model

        # Exécuter la requête selon la version
        if request.rag_version == "v1":
            # v1 : simple retrieval
            answer, retrieval_results = rag.query(
                question=request.question,
                k=request.k
            )

        elif request.rag_version == "v2":
            # v2 : hybrid + reranking + quantitatif
            answer, retrieval_results = rag.query(
                question=request.question,
                k=request.k,
                use_reranking=request.use_reranking,
                include_quantitative=request.include_quantitative,
                commune_filter=request.commune_filter
            )

        elif request.rag_version == "v2.1":
            # v2.1 : LlamaIndex équivalent v2
            result = rag.query(
                question=request.question,
                commune=request.commune_filter
            )
            answer = result["answer"]
            # Convertir les sources au format attendu
            retrieval_results = []
            for source in result["sources"]:
                retrieval_results.append(type('obj', (object,), {
                    'text': source['text'],
                    'score': source['score'] if source['score'] is not None else 0.0,
                    'metadata': source['metadata']
                })())

        elif request.rag_version == "v2.2":
            # v2.2 : Portrait avec filtres démographiques
            # Construire les filtres portrait explicites si fournis
            portrait_filters = None
            if not request.auto_detect_portrait:
                # Mode manuel: utiliser les filtres fournis par l'utilisateur
                has_explicit_filter = any([
                    request.portrait_age_min,
                    request.portrait_age_max,
                    request.portrait_genre,
                    request.portrait_profession,
                    request.portrait_dimension
                ])
                if has_explicit_filter:
                    portrait_filters = {
                        'has_portrait_filter': True,
                        'age_min': request.portrait_age_min,
                        'age_max': request.portrait_age_max,
                        'genre': request.portrait_genre,
                        'profession': request.portrait_profession,
                        'dimension': request.portrait_dimension
                    }

            answer, retrieval_results, detected_filters = rag.query(
                question=request.question,
                k=request.k,
                use_reranking=request.use_reranking,
                include_quantitative=request.include_quantitative,
                commune_filter=request.commune_filter,
                portrait_filters=portrait_filters,
                auto_detect_filters=request.auto_detect_portrait
            )

        elif request.rag_version == "v3":
            # v3 : v2 + ontologie
            answer, retrieval_results = rag.query(
                question=request.question,
                k=request.k,
                use_reranking=request.use_reranking,
                include_quantitative=request.include_quantitative,
                commune_filter=request.commune_filter,
                use_ontology_enrichment=request.use_ontology_enrichment
            )

        elif request.rag_version == "v4":
            # v4 : v2 + cross-analysis
            if request.use_cross_analysis:
                answer, retrieval_results = rag.query_with_cross_analysis(
                    question=request.question,
                    k=request.k,
                    use_reranking=request.use_reranking,
                    include_quantitative=request.include_quantitative,
                    commune_filter=request.commune_filter
                )
            else:
                # Fallback to regular query if cross-analysis is disabled
                answer, retrieval_results = rag.query(
                    question=request.question,
                    k=request.k,
                    use_reranking=request.use_reranking,
                    include_quantitative=request.include_quantitative,
                    commune_filter=request.commune_filter
                )

        elif request.rag_version == "v5":
            # v5 : Graph-RAG avec Neo4j
            answer, retrieval_results = rag.query(
                question=request.question,
                k=request.k,
                use_graph=True,
                use_reranking=request.use_reranking,
                commune_filter=request.commune_filter
            )

        elif request.rag_version == "v6":
            # v6 : G-Retriever avec GNN
            answer, retrieval_results = rag.query(
                question=request.question,
                k=request.k,
                use_gnn=True,
                use_reranking=request.use_reranking,
                commune_filter=request.commune_filter
            )

        elif request.rag_version == "v7":
            # v7 : LlamaIndex avec 3 modes (router, sub_question, hybrid) - graphes optionnels
            answer, retrieval_results = rag.query(
                question=request.question,
                mode=request.query_mode,  # router, sub_question, ou hybrid
                commune_filter=request.commune_filter,
                k=request.k
            )

        elif request.rag_version == "v8":
            # v8 : LlamaIndex FULL avec 3 modes (router, sub_question, hybrid) - graphes OBLIGATOIRES
            answer, retrieval_results = rag.query(
                question=request.question,
                mode=request.query_mode,  # router, sub_question, ou hybrid
                commune_filter=request.commune_filter,
                k=request.k
            )

        elif request.rag_version == "v9":
            # v9 : RAPTOR-lite — synthèses analytiques pré-calculées
            # Le retriever retourne (context_str, sources_list)
            # On passe le contexte à un LLM pour générer la réponse
            context_str, raptor_sources = rag.query(
                question=request.question,
                k=request.k
            )
            # Enrichissement OppChoVec si la question y fait référence
            opp_docs = []
            if _is_oppchovec_question(request.question):
                opp_docs = rag.query_oppchovec(request.question, k=3)
                if opp_docs:
                    opp_block = "\n\n[Scores OppChoVec (objectif/quanti) par commune]\n" + "\n\n".join(
                        d["text"] for d in opp_docs
                    )
                    context_str = context_str + opp_block
            # Générer une réponse via LLM
            from openai import OpenAI
            llm_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
            llm_response = llm_client.chat.completions.create(
                model=request.llm_model,
                messages=[
                    {"role": "system", "content": (
                        "Tu es un assistant spécialisé dans l'analyse de la qualité de vie en Corse. "
                        "Réponds à la question en te basant UNIQUEMENT sur le contexte fourni. "
                        "Le contexte contient plusieurs types de sources : "
                        "les blocs [Synthèse RAPTOR ...] et [Scores enquête ...] proviennent de l'enquête citoyenne (perceptions et satisfaction déclarées par les habitants) ; "
                        "les blocs [Scores OppChoVec ...] et [Données OppChoVec ...] sont LES INDICATEURS TERRITORIAUX OBJECTIFS DE RÉFÉRENCE "
                        "(scores agrégés 0-10 : Opportunités / Choix / Vécu, indépendants des opinions) ; "
                        "les blocs [Équipements et services commune ...] sont aussi des indicateurs objectifs factuels "
                        "(médecins, écoles, commerces, taux d'activité, taux de pauvreté, prix immobilier, etc.) ; "
                        "les blocs [Structure territoriale ...] et [Géographie ...] sont des données géographiques factuelles. "
                        "RÈGLES DE RÉDACTION — à respecter absolument : "
                        "N'utilise JAMAIS les termes techniques internes dans ta réponse : ne mentionne pas 'RAPTOR', 'subjectif/quali', 'objectif/quanti', 'OppChoVec intégré', etc. "
                        "Pour citer tes sources, utilise des formulations naturelles : 'selon l'enquête citoyenne', 'les habitants interrogés estiment que', 'les indicateurs territoriaux montrent que', 'selon les données géographiques'. "
                        "Pour une question de perception/satisfaction → appuie-toi sur les données d'enquête. "
                        "Pour une question d'indicateurs objectifs ou de rang territorial → appuie-toi sur les scores OppChoVec. "
                        "RÈGLE ABSOLUE : si le contexte contient un bloc [Scores OppChoVec ...] ou [Données OppChoVec ...], "
                        "tu DOIS le commenter — ne jamais écrire 'aucun indicateur objectif disponible' quand ce bloc est présent. "
                        "ATTENTION — définitions opérationnelles des sous-indicateurs OppChoVec (ne pas surinterprèter) : "
                        "Opp = éducation moyenne + diversité CSP + accessibilité mobilité + couverture TIC/haut débit. "
                        "Cho = % population avec droit de vote + absence de quartiers prioritaires (QPV) — "
                        "PAS une mesure de libertés individuelles au sens large. "
                        "Vec = revenu fiscal moyen + qualité du logement + stabilité de l'emploi + accès aux services en <20 min. "
                        "Ces scores sont des proxies statistiques (0-10, relatif aux 360 communes corses uniquement). "
                        "Ne pas extrapoler leur signification au-delà de ces composantes concrètes. "
                        "Pour une question mixte quanti/quali → croise scores OppChoVec + données d'enquête (et équipements si présents). "
                        "RÈGLE GÉOGRAPHIQUE STRICTE : réponds UNIQUEMENT sur la/les commune(s) mentionnée(s) dans la question. "
                        "Ne cite JAMAIS d'autres communes comme exemples si elles n'ont pas été demandées. "
                        "Si un bloc de contexte précise 'tous résidents' ou 'Corse entière', indique-le explicitement dans ta réponse. "
                        "Si les données sont limitées (pas de données CSP-spécifiques pour cette commune), sois sobre : "
                        "dis-le en une phrase et utilise ce que tu as (score OppChoVec communal, données globales CSP), sans te diluer sur d'autres communes. "
                        "Cite des extraits courts quand c'est pertinent. Sois factuel et nuancé. "
                        "IMPORTANT — honnêteté sur les lacunes : si le contexte fourni ne contient pas l'information nécessaire pour répondre à un point précis, dis-le clairement ('je ne dispose pas de données sur ce point'). "
                        "Ne brode jamais, n'invente jamais de chiffres, de noms ou de faits absents du contexte. Une réponse partielle honnête vaut mieux qu'une réponse complète inventée."
                    )},
                    {"role": "user", "content": f"Contexte :\n{context_str}\n\nQuestion : {request.question}"}
                ],
                temperature=0.3,
                max_tokens=1500
            )
            answer = llm_response.choices[0].message.content
            # Fusionner sources RAPTOR + sources OppChoVec
            opp_sources = [
                {
                    "rank": len(raptor_sources) + i,
                    "source_type": "oppchovec_score",
                    "commune": d["meta"].get("commune"),
                    "oppchovec_0_10": d["meta"].get("oppchovec_0_10"),
                    "opp_0_10": d["meta"].get("opp_0_10"),
                    "cho_0_10": d["meta"].get("cho_0_10"),
                    "vec_0_10": d["meta"].get("vec_0_10"),
                    "extrait": d["text"][:1500],
                }
                for i, d in enumerate(opp_docs)
            ]
            retrieval_results = raptor_sources + opp_sources

        elif request.rag_version == "v11":
            # v11 : Agentic RAG — boucle ReAct + CRAG gate + fast path v9
            answer, retrieval_results, v10_sub_qa = rag.query(
                question=request.question,
                k=request.k,
                max_iterations=request.max_iterations,
                use_fast_path=request.use_fast_path,
            )
            v10_scoring = {"applicable": False}

        elif request.rag_version == "v10":
            # v10 : RAPTOR + Sous-questions + Notation
            # OppChoVec est maintenant récupéré par sous-question dans RaptorRetriever.query().
            # On conserve un pré-fetch léger uniquement pour guider le décomposeur (extra_context)
            # quand la question est sur le bien-être global ou sur OppChoVec explicitement.
            enquete_docs_v10 = []
            extra_ctx_v10 = ""
            is_bieneetre_v10 = _is_bieneetre_question(request.question)

            if is_bieneetre_v10:
                # Injecter les scores d'enquête par commune pour guider la décomposition
                enquete_col = rag.retriever._extra_cols.get("enquete_scores_commune")
                if enquete_col:
                    try:
                        q_emb = rag.retriever._encode_query(request.question)
                        res = enquete_col.query(
                            query_embeddings=[q_emb],
                            n_results=min(3, enquete_col.count()),
                            include=["documents", "metadatas"],
                        )
                        enquete_docs_v10 = res["documents"][0] if res["documents"] else []
                    except Exception:
                        enquete_docs_v10 = []
                if enquete_docs_v10:
                    extra_ctx_v10 = "[Scores enquête par commune (subjectif/quanti)]\n" + "\n\n".join(enquete_docs_v10)

            answer, retrieval_results, v10_scoring, v10_sub_qa = rag.query(
                question=request.question,
                k=request.k,
                n_subquestions=request.n_subquestions,
                extra_context=extra_ctx_v10,
                force_mixed=is_bieneetre_v10,
            )

        elif request.rag_version == "v_decomp_raptor":
            answer, retrieval_results, v10_scoring, v10_sub_qa = rag.query(
                question=request.question,
                k=request.k,
                n_subquestions=request.n_subquestions,
                use_bilan=False,
            )

        elif request.rag_version == "v_decomp":
            answer, retrieval_results, v10_scoring, v10_sub_qa = rag.query(
                question=request.question,
                k=request.k,
                n_subquestions=request.n_subquestions,
            )

        elif request.rag_version == "v_vanilla_k10":
            answer, retrieval_results = rag.query(request.question, k=10)
            v10_sub_qa = None

        elif request.rag_version == "v_vanilla_k25":
            answer, retrieval_results = rag.query(request.question, k=25)
            v10_sub_qa = None

        # Convertir les résultats en format API
        if request.rag_version in ("v_vanilla_k10", "v_vanilla_k25", "v_decomp"):
            # ablations retournent List[Dict] avec clés: content, metadata, source_type, label
            sources = [
                Source(
                    content=result.get('content', ''),
                    score=1.0,
                    metadata={**result.get('metadata', {}),
                              "source_type": result.get('source_type', ''),
                              "label": result.get('label', '')}
                )
                for result in retrieval_results
            ]
        elif request.rag_version in ("v9", "v10", "v11", "v_decomp_raptor"):
            # v9/v10/v11 retournent List[Dict] avec clés: rank, type/commune, extrait, etc.
            sources = [
                Source(
                    content=result.get('extrait', result.get('content', '')),
                    score=1.0 - result.get('rank', 0) * 0.1,
                    metadata={k: v for k, v in result.items() if k not in ('extrait', 'content')}
                )
                for result in retrieval_results
            ]
        elif request.rag_version in ["v7", "v8"]:
            # v7/v8 retourne List[Dict] avec clés: content, score, metadata, source_type
            sources = [
                Source(
                    content=result['content'],
                    score=result['score'],
                    metadata=result['metadata']
                )
                for result in retrieval_results
            ]
        else:
            # v1-v6 retournent List[RetrievalResult]
            sources = [
                Source(
                    content=result.text,
                    score=result.score,
                    metadata=result.metadata
                )
                for result in retrieval_results
            ]

        # Construire le contexte complet (ce qui est passé au LLM)
        context_parts = [
            f"=== CONTEXTE PASSÉ AU LLM ({request.rag_version.upper()}) ===",
            f"Question: {request.question}",
            f"Date: {datetime.now().isoformat()}",
            f"Nombre de sources: {len(sources)}",
            "",
            "=== SOURCES UTILISÉES ===",
            ""
        ]

        for i, source in enumerate(sources, 1):
            context_parts.append(f"--- Source {i} (score: {source.score:.4f}) ---")
            # Ajouter les métadonnées pertinentes
            if source.metadata:
                meta_str = ", ".join(f"{k}: {v}" for k, v in source.metadata.items() if v and k not in ['content'])
                if meta_str:
                    context_parts.append(f"Métadonnées: {meta_str}")
            context_parts.append(f"Contenu:\n{source.content}")
            context_parts.append("")

        context_parts.append("=== FIN DU CONTEXTE ===")
        full_context = "\n".join(context_parts)

        # Construire les metadonnees (scoring v10 si applicable)
        metadata = {
            "k": request.k,
            "use_reranking": request.use_reranking if request.rag_version != "v1" else False,
            "use_ontology": request.use_ontology_enrichment if request.rag_version == "v3" else False,
            "use_cross_analysis": request.use_cross_analysis if request.rag_version == "v4" else False,
            "include_quantitative": request.include_quantitative if request.rag_version != "v1" else False,
            "commune_filter": request.commune_filter,
            "num_sources": len(sources)
        }
        if request.rag_version == "v10" and v10_scoring.get("applicable"):
            metadata["scoring"] = {
                "dimension": v10_scoring["dimension"],
                "score": v10_scoring["score"],
                "score_max": 5,
                "justification": v10_scoring["justification"],
            }

        # Construire la réponse
        response = QueryResponse(
            answer=answer,
            sources=sources,
            context=full_context,
            metadata=metadata,
            rag_version_used=request.rag_version,
            timestamp=datetime.now().isoformat(),
            sub_questions=v10_sub_qa if request.rag_version in ("v10", "v11", "v_decomp", "v_decomp_raptor") else None
        )

        print(f"[{datetime.now().isoformat()}] Réponse générée ({request.rag_version}) avec {len(sources)} sources")

        return response

    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Erreur lors du traitement avec {request.rag_version}: {str(e)}"
        )


# === ENDPOINTS EXPLORATION CORPUS ===

@app.get("/api/browse/filters", tags=["Browse"])
async def get_browse_filters():
    """Retourne les valeurs distinctes disponibles pour filtrer les verbatims."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("portrait_verbatims")
        all_metas = col.get(include=["metadatas"])["metadatas"]

        age_order = ["15-24", "25-34", "35-49", "50-64", "65+"]

        def sort_age(lst):
            return sorted(lst, key=lambda x: age_order.index(x) if x in age_order else 99)

        return {
            "communes":    sorted(set(m["nom"]        for m in all_metas if m.get("nom"))),
            "genres":      sorted(set(m["genre"]      for m in all_metas if m.get("genre"))),
            "age_ranges":  sort_age(list(set(m["age_range"]   for m in all_metas if m.get("age_range")))),
            "professions": sorted(set(m["profession"] for m in all_metas if m.get("profession"))),
            "dimensions":  sorted(set(m["dimension"]  for m in all_metas if m.get("dimension"))),
            "total":       len(all_metas),
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/verbatims", tags=["Browse"])
async def browse_verbatims(
    commune:    Optional[str] = None,
    genre:      Optional[str] = None,
    age_range:  Optional[str] = None,
    profession: Optional[str] = None,
    dimension:  Optional[str] = None,
):
    """Retourne tous les verbatims correspondant aux filtres (non paginé — 690 docs max)."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("portrait_verbatims")

        conditions = []
        if commune:    conditions.append({"nom":        {"$eq": commune}})
        if genre:      conditions.append({"genre":      {"$eq": genre}})
        if age_range:  conditions.append({"age_range":  {"$eq": age_range}})
        if profession: conditions.append({"profession": {"$eq": profession}})
        if dimension:  conditions.append({"dimension":  {"$eq": dimension}})

        kwargs: dict = {"include": ["documents", "metadatas"]}
        if len(conditions) == 1:
            kwargs["where"] = conditions[0]
        elif len(conditions) > 1:
            kwargs["where"] = {"$and": conditions}

        res = col.get(**kwargs)

        verbatims = [
            {"content": doc, "metadata": meta}
            for doc, meta in zip(res["documents"], res["metadatas"])
        ]
        verbatims.sort(key=lambda v: (
            v["metadata"].get("nom", ""),
            v["metadata"].get("age_exact", 0) or 0,
        ))

        return {"total": len(verbatims), "verbatims": verbatims}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/summaries", tags=["Browse"])
async def browse_summaries():
    """Retourne toutes les synthèses RAPTOR groupées par vue analytique."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("raptor_summaries")

        res = col.get(include=["documents", "metadatas"])

        grouped: Dict[str, list] = {}
        for doc, meta in zip(res["documents"], res["metadatas"]):
            view = meta.get("view_name", "unknown")
            grouped.setdefault(view, []).append({"content": doc, "metadata": meta})

        for view in grouped:
            grouped[view].sort(key=lambda x: (
                x["metadata"].get("dim1_value", ""),
                x["metadata"].get("dim2_value", ""),
            ))

        view_order = [
            "age_range*profession", "age_range*commune", "profession*commune",
            "age_range", "profession", "commune",
        ]
        ordered: Dict[str, list] = {v: grouped[v] for v in view_order if v in grouped}
        ordered.update({v: grouped[v] for v in grouped if v not in view_order})

        return {"total": sum(len(v) for v in ordered.values()), "views": ordered}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/quantitative", tags=["Browse"])
async def browse_quantitative(
    commune:   Optional[str] = None,
    genre:     Optional[str] = None,
    age_range: Optional[str] = None,
):
    """Retourne les données quantitatives du questionnaire (273 répondants)."""
    import pandas as pd
    import numpy as np

    CSV_PATH = "./donnees_brutes/sortie_questionnaire_traited.csv"
    try:
        df = pd.read_csv(CSV_PATH, index_col=0)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Fichier CSV introuvable")

    col_a = "Dans quelle commune résidez-vous ? ( A à S)"
    col_b = "Dans quelle commune résidez-vous ? (T à Z)"
    df["_commune"] = (df[col_a].fillna("") + df[col_b].fillna("")).str.strip()

    if commune:   df = df[df["_commune"] == commune]
    if genre:     df = df[df["genre"] == genre]
    if age_range: df = df[df["Catégorie age"] == age_range]

    score_map = {
        "Les services de transports":                                          "Transports",
        "L'accès à l'éducation":                                               "Education",
        "La couverture des réseaux téléphoniques":                             "Réseaux",
        "Les institutions étatiques (niveau de confiance)":                    "Institutions",
        "Le tourisme (ressenti localement)":                                   "Tourisme",
        "La sécurité":                                                         "Sécurité",
        "L'offre de santé ":                                                   "Santé",
        "Votre situation professionnelle":                                     "Emploi",
        "Vos revenus":                                                         "Revenus",
        "La répartition de votre temps entre travail et temps personnel":      "Temps pro/perso",
        "Votre logement":                                                      "Logement",
        "L'offre de services autour de chez vous":                             "Services",
        "Votre accès à la culture":                                            "Culture",
        "Vous sentez-vous bien entouré ?":                                     "Entourage",
        "Vous sentez-vous impliqué dans la vie locale de votre commune ?":     "Implication locale",
    }

    def safe(val):
        if val is None: return None
        if isinstance(val, float) and np.isnan(val): return None
        s = str(val).strip()
        return None if s in ("", "nan") else val

    rows = []
    for _, row in df.iterrows():
        r = {
            "id":                  safe(row.get("ID")),
            "commune":             row["_commune"] or None,
            "genre":               safe(row.get("genre")),
            "age":                 safe(row.get("Age")),
            "age_range":           safe(row.get("Catégorie age")),
            "profession":          safe(row.get("situation socioprofessionnelle")),
            "dimensions_choisies": safe(row.get(
                "Pour vous, qu'est-ce qui est important pour votre qualité de vie ? Choisissez 3 images"
            )),
            "bonheur":          safe(row.get(
                "Sur une échelle de 1 à 5, pourriez-vous estimer à quel point vous êtes heureux ces derniers temps ?"
            )),
            "qualite_vie":      safe(row.get(
                "Sur une échelle de 1 à 5, pourriez-vous évaluer votre qualité de vie ces derniers temps ?"
            )),
            "confiance_avenir": safe(row.get(
                "Sur une échelle de 1 à 5, pourriez-vous évaluer votre confiance en l'avenir ?"
            )),
            "scores": {short: safe(row.get(long)) for long, short in score_map.items()},
        }
        rows.append(r)

    def mean_score(key):
        vals = [r[key] for r in rows if r[key] is not None]
        return round(sum(float(v) for v in vals) / len(vals), 2) if vals else None

    stats = {
        "bonheur_moyen":         mean_score("bonheur"),
        "qualite_vie_moyenne":   mean_score("qualite_vie"),
        "confiance_avenir_moy":  mean_score("confiance_avenir"),
        "communes_disponibles":  sorted(set(r["commune"] for r in rows if r["commune"])),
        "age_ranges_disponibles": sorted(set(r["age_range"] for r in rows if r["age_range"])),
    }

    return {"total": len(rows), "stats": stats, "rows": rows}


@app.get("/api/browse/quanti_summaries", tags=["Browse"])
async def browse_quanti_summaries():
    """Retourne les synthèses RAPTOR enquête citoyenne groupées par vue analytique."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("raptor_enquete_summaries")
        # Exclure les docs spéciaux (methodology, global) — garder uniquement les synthèses groupées
        res = col.get(
            where={"source_type": {"$eq": "enquete_responses"}},
            include=["documents", "metadatas"],
        )

        grouped: Dict[str, list] = {}
        for doc, meta in zip(res["documents"], res["metadatas"]):
            view = meta.get("view_name", "unknown")
            grouped.setdefault(view, []).append({"content": doc, "metadata": meta})

        for view in grouped:
            grouped[view].sort(key=lambda x: (
                x["metadata"].get("dim1_value", ""),
                x["metadata"].get("dim2_value", ""),
            ))

        view_order = [
            "enquete_age_range*profession", "enquete_age_range*commune", "enquete_profession*commune",
            "enquete_age_range", "enquete_profession", "enquete_commune",
            "enquete_dimension*commune", "enquete_dimension*age_range", "enquete_dimension*profession",
            "enquete_dimension",
        ]
        ordered: Dict[str, list] = {v: grouped[v] for v in view_order if v in grouped}
        ordered.update({v: grouped[v] for v in grouped if v not in view_order})

        return {"total": sum(len(v) for v in ordered.values()), "views": ordered}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/equipements", tags=["Browse"])
async def browse_equipements(commune: Optional[str] = None):
    """Retourne les données équipements/emploi/services par commune."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("communes_equipements")
        kwargs: dict = {"include": ["documents", "metadatas"]}
        if commune:
            kwargs["where"] = {"commune": {"$eq": commune}}
        res = col.get(**kwargs)
        docs = [
            {"commune": m.get("commune", ""), "content": doc}
            for doc, m in zip(res["documents"], res["metadatas"])
        ]
        docs.sort(key=lambda x: x["commune"])
        communes = sorted(set(d["commune"] for d in docs if d["commune"]))
        return {"total": len(docs), "docs": docs, "communes": communes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/communes_profil", tags=["Browse"])
async def browse_communes_profil(commune: Optional[str] = None):
    """Retourne le profil démographique des répondants enquête par commune."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("communes_profil")
        kwargs: dict = {"include": ["documents", "metadatas"]}
        if commune:
            kwargs["where"] = {"commune": {"$eq": commune}}
        res = col.get(**kwargs)
        docs = [
            {
                "commune":      m.get("commune", ""),
                "n_repondants": m.get("n_repondants", 0),
                "content":      doc,
            }
            for doc, m in zip(res["documents"], res["metadatas"])
        ]
        docs.sort(key=lambda x: (-x["n_repondants"], x["commune"]))
        communes = sorted(set(d["commune"] for d in docs if d["commune"]))
        return {"total": len(docs), "docs": docs, "communes": communes}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/oppchovec", tags=["Browse"])
async def browse_oppchovec(epci: Optional[str] = None, sort_by: str = "oppchovec_0_10"):
    """Retourne les scores OppChoVec par commune, optionnellement filtrés par EPCI."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")

        # Communes individuelles
        opp = chroma.get_collection("oppchovec_scores")
        communes_res = opp.get(where={"source": "oppchovec_betti_0_10"}, include=["metadatas"])

        # Mapping commune → EPCI depuis communes_geo
        geo = chroma.get_collection("communes_geo")
        geo_res = geo.get(include=["metadatas"])
        epci_map = {m["commune"]: m.get("epci", "") for m in geo_res["metadatas"]}

        communes = []
        for m in communes_res["metadatas"]:
            commune = m["commune"]
            epci_val = epci_map.get(commune, "")
            if epci and epci_val != epci:
                continue
            communes.append({
                "commune":       commune,
                "epci":          epci_val,
                "score_global":  round(m.get("oppchovec_0_10", 0), 4),
                "opp":           round(m.get("opp_0_10", 0), 4),
                "cho":           round(m.get("cho_0_10", 0), 4),
                "vec":           round(m.get("vec_0_10", 0), 4),
                "rank_total":    m.get("rank_total"),
                "rank_opp":      m.get("rank_opp"),
                "rank_cho":      m.get("rank_cho"),
                "rank_vec":      m.get("rank_vec"),
            })

        sort_key = {"oppchovec_0_10": "score_global", "opp_0_10": "opp",
                    "cho_0_10": "cho", "vec_0_10": "vec"}.get(sort_by, "score_global")
        communes.sort(key=lambda x: x[sort_key], reverse=True)

        # Agrégats EPCI
        agg_res = opp.get(where={"source": "oppchovec_aggregate"}, include=["metadatas"])
        aggregates = [
            {
                "zone":      m["zone"],
                "epci":      m.get("epci", ""),
                "score_global": round(m.get("oppchovec_0_10", 0), 4),
                "opp":       round(m.get("opp_0_10", 0), 4),
                "cho":       round(m.get("cho_0_10", 0), 4),
                "vec":       round(m.get("vec_0_10", 0), 4),
                "nb_communes": m.get("nb_communes", 0),
            }
            for m in agg_res["metadatas"]
        ]
        aggregates.sort(key=lambda x: x["score_global"], reverse=True)

        # Liste des EPCI disponibles
        all_epcis = sorted(set(v for v in epci_map.values() if v))

        return {
            "total": len(communes),
            "communes": communes,
            "aggregates": aggregates,
            "epcis": all_epcis,
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/entretiens", tags=["Browse"])
async def browse_entretiens(commune: Optional[str] = None):
    """Retourne les entretiens qualitatifs, optionnellement filtrés par commune."""
    try:
        import chromadb as _chromadb
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("portrait_entretiens")

        kwargs: dict = {"include": ["documents", "metadatas"]}
        if commune:
            kwargs["where"] = {"commune": {"$eq": commune}}

        res = col.get(**kwargs)
        entretiens: Dict[str, list] = {}
        for doc, meta in zip(res["documents"], res["metadatas"]):
            key = f"{meta.get('commune', '?')} — Entretien {meta.get('num_entretien', '?')}"
            entretiens.setdefault(key, []).append({
                "chunk_idx": meta.get("chunk_idx", 0),
                "content":   doc,
                "commune":   meta.get("commune", ""),
                "num":       meta.get("num_entretien"),
            })
        # Trier les chunks par index
        for k in entretiens:
            entretiens[k].sort(key=lambda x: x["chunk_idx"])

        communes_list = sorted(set(meta.get("commune", "") for meta in res["metadatas"] if meta.get("commune")))
        return {"total_chunks": len(res["documents"]), "entretiens": entretiens, "communes": communes_list}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.get("/api/browse/classements_communes", tags=["Browse"])
async def browse_classements_communes():
    """Retourne les classements des communes par score perçu (enquête QdV Likert)."""
    import pandas as pd
    import numpy as np
    import unicodedata

    def _norm(s: str) -> str:
        return "".join(c for c in unicodedata.normalize("NFD", s.lower())
                       if unicodedata.category(c) != "Mn")

    def _find_col(columns, kw):
        kw_n = _norm(kw)
        for col in columns:
            if kw_n in _norm(col):
                return col
        return None

    LIKERT_MAP = {
        "Très satisfait": 5.0, "Satisfait": 4.0, "Neutre": 3.0,
        "Peu satisfait": 2.0, "Très peu satisfait": 1.0,
        "Très bien entouré": 5.0, "Bien entouré": 4.0, "Moyennement entouré": 3.0,
        "Peu entouré": 2.0, "Très peu entouré": 1.0,
        "Très impliqué": 5.0, "Impliqué": 4.0, "Moyennement Impliqué": 3.0,
        "Peu impliqué": 2.0, "Très peu impliqué": 1.0,
    }

    DIM_KEYWORDS = {
        "Transports":             ["services de transport", "transports en commun", "seau routier", "encombrement"],
        "Santé":                  ["offre de sant", "decins generalistes", "attente pour avoir", "decins sp"],
        "Éducation":              ["ducation"],
        "Logement":               ["votre logement"],
        "Revenus":                ["vos revenus"],
        "Emploi":                 ["votre situation professionnelle"],
        "Sécurité":               ["curit"],
        "Culture":                ["la culture"],
        "Services de proximité":  ["services autour de chez vous"],
        "Réseau":                 ["couverture"],
        "Ratio vie pro/vie perso":["partition de votre temps"],
        "Communauté et relations":["entour", "impliqu"],
        "Tourisme":               ["tourisme"],
        "Institutions":           ["institutions"],
    }
    MIN_N = 3

    CSV_PATH = "./donnees_brutes/sortie_questionnaire_traited.csv"
    try:
        df = pd.read_csv(CSV_PATH, index_col=0)
    except FileNotFoundError:
        raise HTTPException(status_code=404, detail="Fichier CSV introuvable")

    col_a = _find_col(list(df.columns), "commune")
    col_b = None
    for col in df.columns:
        if col != col_a and _norm("commune") in _norm(col):
            col_b = col
            break
    if col_a and col_b:
        df["_commune"] = df[col_a].fillna(df[col_b]).astype(str).str.strip()
    elif col_a:
        df["_commune"] = df[col_a].astype(str).str.strip()
    else:
        raise HTTPException(status_code=500, detail="Colonne commune introuvable")

    # Résoudre colonnes Likert
    col_map: Dict[str, list] = {}
    for dim, kws in DIM_KEYWORDS.items():
        cols = [c for kw in kws for c in [_find_col(list(df.columns), kw)] if c]
        seen: list = []
        for c in cols:
            if c not in seen:
                seen.append(c)
        if seen:
            col_map[dim] = seen

    col_bonheur   = _find_col(list(df.columns), "heureux")
    col_qdv       = _find_col(list(df.columns), "qualite de vie")
    col_confiance = _find_col(list(df.columns), "confiance en l")

    def to_num(val) -> Optional[float]:
        if not isinstance(val, str):
            try:
                fv = float(val)
                return fv if 1.0 <= fv <= 5.0 else None
            except (ValueError, TypeError):
                return None
        v = val.strip()
        if "moyennement" in v.lower():
            return 3.0
        return LIKERT_MAP.get(v)

    def mean_col(grp, col) -> Optional[float]:
        if not col or col not in grp.columns:
            return None
        vals = [to_num(v) for v in grp[col] if to_num(v) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    def dim_mean(grp, cols) -> Optional[float]:
        vals = [to_num(v) for col in cols if col in grp.columns
                for v in grp[col] if to_num(v) is not None]
        return round(sum(vals) / len(vals), 2) if vals else None

    results: Dict[str, list] = {k: [] for k in ["global", "bien_etre", "confiance_avenir"] + list(col_map.keys())}

    for commune, grp in df.groupby("_commune"):
        if not commune or commune.lower() in ("nan", "inconnue", ""):
            continue
        n = len(grp)
        if n < MIN_N:
            continue

        dim_scores = {dim: dim_mean(grp, cols) for dim, cols in col_map.items()}
        valid_dims = [s for s in dim_scores.values() if s is not None]
        global_sc = round(sum(valid_dims) / len(valid_dims), 2) if valid_dims else None

        be_vals = [v for v in [mean_col(grp, col_bonheur), mean_col(grp, col_qdv)] if v is not None]
        be_sc = round(sum(be_vals) / len(be_vals), 2) if be_vals else None
        conf_sc = mean_col(grp, col_confiance)

        entry = {"commune": commune, "n": n}

        if global_sc is not None:
            results["global"].append({**entry, "score": global_sc})
        if be_sc is not None:
            results["bien_etre"].append({**entry, "score": be_sc})
        if conf_sc is not None:
            results["confiance_avenir"].append({**entry, "score": conf_sc})
        for dim, sc in dim_scores.items():
            if sc is not None:
                results[dim].append({**entry, "score": sc})

    for key in results:
        results[key].sort(key=lambda x: -x["score"])
        for i, row in enumerate(results[key], 1):
            row["rang"] = i

    return {
        "rankings": results,
        "dimensions": list(col_map.keys()),
        "min_n": MIN_N,
    }


@app.get("/api/stats", tags=["Browse"])
async def get_stats():
    """
    Retourne les effectifs réels par dimension d'analyse.
    Toutes les valeurs proviennent des données brutes (ChromaDB + CSV).
    Rien n'est inventé ni estimé.
    """
    from collections import Counter
    import chromadb as _chromadb
    import pandas as pd

    result = {}

    # ── 1. Verbatims (portrait_verbatims, ChromaDB) ───────────────────────────
    try:
        chroma = _chromadb.PersistentClient(path="./chroma_portrait")
        col = chroma.get_collection("portrait_verbatims")
        all_metas = col.get(include=["metadatas"])["metadatas"]

        result["verbatims"] = {
            "total":          len(all_metas),
            "par_commune":    dict(Counter(m["nom"]        for m in all_metas if m.get("nom")).most_common()),
            "par_age_range":  dict(Counter(m["age_range"]  for m in all_metas if m.get("age_range")).most_common()),
            "par_profession": dict(Counter(m["profession"] for m in all_metas if m.get("profession")).most_common()),
            "par_dimension":  dict(Counter(m["dimension"]  for m in all_metas if m.get("dimension")).most_common()),
            "par_genre":      dict(Counter(m["genre"]      for m in all_metas if m.get("genre")).most_common()),
        }
    except Exception as e:
        result["verbatims"] = {"error": str(e)}

    # ── 2. Synthèses RAPTOR (raptor_summaries, ChromaDB) ──────────────────────
    try:
        chroma2 = _chromadb.PersistentClient(path="./chroma_portrait")
        col2 = chroma2.get_collection("raptor_summaries")
        all_sum_metas = col2.get(include=["metadatas"])["metadatas"]

        par_vue: Dict[str, dict] = {}
        for m in all_sum_metas:
            vn = m.get("view_name", "inconnu")
            nc = int(m.get("num_chunks", 0))
            if vn not in par_vue:
                par_vue[vn] = {"total": 0, "chunks_total": 0}
            par_vue[vn]["total"]        += 1
            par_vue[vn]["chunks_total"] += nc

        vue_order = [
            "age_range*profession", "age_range*commune", "profession*commune",
            "dimension*commune", "dimension*age_range", "dimension*profession",
            "age_range", "profession", "commune", "dimension",
        ]
        par_vue_ordered = {}
        for vn in vue_order:
            if vn in par_vue:
                d = par_vue[vn]
                par_vue_ordered[vn] = {
                    "total":        d["total"],
                    "chunks_moyen": round(d["chunks_total"] / d["total"], 1),
                }
        for vn, d in par_vue.items():
            if vn not in par_vue_ordered:
                par_vue_ordered[vn] = {
                    "total":        d["total"],
                    "chunks_moyen": round(d["chunks_total"] / d["total"], 1),
                }

        result["raptor"] = {
            "total":   len(all_sum_metas),
            "par_vue": par_vue_ordered,
        }
    except Exception as e:
        result["raptor"] = {"error": str(e)}

    # ── 3. Répondants enquête (CSV) ───────────────────────────────────────────
    CSV_PATH = "./donnees_brutes/sortie_questionnaire_traited.csv"
    try:
        df = pd.read_csv(CSV_PATH, index_col=0)
        col_a = "Dans quelle commune résidez-vous ? ( A à S)"
        col_b = "Dans quelle commune résidez-vous ? (T à Z)"
        df["_commune"] = (df[col_a].fillna("") + df[col_b].fillna("")).str.strip()

        age_order = ["15-24", "25-34", "35-49", "50-64", "65+"]
        age_counts = dict(Counter(df["Catégorie age"].dropna()))
        age_counts_ordered = {k: age_counts[k] for k in age_order if k in age_counts}
        age_counts_ordered.update({k: v for k, v in age_counts.items() if k not in age_counts_ordered})

        result["enquete"] = {
            "total":          len(df),
            "par_commune":    dict(Counter(df["_commune"].replace("", None).dropna()).most_common()),
            "par_age_range":  age_counts_ordered,
            "par_profession": dict(Counter(df["situation socioprofessionnelle"].dropna()).most_common()),
            "par_genre":      dict(Counter(df["genre"].dropna()).most_common()),
        }
    except Exception as e:
        result["enquete"] = {"error": str(e)}

    # ── 4. OppChoVec (oppchovec_scores, ChromaDB) ─────────────────────────────
    try:
        chroma2 = _chromadb.PersistentClient(path="./chroma_portrait")
        opp_col = chroma2.get_collection("oppchovec_scores")
        opp_metas = opp_col.get(where={"source": "oppchovec_betti_0_10"}, include=["metadatas"])["metadatas"]
        geo_col = chroma2.get_collection("communes_geo")
        geo_metas = geo_col.get(include=["metadatas"])["metadatas"]
        epci_map = {m["commune"]: m.get("epci", "") for m in geo_metas}

        scores = [m.get("oppchovec_0_10", 0) for m in opp_metas]
        par_epci = dict(Counter(epci_map.get(m["commune"], "Inconnu") for m in opp_metas).most_common())

        result["oppchovec"] = {
            "total":       len(opp_metas),
            "score_moyen": round(sum(scores) / len(scores), 4) if scores else 0,
            "score_max":   round(max(scores), 4) if scores else 0,
            "score_min":   round(min(scores), 4) if scores else 0,
            "par_epci":    par_epci,
        }
    except Exception as e:
        result["oppchovec"] = {"error": str(e)}

    # ── 5. Entretiens qualitatifs (portrait_entretiens, ChromaDB) ─────────────
    try:
        chroma3 = _chromadb.PersistentClient(path="./chroma_portrait")
        ent_col = chroma3.get_collection("portrait_entretiens")
        ent_metas = ent_col.get(include=["metadatas"])["metadatas"]
        entretiens_par_commune = dict(Counter(m.get("commune", "?") for m in ent_metas).most_common())
        n_entretiens = len(set((m.get("commune"), m.get("num_entretien")) for m in ent_metas))
        result["entretiens"] = {
            "total_chunks":       len(ent_metas),
            "nb_entretiens":      n_entretiens,
            "par_commune":        entretiens_par_commune,
        }
    except Exception as e:
        result["entretiens"] = {"error": str(e)}

    return result


# ════════════════════════════════════════════════════════════════
# ENDPOINTS ÉVALUATION (eval_from_excel.py)
# ════════════════════════════════════════════════════════════════

# Stockage des jobs dans des fichiers temporaires (partagé entre workers uvicorn)
_EVAL_JOBS_DIR = tempfile.gettempdir()


def _job_state_path(job_id: str) -> str:
    return os.path.join(_EVAL_JOBS_DIR, f"eval_job_{job_id}.json")


def _job_lines_path(job_id: str) -> str:
    return os.path.join(_EVAL_JOBS_DIR, f"eval_job_{job_id}_lines.txt")


def _read_job_state(job_id: str) -> dict | None:
    path = _job_state_path(job_id)
    if not os.path.exists(path):
        return None
    try:
        with open(path, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _write_job_state(job_id: str, state: dict):
    path = _job_state_path(job_id)
    tmp = path + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(state, f)
    os.replace(tmp, path)


def _append_job_line(job_id: str, line: str):
    with open(_job_lines_path(job_id), "a", encoding="utf-8") as f:
        f.write(line + "\n")


def _read_job_lines(job_id: str) -> list:
    path = _job_lines_path(job_id)
    if not os.path.exists(path):
        return []
    with open(path, encoding="utf-8") as f:
        return [ln.rstrip("\n") for ln in f.readlines()]


class EvalStartRequest(BaseModel):
    excel_path: str
    rag_version: str = "v10"
    k: int = 7
    rows: List[int] = []       # liste de numéros de ligne Excel (vide = toutes)
    no_judge: bool = False
    no_robust: bool = False
    output_dir: str = "comparaisons_rag"


async def _run_eval_subprocess(job_id: str, req: EvalStartRequest):
    """Lance eval_from_excel.py en sous-processus et capture la sortie ligne par ligne."""
    state = _read_job_state(job_id) or {"status": "running", "result_path": None, "error": None, "pid": None}
    try:
        python = sys.executable
        script = os.path.join(os.path.dirname(__file__), "eval_from_excel.py")
        cmd = [
            python, script,
            "--input", req.excel_path,
            "--version", req.rag_version,
            "--k", str(req.k),
            "--output", req.output_dir,
        ]
        if req.rows:
            cmd += ["--rows", ",".join(str(r) for r in req.rows)]
        if req.no_judge:
            cmd.append("--no-judge")
        if req.no_robust:
            cmd.append("--no-robust")

        proc = await asyncio.create_subprocess_exec(
            *cmd,
            stdout=asyncio.subprocess.PIPE,
            stderr=asyncio.subprocess.STDOUT,
            cwd=os.path.dirname(__file__),
        )
        state["pid"] = proc.pid
        _write_job_state(job_id, state)

        async for raw_line in proc.stdout:
            line = raw_line.decode("utf-8", errors="replace").rstrip()
            _append_job_line(job_id, line)
            # Détecter le chemin Excel en sortie
            if line.startswith("Markdown sauvegarde :"):
                state["result_path"] = line.split(":", 1)[1].strip()
                _write_job_state(job_id, state)

        await proc.wait()
        state["status"] = "done" if proc.returncode == 0 else "error"
        if proc.returncode != 0:
            state["error"] = f"Processus terminé avec code {proc.returncode}"
        _write_job_state(job_id, state)

    except Exception as e:
        state["status"] = "error"
        state["error"] = str(e)
        _append_job_line(job_id, f"[ERREUR] {e}")
        _write_job_state(job_id, state)


def _read_eval_excel(path: str) -> list:
    """Lit un Excel d'évaluation et retourne la liste des questions avec leurs métriques."""
    import openpyxl
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    questions = []
    COL_METRIC_MAP = {4: "factual", 5: "binary", 6: "judge",
                      7: "refusal", 8: "halluc", 9: "overconf", 10: "robust"}
    for r in range(2, ws.max_row + 1):
        q = ws.cell(r, 3).value
        if not q:
            continue
        metrics = [name for col, name in COL_METRIC_MAP.items()
                   if ws.cell(r, col).value == "X"]
        questions.append({
            "excel_row":  r,
            "section":    ws.cell(r, 1).value or "",
            "subsection": ws.cell(r, 2).value or "",
            "question":   str(q).strip(),
            "metrics":    metrics,
        })
    return questions


class EvalPreviewRequest(BaseModel):
    path: str


@app.post("/api/eval/preview", tags=["Evaluation"])
def eval_preview(req: EvalPreviewRequest):
    """Retourne la liste des questions d'un fichier Excel d'évaluation (chemin serveur)."""
    path = req.path.strip()
    if not os.path.exists(path):
        raise HTTPException(status_code=404, detail=f"Fichier introuvable : {path}")
    try:
        questions = _read_eval_excel(path)
        return {"questions": questions, "total": len(questions)}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/eval/upload", tags=["Evaluation"])
async def eval_upload(file: UploadFile = File(...)):
    """Upload un fichier Excel d'évaluation — le stocke temporairement et retourne son chemin serveur."""
    suffix = os.path.splitext(file.filename or "eval.xlsx")[1] or ".xlsx"
    tmp_dir = os.path.join(tempfile.gettempdir(), "rag_eval_uploads")
    os.makedirs(tmp_dir, exist_ok=True)
    dest = os.path.join(tmp_dir, f"{uuid.uuid4().hex[:8]}{suffix}")
    try:
        content = await file.read()
        with open(dest, "wb") as f:
            f.write(content)
        questions = _read_eval_excel(dest)
        return {"server_path": dest, "questions": questions, "total": len(questions)}
    except Exception as e:
        if os.path.exists(dest):
            os.remove(dest)
        raise HTTPException(status_code=500, detail=str(e))


@app.post("/api/eval/start", tags=["Evaluation"])
async def eval_start(req: EvalStartRequest, background_tasks: BackgroundTasks):
    """Démarre un job d'évaluation en arrière-plan."""
    if not os.path.exists(req.excel_path):
        raise HTTPException(status_code=400, detail=f"Fichier introuvable : {req.excel_path}")
    job_id = str(uuid.uuid4())[:8]
    _write_job_state(job_id, {"status": "running", "result_path": None, "error": None, "pid": None})
    open(_job_lines_path(job_id), "w").close()  # fichier de logs vide
    background_tasks.add_task(_run_eval_subprocess, job_id, req)
    return {"job_id": job_id}


@app.get("/api/eval/{job_id}/stream", tags=["Evaluation"])
async def eval_stream(job_id: str):
    """SSE : flux de progression du job d'évaluation."""
    if _read_job_state(job_id) is None:
        raise HTTPException(status_code=404, detail="Job introuvable")

    async def generate():
        sent = 0
        while True:
            lines = _read_job_lines(job_id)
            while sent < len(lines):
                yield f"data: {lines[sent]}\n\n"
                sent += 1
            state = _read_job_state(job_id)
            if state and state["status"] in ("done", "error"):
                yield f"data: __STATUS__{state['status']}__\n\n"
                if state.get("result_path"):
                    yield f"data: __RESULT_PATH__{state['result_path']}__\n\n"
                break
            await asyncio.sleep(0.5)

    return StreamingResponse(generate(), media_type="text/event-stream",
                             headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


@app.get("/api/eval/{job_id}/status", tags=["Evaluation"])
async def eval_status(job_id: str):
    """Statut d'un job d'évaluation."""
    state = _read_job_state(job_id)
    if state is None:
        raise HTTPException(status_code=404, detail="Job introuvable")
    return {
        "job_id": job_id,
        "status": state["status"],
        "lines_count": len(_read_job_lines(job_id)),
        "result_path": state.get("result_path"),
        "error": state.get("error"),
    }


@app.get("/api/eval/{job_id}/result", tags=["Evaluation"])
async def eval_result(job_id: str):
    """Télécharge le rapport Markdown d'un job terminé."""
    state = _read_job_state(job_id)
    if state is None:
        raise HTTPException(status_code=404, detail="Job introuvable")
    if state["status"] != "done":
        raise HTTPException(status_code=400, detail=f"Job non terminé (statut : {state['status']})")
    path = state.get("result_path")
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fichier résultat introuvable")
    filename = os.path.basename(path)
    return FileResponse(path, filename=filename, media_type="text/markdown")


@app.get("/api/eval/{job_id}/content", tags=["Evaluation"])
async def eval_content(job_id: str):
    """Retourne le contenu texte du rapport Markdown pour affichage inline."""
    state = _read_job_state(job_id)
    if state is None:
        raise HTTPException(status_code=404, detail="Job introuvable")
    if state["status"] != "done":
        raise HTTPException(status_code=400, detail=f"Job non terminé (statut : {state['status']})")
    path = state.get("result_path")
    if not path or not os.path.exists(path):
        raise HTTPException(status_code=404, detail="Fichier résultat introuvable")
    with open(path, encoding="utf-8") as f:
        content = f.read()
    return {"content": content, "filename": os.path.basename(path)}


# ════════════════════════════════════════════════════════════════
# ENDPOINTS DEBUG RETRIEVAL
# ════════════════════════════════════════════════════════════════

def _get_eval_functions():
    """Importe les fonctions de scoring depuis eval_from_excel.py."""
    import importlib.util, sys
    eval_path = os.path.join(os.path.dirname(__file__), "eval_from_excel.py")
    spec = importlib.util.spec_from_file_location("eval_from_excel", eval_path)
    mod  = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


@app.get("/api/eval/questions", tags=["Debug"])
def get_eval_questions(excel_path: str):
    """Lit le fichier Excel et retourne les questions avec leur ground-truth retrieval."""
    if not os.path.exists(excel_path):
        raise HTTPException(status_code=404, detail=f"Fichier introuvable : {excel_path}")
    try:
        mod = _get_eval_functions()
        questions = mod.load_questions(excel_path)
        return {
            "count": len(questions),
            "questions": [
                {
                    "excel_row":    q["excel_row"],
                    "section":      q["section"],
                    "subsection":   q["subsection"],
                    "question":     q["question"],
                    "retrieval_gt": q["retrieval_gt"],
                    "do_retrieval": q["do_retrieval"],
                }
                for q in questions
            ]
        }
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


class DebugRetrievalRequest(BaseModel):
    question:     str
    retrieval_gt: Optional[str] = None
    rag_version:  str = "v10"
    k:            int = 7


@app.post("/api/eval/debug_retrieval", tags=["Debug"])
def debug_retrieval(req: DebugRetrievalRequest):
    """
    Lance une requête RAG et retourne les sources classifiées + le score retrieval.
    Permet de diagnostiquer recall/precision question par question.
    """
    rag = rag_pipelines.get(req.rag_version)
    if rag is None:
        raise HTTPException(status_code=400, detail=f"Version {req.rag_version} non disponible")

    try:
        mod = _get_eval_functions()

        # 1. Retrieval uniquement (pas de LLM) — on va chercher les sources
        # directement via le RaptorRetriever commun à v9/v10/v11
        retriever = getattr(rag, 'retriever', None) or rag
        if hasattr(retriever, 'query'):
            _, retrieval_results = retriever.query(question=req.question, k=req.k)
        else:
            raise HTTPException(status_code=400, detail="Version non supportée pour le debug")

        # Normaliser v1 (RetrievalResult dataclass) en dicts compatibles classify_source
        def _normalize_source(s):
            if isinstance(s, dict):
                return s
            meta = getattr(s, 'metadata', {}) or {}
            return {
                "source_type": "verbatim_evidence",
                "commune":     meta.get("nom", ""),
                "dimension":   meta.get("dimension", ""),
                "genre":       meta.get("genre", ""),
                "age":         meta.get("age_exact", ""),
                "extrait":     getattr(s, 'text', "")[:400],
            }
        retrieval_results = [_normalize_source(s) for s in retrieval_results]

        # 2. Classer chaque source
        classified = []
        for s in retrieval_results:
            cat = mod.classify_source(s)
            classified.append({
                "category":    cat,
                "type":        s.get("type", ""),
                "source_type": s.get("source_type", ""),
                "commune":     s.get("commune", ""),
                "view":        s.get("view", ""),
                "rank":        s.get("rank", 0),
                "extrait":     (s.get("extrait", "") or "")[:200],
            })

        # 3. Scorer le retrieval
        score = mod.score_retrieval(retrieval_results, req.retrieval_gt)

        # 4. Parser le GT pour affichage
        gt_items = mod.parse_retrieval_ground_truth(req.retrieval_gt) if req.retrieval_gt else []

        return {
            "question":    req.question,
            "score":       score,
            "gt_items":    gt_items,
            "classified":  classified,
            "n_sources":   len(retrieval_results),
        }

    except HTTPException:
        raise
    except Exception as e:
        import traceback; traceback.print_exc()
        raise HTTPException(status_code=500, detail=str(e))


# === POINT D'ENTRÉE ===

if __name__ == "__main__":
    # Configuration du serveur
    host = os.getenv("API_HOST", "0.0.0.0")
    port = int(os.getenv("API_PORT", "8000"))
    reload = os.getenv("API_RELOAD", "false").lower() == "true"

    print("\n" + "="*60)
    print("DÉMARRAGE DU SERVEUR API MULTI-VERSION")
    print("="*60)
    print(f"Host: {host}")
    print(f"Port: {port}")
    print(f"Documentation: http://localhost:{port}/docs")
    print("="*60 + "\n")

    # Démarrer le serveur
    uvicorn.run(
        "api_server_multi_version:app",
        host=host,
        port=port,
        reload=reload,
        log_level="info"
    )
