"""
Script de comparaison entre RAG v2 (fait main) et v2.1 (LlamaIndex)
Permet de comparer les résultats, performances et qualité des réponses

Format de sortie identique à compare_rag_api.py pour faciliter la comparaison
"""

import os
import sys
import json
import time
import requests
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================================
# CONFIGURATION
# ============================================================================

API_URL = "http://localhost:8000/api/query"
VERSIONS_URL = "http://localhost:8000/api/versions"

# Configurations des systèmes à comparer
SYSTEMS_CONFIG = [
    {"version": "v2", "mode": None, "name": "RAG v2 (Hybrid+Reranking)"},
    {"version": "v2.1", "mode": None, "name": "RAG v2.1 (LlamaIndex)"},
]

# Questions par défaut
DEFAULT_QUESTIONS = [
    "Quels sont les différents types de bien-être ?",
    "Comment est la qualité de vie à Ajaccio ?",
    "Quels sont les problèmes de transport en Corse ?",
    "Que pensent les habitants de Lozzi de leur cadre de vie ?",
    "Quels sont les indicateurs de santé et d'éducation en Corse ?",
    "Comment les Corses perçoivent-ils l'accès aux services publics ?",
    "Quels sont les enjeux environnementaux en Corse ?",
]


# ============================================================================
# FONCTIONS D'INTERROGATION
# ============================================================================

def check_server_availability() -> Dict:
    """Vérifie si le serveur est disponible et quelles versions sont actives"""
    try:
        response = requests.get(VERSIONS_URL, timeout=5)
        if response.status_code == 200:
            versions = response.json()
            available = {v["version"]: v["available"] for v in versions}
            return {"status": "ok", "versions": available}
        else:
            return {"status": "error", "message": f"HTTP {response.status_code}"}
    except requests.exceptions.ConnectionError:
        return {"status": "error", "message": "Impossible de se connecter au serveur. Lancez api_server_multi_version.py"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def query_rag_api(question: str, version: str, mode: Optional[str] = None) -> Dict:
    """
    Interroge le système RAG via l'API HTTP
    """
    start_time = time.time()

    payload = {
        "question": question,
        "rag_version": version,
        "k": 5,
        "use_reranking": True,
    }

    if mode:
        payload["query_mode"] = mode

    try:
        response = requests.post(API_URL, json=payload, timeout=120)
        elapsed_time = time.time() - start_time

        if response.status_code == 200:
            data = response.json()
            return {
                "success": True,
                "answer": data.get("answer", ""),
                "sources": data.get("sources", []),
                "num_sources": len(data.get("sources", [])),
                "elapsed_time": elapsed_time,
                "metadata": data.get("metadata", {}),
                "context": data.get("context", ""),
                "error": None
            }
        else:
            return {
                "success": False,
                "answer": None,
                "sources": [],
                "num_sources": 0,
                "elapsed_time": elapsed_time,
                "metadata": {},
                "context": "",
                "error": f"HTTP {response.status_code}: {response.text[:200]}"
            }
    except Exception as e:
        elapsed_time = time.time() - start_time
        return {
            "success": False,
            "answer": None,
            "sources": [],
            "num_sources": 0,
            "elapsed_time": elapsed_time,
            "metadata": {},
            "context": "",
            "error": str(e)
        }


def format_sources(sources: List[Dict], max_sources: int = 3) -> str:
    """Formate les sources pour l'affichage"""
    if not sources:
        return "Aucune source"

    formatted = []
    for i, source in enumerate(sources[:max_sources], 1):
        metadata = source.get("metadata", {})
        score = source.get("score", 0)
        nom = metadata.get("nom", "?")
        src_type = metadata.get("source", "?")
        formatted.append(f"  {i}. [{nom}/{src_type}] (score: {score:.3f})")

    if len(sources) > max_sources:
        formatted.append(f"  ... et {len(sources) - max_sources} autres sources")

    return "\n".join(formatted)


# ============================================================================
# EXPORT JSON
# ============================================================================

def save_results_json(results: List[Dict], output_path: str):
    """Sauvegarde les résultats en JSON"""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(results, f, ensure_ascii=False, indent=2)


# ============================================================================
# EXPORT TEXTE
# ============================================================================

def save_results_text(results: List[Dict], output_path: str, systems: List[Dict]):
    """Sauvegarde les résultats en texte"""
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write("=" * 80 + "\n")
        f.write("COMPARAISON RAG v2 vs v2.1 (LlamaIndex)\n")
        f.write(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
        f.write("=" * 80 + "\n\n")

        for result in results:
            f.write("\n" + "=" * 80 + "\n")
            f.write(f"QUESTION: {result['question']}\n")
            f.write("=" * 80 + "\n\n")

            for sys_name, sys_result in result["results"].items():
                f.write(f"\n--- {sys_name} ---\n")

                if sys_result["success"]:
                    f.write(f"Temps: {sys_result['elapsed_time']:.2f}s | Sources: {sys_result['num_sources']}\n\n")
                    f.write(f"RÉPONSE:\n{sys_result['answer']}\n\n")
                    f.write(f"SOURCES:\n{format_sources(sys_result['sources'])}\n")
                else:
                    f.write(f"ERREUR: {sys_result['error']}\n")

                f.write("\n")

        # Résumé
        f.write("\n" + "=" * 80 + "\n")
        f.write("RÉSUMÉ\n")
        f.write("=" * 80 + "\n\n")

        for sys in systems:
            sys_name = sys["name"]
            sys_results = [r["results"].get(sys_name, {}) for r in results]
            success_count = sum(1 for r in sys_results if r.get("success", False))
            avg_time = sum(r.get("elapsed_time", 0) for r in sys_results if r.get("success", False))
            avg_time = avg_time / success_count if success_count > 0 else 0

            f.write(f"{sys_name}:\n")
            f.write(f"  - Succès: {success_count}/{len(results)}\n")
            f.write(f"  - Temps moyen: {avg_time:.2f}s\n\n")


# ============================================================================
# EXPORT EXCEL
# ============================================================================

def save_results_excel(results: List[Dict], output_path: str, systems: List[Dict]):
    """Sauvegarde les résultats en Excel avec mise en forme"""
    wb = Workbook()

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Feuille de résumé
    ws_summary = wb.active
    ws_summary.title = "Résumé"

    # En-têtes résumé
    summary_headers = ["Question", "v2 Temps (s)", "v2.1 Temps (s)", "Diff (s)", "Plus rapide",
                       "v2 Sources", "v2.1 Sources", "Sources communes"]
    for col, header in enumerate(summary_headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    # Données résumé
    for row, result in enumerate(results, 2):
        v2_result = result["results"].get("RAG v2 (Hybrid+Reranking)", {})
        v2_1_result = result["results"].get("RAG v2.1 (LlamaIndex)", {})

        v2_time = v2_result.get("elapsed_time", 0) if v2_result.get("success") else -1
        v2_1_time = v2_1_result.get("elapsed_time", 0) if v2_1_result.get("success") else -1

        diff = v2_time - v2_1_time if v2_time > 0 and v2_1_time > 0 else 0
        faster = "v2.1" if diff > 0 else "v2" if diff < 0 else "="

        # Sources communes
        v2_sources = set(s.get("metadata", {}).get("id", "") for s in v2_result.get("sources", []))
        v2_1_sources = set(s.get("metadata", {}).get("id", "") for s in v2_1_result.get("sources", []))
        common = len(v2_sources.intersection(v2_1_sources))

        values = [
            result["question"][:50] + "..." if len(result["question"]) > 50 else result["question"],
            round(v2_time, 2) if v2_time > 0 else "Erreur",
            round(v2_1_time, 2) if v2_1_time > 0 else "Erreur",
            round(diff, 2),
            faster,
            v2_result.get("num_sources", 0),
            v2_1_result.get("num_sources", 0),
            common
        ]

        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = wrap_alignment

    # Ajuster les largeurs
    ws_summary.column_dimensions['A'].width = 40
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws_summary.column_dimensions[col].width = 15

    # Feuille de comparaison détaillée
    ws_compare = wb.create_sheet("Comparaison détaillée")

    row_num = 1
    for result in results:
        # Question
        cell = ws_compare.cell(row=row_num, column=1, value=f"QUESTION: {result['question']}")
        cell.font = Font(bold=True, size=12)
        ws_compare.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=4)
        row_num += 2

        # En-têtes
        headers = ["Système", "Temps (s)", "Sources", "Réponse"]
        for col, header in enumerate(headers, 1):
            cell = ws_compare.cell(row=row_num, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row_num += 1

        # Données pour chaque système
        for sys in systems:
            sys_name = sys["name"]
            sys_result = result["results"].get(sys_name, {})

            if sys_result.get("success"):
                values = [
                    sys_name,
                    round(sys_result.get("elapsed_time", 0), 2),
                    sys_result.get("num_sources", 0),
                    sys_result.get("answer", "")[:500] + "..." if len(sys_result.get("answer", "")) > 500 else sys_result.get("answer", "")
                ]
            else:
                values = [sys_name, "Erreur", 0, sys_result.get("error", "Erreur inconnue")]

            for col, value in enumerate(values, 1):
                cell = ws_compare.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                cell.alignment = wrap_alignment

            row_num += 1

        row_num += 2  # Espace entre questions

    # Ajuster les largeurs
    ws_compare.column_dimensions['A'].width = 30
    ws_compare.column_dimensions['B'].width = 12
    ws_compare.column_dimensions['C'].width = 10
    ws_compare.column_dimensions['D'].width = 80

    # Feuille de réponses complètes
    ws_answers = wb.create_sheet("Réponses complètes")

    row_num = 1
    for result in results:
        # Question
        cell = ws_answers.cell(row=row_num, column=1, value=f"QUESTION: {result['question']}")
        cell.font = Font(bold=True, size=12, color="FFFFFF")
        cell.fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        ws_answers.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
        row_num += 2

        for sys in systems:
            sys_name = sys["name"]
            sys_result = result["results"].get(sys_name, {})

            # Nom du système
            cell = ws_answers.cell(row=row_num, column=1, value=sys_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D6DCE5", end_color="D6DCE5", fill_type="solid")

            if sys_result.get("success"):
                cell = ws_answers.cell(row=row_num, column=2,
                                       value=f"Temps: {sys_result['elapsed_time']:.2f}s | Sources: {sys_result['num_sources']}")
            else:
                cell = ws_answers.cell(row=row_num, column=2, value="ERREUR")
                cell.font = Font(color="FF0000")

            row_num += 1

            # Réponse
            if sys_result.get("success"):
                cell = ws_answers.cell(row=row_num, column=1, value=sys_result.get("answer", ""))
                ws_answers.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=2)
                cell.alignment = wrap_alignment
            else:
                cell = ws_answers.cell(row=row_num, column=1, value=sys_result.get("error", "Erreur inconnue"))
                cell.font = Font(color="FF0000")

            row_num += 2

        row_num += 1  # Espace entre questions

    ws_answers.column_dimensions['A'].width = 80
    ws_answers.column_dimensions['B'].width = 40

    wb.save(output_path)
    print(f"  OK Fichier Excel créé: {output_path}")


# ============================================================================
# FONCTION PRINCIPALE
# ============================================================================

def run_comparison(questions: List[str]) -> List[Dict]:
    """Exécute la comparaison sur toutes les questions"""
    results = []

    for i, question in enumerate(questions, 1):
        print(f"\n[{i}/{len(questions)}] Question: {question}...")
        print("-" * 80)

        question_results = {"question": question, "results": {}}

        for sys_config in SYSTEMS_CONFIG:
            sys_name = sys_config["name"]
            version = sys_config["version"]
            mode = sys_config.get("mode")

            print(f"  - {sys_name}...", end=" ", flush=True)

            result = query_rag_api(question, version, mode)

            if result["success"]:
                print(f"OK ({result['elapsed_time']:.2f}s, {result['num_sources']} sources)")
            else:
                print(f"ERREUR: {result['error'][:50]}...")

            question_results["results"][sys_name] = result

        results.append(question_results)

    return results


def print_summary(results: List[Dict]):
    """Affiche un résumé des résultats"""
    print("\n" + "=" * 80)
    print("RÉSUMÉ DE LA COMPARAISON v2 vs v2.1")
    print("=" * 80)

    for sys_config in SYSTEMS_CONFIG:
        sys_name = sys_config["name"]
        sys_results = [r["results"].get(sys_name, {}) for r in results]

        success_count = sum(1 for r in sys_results if r.get("success", False))
        total_time = sum(r.get("elapsed_time", 0) for r in sys_results if r.get("success", False))
        avg_time = total_time / success_count if success_count > 0 else 0
        total_sources = sum(r.get("num_sources", 0) for r in sys_results if r.get("success", False))
        avg_sources = total_sources / success_count if success_count > 0 else 0
        error_count = len(results) - success_count

        print(f"\n{sys_name}:")
        print(f"  - Temps moyen: {avg_time:.2f}s")
        print(f"  - Temps total: {total_time:.2f}s")
        print(f"  - Sources moyennes: {avg_sources:.1f}")
        print(f"  - Erreurs: {error_count}/{len(results)}")


def main():
    """Fonction principale"""
    print("=" * 80)
    print("COMPARAISON DES SYSTÈMES RAG v2 vs v2.1")
    print("=" * 80)

    # Vérifier la disponibilité du serveur
    print("\nVérification du serveur API...")
    server_status = check_server_availability()

    if server_status["status"] != "ok":
        print(f"ERREUR: {server_status['message']}")
        sys.exit(1)

    # Vérifier que v2 et v2.1 sont disponibles
    versions = server_status["versions"]
    print("Serveur OK. Versions disponibles:")
    for sys_config in SYSTEMS_CONFIG:
        version = sys_config["version"]
        available = versions.get(version, False)
        status = "OK" if available else "NON DISPONIBLE"
        print(f"  - {sys_config['name']}: {status}")

        if not available:
            print(f"\nERREUR: {sys_config['name']} n'est pas disponible!")
            sys.exit(1)

    # Questions à tester
    questions = DEFAULT_QUESTIONS

    print(f"\nQuestions à tester ({len(questions)}):")
    for i, q in enumerate(questions, 1):
        print(f"  {i}. {q}")

    print("\nLancement de la comparaison...")

    # Exécuter la comparaison
    results = run_comparison(questions)

    # Sauvegarder les résultats
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_dir = "comparaisons_rag"
    os.makedirs(output_dir, exist_ok=True)

    # Fichiers de sortie
    json_path = os.path.join(output_dir, f"comparison_v2_vs_v2_1_{timestamp}.json")
    txt_path = os.path.join(output_dir, f"comparison_v2_vs_v2_1_{timestamp}.txt")
    xlsx_path = os.path.join(output_dir, f"comparison_v2_vs_v2_1_{timestamp}.xlsx")

    save_results_json(results, json_path)
    print(f"Résultats JSON: {json_path}")

    save_results_text(results, txt_path, SYSTEMS_CONFIG)
    print(f"Résultats texte: {txt_path}")

    save_results_excel(results, xlsx_path, SYSTEMS_CONFIG)

    # Afficher le résumé
    print_summary(results)

    print("\n" + "=" * 80)
    print("Comparaison terminée!")
    print("=" * 80)


if __name__ == "__main__":
    main()
