"""
Recall / Précision / F1 sur 109 questions × 4 configs.

Convention RAPTOR (configs sans RAPTOR) :
  - Si GT requiert un résumé RAPTOR et que la config ne peut pas le
    récupérer directement, crédit accordé si ≥30% des source_chunk_ids
    du résumé RAPTOR sont parmi les chunks récupérés.
  - Chunk ID reconstruit depuis metadata : portrait_{commune}_{chunk_idx}

Limites architecturales (Q40, Q104+) : skip (aucun GT attendu).
"""
import json, logging, re, sys, io
from collections import defaultdict
from datetime import datetime
from pathlib import Path

import chromadb
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
logging.basicConfig(level=logging.INFO, format='%(levelname)s %(message)s',
                    stream=sys.stdout)
log = logging.getLogger(__name__)

# ── Config ────────────────────────────────────────────────────────────────────
XLSX    = r'C:\Users\comiti_g\Downloads\rag_evaluation_with_metrics_full.xlsx'
JSON    = 'comparaisons_rag/ablations_103q_v43_gpt4o_COMPLET.json'
CHROMA  = 'chroma_portrait'
RAPTOR_THRESHOLD = 0.30          # convention 30%
CONFIGS = ['v_vanilla_k10', 'v_vanilla_k25', 'v_decomp', 'v_decomp_raptor']
LABELS  = {'v_vanilla_k10': 'Vanilla k10', 'v_vanilla_k25': 'Vanilla k25',
           'v_decomp': 'Decomp', 'v_decomp_raptor': 'Decomp+Raptor'}
RAPTOR_CONFIGS  = {'v_decomp_raptor'}           # accès direct aux résumés RAPTOR
NO_RAPTOR_CONFIGS = set(CONFIGS) - RAPTOR_CONFIGS

ARCH_SECTION = 'limites architecturales'        # skip

SEC_ORDER = [
    'Retrieval mono-commune',
    'Raisonnement comparatif',
    'Raisonnement causal et contre-intuitif',
    "Gestion de l'absence d'information",
    "Gestion de l'incertitude et des biais",
    'Robustesse sémantique',
    'Limites architecturales',
]

ts = datetime.now().strftime('%Y%m%d_%H%M%S')
CSV_OUT  = Path(f'comparaisons_rag/retrieval_metrics_109q_{ts}.csv')
HTML_OUT = Path(f'comparaisons_rag/retrieval_metrics_109q_{ts}.html')

# ── 1. Charger RAPTOR → source_chunk_ids depuis ChromaDB ─────────────────────
log.info('Chargement des résumés RAPTOR depuis ChromaDB...')
chroma_client = chromadb.PersistentClient(path=CHROMA)

def _load_raptor_index(collection_name: str) -> list:
    """Retourne liste de {id, meta, source_ids: set}."""
    try:
        col = chroma_client.get_collection(collection_name)
        total = col.count()
        entries = []
        offset = 0
        while offset < total:
            batch = col.get(limit=500, offset=offset, include=['metadatas'])
            for doc_id, meta in zip(batch['ids'], batch['metadatas']):
                raw_ids = meta.get('source_chunk_ids', '[]')
                try:
                    ids = set(json.loads(raw_ids))
                except Exception:
                    ids = set()
                entries.append({'id': doc_id, 'meta': meta, 'source_ids': ids})
            offset += len(batch['ids'])
            if len(batch['ids']) < 500:
                break
        log.info(f'  {collection_name}: {len(entries)} résumés')
        return entries
    except Exception as e:
        log.warning(f'  {collection_name}: {e}')
        return []

raptor_quali_index  = _load_raptor_index('raptor_summaries')
raptor_quanti_index = _load_raptor_index('raptor_quanti_summaries')
raptor_entretiens_index = _load_raptor_index('raptor_entretiens_summaries')

def _all_raptor(): return raptor_quali_index + raptor_quanti_index + raptor_entretiens_index

# ── 2. Parser GT → catégories structurées ─────────────────────────────────────
_COMMUNES_KNOWN = {
    'ajaccio', 'bastia', 'lozzi', 'corte', 'bonifacio', 'piedicorte',
    'pedicorte', 'aïti', 'aiti', 'niolu', 'guarguale', 'corse',
    'calvi', 'porto-vecchio', 'ghisonaccia', 'bastelicaccia', 'venzolasca',
    'lucciana', 'olmeto', 'moltifao',
}
_STRIP = re.compile(r"[àáâãäåèéêëìíîïòóôõöùúûüýÿ]")

def _norm(s):
    s = s.lower().strip()
    s = s.replace('â', 'a').replace('à', 'a').replace('é', 'e').replace('è', 'e')
    s = s.replace('ê', 'e').replace('î', 'i').replace('ô', 'o').replace('û', 'u')
    s = s.replace('ï', 'i').replace('ü', 'u').replace('ù', 'u').replace('ç', 'c')
    return s

def _extract_commune(text: str):
    t = _norm(text)
    for c in sorted(_COMMUNES_KNOWN, key=len, reverse=True):
        if c in t:
            # capitaliser proprement
            orig = {'ajaccio': 'Ajaccio', 'bastia': 'Bastia', 'lozzi': 'Lozzi',
                    'corte': 'Corte', 'bonifacio': 'Bonifacio',
                    'piedicorte': 'Piedicorte-di-Gaggio', 'pedicorte': 'Piedicorte-di-Gaggio',
                    'aïti': 'Aïti', 'aiti': 'Aïti', 'niolu': 'Niolu',
                    'guarguale': 'Guargualè', 'lucciana': 'Lucciana',
                    'olmeto': 'Olmeto', 'calvi': 'Calvi'}.get(c, c.capitalize())
            return orig
    return None

def _extract_csp(text: str):
    t = _norm(text)
    for kw, label in [('entrepreneur', 'entrepreneurs'), ('artisan', 'artisans'),
                      ('retraite', 'retraités'), ('senior', 'retraités'),
                      ('etudiant', 'étudiants'), ('jeune', '18-25'),
                      ('25-', '25-34'), ('18-25', '18-25'), ('15-24', '15-24')]:
        if kw in t:
            return label
    return None

def _extract_age(text: str):
    m = re.search(r'(\d{2}[-–]\d{2,3})', text)
    return m.group(1) if m else None

def parse_gt_token(token: str):
    """
    Parse un token GT en catégorie structurée.
    Retourne dict avec keys: type, commune, csp, age, collection, optional, raw
    type: 'oppchovec' | 'raptor' | 'entretiens' | 'verbatims' | 'stats_enquete' |
          'stats_repondants' | 'wiki' | 'equipements' | 'methodology' |
          'classement' | 'rien' | 'epci' | 'unknown'
    """
    t = token.strip()
    tn = _norm(t)
    optional = 'facultatif' in tn or 'ou rien' in tn or 'wiki' == tn.split()[0] if tn else False

    if not t or tn in ('', 'nan', 'none'):
        return None
    if tn.startswith('rien') or 'pas de donnee' in tn or 'pas de data' in tn:
        return {'type': 'rien', 'optional': False, 'raw': t}

    res = {'optional': optional, 'raw': t,
           'commune': _extract_commune(t),
           'csp': _extract_csp(t),
           'age': _extract_age(t)}

    if 'raptor' in tn:
        quali = 'quali' in tn or 'portrait' in tn or 'entretien' in tn
        quanti = 'quanti' in tn or 'dimension' in tn or 'global' in tn
        if quali and not quanti:
            res['collection'] = 'raptor_summaries'
        elif quanti and not quali:
            res['collection'] = 'raptor_quanti_summaries'
        else:
            res['collection'] = 'both'
        res['type'] = 'raptor'
        return res

    if 'tous les raptor' in tn:
        res['type'] = 'raptor'; res['collection'] = 'both'; return res

    if 'entretien' in tn and 'raptor' not in tn:
        res['type'] = 'entretiens'; return res
    if 'verbatim' in tn:
        res['type'] = 'verbatims'; return res
    if 'oppchovec' in tn or 'opp chovec' in tn:
        if 'explication' in tn or 'methodo' in tn:
            res['type'] = 'methodology'; return res
        res['type'] = 'oppchovec'; return res
    if 'classement' in tn:
        res['type'] = 'classement'; return res
    if 'wiki' in tn:
        res['type'] = 'wiki'; optional = True; res['optional'] = True; return res
    if 'equipement' in tn:
        res['type'] = 'equipements'; return res
    if 'epci' in tn or 'zones_epci' in tn or 'docu epci' in tn:
        res['type'] = 'epci'; return res
    if 'stats repondant' in tn or 'profil repondant' in tn:
        res['type'] = 'stats_repondants'; return res
    if 'stats' in tn or 'resultats enquete' in tn or 'score' in tn and 'enquete' in tn:
        res['type'] = 'stats_enquete'; return res
    if 'explication' in tn or 'methodo' in tn:
        res['type'] = 'methodology'; return res
    if 'resume enquete' in tn or 'donnees enquete' in tn:
        res['type'] = 'stats_enquete'; return res

    res['type'] = 'unknown'
    return res

def parse_gt(gt_text) -> list:
    """Parse le texte GT en liste de tokens structurés."""
    if not gt_text or str(gt_text).strip().lower() in ('nan', 'none', ''):
        return []
    tokens = []
    for part in str(gt_text).split(','):
        p = parse_gt_token(part.strip())
        if p:
            tokens.append(p)
    return tokens

# ── 3. Mapping GT RAPTOR → source_chunk_ids via ChromaDB ─────────────────────
def _meta_matches(meta: dict, tok: dict) -> bool:
    """Est-ce que ce résumé RAPTOR correspond au token GT ?"""
    vn = meta.get('view_name', '')
    d1 = _norm(meta.get('dim1_value', '') or '')
    d2 = _norm(meta.get('dim2_value', '') or '')

    commune = tok.get('commune')
    csp     = tok.get('csp')
    age     = tok.get('age')

    if commune:
        cn = _norm(commune)
        if cn not in d1 and cn not in d2:
            return False
    if csp:
        cn = _norm(csp)
        if cn not in d1 and cn not in d2:
            return False
    if age:
        an = _norm(age)
        if an not in d1 and an not in d2:
            return False
    return True

def get_raptor_source_ids(tok: dict) -> set:
    """Retourne l'union des source_chunk_ids de tous les résumés correspondant au token."""
    col = tok.get('collection', 'both')
    if col == 'raptor_summaries':
        indices = raptor_quali_index
    elif col == 'raptor_quanti_summaries':
        indices = raptor_quanti_index
    else:  # both
        indices = raptor_quali_index + raptor_quanti_index + raptor_entretiens_index

    result = set()
    for entry in indices:
        if _meta_matches(entry['meta'], tok):
            result |= entry['source_ids']
    return result

# ── 4. Classifier les sources récupérées ─────────────────────────────────────
def _source_type_of(s_meta: dict) -> str:
    """Catégorie d'une source récupérée."""
    st = s_meta.get('source_type', '') or ''
    t  = s_meta.get('type', '') or ''
    src = s_meta.get('source', '') or ''
    vn  = s_meta.get('view_name', '') or ''

    if 'raptor_entretiens' in st:  return 'raptor_portrait'
    if 'raptor_summaries' in st or 'raptor_portrait' in st: return 'raptor_portrait'
    if 'raptor_quanti' in st or 'raptor_enquete' in st:     return 'raptor_enquete'
    if t == 'raptor_summary':  return 'raptor_portrait'
    if t == 'raptor_enquete_summary': return 'raptor_enquete'
    if st in ('oppchovec_scores',) or 'oppchovec' in src.lower(): return 'oppchovec'
    if t == 'methodology' or st == 'methodology': return 'methodology'
    if t == 'classement_dimensions' or 'classement' in st: return 'classement'
    if st == 'enquete_scores_commune' or 'enquete_score' in st: return 'stats_enquete'
    if st == 'communes_profil' or 'profil' in st: return 'stats_repondants'
    if st == 'portrait_entretiens': return 'entretiens'
    if st == 'portrait_verbatims':  return 'verbatims'
    if st == 'communes_wiki':       return 'wiki'
    if st in ('communes_equipements', 'equipements'): return 'equipements'
    if st in ('zones_epci', 'communes_geo'): return 'epci'
    return 'unknown'

def _chunk_id_of(s_meta: dict):
    """Reconstruit l'ID ChromaDB d'un chunk portrait si possible."""
    st = s_meta.get('source_type', '')
    if st not in ('portrait_entretiens', 'portrait_verbatims'):
        return None
    commune = (s_meta.get('commune') or s_meta.get('nom') or '').replace(' ', '-')
    idx = s_meta.get('chunk_idx')
    if commune and idx is not None:
        return f'portrait_{commune}_{idx}'
    return None

# ── 5. Calcul recall/précision/F1 par question × config ──────────────────────
def compute_metrics(sources: list, gt_tokens: list, is_raptor_config: bool) -> dict:
    """
    sources      : liste de dicts metadata des sources récupérées
    gt_tokens    : liste de tokens GT parsés
    is_raptor_config : True pour decomp_raptor
    """
    # Filtrer les tokens GT
    required_toks = [t for t in gt_tokens if not t.get('optional') and t.get('type') != 'rien']
    optional_toks = [t for t in gt_tokens if t.get('optional')]
    is_refusal    = any(t.get('type') == 'rien' for t in gt_tokens)

    if is_refusal:
        n_retrieved = len([s for s in sources if s])
        return {'recall': None, 'precision': None, 'f1': None,
                'is_refusal': True, 'tnrr': 1 if n_retrieved == 0 else 0,
                'n_required': 0, 'n_hits': 0}

    if not required_toks:
        return {'recall': None, 'precision': None, 'f1': None,
                'is_refusal': False, 'tnrr': None,
                'n_required': 0, 'n_hits': 0, 'skip_reason': 'no_required_gt'}

    # Catégories récupérées
    retrieved_cats = set()
    retrieved_chunk_ids = set()
    for s in sources:
        meta = s if isinstance(s, dict) else {}
        cat = _source_type_of(meta)
        retrieved_cats.add(cat)
        cid = _chunk_id_of(meta)
        if cid:
            retrieved_chunk_ids.add(cid)

    # Évaluer chaque token GT requis
    hits = 0
    details = []
    for tok in required_toks:
        ttype = tok['type']
        satisfied = False

        if ttype == 'raptor':
            if is_raptor_config:
                # Évaluation directe : cherche raptor_portrait ou raptor_enquete
                col = tok.get('collection', 'both')
                if col == 'raptor_summaries':
                    satisfied = 'raptor_portrait' in retrieved_cats
                elif col == 'raptor_quanti_summaries':
                    satisfied = 'raptor_enquete' in retrieved_cats
                else:
                    satisfied = bool(retrieved_cats & {'raptor_portrait', 'raptor_enquete'})
            else:
                # Convention 30% : récupère les source_chunk_ids du RAPTOR correspondant
                raptor_ids = get_raptor_source_ids(tok)
                if raptor_ids:
                    overlap = len(raptor_ids & retrieved_chunk_ids)
                    ratio = overlap / len(raptor_ids)
                    satisfied = ratio >= RAPTOR_THRESHOLD
                    details.append(f"RAPTOR 30%: {overlap}/{len(raptor_ids)}={ratio:.0%} {'✓' if satisfied else '✗'}")
                else:
                    # Pas de résumé trouvé → fallback catégorie générale
                    col = tok.get('collection', 'both')
                    if col == 'raptor_summaries':
                        satisfied = bool(retrieved_cats & {'entretiens', 'verbatims', 'raptor_portrait'})
                    elif col == 'raptor_quanti_summaries':
                        satisfied = bool(retrieved_cats & {'stats_enquete', 'stats_repondants', 'raptor_enquete'})
                    else:
                        satisfied = bool(retrieved_cats & {'entretiens', 'verbatims', 'stats_enquete',
                                                           'raptor_portrait', 'raptor_enquete'})
                    details.append(f"RAPTOR fallback→catégorie {'✓' if satisfied else '✗'}")

        elif ttype == 'oppchovec':
            satisfied = 'oppchovec' in retrieved_cats
        elif ttype == 'entretiens':
            satisfied = bool(retrieved_cats & {'entretiens', 'raptor_portrait'})
        elif ttype == 'verbatims':
            satisfied = bool(retrieved_cats & {'verbatims', 'entretiens', 'raptor_portrait'})
        elif ttype == 'stats_enquete':
            satisfied = bool(retrieved_cats & {'stats_enquete', 'stats_repondants', 'raptor_enquete'})
        elif ttype == 'stats_repondants':
            satisfied = 'stats_repondants' in retrieved_cats
        elif ttype == 'methodology':
            satisfied = bool(retrieved_cats & {'methodology', 'oppchovec'})
        elif ttype == 'classement':
            satisfied = bool(retrieved_cats & {'classement', 'oppchovec'})
        elif ttype == 'wiki':
            satisfied = 'wiki' in retrieved_cats
        elif ttype == 'equipements':
            satisfied = bool(retrieved_cats & {'equipements', 'epci'})
        elif ttype == 'epci':
            satisfied = bool(retrieved_cats & {'epci', 'equipements'})
        else:
            pass  # unknown → not satisfied

        if satisfied:
            hits += 1

    recall = hits / len(required_toks) if required_toks else None

    # Précision : parmi les catégories récupérées, combien sont pertinentes ?
    valid_cats = set()
    for tok in required_toks + optional_toks:
        ttype = tok['type']
        if ttype == 'raptor':
            valid_cats |= {'raptor_portrait', 'raptor_enquete', 'entretiens', 'verbatims', 'stats_enquete'}
        elif ttype == 'oppchovec':    valid_cats.add('oppchovec')
        elif ttype in ('entretiens', 'verbatims'): valid_cats |= {'entretiens', 'verbatims', 'raptor_portrait'}
        elif ttype == 'stats_enquete': valid_cats |= {'stats_enquete', 'stats_repondants', 'raptor_enquete'}
        elif ttype == 'stats_repondants': valid_cats.add('stats_repondants')
        elif ttype == 'methodology':  valid_cats |= {'methodology', 'oppchovec'}
        elif ttype == 'classement':   valid_cats |= {'classement', 'oppchovec'}
        elif ttype == 'wiki':         valid_cats.add('wiki')
        elif ttype == 'equipements':  valid_cats |= {'equipements', 'epci'}
        elif ttype == 'epci':         valid_cats |= {'epci', 'equipements'}

    useful = retrieved_cats & valid_cats
    precision = len(useful) / max(len(retrieved_cats - {'unknown'}), 1) if retrieved_cats else None

    if recall is not None and precision is not None and (recall + precision) > 0:
        f1 = 2 * recall * precision / (recall + precision)
    else:
        f1 = None

    return {
        'recall': round(recall, 3) if recall is not None else None,
        'precision': round(precision, 3) if precision is not None else None,
        'f1': round(f1, 3) if f1 is not None else None,
        'is_refusal': False, 'tnrr': None,
        'n_required': len(required_toks), 'n_hits': hits,
        'details': '; '.join(details),
    }

# ── 6. Charger données ────────────────────────────────────────────────────────
log.info('Chargement Excel GT...')
wb = openpyxl.load_workbook(XLSX)
ws = wb.active
gt_by_row = {}
for r in range(2, ws.max_row + 1):
    q = ws.cell(r, 3).value
    if q and str(q).strip():
        gt_by_row[r - 1] = ws.cell(r, 4).value

log.info('Chargement JSON résultats RAG...')
with open(JSON, encoding='utf-8') as f:
    data = json.load(f)

index = {}
for cfg, entries in data.items():
    for e in entries:
        index[(e['excel_row'], cfg)] = e

# ── 7. Calcul principal ───────────────────────────────────────────────────────
log.info('Calcul des métriques...')
all_rows = sorted({e['excel_row'] for e in data['v_vanilla_k10']})

records = []   # une ligne par (row, cfg)
n_skip_arch = 0
n_skip_no_gt = 0
n_null_gt = 0
n_refusal = 0

for row in all_rows:
    e_ref = index.get((row, 'v_vanilla_k10'), {})
    sec   = (e_ref.get('section') or '').strip()
    subsec = (e_ref.get('subsection') or '').strip()
    question = e_ref.get('question', '')

    if ARCH_SECTION in sec.lower():
        n_skip_arch += 1
        continue

    gt_text  = gt_by_row.get(row)
    gt_tokens = parse_gt(gt_text)

    if not gt_tokens:
        n_skip_no_gt += 1
        continue

    for cfg in CONFIGS:
        e = index.get((row, cfg), {})
        sources_raw = e.get('sources', [])
        sources_meta = [s.get('metadata', s) if isinstance(s, dict) else {} for s in sources_raw]

        is_raptor = cfg in RAPTOR_CONFIGS
        m = compute_metrics(sources_meta, gt_tokens, is_raptor)
        m['row'] = row; m['cfg'] = cfg
        m['section'] = sec; m['subsection'] = subsec
        m['question'] = question[:80]
        m['gt_raw'] = str(gt_text)[:100]
        records.append(m)

        if m.get('is_refusal'):
            n_refusal += 1

log.info(f'  {len(all_rows)} questions totales')
log.info(f'  {n_skip_arch} skippées (limites architecturales)')
log.info(f'  {n_skip_no_gt} skippées (GT vide)')
log.info(f'  {n_refusal / len(CONFIGS):.0f} questions refus')
log.info(f'  {len(records)} calculs effectués ({len(records)//len(CONFIGS)} questions × {len(CONFIGS)} configs)')

# ── 8. CSV ────────────────────────────────────────────────────────────────────
log.info(f'Export CSV → {CSV_OUT}')
csv_lines = ['row,config,section,subsection,recall,precision,f1,is_refusal,tnrr,n_required,n_hits,gt_raw,question']
for r in records:
    def fv(v): return '' if v is None else f'{v:.3f}'
    csv_lines.append(
        f"{r['row']},{r['cfg']},{r['section'].replace(',','')},{r['subsection'].replace(',','')},"
        f"{fv(r['recall'])},{fv(r['precision'])},{fv(r['f1'])},"
        f"{r.get('is_refusal',False)},{fv(r.get('tnrr'))},"
        f"{r.get('n_required',0)},{r.get('n_hits',0)},"
        f"\"{r['gt_raw']}\",\"{r['question']}\""
    )
CSV_OUT.write_text('\n'.join(csv_lines), encoding='utf-8')

# ── 9. Agrégation ─────────────────────────────────────────────────────────────
def mean(lst): return sum(lst) / len(lst) if lst else None

def agg(records_subset, cfg):
    sub = [r for r in records_subset if r['cfg'] == cfg and not r.get('is_refusal')]
    rs = [r['recall'] for r in sub if r.get('recall') is not None]
    ps = [r['precision'] for r in sub if r.get('precision') is not None]
    fs = [r['f1'] for r in sub if r.get('f1') is not None]
    refusals = [r for r in records_subset if r['cfg'] == cfg and r.get('is_refusal')]
    tnrr_vals = [r.get('tnrr', 0) for r in refusals if r.get('tnrr') is not None]
    return {
        'recall': mean(rs), 'precision': mean(ps), 'f1': mean(fs),
        'n': len(sub), 'tnrr': mean(tnrr_vals),
        'n_null_precision': sum(1 for r in sub if r.get('precision') is None),
    }

# ── 10. Affichage console ─────────────────────────────────────────────────────
print()
print('=' * 75)
print('TABLEAU 1 — Global par config')
print('=' * 75)
print(f"{'Config':<18} {'N':>4}  {'Recall':>7}  {'Précision':>9}  {'F1':>7}  {'TNRR':>6}")
print('-' * 75)
global_agg = {}
for cfg in CONFIGS:
    a = agg(records, cfg)
    global_agg[cfg] = a
    tnrr_s = f"{a['tnrr']:.0%}" if a['tnrr'] is not None else '—'
    print(f"{LABELS[cfg]:<18} {a['n']:>4}  "
          f"{a['recall']:.1%}    {a['precision']:.1%}     {a['f1']:.1%}  {tnrr_s}")

print()
print('=' * 75)
print('TABLEAU 2 — Par section')
print('=' * 75)

def clean_sec(s):
    s = (s or '').replace(''', '').replace(''', "'")
    if 'absence' in s.lower() and 'information' in s.lower():
        return "Gestion de l'absence d'information"
    return s

# Normaliser sections
for r in records:
    r['section_norm'] = clean_sec(r['section'])

by_sec = defaultdict(list)
for r in records:
    by_sec[r['section_norm']].append(r)

for sec in SEC_ORDER:
    recs = by_sec.get(sec, [])
    if not recs: continue
    n_q = len({r['row'] for r in recs if r['cfg'] == CONFIGS[0]})
    print(f"\n  {sec} (N={n_q})")
    print(f"  {'Métrique':<12} " + '  '.join(f"{LABELS[c]:>12}" for c in CONFIGS))
    for metric in ('recall', 'precision', 'f1'):
        row_s = f"  {metric.capitalize():<12}"
        for cfg in CONFIGS:
            a = agg(recs, cfg)
            v = a[metric]
            row_s += f"  {v:.1%}" if v is not None else '       —    '
        print(row_s)

# ── 11. Comparaison raptor vs vanilla_k10 ─────────────────────────────────────
print()
print('=' * 75)
print('COMPARAISON Decomp+Raptor vs Vanilla k10')
print('=' * 75)
r_ref = global_agg['v_vanilla_k10']
r_rap = global_agg['v_decomp_raptor']
print(f"  Écart Recall  global : {(r_rap['recall'] or 0) - (r_ref['recall'] or 0):+.1%}")
print(f"  Écart F1      global : {(r_rap['f1'] or 0) - (r_ref['f1'] or 0):+.1%}")
print()
print('  Sections où avantage Recall Raptor > +10% :')
for sec in SEC_ORDER:
    recs = by_sec.get(sec, [])
    if not recs: continue
    a_ref = agg(recs, 'v_vanilla_k10')
    a_rap = agg(recs, 'v_decomp_raptor')
    if a_ref['recall'] is not None and a_rap['recall'] is not None:
        delta = a_rap['recall'] - a_ref['recall']
        if delta > 0.10:
            print(f"    {sec}: +{delta:.1%} ({a_ref['recall']:.1%} → {a_rap['recall']:.1%})")

# ── 12. HTML ──────────────────────────────────────────────────────────────────
log.info(f'Génération HTML → {HTML_OUT}')

COLORS = {'v_vanilla_k10': '#c0392b', 'v_vanilla_k25': '#e67e22',
          'v_decomp': '#27ae60', 'v_decomp_raptor': '#2980b9'}

def bg(v):
    if v is None: return '#f5f5f5'
    if v >= 0.80: return '#d5f5e3'
    if v >= 0.65: return '#eafaf1'
    if v >= 0.50: return '#fef9e7'
    if v >= 0.35: return '#fef5e4'
    return '#fdecea'
def fg(v):
    if v is None: return '#aaa'
    if v >= 0.80: return '#1a7a40'
    if v >= 0.65: return '#27ae60'
    if v >= 0.50: return '#b7950b'
    if v >= 0.35: return '#ca6f1e'
    return '#c0392b'
def cell(v, bold=False):
    if v is None:
        return '<td style="background:#f5f5f5;color:#aaa;text-align:center;font-size:0.82em">—</td>'
    fw = 'bold' if bold else 'normal'
    return (f'<td style="background:{bg(v)};color:{fg(v)};font-weight:{fw};'
            f'text-align:center;font-size:0.85em">{v:.0%}</td>')

# Summary table
sum_rows = ''
for cfg in CONFIGS:
    a = global_agg[cfg]
    c = COLORS[cfg]
    tnrr_s = f"{a['tnrr']:.0%}" if a['tnrr'] is not None else '—'
    sum_rows += (f'<tr><td style="color:{c};font-weight:bold">{LABELS[cfg]}</td>'
                 f'<td style="text-align:center">{a["n"]}</td>'
                 + cell(a['recall']) + cell(a['precision']) + cell(a['f1'])
                 + f'<td style="text-align:center;color:#555">{tnrr_s}</td></tr>\n')

# Section table
sec_html = ''
for i, sec in enumerate(SEC_ORDER):
    recs = by_sec.get(sec, [])
    if not recs: continue
    n_q = len({r['row'] for r in recs if r['cfg'] == CONFIGS[0]})
    sec_html += (f'<tr id="sec{i}" style="background:#2c3e50">'
                 f'<td colspan="13" style="color:white;font-weight:bold;padding:8px 14px">'
                 f'{sec} (N={n_q})</td></tr>\n')

    # Sous-sections
    by_sub = defaultdict(list)
    for r in recs:
        by_sub[r['subsection'] or '—'].append(r)

    for sub in sorted(by_sub.keys()):
        sub_recs = by_sub[sub]
        n_sub = len({r['row'] for r in sub_recs if r['cfg'] == CONFIGS[0]})
        # 3 lignes : R, P, F1
        for metric, mlabel in (('recall','R'), ('precision','P'), ('f1','F1')):
            if metric == 'recall':
                label_cell = f'<td rowspan="3" style="padding:4px 14px 4px 26px;vertical-align:middle">{sub}</td><td rowspan="3" style="text-align:center;color:#888;vertical-align:middle">{n_sub}</td>'
            else:
                label_cell = ''
            row_s = f'<tr><td style="color:#777;font-size:0.82em;padding:2px 8px">{mlabel}</td>'
            for cfg in CONFIGS:
                a = agg(sub_recs, cfg)
                row_s += cell(a[metric])
            row_s += '</tr>\n'
            if metric == 'recall':
                sec_html += f'<tr>{label_cell}{row_s[4:]}'
            else:
                sec_html += row_s

    # Ligne moyenne section
    sec_html += '<tr style="border-top:2px solid #bdc3c7">'
    sec_html += f'<td colspan="2" style="padding:5px 14px;font-style:italic;color:#555;background:#fafafa">Moyenne {sec}</td>'
    for metric, mlabel in (('recall','R'), ('precision','P'), ('f1','F1')):
        sec_html += f'<td style="color:#777;font-size:0.8em;text-align:center;background:#fafafa">{mlabel}</td>'
        for cfg in CONFIGS:
            a = agg(recs, cfg)
            sec_html += cell(a[metric], bold=True)
        sec_html += '</tr><tr style="border-top:0">'
    sec_html += '</tr><tr><td colspan="13" style="height:6px;background:#f0f2f5;border:none"></td></tr>'

nav_html = ' &nbsp;·&nbsp; '.join(
    f'<a href="#sec{i}">{sec}</a>' for i, sec in enumerate(SEC_ORDER) if sec in by_sec)

header_4cfg = ''.join(
    f'<th colspan="4" style="background:{COLORS[c]};color:white">{LABELS[c]}</th>'
    for c in CONFIGS)
subheader_cfg = ''.join(
    '<th style="background:#ecf0f1;font-size:0.78em">R</th>'
    '<th style="background:#ecf0f1;font-size:0.78em">P</th>'
    '<th style="background:#ecf0f1;font-size:0.78em">F1</th>'
    '<th style="background:#ecf0f1;font-size:0.78em;color:#888">—</th>'
    for _ in CONFIGS)

html = f"""<!DOCTYPE html>
<html lang="fr"><head><meta charset="utf-8">
<title>Retrieval R/P/F1 — 109q — {ts}</title>
<style>
  body {{ font-family:-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;
         font-size:13px;margin:0;background:#f0f2f5; }}
  .topbar {{ background:#2c3e50;color:white;padding:12px 24px;position:sticky;
             top:0;z-index:100;box-shadow:0 2px 6px rgba(0,0,0,.3); }}
  .topbar h1 {{ margin:0 0 5px;font-size:1.1em; }}
  .topbar nav {{ font-size:0.82em;opacity:.8; }}
  .topbar nav a {{ color:#aed6f1;text-decoration:none; }}
  .topbar nav a:hover {{ text-decoration:underline; }}
  .container {{ max-width:1300px;margin:0 auto;padding:20px 24px; }}
  h3 {{ margin:20px 0 8px;color:#2c3e50; }}
  table {{ border-collapse:collapse;width:100%;background:white;
           border-radius:8px;overflow:hidden;box-shadow:0 1px 4px rgba(0,0,0,.12);
           margin-bottom:24px; }}
  th,td {{ border:1px solid #e0e0e0;padding:5px 7px; }}
  th {{ font-size:0.82em;position:sticky;top:58px;z-index:10; }}
  tr:hover td {{ filter:brightness(0.96); }}
  .legend {{ display:flex;gap:10px;flex-wrap:wrap;margin-bottom:14px;font-size:0.82em; }}
  .leg {{ padding:3px 10px;border-radius:4px; }}
  .note {{ color:#666;font-size:0.85em;margin-bottom:12px; }}
</style></head><body>
<div class="topbar">
  <h1>Retrieval Recall / Précision / F1 — 109 questions × 4 configs — {ts}</h1>
  <nav>{nav_html}</nav>
</div>
<div class="container">
  <div class="legend">
    <span class="leg" style="background:#d5f5e3;color:#1a7a40">≥ 80%</span>
    <span class="leg" style="background:#eafaf1;color:#27ae60">65–79%</span>
    <span class="leg" style="background:#fef9e7;color:#b7950b">50–64%</span>
    <span class="leg" style="background:#fef5e4;color:#ca6f1e">35–49%</span>
    <span class="leg" style="background:#fdecea;color:#c0392b">&lt; 35%</span>
  </div>
  <p class="note">
    Convention RAPTOR : configs sans RAPTOR créditées si ≥ {RAPTOR_THRESHOLD:.0%} des chunks sources du résumé sont récupérés.<br>
    Limites architecturales exclues. — = question skip (GT vide ou refus).
    CSV : <code>{CSV_OUT.name}</code>
  </p>

  <h3>Résumé global par config</h3>
  <table style="max-width:700px">
    <thead>
      <tr>
        <th style="background:#ecf0f1;text-align:left">Config</th>
        <th style="background:#ecf0f1">N questions</th>
        <th style="background:#ecf0f1">Recall</th>
        <th style="background:#ecf0f1">Précision</th>
        <th style="background:#ecf0f1">F1</th>
        <th style="background:#ecf0f1">TNRR (refus)</th>
      </tr>
    </thead>
    <tbody>{sum_rows}</tbody>
  </table>

  <h3>Détail par section</h3>
  <table>
    <thead>
      <tr>
        <th style="background:#ecf0f1;text-align:left;min-width:200px">Sous-section</th>
        <th style="background:#ecf0f1;color:#888">N</th>
        <th style="background:#ecf0f1;font-size:0.78em">Métrique</th>
        {header_4cfg}
      </tr>
    </thead>
    <tbody>{sec_html}</tbody>
  </table>
</div></body></html>"""

HTML_OUT.write_text(html, encoding='utf-8')
log.info(f'Terminé. CSV={CSV_OUT.name}  HTML={HTML_OUT.name}')
