"""
Script de comparaison des systèmes RAG via l'API HTTP

Compare les versions:
- v2: RAG amélioré avec hybrid retrieval + reranking
- v6: G-Retriever (GNN)
- v7 router: LlamaIndex avec routing intelligent
- v7 sub_question: LlamaIndex avec décomposition de questions

Usage:
    python compare_rag_api.py                    # Questions par défaut
    python compare_rag_api.py questions.json     # Questions depuis un fichier JSON
    python compare_rag_api.py --questions "Q1" "Q2" "Q3"  # Questions en ligne de commande

Le serveur API doit être lancé au préalable:
    python api_server_multi_version.py
"""

import os
import sys
import json
import time
import argparse
import requests
from datetime import datetime
from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side


# ============================================================================
# CONFIGURATION
# ============================================================================

API_URL = "http://localhost:8000/api/query"
VERSIONS_URL = "http://localhost:8000/api/versions"

# Configurations des systèmes à comparer
SYSTEMS_CONFIG = [
    {"version": "v2", "mode": None, "name": "RAG v2 (Hybrid+Reranking)"},
    {"version": "v2.1", "mode": None, "name": "RAG v2.1 (LlamaIndex)"},
    {"version": "v6", "mode": None, "name": "RAG v6 (G-Retriever GNN)"},
    {"version": "v7", "mode": "router", "name": "RAG v7 (Router)"},
    {"version": "v7", "mode": "sub_question", "name": "RAG v7 (Sub-Question)"},
]

# Questions par défaut
DEFAULT_QUESTIONS = [
    "Quels sont les différents types de bien-être ?",
    "Comment est la qualité de vie à Ajaccio ?",
    "Quels sont les problèmes de transport en Corse ?",
    "Que pensent les habitants de Lozzi de leur cadre de vie ?",
    "Quels sont les indicateurs de santé et d'éducation en Corse ?",
    "Comment les Corses perçoivent-ils l'accès aux services publics ?",
    "Quels sont les enjeux environnementaux en Corse ?",
]


# ============================================================================
# FONCTIONS D'INTERROGATION
# ============================================================================

def check_server_availability() -> Dict:
    """Vérifie si le serveur est disponible et quelles versions sont actives"""
    try:
        response = requests.get(VERSIONS_URL, timeout=5)
        if response.status_code == 200:
            versions = response.json()
            available = {v["version"]: v["available"] for v in versions}
            return {"status": "ok", "versions": available}
        else:
            return {"status": "error", "message": f"HTTP {response.status_code}"}
    except requests.exceptions.ConnectionError:
        return {"status": "error", "message": "Impossible de se connecter au serveur. Lancez api_server_multi_version.py"}
    except Exception as e:
        return {"status": "error", "message": str(e)}


def query_rag_api(question: str, version: str, mode: Optional[str] = None) -> Dict:
    """
    Interroge le système RAG via l'API HTTP

    Args:
        question: Question à poser
        version: Version du RAG (v2, v6, v7, etc.)
        mode: Mode pour v7/v8 (router, sub_question, hybrid)

    Returns:
        Dict avec la réponse, les sources, le temps d'exécution
    """
    start_time = time.time()

    payload = {
        "question": question,
        "rag_version": version,
        "k": 5,
        "use_reranking": True,
    }

    if mode:
        payload["query_mode"] = mode

    try:
        response = requests.post(API_URL, json=payload, timeout=120)
        execution_time = time.time() - start_time

        if response.status_code == 200:
            data = response.json()
            return {
                "version": version,
                "mode": mode,
                "answer": data.get("answer", ""),
                "sources": data.get("sources", []),
                "num_sources": len(data.get("sources", [])),
                "execution_time": execution_time,
                "context": data.get("context", ""),
                "metadata": data.get("metadata", {}),
                "success": True
            }
        else:
            error_detail = response.json().get("detail", response.text) if response.text else f"HTTP {response.status_code}"
            return {
                "version": version,
                "mode": mode,
                "answer": f"Erreur: {error_detail}",
                "sources": [],
                "num_sources": 0,
                "execution_time": execution_time,
                "error": error_detail,
                "success": False
            }
    except requests.exceptions.Timeout:
        return {
            "version": version,
            "mode": mode,
            "answer": "Erreur: Timeout (>120s)",
            "sources": [],
            "num_sources": 0,
            "execution_time": time.time() - start_time,
            "error": "Timeout",
            "success": False
        }
    except Exception as e:
        return {
            "version": version,
            "mode": mode,
            "answer": f"Erreur: {str(e)}",
            "sources": [],
            "num_sources": 0,
            "execution_time": time.time() - start_time,
            "error": str(e),
            "success": False
        }


# ============================================================================
# FONCTIONS DE COMPARAISON
# ============================================================================

def run_comparison(questions: List[str], systems: List[Dict] = None) -> List[Dict]:
    """
    Exécute la comparaison pour toutes les questions et tous les systèmes

    Args:
        questions: Liste des questions à tester
        systems: Liste des configurations de systèmes (utilise SYSTEMS_CONFIG par défaut)

    Returns:
        Liste des résultats de comparaison
    """
    if systems is None:
        systems = SYSTEMS_CONFIG

    all_results = []

    for q_idx, question in enumerate(questions, 1):
        print(f"\n[{q_idx}/{len(questions)}] Question: {question[:60]}...")
        print("-" * 80)

        question_results = {
            "question": question,
            "timestamp": datetime.now().isoformat(),
            "systems": {}
        }

        for sys_config in systems:
            version = sys_config["version"]
            mode = sys_config.get("mode")
            name = sys_config["name"]

            sys_key = f"{version}_{mode}" if mode else version

            print(f"  - {name}...", end=" ", flush=True)
            result = query_rag_api(question, version, mode)

            if result["success"]:
                print(f"OK ({result['execution_time']:.2f}s, {result['num_sources']} sources)")
            else:
                print(f"ERREUR: {result.get('error', 'Unknown')}")

            question_results["systems"][sys_key] = {
                "name": name,
                "version": version,
                "mode": mode,
                **result
            }

        all_results.append(question_results)

    return all_results


def format_comparison_text(all_results: List[Dict]) -> str:
    """Formate les résultats en texte lisible"""
    lines = []
    lines.append("=" * 100)
    lines.append("COMPARAISON DES SYSTÈMES RAG")
    lines.append(f"Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * 100)
    lines.append("")

    for result in all_results:
        lines.append("=" * 100)
        lines.append(f"QUESTION: {result['question']}")
        lines.append("=" * 100)
        lines.append("")

        for sys_key, sys_result in result["systems"].items():
            lines.append("-" * 100)
            lines.append(f"{sys_result['name']}")
            lines.append("-" * 100)
            lines.append(f"Temps d'exécution: {sys_result['execution_time']:.2f}s")
            lines.append(f"Nombre de sources: {sys_result['num_sources']}")

            if sys_result.get("error"):
                lines.append(f"ERREUR: {sys_result['error']}")

            lines.append("")
            lines.append("RÉPONSE:")
            lines.append(sys_result["answer"][:2000] + "..." if len(sys_result.get("answer", "")) > 2000 else sys_result.get("answer", ""))
            lines.append("")

            # Afficher les premières sources
            if sys_result.get("sources"):
                lines.append("SOURCES (3 premières):")
                for i, source in enumerate(sys_result["sources"][:3], 1):
                    score = source.get("score", 0)
                    content = source.get("content", "")[:200]
                    meta = source.get("metadata", {})
                    source_type = meta.get("source", meta.get("type", "unknown"))
                    lines.append(f"  {i}. [{source_type}] (score: {score:.3f})")
                    lines.append(f"     {content}...")
            lines.append("")

        lines.append("")

    # Statistiques globales
    lines.append("=" * 100)
    lines.append("STATISTIQUES GLOBALES")
    lines.append("=" * 100)

    system_stats = {}
    for result in all_results:
        for sys_key, sys_result in result["systems"].items():
            if sys_key not in system_stats:
                system_stats[sys_key] = {
                    "name": sys_result["name"],
                    "times": [],
                    "sources": [],
                    "errors": 0
                }
            if sys_result.get("success"):
                system_stats[sys_key]["times"].append(sys_result["execution_time"])
                system_stats[sys_key]["sources"].append(sys_result["num_sources"])
            else:
                system_stats[sys_key]["errors"] += 1

    for sys_key, stats in system_stats.items():
        lines.append(f"\n{stats['name']}:")
        if stats["times"]:
            avg_time = sum(stats["times"]) / len(stats["times"])
            avg_sources = sum(stats["sources"]) / len(stats["sources"])
            lines.append(f"  - Temps moyen: {avg_time:.2f}s")
            lines.append(f"  - Sources moyennes: {avg_sources:.1f}")
        lines.append(f"  - Erreurs: {stats['errors']}")

    lines.append("")
    return "\n".join(lines)


def export_to_excel(all_results: List[Dict], output_file: str):
    """Exporte les résultats vers un fichier Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparaison RAG"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    wrap_alignment = Alignment(vertical="top", wrap_text=True)

    # Couleurs par système
    colors = {
        "v2": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),  # Vert clair
        "v2.1": PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid"),  # Vert menthe
        "v6": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),  # Orange clair
        "v7_router": PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),  # Bleu clair
        "v7_sub_question": PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),  # Jaune clair
    }

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Récupérer les clés des systèmes depuis le premier résultat
    if not all_results:
        wb.save(output_file)
        return

    system_keys = list(all_results[0]["systems"].keys())

    # Créer les en-têtes
    headers = ["Question"]
    for sys_key in system_keys:
        sys_name = all_results[0]["systems"][sys_key]["name"]
        headers.extend([
            f"{sys_name} - Réponse",
            f"{sys_name} - Sources",
            f"{sys_name} - Temps (s)"
        ])

    # Écrire les en-têtes
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = thin_border

    ws.row_dimensions[1].height = 40

    # Remplir les données
    for row_idx, result in enumerate(all_results, 2):
        col_idx = 1

        # Question
        cell = ws.cell(row=row_idx, column=col_idx, value=result["question"])
        cell.alignment = wrap_alignment
        cell.border = thin_border
        col_idx += 1

        # Données de chaque système
        for sys_key in system_keys:
            sys_result = result["systems"].get(sys_key, {})
            fill_color = colors.get(sys_key, None)

            # Réponse
            answer = sys_result.get("answer", "N/A")
            if len(answer) > 500:
                answer = answer[:500] + "..."
            cell = ws.cell(row=row_idx, column=col_idx, value=answer)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if fill_color:
                cell.fill = fill_color
            col_idx += 1

            # Sources (résumé)
            sources = sys_result.get("sources", [])
            sources_text = f"{len(sources)} sources"
            if sources:
                source_types = [s.get("metadata", {}).get("source", "unknown") for s in sources[:3]]
                sources_text += f"\n({', '.join(source_types)})"
            cell = ws.cell(row=row_idx, column=col_idx, value=sources_text)
            cell.alignment = wrap_alignment
            cell.border = thin_border
            if fill_color:
                cell.fill = fill_color
            col_idx += 1

            # Temps
            exec_time = sys_result.get("execution_time", 0)
            cell = ws.cell(row=row_idx, column=col_idx, value=f"{exec_time:.2f}")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
            if fill_color:
                cell.fill = fill_color
            col_idx += 1

        ws.row_dimensions[row_idx].height = 100

    # Ajuster la largeur des colonnes
    ws.column_dimensions['A'].width = 40  # Question

    col_idx = 2
    for sys_key in system_keys:
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 60  # Réponse
        col_idx += 1
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 20  # Sources
        col_idx += 1
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 12  # Temps
        col_idx += 1

    wb.save(output_file)
    print(f"  OK Fichier Excel créé: {output_file}")


# ============================================================================
# FONCTION PRINCIPALE
# ============================================================================

def main():
    parser = argparse.ArgumentParser(
        description="Compare les systèmes RAG via l'API HTTP",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Exemples:
  python compare_rag_api.py                           # Questions par défaut
  python compare_rag_api.py questions.json            # Depuis fichier JSON
  python compare_rag_api.py --questions "Q1" "Q2"     # Questions en args
        """
    )

    parser.add_argument(
        "questions_file",
        nargs="?",
        help="Fichier JSON contenant une liste de questions"
    )

    parser.add_argument(
        "--questions", "-q",
        nargs="+",
        help="Questions à tester (en ligne de commande)"
    )

    parser.add_argument(
        "--output-dir", "-o",
        default="comparaisons_rag",
        help="Répertoire de sortie (défaut: comparaisons_rag)"
    )

    args = parser.parse_args()

    print("=" * 80)
    print("COMPARAISON DES SYSTÈMES RAG VIA API")
    print("=" * 80)
    print("")

    # Vérifier le serveur
    print("Vérification du serveur API...")
    server_status = check_server_availability()

    if server_status["status"] != "ok":
        print(f"ERREUR: {server_status['message']}")
        print("Lancez d'abord: python api_server_multi_version.py")
        sys.exit(1)

    print(f"Serveur OK. Versions disponibles:")
    available_systems = []
    for sys_config in SYSTEMS_CONFIG:
        version = sys_config["version"]
        is_available = server_status["versions"].get(version, False)
        status = "OK" if is_available else "Non disponible"
        print(f"  - {sys_config['name']}: {status}")
        if is_available:
            available_systems.append(sys_config)

    if not available_systems:
        print("\nAucun système disponible!")
        sys.exit(1)

    print("")

    # Charger les questions
    if args.questions:
        questions = args.questions
    elif args.questions_file:
        with open(args.questions_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            questions = data if isinstance(data, list) else data.get("questions", [])
    else:
        questions = DEFAULT_QUESTIONS

    print(f"Questions à tester ({len(questions)}):")
    for i, q in enumerate(questions, 1):
        print(f"  {i}. {q[:70]}{'...' if len(q) > 70 else ''}")
    print("")

    # Créer le répertoire de sortie
    os.makedirs(args.output_dir, exist_ok=True)

    # Lancer la comparaison
    print("Lancement de la comparaison...")
    all_results = run_comparison(questions, available_systems)

    # Sauvegarder les résultats
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

    # Fichier texte
    txt_file = os.path.join(args.output_dir, f"comparison_api_{timestamp}.txt")
    with open(txt_file, 'w', encoding='utf-8') as f:
        f.write(format_comparison_text(all_results))
    print(f"\nRésultats texte: {txt_file}")

    # Fichier JSON
    json_file = os.path.join(args.output_dir, f"comparison_api_{timestamp}.json")
    with open(json_file, 'w', encoding='utf-8') as f:
        json.dump(all_results, f, indent=2, ensure_ascii=False)
    print(f"Résultats JSON: {json_file}")

    # Fichier Excel
    excel_file = os.path.join(args.output_dir, f"comparison_api_{timestamp}.xlsx")
    export_to_excel(all_results, excel_file)

    # Afficher un résumé
    print("\n" + "=" * 80)
    print("RÉSUMÉ")
    print("=" * 80)

    for sys_config in available_systems:
        sys_key = f"{sys_config['version']}_{sys_config.get('mode')}" if sys_config.get('mode') else sys_config['version']
        times = []
        sources_count = []
        errors = 0

        for result in all_results:
            sys_result = result["systems"].get(sys_key, {})
            if sys_result.get("success"):
                times.append(sys_result["execution_time"])
                sources_count.append(sys_result["num_sources"])
            else:
                errors += 1

        print(f"\n{sys_config['name']}:")
        if times:
            print(f"  - Temps moyen: {sum(times)/len(times):.2f}s")
            print(f"  - Temps total: {sum(times):.2f}s")
            print(f"  - Sources moyennes: {sum(sources_count)/len(sources_count):.1f}")
        print(f"  - Erreurs: {errors}/{len(questions)}")

    print("\n" + "=" * 80)
    print("Comparaison terminée!")
    print("=" * 80)

    return all_results


if __name__ == "__main__":
    main()
