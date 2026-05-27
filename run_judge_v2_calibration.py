"""
run_judge_v2_calibration.py

Lance le judge V2 (GPT-4o, prompt few-shot, barèmes explicites) sur les
20 questions de calibration issues du fichier d'annotation humaine.

Sorties :
  - comparaisons_rag/judge_scores_v2_gpt4o.json   : notes brutes V2
  - comparaisons_rag/judge_agreement_v2_report.md  : accord V2 vs humain

Usage :
    python run_judge_v2_calibration.py
    python run_judge_v2_calibration.py --input "C:/path/to/annotation.xlsx"
"""

import argparse
import json
import logging
import os
import sys
import io
import time
from datetime import datetime
from pathlib import Path

import openpyxl
from dotenv import load_dotenv

load_dotenv()

if hasattr(sys.stdout, "buffer"):
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8", errors="replace")

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler(sys.stdout)],
)
log = logging.getLogger(__name__)

# Colonnes 1-based dans le fichier annotation_humaine_20q_final_v3_avec_juge.xlsx
# (même structure que build_annotation_with_judge.py)
COL_LIGNE       = 1   # A
COL_SECTION     = 2   # B
COL_SUBSECTION  = 3   # C
COL_QUESTION    = 4   # D
COL_REPONSE     = 5   # E
COL_SOURCES     = 6   # F
COL_PERT_H      = 7   # G  pertinence humain
COL_FOND_H      = 11  # K  fondement humain
COL_NUAN_H      = 15  # O  nuance humain
COL_COHE_H      = 19  # S  cohérence humain

DEFAULT_INPUT = r"C:\Users\comiti_g\Downloads\annotation_editee_2026-05-12 (2).xlsx"
OUTPUT_DIR    = Path("comparaisons_rag")


# ── Importer score_judge et _build_sources_text depuis eval_from_excel ──────
def _import_judge():
    sys.path.insert(0, str(Path(__file__).parent))
    from eval_from_excel import score_judge, _build_sources_text
    return score_judge, _build_sources_text


def _parse_sources_from_cell(raw: str) -> list:
    """
    Convertit le texte brut '--- Source N [meta] ---\ncontenu' en liste
    de dicts {content, metadata} compatibles avec score_judge.
    """
    if not raw:
        return []
    import re
    parts = re.split(r"--- Source \d+ \[([^\]]*)\] ---\n?", str(raw).replace("\r\n", "\n"))
    sources = []
    for i in range(1, len(parts), 2):
        meta_str = parts[i]
        body = parts[i + 1].strip() if i + 1 < len(parts) else ""
        if not body:
            continue
        meta = {}
        for kv in meta_str.split(","):
            kv = kv.strip()
            if "=" in kv:
                k, _, v = kv.partition("=")
                meta[k.strip()] = v.strip()
        sources.append({"content": body, "metadata": meta})
    return sources


def run_calibration(input_path: str) -> list:
    log.info("Chargement du fichier : %s", input_path)
    wb = openpyxl.load_workbook(input_path, data_only=True)
    ws = wb.active

    results = []
    total = sum(1 for r in range(2, ws.max_row + 1) if ws.cell(r, COL_QUESTION).value)
    done = 0

    score_judge, _ = _import_judge()

    for r in range(2, ws.max_row + 1):
        question = ws.cell(r, COL_QUESTION).value
        if not question:
            continue
        done += 1
        question = str(question).strip()
        section  = str(ws.cell(r, COL_SECTION).value or "")
        reponse  = str(ws.cell(r, COL_REPONSE).value or "")
        sources_raw = ws.cell(r, COL_SOURCES).value

        # Notes humaines pour comparaison ultérieure
        def to_num(v):
            try:
                return float(v) if v not in (None, "") else None
            except (TypeError, ValueError):
                return None

        pert_h = to_num(ws.cell(r, COL_PERT_H).value)
        fond_h = to_num(ws.cell(r, COL_FOND_H).value)
        nuan_h = to_num(ws.cell(r, COL_NUAN_H).value)
        cohe_h = to_num(ws.cell(r, COL_COHE_H).value)
        dims_h = [x for x in [pert_h, fond_h, nuan_h, cohe_h] if x is not None]
        sg_h   = round(sum(dims_h) / len(dims_h), 2) if dims_h else None

        sources = _parse_sources_from_cell(sources_raw)
        log.info("[%d/%d] R%d : %s", done, total, r, question[:70])

        j = score_judge(question, reponse, sources, section)

        if j.get("error"):
            log.warning("   ÉCHEC juge : %s", j["error"])
        else:
            log.info("   score_global juge=%s  |  humain=%s",
                     j.get("score_global"), sg_h)

        results.append({
            "row":      r,
            "question": question,
            "section":  section,
            "scores_humain": {
                "pertinence":            pert_h,
                "fondement_factuel":     fond_h,
                "nuance_incertitude":    nuan_h,
                "coherence_qualiquanti": cohe_h,
                "score_global":          sg_h,
            },
            "scores_juge_v2": j,
            "timestamp": datetime.now().isoformat(),
        })
        time.sleep(1)  # éviter le rate-limit

    return results


def compute_agreement(results: list) -> dict:
    """Métriques d'accord Pearson / Spearman / MAE sur score_global."""
    try:
        from scipy import stats
        import numpy as np
    except ImportError:
        log.warning("scipy/numpy non disponibles — métriques d'accord ignorées")
        return {}

    dims = ("pertinence", "fondement_factuel", "nuance_incertitude",
            "coherence_qualiquanti", "score_global")
    metrics = {}
    for d in dims:
        h_vals, j_vals = [], []
        for r in results:
            h = r["scores_humain"].get(d)
            j = r["scores_juge_v2"].get(d)
            if h is not None and j is not None and not r["scores_juge_v2"].get("error"):
                h_vals.append(float(h))
                j_vals.append(float(j))
        if len(h_vals) < 3:
            continue
        h = np.array(h_vals)
        j = np.array(j_vals)
        pearson_r, pearson_p = stats.pearsonr(h, j)
        spearman_r, _ = stats.spearmanr(h, j)
        mae = float(np.mean(np.abs(h - j)))
        bias = float(np.mean(h - j))
        exact = float(np.mean(h == j))
        within1 = float(np.mean(np.abs(h - j) <= 1))
        metrics[d] = {
            "n":         len(h_vals),
            "pearson_r": round(pearson_r, 3),
            "pearson_p": round(pearson_p, 3),
            "spearman":  round(spearman_r, 3),
            "mae":       round(mae, 3),
            "bias_h_minus_j": round(bias, 3),
            "exact_pct": round(exact * 100, 1),
            "within1_pct": round(within1 * 100, 1),
        }
    return metrics


def build_report(results: list, metrics: dict, ts: str) -> str:
    lines = [
        "# Accord juge V2 (GPT-4o few-shot) vs annotations humaines",
        "",
        f"**Date** : {ts}  ",
        f"**Modèle juge** : gpt-4o (prompt V2 — few-shot, barèmes explicites)  ",
        f"**Questions** : {len(results)}",
        "",
        "## Métriques d'accord",
        "",
        "| Dimension | n | Pearson r | p | Spearman | MAE | Biais (H−J) | Exact % | ±1 % |",
        "|-----------|---|-----------|---|----------|-----|-------------|---------|------|",
    ]
    for d, m in metrics.items():
        lines.append(
            f"| {d} | {m['n']} | {m['pearson_r']} | {m['pearson_p']} "
            f"| {m['spearman']} | {m['mae']} | {m['bias_h_minus_j']} "
            f"| {m['exact_pct']} | {m['within1_pct']} |"
        )
    lines += ["", "## Détail par question", ""]
    for r in results:
        j = r["scores_juge_v2"]
        h = r["scores_humain"]
        err = j.get("error")
        lines.append(f"### R{r['row']} — {r['question'][:80]}")
        if err:
            lines.append(f"> ⚠️ Erreur juge : {err}")
        else:
            lines.append(
                f"| | Pertinence | Fondement | Nuance | Cohérence | Score global |"
            )
            lines.append("|---|---|---|---|---|---|")
            lines.append(
                f"| 👤 Humain | {h.get('pertinence','—')} | {h.get('fondement_factuel','—')} "
                f"| {h.get('nuance_incertitude','—')} | {h.get('coherence_qualiquanti','—')} "
                f"| {h.get('score_global','—')} |"
            )
            lines.append(
                f"| 🤖 Juge V2 | {j.get('pertinence','—')} | {j.get('fondement_factuel','—')} "
                f"| {j.get('nuance_incertitude','—')} | {j.get('coherence_qualiquanti','—')} "
                f"| {j.get('score_global','—')} |"
            )
            if j.get("raisonnement"):
                lines.append(f"\n> *Raisonnement juge* : {j['raisonnement']}")
        lines.append("")
    return "\n".join(lines)


# ── Tests unitaires ──────────────────────────────────────────────────────────
def _run_tests():
    from eval_from_excel import _parse_judge_v2

    # Test 1 : format V2 structuré
    j = {
        "raisonnement": "test",
        "pertinence":            {"note": 4, "justification": "bien"},
        "fondement_factuel":     {"note": 3, "justification": "moyen"},
        "nuance_incertitude":    {"note": 5, "justification": "excellent"},
        "coherence_qualiquanti": {"note": 4, "justification": "bon"},
    }
    r = _parse_judge_v2(j)
    assert r["pertinence"] == 4,          f"pertinence attendu 4, obtenu {r['pertinence']}"
    assert r["pertinence_justif"] == "bien"
    assert r["score_global"] == 4.0,      f"score_global attendu 4.0, obtenu {r['score_global']}"
    assert r["raisonnement"] == "test"
    log.info("Test 1 OK : format V2 structuré")

    # Test 2 : fallback format V1 plat
    j2 = {
        "pertinence": 3, "pertinence_justif": "passable",
        "fondement_factuel": 4, "fondement_factuel_justif": "ok",
        "nuance_incertitude": 2, "nuance_incertitude_justif": "insuffisant",
        "coherence_qualiquanti": 5, "coherence_qualiquanti_justif": "parfait",
    }
    r2 = _parse_judge_v2(j2)
    assert r2["pertinence"] == 3
    assert r2["score_global"] == 3.5, f"score_global attendu 3.5, obtenu {r2['score_global']}"
    log.info("Test 2 OK : fallback format V1")

    # Test 3 : champ raisonnement absent
    j3 = {"pertinence": {"note": 5, "justification": "x"},
          "fondement_factuel": {"note": 5, "justification": "x"},
          "nuance_incertitude": {"note": 5, "justification": "x"},
          "coherence_qualiquanti": {"note": 5, "justification": "x"}}
    r3 = _parse_judge_v2(j3)
    assert r3["raisonnement"] is None
    assert r3["score_global"] == 5.0
    log.info("Test 3 OK : raisonnement absent géré")

    log.info("Tous les tests unitaires passent.")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default=DEFAULT_INPUT,
                        help="Fichier Excel d'annotation (20 questions)")
    parser.add_argument("--test", action="store_true",
                        help="Lance uniquement les tests unitaires")
    args = parser.parse_args()

    if args.test:
        _run_tests()
        return

    OUTPUT_DIR.mkdir(exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    results = run_calibration(args.input)

    # Sauvegarde JSON brut
    json_path = OUTPUT_DIR / f"judge_scores_v2_gpt4o_{ts}.json"
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(results, f, ensure_ascii=False, indent=2)
    log.info("Scores V2 sauvegardés : %s", json_path)

    # Métriques d'accord
    metrics = compute_agreement(results)

    # Rapport Markdown
    ts_readable = datetime.now().strftime("%Y-%m-%d %H:%M")
    md = build_report(results, metrics, ts_readable)
    md_path = OUTPUT_DIR / f"judge_agreement_v2_report_{ts}.md"
    with open(md_path, "w", encoding="utf-8") as f:
        f.write(md)
    log.info("Rapport accord V2 : %s", md_path)

    # Résumé console
    sg = metrics.get("score_global", {})
    if sg:
        print("\n" + "=" * 60)
        print("ACCORD V2 — Score global")
        print(f"  Pearson  : {sg['pearson_r']}  (p={sg['pearson_p']})")
        print(f"  Spearman : {sg['spearman']}")
        print(f"  MAE      : {sg['mae']}")
        print(f"  Biais    : {sg['bias_h_minus_j']} (H − J)")
        print(f"  Exact    : {sg['exact_pct']} %")
        print(f"  ±1       : {sg['within1_pct']} %")
        print("=" * 60)


if __name__ == "__main__":
    main()
