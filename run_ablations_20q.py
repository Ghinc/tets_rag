"""Run 4 ablation configs on 20 annotated questions."""
import openpyxl, requests, json, time, sys
from datetime import datetime

XLSX = r"C:\Users\comiti_g\Downloads\annotation_humaine_20q_final_v3_avec_juge.xlsx"
BASE = "http://localhost:8000/api/query"
HEADERS = {"Content-Type": "application/json"}
VERSIONS = ["v_vanilla_k10", "v_vanilla_k25", "v_decomp", "v_decomp_raptor"]
OUT_DIR = "comparaisons_rag"

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
questions = []
for r in range(2, ws.max_row + 1):
    questions.append({
        "excel_row": ws.cell(r, 1).value,
        "section": ws.cell(r, 2).value,
        "subsection": ws.cell(r, 3).value,
        "question": ws.cell(r, 4).value,
    })

print(f"20 questions chargees. Lancement des {len(VERSIONS)} configs...", flush=True)

results = {}
for ver in VERSIONS:
    print(f"\n{'='*55}", flush=True)
    print(f"Config: {ver}", flush=True)
    print('='*55, flush=True)
    results[ver] = []
    k = 10 if ver == "v_vanilla_k10" else (25 if ver == "v_vanilla_k25" else 5)
    for i, q in enumerate(questions, 1):
        question = q["question"]
        payload = {"question": question, "rag_version": ver, "k": k}
        try:
            t0 = time.time()
            r_resp = requests.post(BASE, json=payload, headers=HEADERS, timeout=300)
            elapsed = time.time() - t0
            if r_resp.status_code == 200:
                data = r_resp.json()
                entry = {
                    "excel_row": q["excel_row"],
                    "section": q["section"],
                    "subsection": q["subsection"],
                    "question": question,
                    "answer": data.get("answer", ""),
                    "n_sources": len(data.get("sources", [])),
                    "n_subquestions": len(data.get("sub_questions") or []),
                    "elapsed_s": round(elapsed, 1),
                    "status": "ok",
                }
                print(f"  [{i:2}/20] R{q['excel_row']} {question[:45]:<45} | {elapsed:.1f}s | {entry['n_sources']}src", flush=True)
            else:
                entry = {
                    "excel_row": q["excel_row"], "question": question,
                    "status": "error", "error": r_resp.text[:200],
                }
                print(f"  [{i:2}/20] R{q['excel_row']} ERREUR {r_resp.status_code}: {r_resp.text[:100]}", flush=True)
        except Exception as e:
            entry = {"excel_row": q["excel_row"], "question": question, "status": "exception", "error": str(e)}
            print(f"  [{i:2}/20] R{q['excel_row']} EXCEPTION: {e}", flush=True)
        results[ver].append(entry)

ts = datetime.now().strftime("%Y%m%d_%H%M%S")
out_path = f"{OUT_DIR}/ablations_20q_{ts}.json"
with open(out_path, "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2)
print(f"\nResultats sauvegardes : {out_path}", flush=True)

print("\n--- RESUME ---", flush=True)
for ver in VERSIONS:
    ok = sum(1 for e in results[ver] if e.get("status") == "ok")
    err = len(results[ver]) - ok
    times = [e.get("elapsed_s", 0) for e in results[ver] if e.get("status") == "ok"]
    avg_t = sum(times) / max(len(times), 1)
    print(f"  {ver:<20} : {ok}/20 OK, {err} err, temps moyen {avg_t:.1f}s", flush=True)
