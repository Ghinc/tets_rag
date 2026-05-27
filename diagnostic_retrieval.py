"""
diagnostic_retrieval.py — Métriques de retrieval (Recall, Precision, F1) par config.

Appelle l'API RAG sans judge (pas d'appel OpenAI/Claude), calcule les métriques
sur les 20 questions de calibration, et ventile par sous-section.

Usage:
    python diagnostic_retrieval.py                  # 20 questions, 4 configs
    python diagnostic_retrieval.py --max 5          # test rapide
    python diagnostic_retrieval.py --skip-api       # recharge depuis cache JSON
"""
import sys, json, re, time, argparse, requests
sys.stdout.reconfigure(encoding="utf-8")
from pathlib import Path
from collections import defaultdict

sys.path.insert(0, str(Path(__file__).parent))
from eval_from_excel import (
    load_questions,
    parse_retrieval_ground_truth,
    classify_source,
    score_retrieval as _score_retrieval_orig,
    f1_score,
    _KNOWN_COMMUNES,
    _strip_acc_lower,
)

# Equivalences catégorie GT → aliases acceptés pour precision aussi
_EQUIVALENCES = {
    "methodology_oppchovec": {"oppchovec"},
    "stats_enquete":         {"stats_repondants"},
}

def score_retrieval(sources: list, retrieval_gt_text) -> dict:
    """Wrapper qui corrige la precision pour les aliases d'équivalence."""
    result = _score_retrieval_orig(sources, retrieval_gt_text)
    if result.get("precision") is not None and result.get("expected"):
        required = set(result["expected"])
        retrieved = set(result.get("retrieved", []))
        optional_cats = {"wiki"}
        # Expand all_valid avec les aliases
        alias_expanded = set()
        for cat in required:
            alias_expanded |= _EQUIVALENCES.get(cat, set())
        all_valid = required | optional_cats | alias_expanded
        precision = len(retrieved & all_valid) / max(len(retrieved), 1)
        result = dict(result)
        result["precision"] = round(precision, 3)
        result["f1"] = f1_score(result["precision"], result["recall"])
        result["detail"] = (
            f"Recall={result['recall']:.0%} ({len(result['hits'])}/{len(required)}) · "
            f"Precision={precision:.0%} · F1={result['f1'] or 0:.0%}"
        )
    return result

XLSX        = r"C:\Users\comiti_g\Downloads\annotation_humaine_20q_final_v3_avec_juge.xlsx"
BASE        = "http://localhost:8000/api/query"
HEADERS     = {"Content-Type": "application/json"}
OUT_JSON    = "comparaisons_rag/diagnostic_retrieval_20q.json"
CACHE_JSON  = "comparaisons_rag/diagnostic_retrieval_cache.json"

VERSIONS = {
    "v_vanilla_k10":   {"rag_version": "v_vanilla_k10",   "k": 10},
    "v_vanilla_k25":   {"rag_version": "v_vanilla_k25",   "k": 25},
    "v_decomp":        {"rag_version": "v_decomp",         "k": 5},
    "v_decomp_raptor": {"rag_version": "v_decomp_raptor",  "k": 5},
}

VANILLA_CONFIGS = {"v_vanilla_k10", "v_vanilla_k25"}
RAPTOR_CATS     = {"raptor_portrait", "raptor_enquete", "raptor_both"}


def _flatten_source(s: dict) -> dict:
    """Aplatit metadata dans le dict source pour que classify_source() fonctionne."""
    meta = s.get("metadata") or {}
    flat = dict(s)
    for field in ("type", "source_type", "view_name", "commune", "label"):
        if field not in flat or not flat[field]:
            flat[field] = meta.get(field, "")
    return flat


def query_rag(question: str, rag_version: str, k: int, retries: int = 3) -> dict:
    payload = {"question": question, "rag_version": rag_version, "k": k}
    for attempt in range(retries):
        try:
            r = requests.post(BASE, json=payload, headers=HEADERS, timeout=300)
            if r.status_code == 200:
                return r.json()
            print(f"    [ERREUR {r.status_code}]", flush=True)
            return {"error": str(r.status_code), "sources": []}
        except Exception as e:
            if attempt < retries - 1:
                wait = 5 * (attempt + 1)
                print(f"    [Exception, attente {wait}s] {e}", flush=True)
                time.sleep(wait)
            else:
                return {"error": str(e), "sources": []}


def _gt_requires_raptor(retrieval_gt_text) -> bool:
    if not retrieval_gt_text:
        return False
    items = parse_retrieval_ground_truth(retrieval_gt_text)
    return any(i["category"] in RAPTOR_CATS and not i["optional"] for i in items)


def _count_pool_verbatims(commune: str, csp: str | None) -> int:
    """Compte les verbatims disponibles dans le pool ChromaDB pour commune (+ CSP)."""
    try:
        import chromadb
        client = chromadb.PersistentClient(
            path="c:/These/Données2/fichiers_pour_rag/chroma_portrait"
        )
        coll = client.get_collection("portrait_entretiens")
        where = {"nom": {"$eq": commune}} if commune else None
        results = coll.get(where=where, include=[])
        return len(results["ids"])
    except Exception:
        return 0


def score_retrieval_vanilla_raptor(sources: list, retrieval_gt_text,
                                   verbatim_pool_size: int = 0) -> dict:
    """
    score_retrieval avec crédit vanilla pour RAPTOR :
    si GT requiert raptor_portrait/enquete et le pool ≥ 1 verbatims,
    crédite le hit raptor si vanilla récupère ≥ 30% du pool.
    """
    gt_items = parse_retrieval_ground_truth(retrieval_gt_text)
    if not gt_items:
        return {"recall": None, "precision": None, "detail": "GT vide après parsing"}

    if any(i["category"] == "rien" for i in gt_items):
        retrieved = [c for c in (classify_source(_flatten_source(s)) for s in sources if s) if c != "autre"]
        return {
            "recall": None, "precision": None,
            "refusal_case": True,
            "retrieved_count": len(retrieved),
            "detail": f"Refusal attendu ({len(retrieved)} sources récupérées)",
        }

    # Appliquer score_retrieval standard avec sources aplaties
    flat_sources = [_flatten_source(s) for s in sources]
    result = score_retrieval(flat_sources, retrieval_gt_text)

    # Crédit RAPTOR vanilla : si recall < 1 et GT requiert raptor
    if result.get("recall") is not None and result["recall"] < 1.0:
        required_raptor = {
            i["category"] for i in gt_items
            if i["category"] in RAPTOR_CATS and not i["optional"]
        }
        if required_raptor and verbatim_pool_size > 0:
            # Compter les verbatims/entretiens récupérés
            n_verbatims = sum(
                1 for s in flat_sources
                if classify_source(s) in ("verbatims", "entretiens", "raptor_portrait", "raptor_enquete")
            )
            threshold = max(1, round(verbatim_pool_size * 0.3))
            if n_verbatims >= threshold:
                # Crédit : on considère les cats raptor requises comme hits
                hits_set = set(result.get("hits", []))
                for cat in required_raptor:
                    hits_set.add(cat)
                expected = set(result.get("expected", []))
                new_recall = len(hits_set & expected) / len(expected) if expected else result["recall"]
                if new_recall > result["recall"]:
                    result = dict(result)
                    result["recall"] = round(new_recall, 3)
                    result["hits"] = sorted(hits_set)
                    result["raptor_vanilla_credit"] = True
                    result["verbatims_retrieved"] = n_verbatims
                    result["verbatims_pool"] = verbatim_pool_size
                    from eval_from_excel import f1_score
                    result["f1"] = f1_score(result["precision"], result["recall"])
    return result


def build_cache(questions: list, args) -> dict:
    """Appelle l'API RAG pour toutes les questions × configs, retourne le cache."""
    cache = {}
    for ver, cfg in VERSIONS.items():
        print(f"\n{'='*58}", flush=True)
        print(f"Config: {ver}  (k={cfg['k']})", flush=True)
        print('='*58, flush=True)
        cache[ver] = []
        for i, q in enumerate(questions, 1):
            t0 = time.time()
            data = query_rag(q["question"], cfg["rag_version"], cfg["k"])
            elapsed = round(time.time() - t0, 1)
            sources = data.get("sources", [])
            entry = {
                "excel_row":   q["excel_row"],
                "section":     q["section"],
                "subsection":  q["subsection"],
                "question":    q["question"],
                "retrieval_gt": str(q.get("retrieval_gt") or ""),
                "rag_ok":      "error" not in data,
                "n_sources":   len(sources),
                "sources":     [_flatten_source(s) for s in sources],
                "elapsed_s":   elapsed,
            }
            status = f"ok ({len(sources)}src, {elapsed}s)" if "error" not in data else f"ERREUR: {data.get('error')}"
            print(f"  [{i:2}/{len(questions)}] R{q['excel_row']} {status}", flush=True)
            cache[ver].append(entry)
            time.sleep(0.5)

    with open(CACHE_JSON, "w", encoding="utf-8") as f:
        json.dump(cache, f, ensure_ascii=False, indent=2)
    print(f"\nCache sauvegardé : {CACHE_JSON}", flush=True)
    return cache


def compute_metrics(cache: dict, questions: list, pool_sizes: dict) -> dict:
    """Calcule les métriques de retrieval par config × question."""
    results = {}

    for ver, entries in cache.items():
        is_vanilla = ver in VANILLA_CONFIGS
        results[ver] = []
        for entry in entries:
            q_row = entry["excel_row"]
            gt_text = entry.get("retrieval_gt") or ""
            sources = entry.get("sources", [])

            # Trouver la question dans la liste pour le pool RAPTOR
            q_meta = next((q for q in questions if q["excel_row"] == q_row), {})
            pool_size = 0
            if is_vanilla and _gt_requires_raptor(gt_text):
                # Prendre la commune depuis le GT parsé (plus fiable que la question)
                gt_items = parse_retrieval_ground_truth(gt_text)
                commune = next(
                    (i["commune"] for i in gt_items
                     if i["category"] in RAPTOR_CATS and not i["optional"] and i["commune"]),
                    None
                )
                pool_size = pool_sizes.get(commune, 0) if commune else 0

            if not entry.get("rag_ok"):
                m = {"recall": None, "precision": None, "f1": None, "detail": "RAG error"}
            elif is_vanilla and _gt_requires_raptor(gt_text):
                m = score_retrieval_vanilla_raptor(sources, gt_text or None, pool_size)
            else:
                flat = [_flatten_source(s) if "type" not in s else s for s in sources]
                m = score_retrieval(flat, gt_text or None)

            results[ver].append({
                "excel_row":   q_row,
                "section":     entry["section"],
                "subsection":  entry["subsection"],
                "question":    entry["question"],
                "retrieval_gt": gt_text,
                "n_sources":   entry["n_sources"],
                "recall":      m.get("recall"),
                "precision":   m.get("precision"),
                "f1":          m.get("f1"),
                "refusal_case": m.get("refusal_case", False),
                "raptor_credit": m.get("raptor_vanilla_credit", False),
                "detail":      m.get("detail", ""),
                "expected":    m.get("expected", []),
                "retrieved":   m.get("retrieved", []),
                "hits":        m.get("hits", []),
            })
    return results


def avg(vals):
    v = [x for x in vals if x is not None]
    return round(sum(v) / len(v), 3) if v else None


def fmt(v, pct=False):
    if v is None:
        return "  —  "
    return f"{v*100:5.1f}%" if pct else f"{v:.3f}"


def print_global_table(metrics: dict):
    print("\n" + "="*72)
    print("TABLEAU GLOBAL — Métriques de retrieval (20 questions)")
    print("="*72)
    header = f"{'Config':<22} {'n_ok':>5} {'Recall':>9} {'Precision':>10} {'F1':>8} {'TNRR':>8}"
    print(header)
    print("-"*72)
    for ver, entries in metrics.items():
        ok      = [e for e in entries if e["recall"] is not None]
        refusal = [e for e in entries if e.get("refusal_case")]
        recalls    = [e["recall"] for e in ok]
        precisions = [e["precision"] for e in ok if e["precision"] is not None]
        f1s        = [e["f1"] for e in ok if e["f1"] is not None]
        # TNRR : refusal_case → on considère True Negative si aucune source non-autre récupérée
        tnrr_scores = [1 if e["n_sources"] == 0 else 0 for e in refusal]
        tnrr = avg(tnrr_scores)
        print(
            f"  {ver:<20} {len(ok):>5} {fmt(avg(recalls),True):>9} "
            f"{fmt(avg(precisions),True):>10} {fmt(avg(f1s),True):>8} "
            f"{fmt(tnrr,True) if tnrr is not None else '  N/A':>8}"
            f"  ({len(refusal)} refus)"
        )


def print_subsection_tables(metrics: dict, versions: list):
    # Regrouper par section/subsection
    subs = defaultdict(list)
    first_ver = versions[0]
    for e in metrics[first_ver]:
        key = (e["section"], e["subsection"])
        subs[key].append(e["excel_row"])

    # Ordre : Retrieval mono-commune en premier
    section_order = [
        "Retrieval mono-commune",
        "Raisonnement comparatif",
        "Raisonnement causal et contre-intuitif",
        "Gestion d'absence d'information",
        "Gestion de l'incertitude et des biais",
        "Robustesse sémantique",
        "Limites architecturales",
    ]
    def section_rank(key):
        sec, _ = key
        try:
            return section_order.index(sec)
        except ValueError:
            return 99

    print("\n" + "="*72)
    print("VENTILATION PAR SOUS-SECTION")
    print("="*72)

    dims_judge = [
        ("score_global",           "Judge global"),
        ("v41_pertinence",         "Pertinence"),
        ("v41_fondement_factuel",  "Fondement"),
        ("v41_nuance_incertitude", "Nuance"),
        ("v41_coherence_qualiquanti", "Cohérence q/q"),
    ]

    # Charger le JSON ablations pour les scores judge
    try:
        with open("comparaisons_rag/ablations_20q_run_judge_merged_20260516.json",
                  encoding="utf-8") as f:
            judge_data = json.load(f)
        judge_idx = {
            ver: {e["excel_row"]: e for e in judge_data[ver]}
            for ver in versions if ver in judge_data
        }
    except Exception as ex:
        print(f"[WARN] Impossible de charger les scores judge : {ex}")
        judge_idx = {}

    for key in sorted(subs.keys(), key=section_rank):
        sec, sub = key
        rows = subs[key]
        n = len(rows)
        print(f"\n{'─'*72}")
        print(f"  {sec} › {sub}  (n={n})")
        print(f"{'─'*72}")

        col_w = 16
        hdr = f"  {'Métrique':<22}" + "".join(f"{v:<{col_w}}" for v in versions)
        print(hdr)
        print("  " + "-"*70)

        # Judge scores
        if judge_idx:
            for field, label in dims_judge:
                row_vals = []
                for ver in versions:
                    idx = judge_idx.get(ver, {})
                    vals = [idx[r][field] for r in rows if r in idx and idx[r].get(field) is not None]
                    row_vals.append(f"{sum(vals)/len(vals):.2f}" if vals else "  —  ")
                print(f"  {label:<22}" + "".join(f"{v:<{col_w}}" for v in row_vals))

        print("  " + "·"*70)

        # Retrieval
        for met, label in [("recall","Recall"), ("precision","Precision"), ("f1","F1")]:
            row_vals = []
            for ver in versions:
                vals = [e[met] for e in metrics[ver]
                        if e["excel_row"] in rows and e[met] is not None]
                row_vals.append(f"{sum(vals)/len(vals)*100:.1f}%" if vals else "  —  ")
            print(f"  {label:<22}" + "".join(f"{v:<{col_w}}" for v in row_vals))


def print_synthesis(metrics: dict, versions: list):
    print("\n" + "="*72)
    print("SYNTHÈSE — Où les configs divergent")
    print("="*72)

    # Charger scores judge
    try:
        with open("comparaisons_rag/ablations_20q_run_judge_merged_20260516.json",
                  encoding="utf-8") as f:
            jd = json.load(f)
    except Exception:
        jd = {}

    van = "v_vanilla_k10"
    rap = "v_decomp_raptor"

    if van not in jd or rap not in jd:
        print("  (données judge non disponibles)")
        return

    # Par sous-section : judge global + recall
    subs = defaultdict(lambda: {"judge_van": [], "judge_rap": [], "recall_van": [], "recall_rap": []})
    for e_van in metrics.get(van, []):
        row = e_van["excel_row"]
        sub = e_van["subsection"] or e_van["section"]
        j_van = next((e["v41_score_global"] for e in jd[van] if e["excel_row"] == row), None)
        j_rap = next((e["v41_score_global"] for e in jd[rap] if e["excel_row"] == row), None)
        e_rap = next((e for e in metrics.get(rap, []) if e["excel_row"] == row), {})
        if j_van is not None:
            subs[sub]["judge_van"].append(j_van)
        if j_rap is not None:
            subs[sub]["judge_rap"].append(j_rap)
        if e_van.get("recall") is not None:
            subs[sub]["recall_van"].append(e_van["recall"])
        if e_rap.get("recall") is not None:
            subs[sub]["recall_rap"].append(e_rap["recall"])

    def avg_l(lst): return sum(lst)/len(lst) if lst else None

    print(f"\n  {'Sous-section':<38} {'ΔJudge':>8} {'ΔRecall':>8}  Note")
    print("  " + "-"*68)
    for sub, v in sorted(subs.items()):
        jv, jr = avg_l(v["judge_van"]), avg_l(v["judge_rap"])
        rv, rr = avg_l(v["recall_van"]), avg_l(v["recall_rap"])
        if jv is None or jr is None:
            continue
        dj = jr - jv
        dr = (rr - rv) if rv is not None and rr is not None else None
        note = ""
        if dr is not None and dr > 0.15 and dj <= 0.0:
            note = "⚠ recall↑ mais judge≈ (judge ne capte pas le gain retrieval)"
        elif dj > 0.1:
            note = "✓ raptor > vanilla"
        elif dj < -0.1:
            note = "✗ vanilla > raptor"
        dr_str = f"{dr*100:+.1f}%" if dr is not None else "  —  "
        print(f"  {sub:<38} {dj:+.2f}    {dr_str:>8}  {note}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--max", type=int, default=20)
    parser.add_argument("--skip-api", action="store_true",
                        help="Recharge depuis le cache JSON (pas d'appel API)")
    args = parser.parse_args()

    # Charger les 20 IDs depuis annotation_humaine (col 1 = ligne Excel ref)
    import openpyxl
    wb20 = openpyxl.load_workbook(XLSX)
    ws20 = wb20.active
    calib_rows = {}   # excel_row → {section, subsection, question}
    for r in range(2, ws20.max_row + 1):
        row_ref = ws20.cell(r, 1).value   # "Ligne Excel" = ref dans le fichier 103q
        q_text  = ws20.cell(r, 4).value   # "Question"
        if not row_ref or not q_text:
            continue
        calib_rows[int(row_ref)] = {
            "section":    ws20.cell(r, 2).value or "",
            "subsection": ws20.cell(r, 3).value or "",
            "question":   str(q_text).strip(),
        }

    # Charger retrieval GT depuis le fichier 103 questions
    XLSX_103 = r"C:\Users\comiti_g\Downloads\rag_evaluation_with_metrics_full.xlsx"
    wb103 = openpyxl.load_workbook(XLSX_103)
    ws103 = wb103.active
    questions = []
    for r in range(2, ws103.max_row + 1):
        if r not in calib_rows:
            continue
        meta = calib_rows[r]
        questions.append({
            "excel_row":    r,
            "section":      meta["section"],
            "subsection":   meta["subsection"],
            "question":     meta["question"],
            "retrieval_gt": ws103.cell(r, 4).value,  # col D = "Documents à retrieve"
        })

    # Trier par excel_row pour garder l'ordre
    questions.sort(key=lambda x: x["excel_row"])
    questions = questions[:args.max]
    print(f"{len(questions)} questions chargées (GT depuis {Path(XLSX_103).name})", flush=True)

    # Cache RAG (sources)
    if args.skip_api and Path(CACHE_JSON).exists():
        print(f"Chargement cache : {CACHE_JSON}", flush=True)
        with open(CACHE_JSON, encoding="utf-8") as f:
            cache = json.load(f)
    else:
        print("Appels API RAG (sans judge)...", flush=True)
        cache = build_cache(questions, args)

    # Taille pool verbatims par commune (pour crédit vanilla RAPTOR)
    print("\nCalcul pools verbatims par commune...", flush=True)
    communes_needed = set()
    for q in questions:
        if _gt_requires_raptor(q.get("retrieval_gt")):
            gt_items = parse_retrieval_ground_truth(q.get("retrieval_gt"))
            for item in gt_items:
                if item["category"] in RAPTOR_CATS and not item["optional"] and item["commune"]:
                    communes_needed.add(item["commune"])
    pool_sizes = {}
    for c in communes_needed:
        pool_sizes[c] = _count_pool_verbatims(c, None)
        print(f"  {c}: {pool_sizes[c]} verbatims/entretiens", flush=True)

    # Calculer métriques
    versions_list = list(VERSIONS.keys())
    metrics = compute_metrics(cache, questions, pool_sizes)

    # Tableaux
    print_global_table(metrics)
    print_subsection_tables(metrics, versions_list)
    print_synthesis(metrics, versions_list)

    # Sauvegarde
    out = {
        "pool_sizes": pool_sizes,
        "metrics_by_config": metrics,
    }
    with open(OUT_JSON, "w", encoding="utf-8") as f:
        json.dump(out, f, ensure_ascii=False, indent=2)
    print(f"\nSauvegardé : {OUT_JSON}", flush=True)


if __name__ == "__main__":
    main()
