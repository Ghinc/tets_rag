"""
stability_harness.py — Harnais de stabilité inter-runs pour le pipeline RAG.

Mesure la variance de la pipeline sur M runs × N questions :
  - Sémantique  : cosinus BGE-M3 pairwise (leave-one-out)
  - Lexical     : BLEU-4 + ROUGE-L (leave-one-out)
  - Récupération: Jaccard + overlap@5 sur source_keys
  - Attribution : Jaccard sur types sources_mobilisees

Usage:
    python stability_harness.py            # run de validation (2 questions × 5 runs)
    python stability_harness.py --yes      # sans confirmation interactive
    python stability_harness.py --resume comparaisons_rag/stability/runs_XXXX.json
"""

# SentenceTransformer DOIT être importé avant sacrebleu/rouge_score (conflit DLL Windows/PyTorch)
from sentence_transformers import SentenceTransformer  # noqa: E402 — ordre obligatoire

import os
import re
import sys

# Forcer UTF-8 sur stdout/stderr (console Windows cp1252 sinon)
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
if hasattr(sys.stderr, "reconfigure"):
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
import json
import time
import argparse
import traceback
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import openpyxl
import requests

# ============================================================
# CONSTANTES — modifier ici pour changer la campagne
# ============================================================

TARGET_ROWS = [6, 11, 41, 44, 48, 68, 83, 98]   # campagne complète
# Run de validation 2 questions — décommenter si besoin :
# TARGET_ROWS = [62, 41]

M_RUNS         = 5
RAG_VERSION    = "v_decomp_raptor"
K              = 5
TEMP_OVERRIDE: Optional[float] = None   # None = températures de production | 0.0 = greedy/déterministe
API_BASE       = "http://localhost:8000"
QUESTION_DELAY = 3.0   # secondes entre questions
RUN_DELAY      = 1.5   # secondes entre runs d'une même question
API_TIMEOUT    = 300   # secondes timeout HTTP
API_RETRIES    = 3     # tentatives avant abandon

EXCEL_PATH     = r"C:\Users\comiti_g\Downloads\rag_evaluation_with_metrics_full.xlsx"
OUTPUT_DIR     = Path("comparaisons_rag/stability")

EMBED_MODEL_NAME = "BAAI/bge-m3"

# ============================================================
# TABLE DE MAPPING : labels bruts sources_mobilisees → types coarse
# ============================================================

COARSE_MAP: Dict[str, str] = {
    # indicateurs_oppchovec
    "Indicateurs OppChoVec — Ajaccio":                         "indicateurs_oppchovec",
    "Indicateurs OppChoVec — Ghisonaccia":                     "indicateurs_oppchovec",
    "Indicateurs OppChoVec — Corscia":                         "indicateurs_oppchovec",
    "Indicateurs OppChoVec — Corte":                           "indicateurs_oppchovec",
    "Indicateurs OppChoVec — Ocana":                           "indicateurs_oppchovec",
    "Indicateurs OppChoVec — Olmo":                            "indicateurs_oppchovec",
    "Indicateurs OppChoVec — Saint-Florent":                   "indicateurs_oppchovec",
    "Indicateurs OppChoVec — Polveroso":                       "indicateurs_oppchovec",
    "Indicateurs OppChoVec":                                   "indicateurs_oppchovec",
    "Indicateurs OppChoVec (données territoriales objectives)": "indicateurs_oppchovec",
    "Indicateurs OppChoVec (Piano)":                           "indicateurs_oppchovec",
    "Indicateurs OppChoVec (commune de Piano)":                "indicateurs_oppchovec",
    "Indicateurs OppChoVec (scores et rangs pour Piano)":      "indicateurs_oppchovec",
    "Indicateurs OppChoVec (scores globaux et sous-scores pour Piano)": "indicateurs_oppchovec",
    "Exemple de la commune de Piano":                          "indicateurs_oppchovec",
    # enquete_citoyenne
    "Données enquête citoyenne — Ajaccio":                             "enquete_citoyenne",
    "Verbatims bruts enquête citoyenne — Ajaccio":                     "enquete_citoyenne",
    "Verbatims bruts enquête citoyenne — Corte":                       "enquete_citoyenne",
    "Verbatims bruts enquête citoyenne — Balogna":                     "enquete_citoyenne",
    "Synthèse scores enquête citoyenne — vue : enquete_commune":       "enquete_citoyenne",
    "Synthèse scores enquête citoyenne — vue : enquete_commune — 51 répondants": "enquete_citoyenne",
    "Synthèse scores enquête citoyenne — vue : enquete_age_range*commune": "enquete_citoyenne",
    "Synthèse scores enquête citoyenne — vue : enquete_global — 246 répondants": "enquete_citoyenne",
    "Synthèse scores enquête citoyenne — vue : enquete_global":        "enquete_citoyenne",
    "Synthèse scores enquête citoyenne — vue : enquete_profession*commune": "enquete_citoyenne",
    "Synthèse verbatims enquête citoyenne — vue : commune":            "enquete_citoyenne",
    "Synthèse verbatims enquête citoyenne — vue : commune — 129 répondants": "enquete_citoyenne",
    "Synthèse verbatims enquête citoyenne — vue : age_range*commune":  "enquete_citoyenne",
    "Synthèse verbatims enquête citoyenne — vue : dimension — 27 répondants": "enquete_citoyenne",
    "Synthèse verbatims enquête citoyenne — vue : dimension*commune":  "enquete_citoyenne",
    "Synthèse verbatims enquête citoyenne — vue : profession*commune": "enquete_citoyenne",
    "Synthèse verbatims enquête citoyenne — vue : profession*commune — 36 répondants": "enquete_citoyenne",
    "Enquête citoyenne (Corse entière, N=246)":                        "enquete_citoyenne",
    "Enquête citoyenne (données Corse entière, N=246)":                "enquete_citoyenne",
    "Enquête citoyenne (données Corse entière, N=246, perceptions sur les services de proximité)": "enquete_citoyenne",
    # entretiens_qualitatifs
    "Entretiens semi-directifs":                                       "entretiens_qualitatifs",
    "Synthèse entretiens qualitatifs — vue : entretien_commune — Lozzi": "entretiens_qualitatifs",
    "Synthèse entretiens qualitatifs — vue : entretien_commune — Lozzi —": "entretiens_qualitatifs",
    "Synthèse entretiens qualitatifs — vue : entretien_commune — Lozzi — 253 répondants": "entretiens_qualitatifs",
    "Synthèse entretiens qualitatifs — vue : entretien_commune — Grossetto Prugna": "entretiens_qualitatifs",
    "Synthèse entretiens qualitatifs — vue : entretien_commune — Grossetto Prugna — 31 répondants": "entretiens_qualitatifs",
    # equipements
    "Équipements et services communaux — Ajaccio":  "equipements",
    "Équipements et services communaux — Corscia":  "equipements",
    # contextuel_methodo
    "methodology — vue : enquete_methodology":      "contextuel_methodo",
    # AUCUNE_SOURCE — déclarations explicites du système
    "Aucune source mobilisable (données absentes)":     "AUCUNE_SOURCE",
    "Aucune source pertinente dans le contexte":        "AUCUNE_SOURCE",
    "Aucune source pertinente (donnée manquante)":      "AUCUNE_SOURCE",
    "Aucune source pertinente (données manquantes)":    "AUCUNE_SOURCE",
    "Aucune source mobilisée (données manquantes)":     "AUCUNE_SOURCE",
}


def _to_coarse_set(sm: List[Dict]) -> Tuple[frozenset, str, set]:
    """
    Classe les sources_mobilisees d'un run en type coarse.
    Retourne (coarse_set, etat, orphans).
    - etat : "ok" | "aucune_source" | "bloc_manquant"
    - orphans : labels absents de COARSE_MAP (à signaler)
    """
    if not sm:
        return frozenset(), "bloc_manquant", set()

    raw_types = [t for e in sm for t in e.get("types", [])]
    if not raw_types:
        return frozenset(), "bloc_manquant", set()

    orphans = {t for t in raw_types if t not in COARSE_MAP}
    mapped = {COARSE_MAP[t] for t in raw_types if t in COARSE_MAP}

    real = mapped - {"AUCUNE_SOURCE"}
    if real:
        return frozenset(real), "ok", orphans
    if "AUCUNE_SOURCE" in mapped:
        return frozenset({"AUCUNE_SOURCE"}), "aucune_source", orphans
    # Seulement des orphelins
    return frozenset(), "bloc_manquant", orphans


# Fallback explicite pour Q62 si absente de l'Excel
_Q62_FALLBACK = {
    "excel_row": 62,
    "section": "Robustesse sémantique",
    "subsection": "Reformulations indirectes",
    "question": "Quelle commune arrive en tête du classement pour OppChoVec ?",
}


# ============================================================
# SECTION 1 — Pre-checks bloquants
# ============================================================

def precheck_packages() -> None:
    missing = []
    try:
        import sacrebleu  # noqa: F401
    except ImportError:
        missing.append("sacrebleu")
    try:
        import rouge_score  # noqa: F401
    except ImportError:
        missing.append("rouge-score")
    if missing:
        sys.exit(
            f"[BLOQUANT] Packages manquants : {', '.join(missing)}\n"
            f"  → pip install {' '.join(missing)}"
        )
    print("[OK] sacrebleu + rouge_score installés")


def precheck_api() -> None:
    try:
        r = requests.get(f"{API_BASE}/", timeout=5)
        r.raise_for_status()
        print(f"[OK] Serveur RAG accessible sur {API_BASE}")
    except Exception as e:
        sys.exit(f"[BLOQUANT] Serveur RAG inaccessible sur {API_BASE}: {e}")


def precheck_embed() -> object:
    try:
        print(f"[...] Chargement BGE-M3 ({EMBED_MODEL_NAME})...")
        model = SentenceTransformer(EMBED_MODEL_NAME)
        _ = model.encode("test")
        print("[OK] BGE-M3 chargé")
        return model
    except Exception as e:
        sys.exit(f"[BLOQUANT] BGE-M3 non chargeable : {e}")


def precheck_source_stability(question: str) -> Dict:
    """2 appels rapides, compare les source_key sets. Retourne rapport."""
    print(f"[...] Vérification stabilité sources (2 appels sur Q facile)...")

    def get_keys_ordered(resp: dict) -> List[str]:
        sources = resp.get("sources", [])
        return [
            f"{s.get('metadata', {}).get('source_type', '')}|"
            f"{s.get('metadata', {}).get('commune', '')}|"
            f"{s.get('metadata', {}).get('view_name', '')}"
            for s in sources
        ]

    r1 = _call_api_raw(question)
    time.sleep(2)
    r2 = _call_api_raw(question)

    k1 = get_keys_ordered(r1)
    k2 = get_keys_ordered(r2)
    s1, s2 = set(k1), set(k2)
    stable = s1 == s2

    report = {
        "stable": stable,
        "n_sources_run1": len(k1),
        "n_sources_run2": len(k2),
        "only_in_run1": sorted(s1 - s2),
        "only_in_run2": sorted(s2 - s1),
        "keys_run1": k1[:10],
        "keys_run2": k2[:10],
    }

    flag = "[OK]" if stable else "[WARN]"
    print(f"  {flag} Stabilité sources : {'STABLE' if stable else 'INSTABLE'}")
    print(f"       Run1: {len(k1)} sources | Run2: {len(k2)} sources")
    if not stable:
        print(f"       Uniquement run1 : {report['only_in_run1']}")
        print(f"       Uniquement run2 : {report['only_in_run2']}")
        print("  [!] ATTENTION : les metriques Jaccard/overlap@5 peuvent etre affectees")

    return report


# ============================================================
# SECTION 2 — Chargement des questions
# ============================================================

def load_questions() -> Dict[int, Dict]:
    """Charge depuis Excel, filtre sur TARGET_ROWS. Fallback Q62 si absent."""
    try:
        wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
        ws = wb.active
        questions: Dict[int, Dict] = {}
        for r in range(2, ws.max_row + 1):
            excel_row = r - 1
            if excel_row not in TARGET_ROWS:
                continue
            q_text = ws.cell(r, 3).value or ""
            if not q_text.strip():
                continue
            questions[excel_row] = {
                "excel_row": excel_row,
                "section": ws.cell(r, 1).value or "",
                "subsection": ws.cell(r, 2).value or "",
                "question": q_text.strip(),
            }
        wb.close()
    except FileNotFoundError:
        print(f"[WARN] Excel non trouvé : {EXCEL_PATH} — utilisation des fallbacks")
        questions = {}

    # Fallback Q62 si absent (question de validation facile)
    if 62 in TARGET_ROWS and 62 not in questions:
        print(f"[INFO] Q62 absente de l'Excel → fallback explicite")
        questions[62] = _Q62_FALLBACK

    missing = [r for r in TARGET_ROWS if r not in questions]
    if missing:
        print(f"[WARN] Lignes introuvables dans l'Excel : {missing}")

    print(f"[OK] {len(questions)} question(s) chargée(s) : rows {sorted(questions)}")
    return questions


# ============================================================
# SECTION 3 — Boucle principale + sauvegarde incrémentale
# ============================================================

def _call_api_raw(question: str) -> dict:
    """Un appel brut, sans retry."""
    payload: dict = {"question": question, "rag_version": RAG_VERSION, "k": K}
    if TEMP_OVERRIDE is not None:
        payload["temperature_override"] = TEMP_OVERRIDE
    r = requests.post(f"{API_BASE}/api/query", json=payload, timeout=API_TIMEOUT)
    r.raise_for_status()
    return r.json()


def call_api(question: str) -> dict:
    """Appel avec retry exponentiel."""
    for attempt in range(API_RETRIES):
        try:
            return _call_api_raw(question)
        except Exception as e:
            wait = 10 * (attempt + 1)
            print(f"  [retry {attempt + 1}/{API_RETRIES}] {e} — attente {wait}s")
            if attempt < API_RETRIES - 1:
                time.sleep(wait)
    return {"error": "echec après retries", "answer": "", "sources": []}


def parse_sources_mobilisees_from_text(answer: str) -> List[Dict]:
    """Parse le bloc ===SOURCES_MOBILISEES=== depuis le texte (fallback si champ absent)."""
    m = re.search(r'===SOURCES_MOBILISEES===(.*?)===FIN_SOURCES===', answer, re.DOTALL)
    if not m:
        return []
    result = []
    for line in m.group(1).strip().split('\n'):
        mm = re.match(r'\*\*SQ(\d+)\*\*\s*:\s*(.*)', line.strip())
        if mm:
            types = [t.strip() for t in mm.group(2).split(';') if t.strip()]
            result.append({"sq": int(mm.group(1)), "types": types})
    return result


def _make_source_key(s: dict) -> str:
    meta = s.get("metadata", {})
    return (
        f"{meta.get('source_type', '')}|"
        f"{meta.get('commune', '')}|"
        f"{meta.get('view_name', '')}"
    )


def run_question(q_meta: dict, run_i: int, ts: str) -> dict:
    """Exécute un run, sauvegarde JSON incrémental, retourne le dict de résultats."""
    question = q_meta["question"]
    excel_row = q_meta["excel_row"]

    t0 = time.time()
    resp = call_api(question)
    elapsed = round(time.time() - t0, 1)

    answer = resp.get("answer", "")
    sources = resp.get("sources", [])

    # sources_mobilisees : champ structuré si dispo, sinon parse depuis texte
    sources_mob = resp.get("sources_mobilisees")
    if not sources_mob:
        sources_mob = parse_sources_mobilisees_from_text(answer)

    source_keys = [_make_source_key(s) for s in sources]

    sources_detail = []
    for rank_i, s in enumerate(sources):
        meta = s.get("metadata", {})
        sources_detail.append({
            "source_type":      meta.get("type") or meta.get("source_type", ""),
            "commune":          meta.get("commune") or meta.get("dim1_value", ""),
            "view_name":        meta.get("view_name") or meta.get("view", ""),
            "sub_question_idx": meta.get("sub_question_idx"),
            "extrait":          (s.get("content") or "")[:500],
            "score":            s.get("score"),
            "rank":             meta.get("rank") if meta.get("rank") is not None else rank_i + 1,
        })

    result = {
        "excel_row":          excel_row,
        "run":                run_i,
        "question":           question,
        "section":            q_meta.get("section", ""),
        "subsection":         q_meta.get("subsection", ""),
        "answer":             answer,
        "sources_mobilisees": sources_mob,
        "source_keys":        source_keys,
        "sources_detail":     sources_detail,
        "n_sources":          len(sources),
        "sub_questions":      resp.get("sub_questions") or [],
        "elapsed_s":          elapsed,
        "timestamp":          datetime.now().isoformat(),
        "api_error":          resp.get("error"),
    }

    # Sauvegarde incrémentale
    out_path = OUTPUT_DIR / f"run_{ts}_{excel_row}_{run_i:02d}.json"
    out_path.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    return result


def collect_runs(questions: Dict[int, Dict], ts: str,
                 existing: Optional[Dict] = None) -> Dict[int, List[dict]]:
    """Boucle principale. existing = résultats déjà chargés (--resume)."""
    runs: Dict[int, List[dict]] = existing or {row: [] for row in sorted(questions)}

    total = sum(M_RUNS - len(runs.get(row, [])) for row in sorted(questions))
    done = 0

    for row in sorted(questions):
        q_meta = questions[row]
        already = len(runs.get(row, []))
        if already >= M_RUNS:
            print(f"\n[SKIP] Q{row} — {M_RUNS} runs déjà présents")
            continue

        print(f"\n{'='*65}")
        print(f"Q{row} [{q_meta['section']}] {q_meta['question'][:70]}")
        print(f"  → {M_RUNS - already} run(s) restant(s)")

        runs.setdefault(row, [])

        for run_i in range(already + 1, M_RUNS + 1):
            print(f"  [run {run_i}/{M_RUNS}] ", end="", flush=True)
            result = run_question(q_meta, run_i, ts)

            if result.get("api_error"):
                print(f"ERREUR : {result['api_error']}")
            else:
                mob_n = len(result["sources_mobilisees"])
                print(f"OK ({result['elapsed_s']}s) | {result['n_sources']} src | {mob_n} SQ mob.")

            runs[row].append(result)
            done += 1

            if run_i < M_RUNS:
                time.sleep(RUN_DELAY)

        if row != sorted(questions)[-1]:
            print(f"\n  [délai inter-questions {QUESTION_DELAY}s...]")
            time.sleep(QUESTION_DELAY)

    return runs


# ============================================================
# SECTION 4 — Calcul des métriques
# ============================================================

def jaccard(s1: set, s2: set) -> float:
    union = s1 | s2
    return len(s1 & s2) / len(union) if union else 1.0


def overlap_at_k(l1: List[str], l2: List[str], k: int = 5) -> float:
    s1, s2 = set(l1[:k]), set(l2[:k])
    return len(s1 & s2) / k if k else 0.0


def type_set_overlap(sm1: List[Dict], sm2: List[Dict]) -> float:
    t1 = {t for e in sm1 for t in e.get("types", [])}
    t2 = {t for e in sm2 for t in e.get("types", [])}
    return jaccard(t1, t2)


def semantic_sim(answers: List[str], embed_model) -> Tuple[float, float]:
    """Cosinus pairwise leave-one-out. Retourne (mean, std)."""
    valid = [a for a in answers if a and not a.startswith("ERREUR")]
    if len(valid) < 2:
        return 0.0, 0.0
    embs = embed_model.encode([f"query: {a}" for a in valid], show_progress_bar=False)
    sims = []
    for i in range(len(embs)):
        for j in range(i + 1, len(embs)):
            cos = float(np.dot(embs[i], embs[j]) /
                        (np.linalg.norm(embs[i]) * np.linalg.norm(embs[j]) + 1e-9))
            sims.append(cos)
    return float(np.mean(sims)), float(np.std(sims))


def bleu_rouge(answers: List[str]) -> Tuple[float, float]:
    """BLEU-4 + ROUGE-L leave-one-out."""
    from sacrebleu.metrics import BLEU
    from rouge_score import rouge_scorer as rs_mod

    valid = [a for a in answers if a and not a.startswith("ERREUR")]
    if len(valid) < 2:
        return 0.0, 0.0

    bleu_metric = BLEU(effective_order=True)
    scorer = rs_mod.RougeScorer(["rougeL"], use_stemmer=False)
    bleu_scores, rouge_scores = [], []

    for i, hyp in enumerate(valid):
        refs = [valid[j] for j in range(len(valid)) if j != i]
        b = bleu_metric.corpus_score([hyp], [[r] for r in refs]).score / 100
        bleu_scores.append(b)
        rl = float(np.mean([scorer.score(r, hyp)["rougeL"].fmeasure for r in refs]))
        rouge_scores.append(rl)

    return float(np.mean(bleu_scores)), float(np.mean(rouge_scores))


def compute_metrics(row_runs: List[dict], embed_model) -> Dict:
    """Calcule toutes les métriques pour les M runs d'une question."""
    answers      = [r["answer"] for r in row_runs]
    source_lists = [r["source_keys"] for r in row_runs]
    mob_lists    = [r["sources_mobilisees"] for r in row_runs]

    # Sémantique
    sem_mean, sem_std = semantic_sim(answers, embed_model)

    # Lexical
    bleu_mean, rouge_mean = bleu_rouge(answers)

    # Récupération : Jaccard + overlap@5 pairwise
    jacc_vals, ovlp_vals = [], []
    for i in range(len(source_lists)):
        for j in range(i + 1, len(source_lists)):
            jacc_vals.append(jaccard(set(source_lists[i]), set(source_lists[j])))
            ovlp_vals.append(overlap_at_k(source_lists[i], source_lists[j], k=5))

    # Attribution fine (ancienne métrique, conservée pour historique)
    mob_vals = []
    for i in range(len(mob_lists)):
        for j in range(i + 1, len(mob_lists)):
            mob_vals.append(type_set_overlap(mob_lists[i], mob_lists[j]))

    # Attribution coarse — TypeOvlp recalculé sur types coarse
    coarse_sets = []
    all_orphans: set = set()
    n_bloc_manquant = 0
    for sm in mob_lists:
        cset, etat, orphans = _to_coarse_set(sm)
        coarse_sets.append((cset, etat))
        all_orphans.update(orphans)
        if etat == "bloc_manquant":
            n_bloc_manquant += 1

    n_valid = len(mob_lists) - n_bloc_manquant
    coarse_pairs = []
    for i in range(len(coarse_sets)):
        for j in range(i + 1, len(coarse_sets)):
            _, ei = coarse_sets[i]
            _, ej = coarse_sets[j]
            if ei == "bloc_manquant" or ej == "bloc_manquant":
                continue
            si, sj = coarse_sets[i][0], coarse_sets[j][0]
            coarse_pairs.append(jaccard(si, sj))

    if coarse_pairs:
        typeovlp_coarse = round(float(np.mean(coarse_pairs)), 4)
    else:
        typeovlp_coarse = None  # tous bloc_manquant ou une seule run valide

    return {
        "sem_cos_mean":         round(sem_mean, 4),
        "sem_cos_std":          round(sem_std, 4),
        "bleu_mean":            round(bleu_mean, 4),
        "rouge_l_mean":         round(rouge_mean, 4),
        "jaccard_mean":         round(float(np.mean(jacc_vals)) if jacc_vals else 0.0, 4),
        "overlap5_mean":        round(float(np.mean(ovlp_vals)) if ovlp_vals else 0.0, 4),
        "typeovlp_mean":        round(float(np.mean(mob_vals)) if mob_vals else 0.0, 4),
        "typeovlp_coarse_mean": typeovlp_coarse,
        "n_valid_runs":         n_valid,
        "n_bloc_manquant":      n_bloc_manquant,
        "orphans":              sorted(all_orphans),
    }


# ============================================================
# SECTION 5 — Rapport HTML
# ============================================================

_CSS = """
body{font-family:system-ui,sans-serif;margin:2rem;background:#f8f9fa;color:#212529}
h1{color:#2c3e50}h2{color:#34495e;border-bottom:2px solid #dee2e6;padding-bottom:.4rem}
table{border-collapse:collapse;width:100%;margin:1rem 0}
th{background:#343a40;color:#fff;padding:.5rem .7rem;text-align:left}
td{padding:.4rem .7rem;border-bottom:1px solid #dee2e6}
tr:hover td{background:#e9ecef}
.good{color:#28a745;font-weight:bold}.warn{color:#fd7e14;font-weight:bold}
.bad{color:#dc3545;font-weight:bold}
details{margin:.5rem 0}summary{cursor:pointer;font-weight:bold;padding:.3rem}
summary:hover{text-decoration:underline}
pre{background:#fff;border:1px solid #dee2e6;padding:1rem;border-radius:4px;
    white-space:pre-wrap;word-break:break-word;font-size:.85rem;max-height:400px;overflow-y:auto}
.mob-table td{font-size:.82rem}.run-header{background:#e9ecef;font-weight:bold}
"""


def _score_class(val: float, metric: str) -> str:
    thresholds = {
        "sem_cos_mean":         (0.95, 0.85),
        "bleu_mean":            (0.30, 0.15),
        "rouge_l_mean":         (0.50, 0.30),
        "jaccard_mean":         (0.80, 0.50),
        "overlap5_mean":        (0.80, 0.50),
        "typeovlp_mean":        (0.90, 0.70),
        "typeovlp_coarse_mean": (0.80, 0.60),
    }
    hi, lo = thresholds.get(metric, (1.0, 0.5))
    if val >= hi:
        return "good"
    if val >= lo:
        return "warn"
    return "bad"


def _mob_table(row_runs: List[dict]) -> str:
    """Tableau run × types mobilisés."""
    all_types: set = set()
    for r in row_runs:
        for e in r.get("sources_mobilisees", []):
            all_types.update(e.get("types", []))
    if not all_types:
        return "<em>Aucune donnée sources_mobilisées</em>"

    cols = sorted(all_types)
    header = "<tr><th>Run</th>" + "".join(f"<th>{c[:30]}</th>" for c in cols) + "</tr>"
    rows_html = []
    for r in row_runs:
        flat = {t for e in r.get("sources_mobilisees", []) for t in e.get("types", [])}
        cells = "".join(
            f"<td>{'✓' if c in flat else ''}</td>" for c in cols
        )
        rows_html.append(f"<tr><td>Run {r['run']}</td>{cells}</tr>")
    return f"<table class='mob-table'>{header}{''.join(rows_html)}</table>"


def _mob_table_coarse(row_runs: List[dict]) -> str:
    """Tableau run × TYPES COARSE (avec état AUCUNE_SOURCE / BLOC_MANQUANT)."""
    _COARSE_ORDER = ["indicateurs_oppchovec", "enquete_citoyenne",
                     "entretiens_qualitatifs", "equipements",
                     "contextuel_methodo", "AUCUNE_SOURCE"]
    _COARSE_LABEL = {
        "indicateurs_oppchovec":  "OppChoVec",
        "enquete_citoyenne":      "Enquête cit.",
        "entretiens_qualitatifs": "Entretiens",
        "equipements":            "Équipements",
        "contextuel_methodo":     "Méthodo",
        "AUCUNE_SOURCE":          "Aucune source",
    }

    all_coarse: set = set()
    run_info = []
    for r in row_runs:
        cset, etat, orphs = _to_coarse_set(r.get("sources_mobilisees", []))
        run_info.append((r["run"], cset, etat, orphs))
        all_coarse.update(cset)

    cols = [c for c in _COARSE_ORDER if c in all_coarse]
    if not cols and all(ei == "bloc_manquant" for _, _, ei, _ in run_info):
        return "<em style='color:#dc3545'>Tous les runs : BLOC_MANQUANT (pas de section sources capturée)</em>"
    if not cols:
        cols = list(all_coarse)

    header = "<tr><th>Run</th><th>État</th>" + "".join(
        f"<th>{_COARSE_LABEL.get(c, c)}</th>" for c in cols
    ) + "</tr>"

    rows_html = []
    for run_n, cset, etat, orphs in run_info:
        if etat == "bloc_manquant":
            etat_cell = "<td style='color:#dc3545'>BLOC_MANQUANT</td>"
            cells = "".join(f"<td>—</td>" for _ in cols)
        elif etat == "aucune_source":
            etat_cell = "<td style='color:#fd7e14'>AUCUNE_SOURCE</td>"
            cells = "".join(f"<td>{'✗' if c == 'AUCUNE_SOURCE' else ''}</td>" for c in cols)
        else:
            etat_cell = "<td style='color:#28a745'>ok</td>"
            cells = "".join(f"<td>{'✓' if c in cset else ''}</td>" for c in cols)
        orphs_cell = (f" <small style='color:#c0392b'>[orphelins: {'; '.join(orphs)}]</small>" if orphs else "")
        rows_html.append(f"<tr><td>Run {run_n}{orphs_cell}</td>{etat_cell}{cells}</tr>")

    return f"<table class='mob-table'>{header}{''.join(rows_html)}</table>"


def build_html(all_metrics: Dict[int, Dict], all_runs: Dict[int, List[dict]],
               stability_report: Dict, ts: str) -> str:
    dim_keys   = ["sem_cos_mean", "sem_cos_std", "bleu_mean", "rouge_l_mean",
                  "jaccard_mean", "overlap5_mean", "typeovlp_coarse_mean"]
    dim_labels = ["Sem.cos µ", "Sem.cos σ", "BLEU", "ROUGE-L",
                  "Jaccard", "Ovlp@5", "TypeOvlp (coarse)"]

    # Tableau récapitulatif
    header_cells = "".join(f"<th>{l}</th>" for l in dim_labels) + "<th>Runs val.</th>"
    recap_rows = ""
    all_vals: Dict[str, List[float]] = {k: [] for k in dim_keys if k != "sem_cos_std" and k != "typeovlp_coarse_mean"}
    all_vals["typeovlp_coarse_mean"] = []  # seulement les non-None

    for row in sorted(all_metrics):
        m   = all_metrics[row]
        q   = (all_runs[row][0]["question"] if all_runs[row] else "")[:60]
        sec = (all_runs[row][0].get("section", "") if all_runs[row] else "")[:25]
        cells = ""
        for k in dim_keys:
            v = m.get(k)
            if k == "sem_cos_std":
                cells += f"<td>{v:.4f}</td>" if v is not None else "<td>—</td>"
            elif k == "typeovlp_coarse_mean":
                if v is None:
                    cells += "<td><em style='color:#6c757d'>n/a (bloc manquant)</em></td>"
                else:
                    all_vals[k].append(v)
                    cells += f"<td class='{_score_class(v, k)}'>{v:.4f}</td>"
            else:
                v = v or 0.0
                all_vals[k].append(v)
                cells += f"<td class='{_score_class(v, k)}'>{v:.4f}</td>"
        n_val = m.get("n_valid_runs", len(all_runs.get(row, [])))
        n_tot = len(all_runs.get(row, []))
        val_color = "" if n_val == n_tot else " style='color:#fd7e14;font-weight:bold'"
        cells += f"<td{val_color}>{n_val}/{n_tot}</td>"
        recap_rows += f"<tr><td>Q{row}</td><td title='{q}'>{sec}<br><small>{q[:50]}…</small></td>{cells}</tr>"

    # Ligne globale
    global_cells = ""
    for k in dim_keys:
        if k == "sem_cos_std":
            global_cells += "<td>—</td>"
        elif k == "typeovlp_coarse_mean":
            vals = all_vals[k]
            if vals:
                v = float(np.mean(vals))
                global_cells += f"<td class='{_score_class(v, k)}'><strong>{v:.4f}</strong></td>"
            else:
                global_cells += "<td><em>n/a</em></td>"
        else:
            v = float(np.mean(all_vals[k])) if all_vals[k] else 0.0
            global_cells += f"<td class='{_score_class(v, k)}'><strong>{v:.4f}</strong></td>"
    n_val_total = sum(m.get("n_valid_runs", 0) for m in all_metrics.values())
    n_tot_total = sum(len(runs) for runs in all_runs.values())
    global_cells += f"<td>{n_val_total}/{n_tot_total}</td>"
    recap_rows += f"<tr class='run-header'><td colspan='2'>GLOBAL ({len(all_metrics)} questions)</td>{global_cells}</tr>"

    # Sections par question
    sections_html = ""
    for row in sorted(all_runs):
        runs = all_runs[row]
        if not runs:
            continue
        q_text = runs[0]["question"]
        m = all_metrics.get(row, {})

        tc = m.get("typeovlp_coarse_mean")
        tc_str = f"{tc:.3f}" if tc is not None else "n/a"
        metric_summary = (
            f"Sem.cos µ: {m.get('sem_cos_mean',0):.3f} | "
            f"Sem.cos σ: {m.get('sem_cos_std',0):.3f} | "
            f"BLEU: {m.get('bleu_mean',0):.3f} | "
            f"ROUGE-L: {m.get('rouge_l_mean',0):.3f} | "
            f"Jaccard: {m.get('jaccard_mean',0):.3f} | "
            f"Ovlp@5: {m.get('overlap5_mean',0):.3f} | "
            f"TypeOvlp(coarse): {tc_str} "
            f"[{m.get('n_valid_runs',0)}/{len(runs)} runs valides"
            + (f", orphelins: {m.get('orphans')}" if m.get('orphans') else "")
            + "]"
        )

        # Tableau sources_mobilisées (coarse en premier, fine en détail)
        mob_html = _mob_table_coarse(runs)

        # Réponses individuelles
        answers_html = ""
        for r in runs:
            ans = r["answer"] or "(vide)"
            n_src = r["n_sources"]
            err = f" <em style='color:red'>[ERREUR API]</em>" if r.get("api_error") else ""
            answers_html += (
                f"<details><summary>Run {r['run']} — {r['elapsed_s']}s "
                f"| {n_src} sources{err}</summary>"
                f"<pre>{_escape(ans)}</pre></details>"
            )

        mob_fine_html = _mob_table(runs)
        sections_html += f"""
<h2>Q{row} — {_escape(q_text[:80])}</h2>
<p><em>[{_escape(runs[0].get('section',''))} / {_escape(runs[0].get('subsection',''))}]</em></p>
<p><code>{metric_summary}</code></p>
<details open><summary>Types coarse mobilisés par run</summary>
{mob_html}
</details>
<details><summary>Types fins mobilisés par run (détail brut)</summary>
{mob_fine_html}
</details>
<details><summary>Réponses complètes ({len(runs)} runs)</summary>
{answers_html}
</details>
"""

    # Rapport stabilité sources
    stab_color = "green" if stability_report.get("stable") else "orange"
    stab_label = "STABLE" if stability_report.get("stable") else "INSTABLE"
    stab_html = (
        f"<p>IDs sources entre 2 appels : <strong style='color:{stab_color}'>{stab_label}</strong> "
        f"(run1: {stability_report.get('n_sources_run1','?')} src | "
        f"run2: {stability_report.get('n_sources_run2','?')} src)</p>"
    )
    if not stability_report.get("stable"):
        stab_html += (
            f"<p>Uniquement run1 : {stability_report.get('only_in_run1','')}<br>"
            f"Uniquement run2 : {stability_report.get('only_in_run2','')}</p>"
        )

    return f"""<!DOCTYPE html>
<html lang="fr">
<head>
<meta charset="utf-8">
<title>Stability Report — {ts}</title>
<style>{_CSS}</style>
</head>
<body>
<h1>Harnais de stabilité RAG</h1>
<p><strong>Date :</strong> {ts} &nbsp;|&nbsp;
   <strong>Version :</strong> {RAG_VERSION} &nbsp;|&nbsp;
   <strong>k :</strong> {K} &nbsp;|&nbsp;
   <strong>M runs :</strong> {M_RUNS} &nbsp;|&nbsp;
   <strong>Température :</strong> {"<span style='color:#c0392b;font-weight:bold'>T=0 (greedy/déterministe)</span>" if TEMP_OVERRIDE == 0.0 else (f"T={TEMP_OVERRIDE}" if TEMP_OVERRIDE is not None else "Production (naturelle)")} &nbsp;|&nbsp;
   <strong>Questions :</strong> {sorted(all_metrics.keys())}</p>

<h2>Stabilité des IDs sources (pre-check)</h2>
{stab_html}

<h2>Tableau récapitulatif</h2>
<table>
<tr><th>Q</th><th>Question</th>{header_cells}</tr>
{recap_rows}
</table>
<p><small>Légende : <span class='good'>■ ≥ seuil haut</span>
<span class='warn'>■ entre seuils</span>
<span class='bad'>■ &lt; seuil bas</span></small></p>

{sections_html}

<hr>
<p><small>JSON bruts : <code>{OUTPUT_DIR.resolve()}</code></small></p>
</body>
</html>"""


def _escape(s: str) -> str:
    return s.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")


# ============================================================
# Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--yes", action="store_true",
                        help="Lancer sans confirmation interactive")
    parser.add_argument("--resume", metavar="JSON",
                        help="Reprendre depuis un fichier runs_*.json existant")
    args = parser.parse_args()

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    # ── Pre-checks ─────────────────────────────────────────
    print("\n" + "=" * 65)
    print("PRE-CHECKS")
    print("=" * 65)
    precheck_packages()
    precheck_api()
    embed_model = precheck_embed()

    # ── Chargement questions ────────────────────────────────
    print("\n" + "=" * 65)
    print("CHARGEMENT DES QUESTIONS")
    print("=" * 65)
    questions = load_questions()
    if not questions:
        sys.exit("[BLOQUANT] Aucune question chargée.")

    # ── Stabilité sources (sur la première question disponible) ─
    first_q = questions[sorted(questions)[0]]["question"]
    stability_report = precheck_source_stability(first_q)

    # ── Résumé + confirmation ───────────────────────────────
    n_total = sum(M_RUNS for row in questions if row in TARGET_ROWS)
    print(f"\n{'='*65}")
    print(f"RÉSUMÉ VALIDATION")
    print(f"{'='*65}")
    print(f"  Questions : {sorted(questions.keys())}")
    print(f"  Runs/question : {M_RUNS}")
    print(f"  Total appels API : {n_total}")
    temp_label = f"T={TEMP_OVERRIDE} (greedy)" if TEMP_OVERRIDE == 0.0 else (f"T={TEMP_OVERRIDE}" if TEMP_OVERRIDE is not None else "production (naturelle)")
    print(f"  Version RAG : {RAG_VERSION} | k={K} | température : {temp_label}")
    print(f"  Sorties : {OUTPUT_DIR.resolve()}")
    print()

    if not args.yes:
        ans = input("Lancer les appels ? [o/N] ").strip().lower()
        if ans not in ("o", "oui", "y", "yes"):
            print("Annulé.")
            return

    # ── Reprise depuis checkpoint ───────────────────────────
    existing = None
    if args.resume:
        resume_path = Path(args.resume)
        if resume_path.exists():
            existing = json.loads(resume_path.read_text(encoding="utf-8"))
            existing = {int(k): v for k, v in existing.items()}
            print(f"[INFO] Reprise depuis {resume_path}")
        else:
            print(f"[WARN] Fichier --resume introuvable : {resume_path}")

    # ── Runs ────────────────────────────────────────────────
    print(f"\n{'='*65}")
    print(f"RUNS ({n_total} appels)")
    print(f"{'='*65}")
    all_runs = collect_runs(questions, ts, existing=existing)

    # Sauvegarde consolidée des runs bruts
    runs_path = OUTPUT_DIR / f"runs_{ts}.json"
    runs_path.write_text(
        json.dumps({str(k): v for k, v in all_runs.items()}, ensure_ascii=False, indent=2),
        encoding="utf-8"
    )
    print(f"\n[OK] Runs sauvegardés → {runs_path}")

    # ── Métriques ───────────────────────────────────────────
    print(f"\n{'='*65}")
    print("CALCUL DES MÉTRIQUES")
    print("=" * 65)
    all_metrics: Dict[int, Dict] = {}
    for row, runs in all_runs.items():
        if not runs:
            continue
        print(f"  Q{row} ... ", end="", flush=True)
        m = compute_metrics(runs, embed_model)
        all_metrics[row] = m
        print(f"sem={m['sem_cos_mean']:.3f}±{m['sem_cos_std']:.3f} | "
              f"BLEU={m['bleu_mean']:.3f} | ROUGE={m['rouge_l_mean']:.3f} | "
              f"Jacc={m['jaccard_mean']:.3f} | TypeOvlp={m['typeovlp_mean']:.3f}")

    # ── Rapport HTML ─────────────────────────────────────────
    print(f"\n{'='*65}")
    print("GÉNÉRATION RAPPORT HTML")
    print("=" * 65)
    html = build_html(all_metrics, all_runs, stability_report, ts)
    html_path = OUTPUT_DIR / f"stability_report_{ts}.html"
    html_path.write_text(html, encoding="utf-8")
    print(f"[OK] Rapport HTML → {html_path}")

    # ── Export PDF via Playwright / Chromium ──────────────────
    pdf_path = html_path.with_suffix(".pdf")
    try:
        from playwright.sync_api import sync_playwright
        print("[...] Export PDF (Playwright Chromium)...")
        with sync_playwright() as _pw:
            _browser = _pw.chromium.launch()
            _page = _browser.new_page()
            _page.goto(html_path.resolve().as_uri())
            _page.wait_for_load_state("networkidle")
            _page.pdf(
                path=str(pdf_path),
                format="A4",
                print_background=True,
                margin={"top": "15mm", "bottom": "15mm", "left": "12mm", "right": "12mm"},
            )
            _browser.close()
        print(f"[OK] Rapport PDF  → {pdf_path}")
    except Exception as _e:
        print(f"[WARN] PDF non généré : {_e}")
        pdf_path = None

    print(f"\n{'='*65}")
    print("TERMINÉ")
    print(f"  JSON runs   : {runs_path}")
    print(f"  Rapport HTML: {html_path}")
    if pdf_path and pdf_path.exists():
        print(f"  Rapport PDF : {pdf_path}")
    print("=" * 65)


if __name__ == "__main__":
    main()
